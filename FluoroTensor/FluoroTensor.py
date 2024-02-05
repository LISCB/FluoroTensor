# FluoroTensor Suite - By Max Wills
# Program for analysing data from sequentially illuminated objective TIRF, step detection of intensity time traces,
# and single-molecule tracking
# Project start date: 19/11/2021 @ 15:12
# First functioning version completed: 25/11/2021 @ 18:06
# Copyright (C) 2023: Max F. K. Wills & Ian C. Eperon

# FluoroTensor: identification and tracking of colocalised molecules and their stoichiometries in multi-colour single molecule imaging via deep learning
# Max F.K. Wills, Carlos Bueno Alejo, Nikolas Hundt, Andrew J. Hudson, Ian C. Eperon
# bioRxiv 2023.11.21.567874; doi: https://doi.org/10.1101/2023.11.21.567874

# LICENCE:
# This program is made available under the Creative Commons NonCommercial 4.0 licence.


#      Disclaimer of Warranties and Limitation of Liability.
#
#   a. UNLESS OTHERWISE SEPARATELY UNDERTAKEN BY THE LICENSOR, TO THE
#      EXTENT POSSIBLE, THE LICENSOR OFFERS THE LICENSED MATERIAL AS-IS
#      AND AS-AVAILABLE, AND MAKES NO REPRESENTATIONS OR WARRANTIES OF
#      ANY KIND CONCERNING THE LICENSED MATERIAL, WHETHER EXPRESS,
#      IMPLIED, STATUTORY, OR OTHER. THIS INCLUDES, WITHOUT LIMITATION,
#      WARRANTIES OF TITLE, MERCHANTABILITY, FITNESS FOR A PARTICULAR
#      PURPOSE, NON-INFRINGEMENT, ABSENCE OF LATENT OR OTHER DEFECTS,
#      ACCURACY, OR THE PRESENCE OR ABSENCE OF ERRORS, WHETHER OR NOT
#      KNOWN OR DISCOVERABLE. WHERE DISCLAIMERS OF WARRANTIES ARE NOT
#      ALLOWED IN FULL OR IN PART, THIS DISCLAIMER MAY NOT APPLY TO YOU.
#
#   b. TO THE EXTENT POSSIBLE, IN NO EVENT WILL THE LICENSOR BE LIABLE
#      TO YOU ON ANY LEGAL THEORY (INCLUDING, WITHOUT LIMITATION,
#      NEGLIGENCE) OR OTHERWISE FOR ANY DIRECT, SPECIAL, INDIRECT,
#      INCIDENTAL, CONSEQUENTIAL, PUNITIVE, EXEMPLARY, OR OTHER LOSSES,
#      COSTS, EXPENSES, OR DAMAGES ARISING OUT OF THIS PUBLIC LICENSE OR
#      USE OF THE LICENSED MATERIAL, EVEN IF THE LICENSOR HAS BEEN
#      ADVISED OF THE POSSIBILITY OF SUCH LOSSES, COSTS, EXPENSES, OR
#      DAMAGES. WHERE A LIMITATION OF LIABILITY IS NOT ALLOWED IN FULL OR
#      IN PART, THIS LIMITATION MAY NOT APPLY TO YOU.
#
#   c. The disclaimer of warranties and limitation of liability provided
#      above shall be interpreted in a manner that, to the extent
#      possible, most closely approximates an absolute disclaimer and
#      waiver of all liability.


version = "6.6.8r"
update_notes = "Added LSTM Recurrent neural networks. TrackXpress extension package integration. " \
               "Spot detection now only relies on spot criteria and is independant of brightness, " \
               "using cumulative histogram thresholding. Histogram bin-counts are optimized resulting " \
               "in much faster field of view rendering. Fixed photon plots export and saving errors in tracking. " \
               "Users can now export enhanced TIF movie from TrackXpress add-in. Users can curate and index a folder " \
               "so automation runs exclude bad tif stacks. Removed Single marker option in favour of selection. " \
               "Marker spots and overlay on secondary channels is now switched on by default. Added support for " \
               "calibration data to be exported to excel optimiser directly. Pixel size set to 160nm. Fixed " \
               "magnitude and direction of chromatic shift transform with respect to wavelength of marker. " \
               "Additionally supports high pass filter enhancement and region selection for colocalization. " \
               "Fixed way in which diffusion coefficients are calculated. Improved spot circle colour contrast. " \
               "Added option for auto-optimization of chromatic shift correction parameters per field of view. " \
               "Calibration now supports 'oligo mode' in addition to the default mode for fluorescent beads. " \
               "New neural networks: 640nm_300fr_CSH_RCNN8, 561nm_300fr_RCNN3, 488nm_300fr_RCNN3. Added option: " \
               "FRET mode where marker and secondary are displayed simultaneously. Attempted bug fix for tracking crash. " \
               "Fixed export error. Changed high pass filter to top-hat filter. Replaced top-hat filter with " \
               "wavelet transform and added potential support for deconvolution. MSD proportion can be set " \
               "manually per fit to ensure correctly fitted MSD leads to accurate diff. coefficient. " \
               "Additional settings are to be exported into the excel output. Added Statistics view. " \
               "Final neural networks added '100ms_300fr_Protein_high_SNR_model2' (GFP) , '100ms_300fr_Protein_low_SNR_model2' " \
               "(mCherry), '100ms_300fr_Organic_low_SNR_model1' (Cy5 / Alexa / Atto @ 20mW no Gloxy), " \
               "100ms_300fr_Organic_high_SNR_model4 (Cy5 / Alexa / Atto @ 50mW+). Various improvements, " \
               "Stokes shift simulator. Tracking program now has diffusion analysis which automates the " \
               "fitting of MSD plots to calculate the diffusion coefficient and error. Gaussian mixture model " \
               "added for diff coeff. fitting. Fixed bleaching time in export. Fully automate tracking analsys " \
               "with histogrm export and graph customization. Global normalization in TrackXpress now affects " \
               "analsysis not display only. Fixed bug with global normalization and the problem where no tracks " \
               "are kept if some tracked frames contain no spots. Changed Default spot detection parameters " \
               "for tracking and added option to lock stop point and disable MSD autofitting in automated mode. " \
               "Added FRET Analysis mode which fits anticorrelated FRET traces from the persistent list providing " \
               "rate coefficients and the equilibrium constant. Added RMS mode calculation for tracking. Stoichiometries " \
               "of markers will be exported as absolute numbers rather than percentages. Progress bar changed so it can " \
               "moved or minimised to a restore button below region selection buttons. Export selection now filters triple " \
               "colocalized spots by marker step count as well. Fitting tool threshold can now be a float. trcs files now " \
               "store additional data about if neural networks were used and if traces have been fitted. Added sigma/D " \
               "test for checking correct number of components fitted in diffusivity histogram. integrated CLDNN " \
               "normalization. Added Distributions button for steps and filtered by 1 marker step. Plot aSNR distributions."
update_date = "06/01/2024"


import hashlib
print("Loading graphical user interface library...")
import tkinter as tk                                                                                                     # GUI library
from tkinter import ttk
from tkscrolledframe import ScrolledFrame as Sframe
print("Loading Matplotlib...")
import matplotlib.pyplot as plt                                                                                          # plotting library
import matplotlib as mplib
print("Loading numpy...")
import numpy as np                                                                                                       # numerical / arrays library
print("Loading excel interface: openpyxl...")
import openpyxl as opxl                                                                                                  # excel file interface library
print("Loading excel cell styles...")
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
print("Loading menu library...")
import easygui                                                                                                           # GUI box library
print("Loading binary serialisation library...")
import pickle                                                                                                            # serialisation library
print("Loading Matplotlib user interface hooks...")
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)                                  # matplotlib - tkinter interface
import matplotlib.gridspec as gridspec
print("Loading numpy random number generator...")
from numpy.random import default_rng
print("Loading SciPy 1 dimensional linear interpolator...")
from scipy.interpolate import interp1d
import scipy.stats as scipy_stats
print("Loading TIRF analysis library by Maximilian Wills...")
import TIRF_lib as tirf
print("Loading tifffile io lib...")
import tifffile
print("Loading operating system hooks...")
import os
import sys
import traceback
import time
import sklearn
from sklearn import mixture
import asnr_calc as asnr

print("Loading Coefficient of Determination module isolated from sklearn.metrics")
# R squared calculator from sklearn.metrics This has been isolated from sklearn so that FluoroTensor can be successfully compiled without the full sklearn package
# Modified for simple linear regression
from sklearn_r2 import r2_score


# from vosk import Model, KaldiRecognizer
# import pyaudio
# from gtts import gTTS
# from io import BytesIO
# import pygame
# from word2number import w2n
# from number2word import convert
#
#
# sample_freq = 16000
# frame_buffer = 8000
#
# model = Model('vosk-model-en-us-0.22')
# recog = KaldiRecognizer(model, sample_freq)
#
#
# mic = pyaudio.PyAudio()
# audio_stream = mic.open(format=pyaudio.paInt16, channels=1, rate=sample_freq, input=True,
#                         frames_per_buffer=frame_buffer)


"""
Ignore Licence key code
"""

# def check_for_licence():
#     try:
#         with open("flt_licence.ftl", "rb") as check_licence:
#             licence_hash = pickle.load(check_licence)
#         with open("ldata.haf", "rb") as load_comparisons:
#             hash_list = pickle.load(load_comparisons)
#         if licence_hash in hash_list:
#             return True
#         else:
#             return False
#     except:
#         try:
#             os.remove("flt_licence.ftl")
#         except:
#             """ File no longer present """
#         easygui.msgbox(title="Licence Error", msg="FluoroTensor licence is absent or invalid")
#         return False
#
#
# def ask_for_key():
#     code = easygui.enterbox(title="Licence Key", msg="Please enter your 16-digit FluoroTensor licence key")
#     if not code:
#         sys.exit()
#     byte_code = bytes(str(code), 'utf-8')
#     h = hashlib.sha256()
#     h.update(byte_code)
#     licence_hash = h.hexdigest()
#
#     with open("ldata.haf", "rb") as load_comparisons:
#         hash_list = pickle.load(load_comparisons)
#
#     if licence_hash in hash_list:
#         with open("flt_licence.ftl", "wb") as save_licence:
#             pickle.dump(licence_hash, save_licence)
#             easygui.msgbox(title="Software Activated", msg="You have successfully activated your FluoroTensor Licence Key.")
#         return True
#     else:
#         return False
#
#
# # check for licence
#
#
# valid_licence = False
# while not valid_licence:
#     valid_licence = check_for_licence()
#     if not valid_licence:
#         valid_licence = ask_for_key()


print("\nLoading TensorFlow Deep neural network library. This may take some time...")
import tensorflow as tf
print("Initialising...")


unique_ref_list = []

cwd = os.getcwd()
rng = default_rng()


root = tk.Tk()                                                                                                           # create main 'root' window.
root.iconbitmap("icon.ico")
root.title(f"FluoroTensor {version} - By Max Wills (2021-2024)")                                     # set window title.
root["bg"] = "#444444"                                                                                                   # set root window background colour to light grey.
root.geometry("+200+20")
root.update()
col_window = False                                                                                                       # initialise global variable for GUI colour selection window.
advanced_window = False

stats_window = False                                                                                                     # initialise global variable for step count statistics window.
colours = ["#ffaa68", "#7799ff", "#22bb88", "#555555"]                                                                   # create a list of RGB formatted colours for GUI.

sheet = False                                                                                                            # initialise global variable for sheet (excel).
xl_workbook = False                                                                                                      # initialise global variable for workbook (excel).
trace_info = False                                                                                                       # initialise global trace_info list, boolean used for existence checks.
all_traces = []                                                                                                          # initialise global list to store all loaded traces.
all_fits = []                                                                                                            # initialise global list to store all fits loaded from excel.

current_trace = None                                                                                                     # initialise global trace index.
current_fluorophore = None                                                                                               # initialise global info variable for fluorophore in current trace.
current_step_count = None                                                                                                # initialise global info variable for step count of current trace.
current_trace_length = None                                                                                              # initialise global info variable for trace length of current trace.

capture_mouse = False
mouse_xpos = None
waiting_for_fit = False
vertical_position = None
fret_mode = False

# lists in preferences for each flurophore are in order: 640, 561, 488
preferences = {                                                                                                          # create dictionary mapping preferences to their values. with the exception
    "Trace skip number": 10,                                                                                             # of the default directory which has its own variable but is still present in
    "Default directory": "N:/",                                                                                          # the dictionary since this is used to populate the selection list in the GUI.
    "GUI colours": "#555555",
    "Subsampling": False,
    "kernel": "gauss",
    "pre-gauss": False,
    "Subtraction amount": 10,
    "488nm model dir": "100ms_300fr_Protein_high_SNR_model2",
    "561nm model dir": "100ms_300fr_Protein_low_SNR_model2",
    "640nm model dir": "100ms_300fr_Organic_high_SNR_model4",
    "Fitting tool window size": 6,
    "Fitting tool threshold": 3,
    "Fitting tool window stride": 2,
    "Fit convolve": True,
    "Fluorophore config": ["Cyanine 5", "mCherry", "mEGFP"],
    "Sum view frame fraction": [0.2, 0.2, 0.2],
    "Intensity target": [4, 3, 6],
    "Custom NeuralNet integration": False,
    "Custom NeuralNet settings": [300, 300, 300],
    "Calibration optimizer settings": {
        "xmin": -256,
        "ymin": -256,
        "xmax": 640,
        "ymax": 640,
        "xstep": 96,
        "ystep": 96,
        "minscale": 80,
        "maxscale": 130,
        "scalestep": 20,
        "maxiter": 100,
    },
    "Restore Defaults": "",
}


# raw fit criteria in order marker, 561nm, 488nm
raw_fit_criteria = {
    "detection threshold": [3.0, 3.0, 3.0],
    "averaging distance": [2.0, 2.0, 2.0],
    "minimum kernel residual": [22.0, 25.0, 25.0],
    "minimum sigma": [0.8, 0.8, 0.8],
    "maximum sigma": [3.5, 3.5, 3.5],
    "minimum intensity": [20, 20, 20],
    "minimum gauss amplitude": [20, 20, 20],
    "eccentricity threshold": [0.6, 0.6, 0.6],
    "minimum gauss residual": [6.0, 6.0, 6.0],
}

raw_setup = {
    "marker frames": [10, 300],
    "561nm frames": [320, 300],
    "488nm frames": [630, 300],
    "background": [96, 96, 96],
    "analyse": [1, 1, 1],
    "brightness": [4, 4, 4],
    "inversion power": [2, 3, 1],
    "convolutions": [8, 12, 6],
}

calibration = {
    "date": 20221014,
    "X0": 0,
    "Y0": 0,
    "Xc": 0,
    "Yc": 0,
    "SFx": 10000,
    "SFy": 10000,
    "criterion": 2,
}


opt_trial_calib = [128, 128, 100, 100]


def fluoro_name(index):
    return preferences["Fluorophore config"][index]


internal_conversion = {
    "Cyanine 5": lambda: fluoro_name(0),
    "mCherry": lambda: fluoro_name(1),
    "GFP": lambda: fluoro_name(2),
}

kernels = {
    "lin": [1, 1, 1, 1, 1],
    "gauss": [0.2, 0.8, 1.2, 0.8, 0.2],
    "trig": [0, 1, 0, -1, 0],
    "neg": [-1, -1, -1, -1, -1],
    "point": [0, 0, 4, 0, 0],
    "lim": [2, 0, 0, 0, 2],
    "cos": [1, 0, -1, 0, 1],
    "exp": [0, 1, 2, 4, 8],
    "quad": [4, 1, 0, 1, 4],
    "sqr": [2, 2, 0, -2, -2],
    "saw": [2, -2, 2, -2, 2],
    "invsaw": [-2, 2, -2, 2, -2],
    "phase": [0, 0, 3, -3, 0],
    "sobel": [2, 5, 0, -5, -2],
}

default_dir = "N:/"                                                                                                         # set initial default directory location.


def load_preferences():                                                                                                  # functions for loading and saving preferences list have to be defined here
    global preferences                                                                                                   # so they can be called on launch
    global default_dir
    global raw_setup
    global raw_fit_criteria

    try:
        with open("config.dat", "rb") as load_config:
            raw_pref = pickle.load(load_config)
            raw_dir = pickle.load(load_config)
            preferences = raw_pref
            default_dir = raw_dir
    except:
        easygui.msgbox(title="Warning!", msg="Preferences configuration: 'config.dat' could not be loaded. Creating"
                                             " default configuration file. Any previous user preferences will be reset.")
        save_preferences()

    try:
        with open("defaults.dat", "rb") as load_defaults:
            raw_dflts = pickle.load(load_defaults)
            raw_setup = raw_dflts
    except:
        easygui.msgbox(title="Warning!", msg="Defaults configuration: 'defaults.dat' could not be loaded. Creating"
                                             " defaults file. Any previous user changes will no longer exist.")
        save_defaults()

    try:
        with open("criteria.dat", "rb") as load_criteria:
            raw_crit = pickle.load(load_criteria)
            raw_fit_criteria = raw_crit
    except:
        easygui.msgbox(title="Warning!", msg="Spot criteria configuration: 'criteria.dat' could not be loaded. Creating"
                                             " default spot criteria file. Any previous user changes will no longer exist")
        save_criteria()


def save_preferences():
    with open("config.dat", "wb") as save_config:
        pickle.dump(preferences, save_config)
        pickle.dump(default_dir, save_config)


def save_defaults():
    with open("defaults.dat", "wb") as save_dflts:
        pickle.dump(raw_setup, save_dflts)


def save_criteria():
    with open("criteria.dat", "wb") as save_crtr:
        pickle.dump(raw_fit_criteria, save_crtr)


def load_neural_network():
    try:
        model488 = tf.keras.models.load_model(preferences["488nm model dir"])
        print("\nLoaded 488nm neural network")
        model561 = tf.keras.models.load_model(preferences["561nm model dir"])
        print("\nLoaded 561nm neural network")
        model640 = tf.keras.models.load_model(preferences["640nm model dir"])
        print("\nLoaded 640nm neural network")
        return model488, model561, model640
    except:
        print("Warning error occured: Full traceback shown below:")
        print(traceback.format_exc())
        easygui.msgbox(msg="An error occurred while trying to load neural networks. Step detection is not available.\n\n" + str(traceback.format_exc()),
                       title="An error occurred!")
        set_GUI_state(tk.NORMAL)


load_preferences()
model_488, model_561, model_640 = load_neural_network()

try:
    position_model = tf.keras.models.load_model("300_fr_position_model")
    print("\nLoaded position detection model")
except:
    print("Warning error occured: Full traceback shown below:")
    print(traceback.format_exc())
    print("\nFailed to load position detection model. AI Position detection "
          "is not available. Use Algorithmic fitting instead.")


data_in = False                                                                                                          # intitialise global variable into which an excel workbook can be loaded as an object
file_path = ""                                                                                                           # initialise global variable containing file path of currently loaded file.
fluoro_filter = "All"                                                                                                    # initialise global fluorophore filter variable

total_trace_count = None                                                                                                 # initialise global count variable for total number of traces loaded.
Cy5_trace_count = None                                                                                                   # initialise global count variable for total number of Cy5 traces in all traces loaded.
mCherry_trace_count = None                                                                                               # initialise global count variable for total number of mCherry traces in all traces loaded.
GFP_trace_count = None                                                                                                   # initialise global count variable for total number of GFP traces in all traces loaded.

Cy5_sublist = []                                                                                                         # initialise global list of indices of all Cy5 traces in <all_traces> list.
mCherry_sublist = []                                                                                                     # initialise global list of indices of all mCherry traces in <all_traces> list.
GFP_sublist = []                                                                                                         # initialise global list of indices of all GFP traces in <all_traces> list.
all_sublist = []                                                                                                         # initialise global list of indices of all traces in <all_traces> list.

active_trace_list = []                                                                                                   # initialise global list of indices of traces can can be navigated through
                                                                                                                         # with the current fluorophore filter.
trim_undo_reference_stack = []
trim_undo_frames_stack = []
auto_trim_backup = []
auto_trim_fit_backup = []


load_from_excel = True
is_data_loaded = False                                                                                                   # global variable - flag for if the program has user data in memory

show_fit = True
auto_cor = False
integral = False
kernel = False
subtraction = 0
has_manual_fit_happened = False
ready_to_export = False
used_neural_network = False


raw_gui = None
raw_defaults = None
spot_criteria = None
coloc_calib = None
display_params = None
spot_warnings = None
view_raw = None
view_traces = None
quality_window = None
auto_window = None
calib_window = None
calib_analysis_win = None
view_colour_window = None
view_mode_win = None
export_selection_win = None
voice_window = None
opt_settings_win = None
custom_network_win = None
fret_win = None
filter_win = None
licence_win = None

tracking_win = None
raw_stack_window = None
proc_stack_window = None
progress_win = None
graph_window = None
enlarge_win = None
msd_win = None
r2_win = None
histwin = None
auto_track_win = None
auto_select_win = None
font_win = None


array_TIF = None
array_TIF_len = None

orgsum_marker = None
orgsum_561 = None
orgsum_488 = None

sum_array_marker = None
sum_array_561 = None
sum_array_488 = None

colocalization_data = None


def create_workbook_object():                                                                                            # function called by "Load excel workbook" <load_button> button to load a workbook
                                                                                                                         # object into the global <data_in> variable.
    global data_in
    global load_from_excel
    global is_data_loaded
    load_from_excel = True
    open_data = load_excel_file()                                                                                        # call load_excel_file function and return workbook object loaded into local variable

    if open_data:                                                                                                        # if a workbook is loaded and the open file window is not closed or cancelled
        data_in = open_data                                                                                              # put the local variable open_data into the global variable data_in and set
        is_data_loaded = True                                                                                            # flag variable is_data_loaded to true.
    if is_data_loaded:
        retrieve_button["state"] = tk.NORMAL


def load_excel_file():
    work_book = False
    global file_path
    status["text"] = "Opening File..."
    status.update()
    set_GUI_state(tk.DISABLED)
    load_button["state"] = tk.DISABLED
    load_pickle_button["state"] = tk.DISABLED
    preferences_button["state"] = tk.DISABLED

    path = easygui.fileopenbox(msg="Open an Excel (.xlsx) workbook exported from Auswerter 3",
                                    title="Open a correctly formatted excel workbook", default=default_dir+"*.xlsx")
    print("File path returned by file explorer: ", file_path)
    status["text"] = "Loading excel workbook, this may take some time..."
    status.update()
    if path:
        file_path = path
        print("Loading workbook, this may take some time...")
        work_book = opxl.load_workbook(file_path)
        status["text"] = "Excel workbook: '" + str(file_path) + "' was successfully loaded."
        status.update()
        retrieve_button["state"] = tk.NORMAL
    else:
        print("Operation was cancelled, workbook will be returned as <bool> (False)")
        status["text"] = "Operation was cancelled."
        status.update()
    load_button["state"] = tk.NORMAL
    load_pickle_button["state"] = tk.NORMAL
    set_GUI_state(tk.NORMAL)
    preferences_button["state"] = tk.NORMAL
    return work_book


def export_to_excel():
    """ export mode selection has been removed, however references and logic based on it have not removed as to avoid
     breaking the export function"""
    try:
        global ready_to_export
        global export_selection_win
        set_GUI_state(tk.DISABLED)
        export_template_path = easygui.fileopenbox(msg="Open Template excel file for export", default=default_dir+"*.xlsx")
        if not export_template_path:
            ready_to_export = True
            set_GUI_state(tk.NORMAL)
            return
        export_mode = 1
        if export_mode == 1:
            pause_for_selection = True
            export_selection_win = ExpSelectWin()
            while pause_for_selection:
                try:
                    export_selection_win.window.update()
                except:
                    """ Window was closed """
                if export_selection_win.final_selection:
                    pause_for_selection = False
            if export_selection_win.cancelled:
                status["text"] = "Export mode selection - user exit. Export was cancelled."
                status.update()
                ready_to_export = True
                set_GUI_state(tk.NORMAL)
                return

        if export_template_path:
            status["text"] = "Opening template file to export traces and statistics to..."
            status.update()
            export_template = opxl.load_workbook(export_template_path)
            status["text"] += "\nDone!"
            if export_mode == 0:
                status["text"] += "\n\nExport mode: 'Single marker step'\n"
            else:
                status["text"] += "\n\nExport mode: 'Custom selection of marker steps'\n"
            status.update()

            sheet_traces = export_template["Traces"]
            sheet_stats = export_template["Statistics"]
            sheet_photons = export_template["Photons"]
            sheet_coloc = export_template["Colocalization and Parameters"]

            if sheet_traces.cell(row=1, column=4).value is not None:
                easygui.msgbox(title="Error!", msg="Cannot Export! This template file is not empty!")
                status["text"] = "Export was cancelled due to invalid template."
                status.update()
                ready_to_export = True
                set_GUI_state(tk.NORMAL)
                return

            trace_list_export = list(all_traces)
            fit_list_export = list(all_fits)
            info_list_export = list(trace_info)
            column = 4
            status["text"] += "\nExporting data..."
            status.update()
            for trace_index in range(len(trace_list_export)):
                if info_list_export[trace_index][0] == "Cyanine 5":
                    cell_fg = "ff7766"
                if info_list_export[trace_index][0] == "mCherry":
                    cell_fg = "ffcc77"
                if info_list_export[trace_index][0] == "GFP":
                    cell_fg = "77ffbb"

                sheet_traces.cell(row=1, column=column).value = trace_index + 1
                sheet_traces.cell(row=1, column=column).fill = PatternFill("solid", fgColor="cccccc")
                sheet_traces.cell(row=2, column=column).value = internal_conversion[info_list_export[trace_index][0]]()
                sheet_traces.cell(row=2, column=column).fill = PatternFill("solid", fgColor=cell_fg)
                try:
                    sheet_traces.cell(row=3, column=column).value = info_list_export[trace_index][1]
                    sheet_traces.cell(row=3, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=4, column=column).value = info_list_export[trace_index][2]
                    sheet_traces.cell(row=4, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=5, column=column).value = info_list_export[trace_index][3]
                    sheet_traces.cell(row=5, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=6, column=column).value = info_list_export[trace_index][5]
                    sheet_traces.cell(row=6, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=7, column=column).value = info_list_export[trace_index][6]
                    sheet_traces.cell(row=7, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=8, column=column).value = info_list_export[trace_index][7]
                    sheet_traces.cell(row=8, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except IndexError:
                    """ Failed """
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                try:
                    sheet_traces.cell(row=9, column=column).value = info_list_export[trace_index][9][0]
                    sheet_traces.cell(row=9, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except:
                    """ No coloc data """
                try:
                    sheet_traces.cell(row=10, column=column).value = info_list_export[trace_index][9][1]
                    sheet_traces.cell(row=10, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except:
                    """ No coloc data """
                try:
                    sheet_traces.cell(row=11, column=column).value = info_list_export[trace_index][9][2][0]
                    sheet_traces.cell(row=11, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except:
                    """ No coloc data """
                try:
                    sheet_traces.cell(row=12, column=column).value = info_list_export[trace_index][9][2][1]
                    sheet_traces.cell(row=12, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except:
                    """ No coloc data """
                try:
                    sheet_traces.cell(row=13, column=column).value = info_list_export[trace_index][9][2][3]
                    sheet_traces.cell(row=13, column=column).fill = PatternFill("solid", fgColor="cccccc")
                except:
                    """ No coloc data """
                for step_pos in range(len(info_list_export[trace_index][4])):
                    sheet_traces.cell(row=15 + step_pos, column=column).value = info_list_export[trace_index][4][step_pos]
                    sheet_traces.cell(row=15 + step_pos, column=column).fill = PatternFill("solid", fgColor="cccccc")
                for step_height in range(len(info_list_export[trace_index][8])):
                    sheet_traces.cell(row=20 + step_height, column=column).value = info_list_export[trace_index][8][step_height]
                    sheet_traces.cell(row=20 + step_height, column=column).fill = PatternFill("solid", fgColor="cccccc")

                sheet_traces.cell(row=26, column=column).value = "Corrected"
                sheet_traces.cell(row=26, column=column + 1).value = "Fit"

                for frame in range(len(trace_list_export[trace_index])):
                    sheet_traces.cell(row=27 + frame, column=column).value = round(trace_list_export[trace_index][frame], 4)
                    sheet_traces.cell(row=27 + frame, column=column).fill = PatternFill("solid", fgColor=cell_fg)
                    sheet_traces.cell(row=27 + frame, column=column + 1).value = round(fit_list_export[trace_index][frame], 4)
                    sheet_traces.cell(row=27 + frame, column=column + 1).fill = PatternFill("solid", fgColor=cell_fg)

                column += 2
                if trace_index > 4000:
                    break
            row = 1
            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "Cyanine 5":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[0]) * heights[step])

                        for write in range(len(photon_counts)):
                            sheet_photons.cell(row=row, column=2).value = photon_counts[write]
                            row += 1

            row = 1
            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "mCherry":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[0]) * heights[step])

                        for write in range(len(photon_counts)):
                            sheet_photons.cell(row=row, column=3).value = photon_counts[write]
                            row += 1

            row = 1
            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "GFP":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[0]) * heights[step])

                        for write in range(len(photon_counts)):
                            sheet_photons.cell(row=row, column=4).value = photon_counts[write]
                            row += 1

            Cy5_integ = []
            mCherry_integ = []
            GFP_integ = []

            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "Cyanine 5":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[step]) * heights[step])

                        for write in range(len(photon_counts)):
                            Cy5_integ.append(photon_counts[write])

            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "mCherry":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[step]) * heights[step])

                        for write in range(len(photon_counts)):
                            mCherry_integ.append(photon_counts[write])

            for index in range(len(info_list_export)):
                fluorophore = info_list_export[index][0]
                positions = list(info_list_export[index][4])
                heights = info_list_export[index][8]
                if fluorophore == "GFP":

                    if len(heights) > 0:
                        positions.insert(0, 0)
                        photon_counts = []
                        for step in range(len(heights)):
                            photon_counts.append((positions[step + 1] - positions[step]) * heights[step])

                        for write in range(len(photon_counts)):
                            GFP_integ.append(photon_counts[write])

            Cy5_bin = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            mCherry_bin = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            GFP_bin = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

            for bins in range(len(Cy5_integ)):
                if 0 < Cy5_integ[bins] < 10000:
                    Cy5_bin[int(Cy5_integ[bins] / 1000)] += 1
            for bins in range(len(mCherry_integ)):
                if 0 < mCherry_integ[bins] < 10000:
                    mCherry_bin[int(mCherry_integ[bins] / 1000)] += 1
            for bins in range(len(GFP_integ)):
                if 0 < GFP_integ[bins] < 10000:
                    GFP_bin[int(GFP_integ[bins] / 1000)] += 1

            for write in range(len(Cy5_bin)):
                sheet_photons.cell(row=1 + write, column=5).value = Cy5_bin[write]
                sheet_photons.cell(row=1 + write, column=6).value = mCherry_bin[write]
                sheet_photons.cell(row=1 + write, column=7).value = GFP_bin[write]

            if Cy5_trace_count > 0:
                Cy5_steps = [0, 0, 0, 0, 0, 0, 0]
                for trace in Cy5_sublist:
                    if trace_info[trace][1] == "Partially bleached":
                        Cy5_steps[6] += 1
                    elif trace_info[trace][1] == 0:
                        Cy5_steps[0] += 1
                    elif trace_info[trace][1] == 1:
                        Cy5_steps[1] += 1
                    elif trace_info[trace][1] == 2:
                        Cy5_steps[2] += 1
                    elif trace_info[trace][1] == 3:
                        Cy5_steps[3] += 1
                    elif trace_info[trace][1] == 4:
                        Cy5_steps[4] += 1
                    elif trace_info[trace][1] >= 5:
                        Cy5_steps[5] += 1

                row, column = 5, 3
                for index in range(len(Cy5_steps)):
                    sheet_stats.cell(row=row, column=column).value = round(Cy5_steps[index], 3)
                    row += 1

                if export_mode == 1:

                    if 0 in export_selection_win.final_selection:
                        sheet_stats.cell(row=5, column=1).value = "True"
                    if 1 in export_selection_win.final_selection:
                        sheet_stats.cell(row=6, column=1).value = "True"
                    if 2 in export_selection_win.final_selection:
                        sheet_stats.cell(row=7, column=1).value = "True"
                    if 3 in export_selection_win.final_selection:
                        sheet_stats.cell(row=8, column=1).value = "True"
                    if 4 in export_selection_win.final_selection:
                        sheet_stats.cell(row=9, column=1).value = "True"
                    if 5 in export_selection_win.final_selection:
                        sheet_stats.cell(row=10, column=1).value = "True"
                    if "Partially bleached" in export_selection_win.final_selection:
                        sheet_stats.cell(row=11, column=1).value = "True"
                else:
                    sheet_stats.cell(row=6, column=1).value = "True"

            GFP_steps = [0, 0, 0, 0, 0, 0, 0]
            mCherry_steps = [0, 0, 0, 0, 0, 0, 0]
            Cy5_GFP = [0, 0, 0, 0, 0, 0, 0]
            Cy5_mCherry = [0, 0, 0, 0, 0, 0, 0]
            for modify in range(len(trace_info)):
                if trace_info[modify][0] == "GFP" and trace_info[modify][1] == "Partially bleached":
                    trace_info[modify][1] = 0.00001
                if trace_info[modify][0] == "mCherry" and trace_info[modify][1] == "Partially bleached":
                    trace_info[modify][1] = 0.00001
            for trace in Cy5_sublist:
                if trace_info[trace][1] == 0:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[0] += trace_info[trace + 1][1]
                            Cy5_GFP[0] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[0] += trace_info[trace + 1][1]
                            Cy5_mCherry[0] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[0] += trace_info[trace + 2][1]
                            Cy5_GFP[0] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[0] += trace_info[trace + 2][1]
                            Cy5_mCherry[0] += 1
                if trace_info[trace][1] == 1:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[1] += trace_info[trace + 1][1]
                            Cy5_GFP[1] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[1] += trace_info[trace + 1][1]
                            Cy5_mCherry[1] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[1] += trace_info[trace + 2][1]
                            Cy5_GFP[1] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[1] += trace_info[trace + 2][1]
                            Cy5_mCherry[1] += 1
                if trace_info[trace][1] == 2:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[2] += trace_info[trace + 1][1]
                            Cy5_GFP[2] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[2] += trace_info[trace + 1][1]
                            Cy5_mCherry[2] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[2] += trace_info[trace + 2][1]
                            Cy5_GFP[2] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[2] += trace_info[trace + 2][1]
                            Cy5_mCherry[2] += 1
                if trace_info[trace][1] == 3:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[3] += trace_info[trace + 1][1]
                            Cy5_GFP[3] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[3] += trace_info[trace + 1][1]
                            Cy5_mCherry[3] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[3] += trace_info[trace + 2][1]
                            Cy5_GFP[3] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[3] += trace_info[trace + 2][1]
                            Cy5_mCherry[3] += 1
                if trace_info[trace][1] == 4:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[4] += trace_info[trace + 1][1]
                            Cy5_GFP[4] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[4] += trace_info[trace + 1][1]
                            Cy5_mCherry[4] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[4] += trace_info[trace + 2][1]
                            Cy5_GFP[4] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[4] += trace_info[trace + 2][1]
                            Cy5_mCherry[4] += 1
                if trace_info[trace][1] == 5:
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[5] += trace_info[trace + 1][1]
                            Cy5_GFP[5] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[5] += trace_info[trace + 1][1]
                            Cy5_mCherry[5] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[5] += trace_info[trace + 2][1]
                            Cy5_GFP[5] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[5] += trace_info[trace + 2][1]
                            Cy5_mCherry[5] += 1
                if trace_info[trace][1] == "Partially bleached":
                    if trace + 1 < len(trace_info):
                        if trace_info[trace + 1][0] == "GFP":
                            GFP_steps[6] += trace_info[trace + 1][1]
                            Cy5_GFP[6] += 1
                        if trace_info[trace + 1][0] == "mCherry":
                            mCherry_steps[6] += trace_info[trace + 1][1]
                            Cy5_mCherry[6] += 1
                    if trace + 2 < len(trace_info):
                        if trace_info[trace + 2][0] == "GFP":
                            GFP_steps[6] += trace_info[trace + 2][1]
                            Cy5_GFP[6] += 1
                        if trace_info[trace + 2][0] == "mCherry":
                            mCherry_steps[6] += trace_info[trace + 2][1]
                            Cy5_mCherry[6] += 1

            for modify in range(len(trace_info)):
                if trace_info[modify][0] == "GFP" and trace_info[modify][1] == 0.00001:
                    trace_info[modify][1] = "Partially bleached"
                if trace_info[modify][0] == "mCherry" and trace_info[modify][1] == 0.00001:
                    trace_info[modify][1] = "Partially bleached"

            row = 5
            for write in range(len(GFP_steps)):
                if Cy5_GFP[write] > 0:
                    GFP_steps[write] = GFP_steps[write] / Cy5_GFP[write]
                if Cy5_mCherry[write] > 0:
                    mCherry_steps[write] = mCherry_steps[write] / Cy5_mCherry[write]

                sheet_stats.cell(row=row, column=4).value = round(GFP_steps[write], 3)
                sheet_stats.cell(row=row, column=5).value = round(mCherry_steps[write], 3)
                row += 1

            defined_steps_list = export_selection_win.final_selection
            single_Cy5_info = []
            for trace in range(len(trace_info)):
                if export_mode == 0:
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] == 1:
                        if trace + 1 < len(trace_info):
                            if trace_info[trace + 1][0] != "Cyanine 5":
                                single_Cy5_info.append(trace_info[trace + 1])
                        if trace + 2 < len(trace_info):
                            if trace_info[trace + 1][0] != "Cyanine 5" and trace_info[trace + 2][0] != "Cyanine 5":
                                single_Cy5_info.append(trace_info[trace + 2])

                elif export_mode == 1:
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] in defined_steps_list:
                        if trace + 1 < len(trace_info):
                            if trace_info[trace + 1][0] != "Cyanine 5":
                                single_Cy5_info.append(trace_info[trace + 1])
                        if trace + 2 < len(trace_info):
                            if trace_info[trace + 1][0] != "Cyanine 5" and trace_info[trace + 2][0] != "Cyanine 5":
                                single_Cy5_info.append(trace_info[trace + 2])

            if Cy5_trace_count == 0:
                single_Cy5_info = list(trace_info)

            if export_selection_win.final_selection[0] == 'unfiltered':
                if GFP_trace_count > 0:
                    GFP_steps = [0, 0, 0, 0, 0, 0, 0]
                    for trace in trace_info:
                        if trace[0] == "GFP":
                            if trace[1] == "Partially bleached":
                                GFP_steps[6] += 1
                            elif trace[1] == 0:
                                GFP_steps[0] += 1
                            elif trace[1] == 1:
                                GFP_steps[1] += 1
                            elif trace[1] == 2:
                                GFP_steps[2] += 1
                            elif trace[1] == 3:
                                GFP_steps[3] += 1
                            elif trace[1] == 4:
                                GFP_steps[4] += 1
                            elif trace[1] >= 5:
                                GFP_steps[5] += 1

                if mCherry_trace_count > 0:
                    mCherry_steps = [0, 0, 0, 0, 0, 0, 0]
                    for trace in trace_info:
                        if trace[0] == "mCherry":
                            if trace[1] == "Partially bleached":
                                mCherry_steps[6] += 1
                            elif trace[1] == 0:
                                mCherry_steps[0] += 1
                            elif trace[1] == 1:
                                mCherry_steps[1] += 1
                            elif trace[1] == 2:
                                mCherry_steps[2] += 1
                            elif trace[1] == 3:
                                mCherry_steps[3] += 1
                            elif trace[1] == 4:
                                mCherry_steps[4] += 1
                            elif trace[1] >= 5:
                                mCherry_steps[5] += 1

            else:
                if GFP_trace_count > 0:
                    GFP_steps = [0, 0, 0, 0, 0, 0, 0]
                    for trace in single_Cy5_info:
                        if trace[0] == "GFP":
                            if trace[1] == "Partially bleached":
                                GFP_steps[6] += 1
                            elif trace[1] == 0:
                                GFP_steps[0] += 1
                            elif trace[1] == 1:
                                GFP_steps[1] += 1
                            elif trace[1] == 2:
                                GFP_steps[2] += 1
                            elif trace[1] == 3:
                                GFP_steps[3] += 1
                            elif trace[1] == 4:
                                GFP_steps[4] += 1
                            elif trace[1] >= 5:
                                GFP_steps[5] += 1

                if mCherry_trace_count > 0:
                    mCherry_steps = [0, 0, 0, 0, 0, 0, 0]
                    for trace in single_Cy5_info:
                        if trace[0] == "mCherry":
                            if trace[1] == "Partially bleached":
                                mCherry_steps[6] += 1
                            elif trace[1] == 0:
                                mCherry_steps[0] += 1
                            elif trace[1] == 1:
                                mCherry_steps[1] += 1
                            elif trace[1] == 2:
                                mCherry_steps[2] += 1
                            elif trace[1] == 3:
                                mCherry_steps[3] += 1
                            elif trace[1] == 4:
                                mCherry_steps[4] += 1
                            elif trace[1] >= 5:
                                mCherry_steps[5] += 1


            mCh_steps_final = mCherry_steps[1:]
            mCh_steps_final.append(mCherry_steps[0])
            GFP_steps_final = GFP_steps[1:]
            GFP_steps_final.append(GFP_steps[0])
            row = 16
            for write in range(len(GFP_steps)):
                sheet_stats.cell(row=row, column=3).value = mCh_steps_final[write]
                sheet_stats.cell(row=row, column=5).value = GFP_steps_final[write]
                row += 1

            single_cy5_secondary_coloc = 0
            double_cy5_secondary_coloc = 0
            multiple_cy5_secondary_coloc = 0

            valid = [2, 3, 4, 5, "Partially bleached"]

            triple_coloc_steps = []
            if defined_steps_list == ['unfiltered']:
                allowed_triple_list = [0, 1, 2, 3, 4, 5, "Partially bleached"]
            else:
                allowed_triple_list = defined_steps_list

            for trace in Cy5_sublist:
                if trace + 2 < len(trace_info):
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] == 1 and trace_info[trace + 1][0] == "GFP" \
                    and trace_info[trace + 2][0] == "mCherry":
                        single_cy5_secondary_coloc += 1
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] == 1 and trace_info[trace + 1][0] == "mCherry" \
                    and trace_info[trace + 2][0] == "GFP":
                        single_cy5_secondary_coloc += 1

                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] == 2 and trace_info[trace + 1][0] == "GFP" \
                    and trace_info[trace + 2][0] == "mCherry":
                        double_cy5_secondary_coloc += 1
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] == 2 and trace_info[trace + 1][0] == "mCherry" \
                    and trace_info[trace + 2][0] == "GFP":
                        double_cy5_secondary_coloc += 1

                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] in valid and trace_info[trace + 1][0] == "GFP" \
                    and trace_info[trace + 2][0] == "mCherry":
                        multiple_cy5_secondary_coloc += 1
                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace][1] in valid and trace_info[trace + 1][0] == "mCherry" \
                    and trace_info[trace + 2][0] == "GFP":
                        multiple_cy5_secondary_coloc += 1

                    if trace_info[trace][0] == "Cyanine 5" and trace_info[trace + 1][0] == "GFP" \
                    and trace_info[trace + 2][0] == "mCherry" and trace_info[trace][1] in allowed_triple_list:
                        triple_coloc_steps.append([trace_info[trace][1], trace_info[trace + 2][1], trace_info[trace + 1][1]])
                    elif trace_info[trace][0] == "Cyanine 5" and trace_info[trace + 1][0] == "mCherry" \
                    and trace_info[trace + 2][0] == "GFP" and trace_info[trace][1] in allowed_triple_list:
                        triple_coloc_steps.append([trace_info[trace][1], trace_info[trace + 1][1], trace_info[trace + 2][1]])

            sheet_stats.cell(row=13, column=3).value = len(Cy5_sublist)

            sheet_stats.cell(row=4, column=8).value = single_cy5_secondary_coloc
            sheet_stats.cell(row=5, column=8).value = double_cy5_secondary_coloc
            sheet_stats.cell(row=6, column=8).value = multiple_cy5_secondary_coloc

            for tripindex, triple in enumerate(triple_coloc_steps):
                if triple[0] != "Partially bleached":
                    sheet_stats.cell(row=5+tripindex, column=12).value = triple[0]
                else:
                    sheet_stats.cell(row=5 + tripindex, column=12).value = "P"
                if triple[1] != "Partially bleached":
                    sheet_stats.cell(row=5+tripindex, column=13).value = triple[1]
                else:
                    sheet_stats.cell(row=5 + tripindex, column=13).value = "P"
                if triple[2] != "Partially bleached":
                    sheet_stats.cell(row=5+tripindex, column=14).value = triple[2]
                else:
                    sheet_stats.cell(row=5 + tripindex, column=14).value = "P"


            sheet_stats.cell(row=29, column=9).value = preferences["Fluorophore config"][2]
            sheet_stats.cell(row=30, column=9).value = preferences["Fluorophore config"][1]
            sheet_stats.cell(row=31, column=9).value = preferences["Fluorophore config"][0]


            if colocalization_data:
                marker_spots = colocalization_data.count_marker_list
                spots_561 = colocalization_data.count_561_list
                spots_488 = colocalization_data.count_488_list
                coloc_561 = colocalization_data.coloc_561_list
                coloc_488 = colocalization_data.coloc_488_list
                params = colocalization_data.calib_params
                bg640 = colocalization_data.backgrounds640
                bg561 = colocalization_data.backgrounds561
                bg488 = colocalization_data.backgrounds488
                nhance = colocalization_data.enhancements
                fhist = colocalization_data.file_history

                for export in range(len(marker_spots)):
                    sheet_coloc.cell(row=3, column=6+export).value = marker_spots[export]
                    sheet_coloc.cell(row=7, column=6+export).value = spots_488[export]
                    sheet_coloc.cell(row=8, column=6+export).value = spots_561[export]

                    if colocalization_data.trace_mode == "coloc":
                        try:
                            sheet_coloc.cell(row=4, column=6 + export).value = coloc_488[export]
                        except IndexError:
                            """ No Colocalized 488 """
                        try:
                            sheet_coloc.cell(row=5, column=6 + export).value = coloc_561[export]
                        except IndexError:
                            """ No colocalized 561 """
                try:
                    for export in range(len(params)):
                        sheet_coloc.cell(row=10, column=6 + export).value = params[export][0]
                        sheet_coloc.cell(row=11, column=6 + export).value = params[export][1]
                        sheet_coloc.cell(row=12, column=6 + export).value = params[export][2]
                        sheet_coloc.cell(row=13, column=6 + export).value = params[export][3]
                except:
                    """ Error """

                try:
                    for export in range(len(fhist)):
                        sheet_coloc.cell(row=1, column=6 + export).value = fhist[export]
                except:
                    """ Error """

                try:
                    for export in range(len(nhance)):
                        sheet_coloc.cell(row=52, column=6 + export).value = nhance[export][0]
                        sheet_coloc.cell(row=53, column=6 + export).value = nhance[export][1]
                except:
                    """ Error """

                try:
                    for export in range(len(bg640)):
                        sheet_coloc.cell(row=47, column=6 + export).value = bg640[export]
                        sheet_coloc.cell(row=48, column=6 + export).value = bg488[export]
                        sheet_coloc.cell(row=49, column=6 + export).value = bg561[export]
                except:
                    """ Error """

                criteria_keys = list(raw_fit_criteria)
                for key_index in range(len(criteria_keys)):
                    sheet_coloc.cell(row=17+key_index, column=3).value = raw_fit_criteria[criteria_keys[key_index]][0]
                    sheet_coloc.cell(row=17+key_index, column=4).value = raw_fit_criteria[criteria_keys[key_index]][1]
                    sheet_coloc.cell(row=17+key_index, column=5).value = raw_fit_criteria[criteria_keys[key_index]][2]

                frame_frac = preferences["Sum view frame fraction"]
                int_target = preferences["Intensity target"]

                for laser_index in range(3):
                    sheet_coloc.cell(row=30, column=3+laser_index).value = frame_frac[laser_index]
                    sheet_coloc.cell(row=31, column=3+laser_index).value = int_target[laser_index]

                export_calibration = list(calibration)
                for key_index in range(len(export_calibration)):
                    sheet_coloc.cell(row=36+key_index, column=3).value = calibration[export_calibration[key_index]]

                sheet_coloc.cell(row=19, column=8).value = str(preferences["640nm model dir"])
                sheet_coloc.cell(row=20, column=8).value = str(preferences["561nm model dir"])
                sheet_coloc.cell(row=21, column=8).value = str(preferences["488nm model dir"])


            status["text"] += "\nSaving document..."
            status.update()
            export_template.save(export_template_path)
            status["text"] += "\nFinished!"
            status.update()
            easygui.msgbox(title="Export Complete", msg="Export to Excel has completed")
        ready_to_export = True
        set_GUI_state(tk.NORMAL)
    except:
        print("Warning error occured: Full traceback shown below:")
        print(traceback.format_exc())
        print("Cannot Complete Export")
        easygui.msgbox(title="Error!", msg="An error occured during export. Export failed. Show the following message"
                                           " to the developer:\n" + str(traceback.format_exc()))
        set_GUI_state(tk.NORMAL)


def pickle_data():
    global trace_info
    global all_traces
    set_GUI_state(tk.DISABLED)
    choices = [preferences["Fluorophore config"][0], preferences["Fluorophore config"][1], preferences["Fluorophore config"][2], "Use current filter"]
    # tts -> traces to save
    tts = None
    while not tts:
        tts = easygui.multchoicebox(title="Pickle traces", msg="Select which traces to save", choices=choices, preselect=3)
        if not tts:
            easygui.msgbox(title="Warning!", msg="You must select at least one set of fluorophore traces",
                           ok_button="Re-pick")
        if tts:
            if choices[3] in tts and len(tts) > 1:
                easygui.msgbox(title="Warning!", msg="Cannot save according to current filter if other fluorophores are selected",
                               ok_button="Re-pick")
                tts = None
    save_reference_list = []
    status["text"] = "Building reference list of traces to pickle..."
    status.update()
    Cy5_selection = []
    mCherry_selection = []
    GFP_selection = []
    if choices[3] not in tts:
        save_mode = easygui.indexbox(title="Save selection mode", msg="Do you wish to save all traces or a specific"
                                     " range of traces? ", choices=["Save all", "Select range"], default_choice=0)
        if save_mode == 0:
            Cy5_selection = Cy5_sublist
            mCherry_selection = mCherry_sublist
            GFP_selection = GFP_sublist
        elif save_mode == 1:
            try:
                if len(Cy5_sublist) > 0 and choices[0] in tts:
                    Cy5_start = easygui.integerbox(title=preferences["Fluorophore config"][0]+" Selection",
                                                   msg="Enter starting trace for "+preferences["Fluorophore config"][0],
                                                   lowerbound=1, upperbound=len(Cy5_sublist))
                    Cy5_range = easygui.integerbox(title=preferences["Fluorophore config"][0]+" Selection",
                                                   msg="Enter length of range for "+preferences["Fluorophore config"][0],
                                                   lowerbound=1, upperbound=len(Cy5_sublist)-Cy5_start+1)
                    Cy5_selection = Cy5_sublist[slice(Cy5_start - 1, Cy5_start + Cy5_range - 1, 1)]
                if len(mCherry_sublist) > 0 and choices[1] in tts:
                    mCherry_start = easygui.integerbox(title=preferences["Fluorophore config"][1]+" Selection",
                                                       msg="Enter starting trace for "+preferences["Fluorophore config"][1],
                                                       lowerbound=1, upperbound=len(mCherry_sublist))
                    mCherry_range = easygui.integerbox(title=preferences["Fluorophore config"][1]+" Selection",
                                                       msg="Enter length of range for "+preferences["Fluorophore config"][1],
                                                       lowerbound=1, upperbound=len(mCherry_sublist)-mCherry_start+1)
                    mCherry_selection = mCherry_sublist[slice(mCherry_start - 1, mCherry_start + mCherry_range - 1, 1)]
                if len(GFP_sublist) > 0 and choices[2] in tts:
                    GFP_start = easygui.integerbox(title=preferences["Fluorophore config"][2]+" Selection",
                                                   msg="Enter starting trace for "+preferences["Fluorophore config"][2],
                                                   lowerbound=1, upperbound=len(GFP_sublist))
                    GFP_range = easygui.integerbox(title=preferences["Fluorophore config"][2]+" Selection",
                                                   msg="Enter length of range for "+preferences["Fluorophore config"][2],
                                                   lowerbound=1, upperbound=len(GFP_sublist)-GFP_start+1)
                    GFP_selection = GFP_sublist[slice(GFP_start - 1, GFP_start + GFP_range - 1, 1)]
            except:
                Cy5_selection = Cy5_sublist
                mCherry_selection = mCherry_sublist
                GFP_selection = GFP_sublist
                status["text"] += "\nError in range selection. Selecting All traces instead."
                status.update()

    if choices[3] in tts:
        save_reference_list = list(active_trace_list)
    elif choices[3] not in tts:
        if choices[0] in tts and Cy5_trace_count > 0:
            save_reference_list += Cy5_selection
        elif choices[0] in tts and Cy5_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="There are no "+preferences["Fluorophore config"][0]+" traces to save.")
            status["text"] += "\nPickling sequence cancelled due to error: Invalid trace type(s) selected"
            status.update()
            set_GUI_state(tk.NORMAL)
            return
        if choices[1] in tts and mCherry_trace_count > 0:
            save_reference_list += mCherry_selection
        elif choices[1] in tts and mCherry_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="There are no "+preferences["Fluorophore config"][1]+" traces to save.")
            status["text"] += "\nPickling sequence cancelled due to error: Invalid trace type(s) selected"
            status.update()
            set_GUI_state(tk.NORMAL)
            return
        if choices[2] in tts and GFP_trace_count > 0:
            save_reference_list += GFP_selection
        elif choices[2] in tts and GFP_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="There are no "+preferences["Fluorophore config"][2]+" traces to save.")
            status["text"] += "\nPickling sequence cancelled due to error: Invalid trace type(s) selected"
            status.update()
            set_GUI_state(tk.NORMAL)
            return
        save_reference_list.sort()
    status["text"] += "\nReference list successfully built.\nExtracting referenced traces into new list..."
    status.update()
    save_label_list = []
    save_traces_list = []
    save_fits_list = []
    for extract_trace in range(len(save_reference_list)):
        save_label_list.append(trace_info[save_reference_list[extract_trace]])
        save_traces_list.append(all_traces[save_reference_list[extract_trace]])
        save_fits_list.append(all_fits[save_reference_list[extract_trace]])
    status["text"] += "\nTraces extracted, pickle lists have been compiled."
    status.update()
    # print(save_label_list)
    # for i in range(len(save_traces_list)):
        # print(save_traces_list[i])

    save_file_path = easygui.filesavebox(title="Pickle traces", msg="Save traces to a file", default=default_dir +
                                         "untitled traces", filetypes=["*.trcs"])
    if save_file_path:
        ext = save_file_path[-5:]
        if ext == ".trcs":
            save_file_path = save_file_path[:-5]
    if save_file_path:
        status["text"] += "\nPickling traces to '" + str(save_file_path) + "'..."
        with open(save_file_path + ".trcs", "wb") as pickle_file:
            pickle.dump(save_label_list, pickle_file)
            pickle.dump(save_traces_list, pickle_file)
            pickle.dump(save_fits_list, pickle_file)
            pickle.dump(ready_to_export, pickle_file)
            pickle.dump(used_neural_network, pickle_file)
    else:
        status["text"] += "\nPickling operation was cancelled."
        status.update()
    status["text"] += "\nPickling successfully completed."
    set_GUI_state(tk.NORMAL)


def unpickle_data():
    global all_fits
    global all_traces
    global all_sublist
    global active_trace_list
    global trace_info
    global Cy5_sublist
    global mCherry_sublist
    global GFP_sublist
    global total_trace_count
    global Cy5_trace_count
    global mCherry_trace_count
    global GFP_trace_count
    global file_path
    global current_trace
    global load_from_excel
    global is_data_loaded
    global has_manual_fit_happened, ready_to_export, used_neural_network
    global trim_undo_reference_stack, trim_undo_frames_stack, auto_trim_backup, auto_trim_fit_backup
    global waiting_for_fit
    global colocalization_data

    set_GUI_state(tk.DISABLED)
    retrieve_button["state"] = tk.DISABLED
    load_pickle_button["state"] = tk.DISABLED
    load_button["state"] = tk.DISABLED
    preferences_button["state"] = tk.DISABLED
    open_file = easygui.fileopenbox(msg="Open a serialised (pickled) data file", default=default_dir+"*.trcs",
                                    filetypes=["*.trcs"])

    if open_file:
        load_from_excel = False
        all_fits = []
        all_traces = []
        all_sublist = []
        active_trace_list = []
        trace_info = []
        Cy5_sublist = []
        mCherry_sublist = []
        GFP_sublist = []
        total_trace_count = 0
        Cy5_trace_count = 0
        mCherry_trace_count = 0
        GFP_trace_count = 0
        file_path = None
        colocalization_data = None

        status["text"] = "Unpickling data from file: '"+str(open_file)+"'..."
        status.update()
        file_path = open_file
        with open(file_path, "rb") as open_pickle:
            labels = pickle.load(open_pickle)
            traces = pickle.load(open_pickle)
            fits = pickle.load(open_pickle)
            try:
                ready_to_export = pickle.load(open_pickle)
                used_neural_network = pickle.load(open_pickle)
            except:
                """ export state could not be read """
        trace_info = list(labels)
        all_traces = list(traces)
        all_fits = fits

        for count in range(len(trace_info)):
            if len(trace_info[count]) < 5:
                trace_info[count].append([])
            if len(trace_info[count]) < 6:
                trace_info[count].append(None)
                trace_info[count].append(None)
                trace_info[count].append(None)
                trace_info[count].append([])

        total_trace_count = len(trace_info)

        for count in range(total_trace_count):
            if trace_info[count][0] == "Cyanine 5":
                Cy5_trace_count += 1
                Cy5_sublist.append(count)
            if trace_info[count][0] == "mCherry":
                mCherry_trace_count += 1
                mCherry_sublist.append(count)
            if trace_info[count][0] == "GFP":
                GFP_trace_count += 1
                GFP_sublist.append(count)
            all_sublist.append(count)
        active_trace_list = all_sublist
        Cy5_trace_count = len(Cy5_sublist)
        mCherry_trace_count = len(mCherry_sublist)
        GFP_trace_count = len(GFP_sublist)
        easygui.msgbox(title="Trace count", msg="Found " + str(total_trace_count) + " traces, of which\n" +
                       str(Cy5_trace_count) + " were "+preferences["Fluorophore config"][0]+" marker spots,\n" + str(mCherry_trace_count) +
                       " were "+preferences["Fluorophore config"][1]+" spots, and \n" + str(GFP_trace_count) +
                       " were "+preferences["Fluorophore config"][2]+" spots.")
        current_trace = 0
        update_infobox()
        plot_trace(active_trace_list[current_trace])
        set_GUI_state(tk.NORMAL)
        retrieve_button["state"] = tk.DISABLED
        status["text"] += "\nFile opened."
        is_data_loaded = True
    load_button["state"] = tk.NORMAL
    load_pickle_button["state"] = tk.NORMAL
    if is_data_loaded:
        set_GUI_state(tk.NORMAL)
        edit_fit_button["bg"] = button_bg
        edit_fit_button["fg"] = button_fg
        edit_fit_button["relief"] = tk.RAISED
        has_manual_fit_happened = True
        trim_undo_reference_stack = []
        trim_undo_frames_stack = []
        auto_trim_backup = []
        auto_trim_fit_backup = []
        undo_trim_button["state"] = tk.DISABLED
        undo_auto_trim_button["state"] = tk.DISABLED
        export_button["state"] = tk.NORMAL
        waiting_for_fit = False
        if capture_mouse:
            trace_figure.canvas.mpl_disconnect(capture_mouse)
    load_pickle_button["state"] = tk.NORMAL
    load_button["state"] = tk.NORMAL
    preferences_button["state"] = tk.NORMAL


def create_metadata_list():
    global trace_info, has_manual_fit_happened
    set_GUI_state(tk.DISABLED)
    try:
        trc_inf = retrieve_data()
        trace_info = trc_inf
        status["text"] += "\nInfo list successfully created.\nCollecting traces..."
        status.update()
        # print(trace_info)
        collect_traces()
    except KeyError:
        if trc_inf != "FluoroImport":
            load_error("Key")
    except:
        print("Warning error occured: Full traceback shown below:")
        print(traceback.format_exc())
        if trc_inf != "FluoroImport":
            load_error("Other")
    finally:
        if trc_inf == "FluoroImport":
            try:
                retrieve_own_data()
            except:
                load_error("Other")
            return


def load_error(_type):
    global is_data_loaded
    if _type == "Key":
        easygui.msgbox(title="Error while loading data!", msg="An error occurred while attempting to retrieve data:"
                                                              " Could not find sheet 'Colocalization and Traces'."
                                                              " Please make sure that the excel sheet loaded is valid."
                       )

    elif _type == "Other":
        easygui.msgbox(title="Error while loading data!", msg="An error occurred while attempting to retrieve data:"
                                                              " An unknown error occurred. Please make sure that the"
                                                              " excel sheet loaded is valid and that a blank one was"
                                                              " not loaded by mistake."
                       )

    status["text"] = "An error occurred while trying to load traces from excel sheet."
    status.update()

    if not trace_info:
        is_data_loaded = False
    set_GUI_state(tk.NORMAL)


def retrieve_data():
    global sheet
    global xl_workbook
    global has_manual_fit_happened, ready_to_export, used_neural_network
    xl_workbook = data_in
    if xl_workbook:
        if "Colocalization and Traces" not in xl_workbook:
            print("Excel workbook is not from Auswerter 3... Attempting FluoroTensor template load")
            status["text"] += "\nExcel workbook is not from Auswerter 3... Attempting FluoroTensor template load"
            status.update()
            trace_metadata = "FluoroImport"
            return trace_metadata
        sheet = xl_workbook["Colocalization and Traces"]
        is_data = True
        start_column = 8
        trace_metadata = []
        status["text"] = "Creating trace description / information list from sheet: 'Colocalization and Traces'"
        status.update()
        while is_data:
            start_frame = str(sheet.cell(row=20, column=start_column).value)
            end_frame = str(sheet.cell(row=20, column=start_column + 1).value)
            if start_frame == "None":
                is_data = False
            else:
                this_trace = []
                frame_count = int(end_frame) - int(start_frame)
                GFP = str(sheet.cell(row=24, column=start_column).value)
                mCherry = str(sheet.cell(row=25, column=start_column).value)
                Cy5 = str(sheet.cell(row=26, column=start_column).value)
                # trace_metadata list will be appended with current_trace list in order:
                # fluorophore type (GFP, mCherry, Cy5), number of steps, number of frames, start of "Corrected" data column
                if GFP != "None":
                    this_trace.append("GFP")
                    this_trace.append(int(GFP))
                    this_trace.append(frame_count)
                    this_trace.append(start_column + 1)
                    this_trace.append([])
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append([])
                    trace_metadata.append(this_trace)
                elif mCherry != "None":
                    this_trace.append("mCherry")
                    this_trace.append(int(mCherry))
                    this_trace.append(frame_count)
                    this_trace.append(start_column + 1)
                    this_trace.append([])
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append([])
                    trace_metadata.append(this_trace)
                elif Cy5 != "None":
                    this_trace.append("Cyanine 5")
                    this_trace.append(int(Cy5))
                    this_trace.append(frame_count)
                    this_trace.append(start_column + 1)
                    this_trace.append([])
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append(None)
                    this_trace.append([])
                    trace_metadata.append(this_trace)

                start_column += 3
            # print(start_frame)
            # print(end_frame)
        has_manual_fit_happened = False
        ready_to_export = False
        used_neural_network = False
        export_button["state"] = tk.DISABLED
        return trace_metadata


def collect_traces():
    global sheet
    global trace_info
    global all_traces
    global all_fits
    global current_trace
    global total_trace_count
    global Cy5_trace_count
    global mCherry_trace_count
    global GFP_trace_count
    global Cy5_sublist
    global mCherry_sublist
    global GFP_sublist
    global active_trace_list
    global all_sublist
    global trim_undo_reference_stack, trim_undo_frames_stack, auto_trim_backup, auto_trim_fit_backup
    global waiting_for_fit
    global colocalization_data
    global ready_to_export

    all_traces = []
    all_fits = []
    for single_trace in trace_info:
        trace = []
        fit = []
        for point_in_trace in range(single_trace[2]):
            if str(sheet.cell(row=30+point_in_trace, column=single_trace[3]).value) != "None":
                trace.append(float(str(sheet.cell(row=30+point_in_trace, column=single_trace[3]).value)))
            if str(sheet.cell(row=30+point_in_trace, column=single_trace[3]+1).value) != "None":
                fit.append(float(str(sheet.cell(row=30+point_in_trace, column=single_trace[3]+1).value)))
        all_traces.append(trace)
        all_fits.append(fit)
    status["text"] += "\nAll traces and fits collected."
    status.update()
    current_trace = 0
    current_info.update()
    plot_trace(current_trace)
    # count traces
    total_trace_count = 0
    Cy5_trace_count = 0
    mCherry_trace_count = 0
    GFP_trace_count = 0

    Cy5_sublist = []
    mCherry_sublist = []
    GFP_sublist = []
    active_trace_list = []
    all_sublist = []

    colocalization_data = None

    total_trace_count = len(trace_info)

    for count in range(total_trace_count):
        if trace_info[count][0] == "Cyanine 5":
            Cy5_trace_count += 1
            Cy5_sublist.append(count)
        if trace_info[count][0] == "mCherry":
            mCherry_trace_count += 1
            mCherry_sublist.append(count)
        if trace_info[count][0] == "GFP":
            GFP_trace_count += 1
            GFP_sublist.append(count)
        all_sublist.append(count)
    active_trace_list = all_sublist
    Cy5_trace_count = len(Cy5_sublist)
    mCherry_trace_count = len(mCherry_sublist)
    GFP_trace_count = len(GFP_sublist)
    easygui.msgbox(title="Trace count", msg="Found " + str(total_trace_count) + " traces, of which\n" +
                   str(Cy5_trace_count) + " were "+preferences["Fluorophore config"][0]+" marker spots,\n" + str(mCherry_trace_count) +
                   " were "+preferences["Fluorophore config"][1]+" spots, and \n" + str(GFP_trace_count) +
                   " were "+preferences["Fluorophore config"][2]+" spots.")
    ready_to_export = True
    set_GUI_state(tk.NORMAL)
    update_infobox()
    current_fit_button["state"] = tk.DISABLED
    advanced_fit_button["state"] = tk.DISABLED
    edit_fit_button["state"] = tk.DISABLED
    edit_fit_button["bg"] = button_bg
    edit_fit_button["fg"] = button_fg
    edit_fit_button["relief"] = tk.RAISED
    trim_undo_reference_stack = []
    trim_undo_frames_stack = []
    auto_trim_backup = []
    auto_trim_fit_backup = []
    undo_trim_button["state"] = tk.DISABLED
    undo_auto_trim_button["state"] = tk.DISABLED
    waiting_for_fit = False
    if capture_mouse:
        trace_figure.canvas.mpl_disconnect(capture_mouse)


def retrieve_own_data():
    global all_fits
    global all_traces
    global all_sublist
    global active_trace_list
    global trace_info
    global Cy5_sublist
    global mCherry_sublist
    global GFP_sublist
    global total_trace_count
    global Cy5_trace_count
    global mCherry_trace_count
    global GFP_trace_count
    global file_path
    global current_trace
    global load_from_excel
    global is_data_loaded
    global has_manual_fit_happened, ready_to_export, used_neural_network
    global trim_undo_reference_stack, trim_undo_frames_stack, auto_trim_backup, auto_trim_fit_backup
    global waiting_for_fit
    global colocalization_data

    load_from_excel = True
    all_fits = []
    all_traces = []
    all_sublist = []
    active_trace_list = []
    trace_info = []
    Cy5_sublist = []
    mCherry_sublist = []
    GFP_sublist = []
    total_trace_count = 0
    Cy5_trace_count = 0
    mCherry_trace_count = 0
    GFP_trace_count = 0
    file_path = None
    colocalization_data = None
    current_trace = 0
    is_data_loaded = True
    has_manual_fit_happened = False
    ready_to_export = True
    used_neural_network = False
    trim_undo_reference_stack = []
    trim_undo_frames_stack = []
    auto_trim_backup = []
    auto_trim_fit_backup = []
    waiting_for_fit = False

    if capture_mouse:
        trace_figure.canvas.mpl_disconnect(capture_mouse)

    try:
        trace_sheet = xl_workbook["Traces"]
        settings_sheet = xl_workbook["Colocalization and Parameters"]
    except:
        load_error("Other")

    if settings_sheet.cell(row=17, column=3).value is not None:
        choice = easygui.ccbox(msg="FluoroTensor has found preferences, criteria and calibration settings"
                               " in the template file. Would you like to use this file to restore these settings"
                               " for re-analysis of raw data?\n\nLoad and use settings?", choices=["Yes", "No "],
                               default_choice="No", cancel_choice="No", title="Found settings in Template!")

        if choice:
            import_settings_from_template()

            exit_load = easygui.ccbox(title="Continue Loading?", msg="Settings have been applied. Do you still wish"
                                      " to load traces?", choices=["Yes", "No"], default_choice="No", cancel_choice="No")
            if not exit_load:
                is_data_loaded = False
                set_GUI_state(tk.NORMAL)
                trace_fig2.clf()
                trace_canv2.draw()
                trace_figure.clf()
                trace_canvas.draw()
                current_info["text"] = ""
                current_info.update()
                status["text"] += "\n\nTrace loading was cancelled!"
                return

    is_data = True
    column = 4

    while is_data:
        imp_fluoro = trace_sheet.cell(row=2, column=column).value
        if imp_fluoro is None:
            is_data = False
            break

        colourhex = trace_sheet.cell(row=2, column=column).fill.start_color.index

        this_trace_info = []
        if colourhex == "00ff7766" or colourhex == "FFFF7766":
            this_trace_info.append("Cyanine 5")
        elif colourhex == "0077ffbb" or colourhex == "FF77FFBB":
            this_trace_info.append("GFP")
        elif colourhex == "00ffcc77" or colourhex == "FFFFCC77":
            this_trace_info.append("mCherry")
        this_trace_info.append(trace_sheet.cell(row=3, column=column).value)
        this_trace_info.append(trace_sheet.cell(row=4, column=column).value)
        this_trace_info.append(trace_sheet.cell(row=5, column=column).value)
        this_trace_info.append([])
        this_trace_info.append(trace_sheet.cell(row=6, column=column).value)
        this_trace_info.append(trace_sheet.cell(row=7, column=column).value)
        this_trace_info.append(trace_sheet.cell(row=8, column=column).value)
        this_trace_info.append([])

        this_trace = []
        this_fit = []

        for frame in range(int(this_trace_info[2])):
            if trace_sheet.cell(row=27+frame, column=column).value is not None:
                this_trace.append(trace_sheet.cell(row=27+frame, column=column).value)
            if trace_sheet.cell(row=27+frame, column=column+1).value is not None:
                this_fit.append(trace_sheet.cell(row=27+frame, column=column+1).value)

        trace_info.append(this_trace_info)
        all_traces.append(this_trace)
        all_fits.append(this_fit)

        column += 2

    total_trace_count = len(trace_info)

    for count in range(total_trace_count):
        if trace_info[count][0] == "Cyanine 5":
            Cy5_trace_count += 1
            Cy5_sublist.append(count)
        if trace_info[count][0] == "mCherry":
            mCherry_trace_count += 1
            mCherry_sublist.append(count)
        if trace_info[count][0] == "GFP":
            GFP_trace_count += 1
            GFP_sublist.append(count)
        all_sublist.append(count)
    active_trace_list = all_sublist
    Cy5_trace_count = len(Cy5_sublist)
    mCherry_trace_count = len(mCherry_sublist)
    GFP_trace_count = len(GFP_sublist)
    easygui.msgbox(title="Trace count", msg="Found " + str(total_trace_count) + " traces, of which\n" +
                                            str(Cy5_trace_count) + " were " + preferences["Fluorophore config"][0] +
                                            " marker spots,\n" + str(mCherry_trace_count) + " were " +
                                            preferences["Fluorophore config"][1] + " spots, and \n" + str(GFP_trace_count) +
                                            " were " + preferences["Fluorophore config"][2] + " spots.")

    plot_trace(active_trace_list[0])
    current_trace = 0
    update_infobox()
    set_GUI_state(tk.NORMAL)
    status["text"] += "\nDone!"
    status.update()


def import_settings_from_template():
        settings_sheet = xl_workbook["Colocalization and Parameters"]
        criteria = []
        for crit in range(9):
            row_data = []
            row_data.append(settings_sheet.cell(row=17+crit, column=3).value)
            row_data.append(settings_sheet.cell(row=17+crit, column=4).value)
            row_data.append(settings_sheet.cell(row=17+crit, column=5).value)

            criteria.append(row_data)

        keys = list(raw_fit_criteria)
        for ind, key in enumerate(keys):
            if criteria[ind][0] is not None:
                raw_fit_criteria[key][0] = criteria[ind][0]
            if criteria[ind][1] is not None:
                raw_fit_criteria[key][1] = criteria[ind][1]
            if criteria[ind][2] is not None:
                raw_fit_criteria[key][2] = criteria[ind][2]

        if settings_sheet.cell(row=30, column=3).value is not None:
            preferences["Sum view frame fraction"][0] = settings_sheet.cell(row=30, column=3).value
        if settings_sheet.cell(row=30, column=4).value is not None:
            preferences["Sum view frame fraction"][1] = settings_sheet.cell(row=30, column=4).value
        if settings_sheet.cell(row=30, column=5).value is not None:
            preferences["Sum view frame fraction"][2] = settings_sheet.cell(row=30, column=5).value
        if settings_sheet.cell(row=31, column=3).value is not None:
            preferences["Intensity target"][0] = settings_sheet.cell(row=31, column=3).value
        if settings_sheet.cell(row=31, column=4).value is not None:
            preferences["Intensity target"][1] = settings_sheet.cell(row=31, column=4).value
        if settings_sheet.cell(row=31, column=5).value is not None:
            preferences["Intensity target"][2] = settings_sheet.cell(row=31, column=5).value

        calib = []

        for cal in range(8):
            calib.append(settings_sheet.cell(row=36+cal, column=3).value)

        keys = list(calibration)

        for ind, key in enumerate(keys):
            if calib[ind] is not None:
                calibration[key] = calib[ind]


def update_infobox():
    global trace_info
    global current_trace
    global current_fluorophore
    global current_step_count
    global current_trace_length
    global active_trace_list
    trace_label = trace_info[active_trace_list[current_trace]]
    current_fluorophore = trace_label[0]
    current_step_count = trace_label[1]
    current_trace_length = trace_label[2]
    step_label = current_step_count
    if current_step_count == 5:
        step_label = "≥5"
    current_info["text"] = "Current trace: " + str(current_trace + 1) + " / " + str(len(active_trace_list)) + \
                           " in current selection" + "\nFluorophore: " + str(internal_conversion[str(current_fluorophore)]()) + "\nStep count: " + \
                           str(step_label) + "\nTrace length: " + str(current_trace_length) + \
                           "\n\nCurrent file: " + str(file_path)
    if len(trace_label) > 4:
        positions = trace_label[4]
        current_info["text"] = "Current trace: " + str(current_trace + 1) + " / " + str(len(active_trace_list)) + \
                               " in current selection" + "\nFluorophore: " + str(internal_conversion[str(current_fluorophore)]()) + "\nStep count: " + \
                               str(step_label) + "\nTrace length: " + str(current_trace_length) + \
                               "\nStep positions: " + str(positions) + "\n\nCurrent file: " + str(file_path)
    if len(trace_label) > 5:
        positions = trace_label[4]
        weighted_act = trace_label[5]
        conf = trace_label[6]
        mean_dev = trace_label[7]
        heights = trace_label[8]
        current_info["text"] = "Current trace: " + str(current_trace + 1) + " / " + str(len(active_trace_list)) + \
                               " in current selection" + "\nFluorophore: " + str(internal_conversion[str(current_fluorophore)]()) + "\nStep count: " + \
                               str(step_label) + "\nTrace length: " + str(current_trace_length) + \
                               "\nStep positions: " + str(positions) + "\nStep heights: " + str(heights) + \
                               "\nWeighted activation: " + str(weighted_act) + \
                               "\nPrediction Confidence: " + str(conf) + "\nOutput neuron mean deviation: " + \
                               str(mean_dev) + "\nCurrent file: " + str(file_path)
    current_info.update()
    step_box["text"] = str(current_step_count)
    if current_step_count == "Partially bleached":
        step_box["text"] = "P"
    if current_step_count == 5:
        step_box["text"] = "≥5"
    step_box.update()
    pass


def plot_trace(trace_index):
    global trace_canvas
    global preferences
    global auto_cor
    global integral
    integral = False
    integration_button["relief"] = tk.RAISED
    integration_button["bg"] = button_bg
    integration_button["fg"] = button_fg
    auto_cor = False
    # correlation_button["relief"] = tk.RAISED
    # correlation_button["bg"] = "#eeeeee"
    trace_figure.clf()
    trace_canvas.draw()
    trace_fig2.clf()
    trace_canv2.draw()
    trace_label = trace_info[trace_index]
    trace_series = all_traces[trace_index]
    try:
        fret_label = trace_info[trace_index + 1]
        fret_trace = all_traces[trace_index + 1]
    except IndexError:
        print("No FRET trace!")
    if len(all_fits) > 0:
        fit_series = all_fits[trace_index]

    backup_plot = trace_fig2.add_subplot(111)
    backup_plot.set_facecolor("#222222")
    backup_plot.spines['bottom'].set_color('blue')
    backup_plot.spines['top'].set_color('blue')
    backup_plot.spines['left'].set_color('blue')
    backup_plot.spines['right'].set_color('blue')
    backup_plot.xaxis.label.set_color('white')
    backup_plot.yaxis.label.set_color('white')
    backup_plot.tick_params(axis='x', colors='white')
    backup_plot.tick_params(axis='y', colors='white')
    backup_plot.grid(color="#333333")
    if not fret_mode:
        backup_plot.plot(trace_series, linewidth=1)
    else:
        backup_plot.plot(fret_trace, linewidth=1)
        backup_plot.set_title("Trace "+str(trace_index + 2)+", " + internal_conversion[str(fret_label[0])](), color="white")
    if len(all_fits) > 0 and not fret_mode:
        backup_plot.plot(fit_series, linewidth=1)

    plot_area = trace_figure.add_subplot(111)
    plot_area.set_facecolor("#222222")
    plot_area.spines['bottom'].set_color('blue')
    plot_area.spines['top'].set_color('blue')
    plot_area.spines['left'].set_color('blue')
    plot_area.spines['right'].set_color('blue')
    plot_area.xaxis.label.set_color('white')
    plot_area.yaxis.label.set_color('white')
    plot_area.tick_params(axis='x', colors='white')
    plot_area.tick_params(axis='y', colors='white')
    plot_area.grid(color="#333333")

    if not preferences["Subsampling"] and not kernel:
        plot_area.plot(trace_series)
        if len(all_fits) > 0 and show_fit:
            plot_area.plot(fit_series)
        if waiting_for_fit:
            plot_area.axvline(vertical_position, color="w", linestyle="--")
    elif preferences["Subsampling"] and not kernel:
        sub_trace_1 = np.array(trace_series[slice(0, -1, 2)])
        sub_trace_2 = np.array(trace_series[slice(1, -1, 2)])
        if len(all_fits) > 0:
            sub_fit = fit_series[slice(0, -1, 2)]
        if len(sub_trace_2) < len(sub_trace_1):
            sub_trace_2 = np.append(sub_trace_2, 0)
        sub_trace = np.array((sub_trace_1 + sub_trace_2) / 2)
        plot_area.plot(sub_trace)
        if len(all_fits) > 0 and show_fit:
            plot_area.plot(sub_fit)
    elif kernel:
        if preferences["pre-gauss"]:
            trace_series = convolution_kernel(trace_series, "gauss")
        trace_series = convolution_kernel(trace_series, preferences["kernel"])
        plot_area.plot(trace_series)
        if len(all_fits) > 0 and show_fit:
            plot_area.plot(fit_series)
        if waiting_for_fit:
            plot_area.axvline(vertical_position, color="w", linestyle="--")
    plot_area.set_title("Trace "+str(trace_index + 1)+", " + internal_conversion[str(trace_label[0])]() + ", Step count: "+str(trace_label[1]), color="white")
    if trace_label[1] == 5:
        plot_area.set_title(
            "Trace " + str(trace_index + 1) + ", " + internal_conversion[str(trace_label[0])]() + ", Step count: ≥5", color="white")
    plot_area.set_xlabel("Time (frames)", color="white")
    plot_area.set_ylabel("Fluoroescence Intensity (Counts / A.U.)", color="white")
    trace_canv2.draw()
    trace_canvas.draw()


def convolution_kernel(trace_in, kern_type):
    _kernel = kernels[kern_type]
    convolved_trace = []
    for convolve in range(2, len(trace_in) - 2):
        c1 = trace_in[convolve - 2] * _kernel[0]
        c2 = trace_in[convolve - 1] * _kernel[1]
        c3 = trace_in[convolve] * _kernel[2]
        c4 = trace_in[convolve + 1] * _kernel[3]
        c5 = trace_in[convolve + 2] * _kernel[4]
        if sum(_kernel) != 0:
            c_total = (c1 + c2 + c3 + c4 + c5) / sum(_kernel)
        else:
            c_total = (c1 + c2 + c3 + c4 + c5)
        convolved_trace.append(c_total)
    return convolved_trace


def change_filter():
    global active_trace_list
    global current_trace
    global fluoro_filter
    global filter_win
    filter_choices = [preferences["Fluorophore config"][0], preferences["Fluorophore config"][1], preferences["Fluorophore config"][2], "Advanced Filter"]
    set_GUI_state(tk.DISABLED)
    selections = easygui.multchoicebox(title="Filter by fluorophore", msg="Select fluorophore(s) to filter by",
                                       choices=filter_choices)
    set_GUI_state(tk.NORMAL)
    if selections:
        active_trace_list = []
        bad_selection = False
        if preferences["Fluorophore config"][0] in selections and Cy5_trace_count > 0:
            active_trace_list += Cy5_sublist
        elif preferences["Fluorophore config"][0] in selections and Cy5_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="No "+preferences["Fluorophore config"][0]+" traces in dataset.")
            bad_selection = True
        if preferences["Fluorophore config"][1] in selections and mCherry_trace_count > 0:
            active_trace_list += mCherry_sublist
        elif preferences["Fluorophore config"][1] in selections and mCherry_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="No "+preferences["Fluorophore config"][1]+" traces in dataset.")
            bad_selection = True
        if preferences["Fluorophore config"][2] in selections and GFP_trace_count > 0:
            active_trace_list += GFP_sublist
        elif preferences["Fluorophore config"][2] in selections and GFP_trace_count == 0:
            easygui.msgbox(title="Warning!", msg="No "+preferences["Fluorophore config"][2]+" traces in dataset.")
            bad_selection = True
        if "Advanced Filter" in selections:
            active_trace_list = []
            filter_win = FilterWin()
            return
        if bad_selection:
            active_trace_list = all_sublist
            status["text"] = "Unable to update fluorophore filter due to invalid selection. Returning to default: ['All']"
            status.update()
        else:
            status["text"] = "Updated fluorophore filter to " + str(selections)
            status.update()
        fluoro_filter = selections
        active_trace_list.sort()
        current_trace = 0
        update_infobox()
        plot_trace(active_trace_list[current_trace])
    else:
        fluoro_filter = ['All']
        active_trace_list = list(all_sublist)
        current_trace = 0
        update_infobox()
        plot_trace(active_trace_list[current_trace])
        status["text"] = "Updated fluorophore filter to ['All']"
        status.update()


def make_correction():
    global trace_info
    set_GUI_state(tk.DISABLED)
    def_val = trace_info[active_trace_list[current_trace]][1]
    if def_val == "Partially bleached":
        def_val = "P"
    number_of_steps = easygui.enterbox(title="Amend step count", msg="Enter corrected step count. Set to 5 for 5 or more, set to P for 'partially bleached",
                                       default=None)
    set_GUI_state(tk.NORMAL)
    if number_of_steps != "P":
        try:
            num_steps = int(number_of_steps)
            if num_steps > 5:
                num_steps = 5
            trace_info[active_trace_list[current_trace]][1] = num_steps
            status["text"] = "Updated step count to " + str(num_steps)
            status.update()
            trace_info[active_trace_list[current_trace]][5] = "N/A, manually assigned"
            trace_info[active_trace_list[current_trace]][6] = "N/A, manually assigned"
            trace_info[active_trace_list[current_trace]][7] = "N/A, manually assigned"
        except:
            status["text"] = "Value entered is an invalid step count"
            status.update()
    else:
        trace_info[active_trace_list[current_trace]][1] = "Partially bleached"
        status["text"] = "Updated step count to 'Partially bleached'"
        status.update()
        trace_info[active_trace_list[current_trace]][5] = "N/A, manually assigned"
        trace_info[active_trace_list[current_trace]][6] = "N/A, manually assigned"
        trace_info[active_trace_list[current_trace]][7] = "N/A, manually assigned"
    update_infobox()
    plot_trace(active_trace_list[current_trace])


def trim_trace():
    set_GUI_state(tk.DISABLED)
    quantity = easygui.integerbox(title="Delete first n frames of trace",
                                  msg="Enter number of frames to remove from trace. value must be between 1 and 200",
                                  default=None, lowerbound=1, upperbound=200)
    set_GUI_state(tk.NORMAL)
    if quantity is not None:
        deleted_frames = []
        for delete in range(quantity):
            deleted_frames.append(all_traces[active_trace_list[current_trace]].pop(0))
        deleted_frames = list(reversed(deleted_frames))
        trim_undo_frames_stack.append(deleted_frames)
        trim_undo_reference_stack.append(active_trace_list[current_trace])
        undo_trim_button["state"] = tk.NORMAL
        trace_info[active_trace_list[current_trace]][2] = len(all_traces[active_trace_list[current_trace]])
        update_infobox()
        plot_trace(active_trace_list[current_trace])
        status["text"] = "Trimmed " + str(quantity) + " frames from the beginning of trace " + str(active_trace_list[current_trace] + 1)
        status["text"] += ".\nThe number of trim actions this session is " + str(len(trim_undo_reference_stack)) + ". These actions can be undone."
        status.update()


def undo_trim():
    if len(trim_undo_reference_stack) > 0:
        frames_to_restore = trim_undo_frames_stack.pop(-1)
        trace_index = trim_undo_reference_stack.pop(-1)
        for restore in range(len(frames_to_restore)):
            all_traces[trace_index].insert(0, frames_to_restore[restore])
        trace_info[trace_index][2] = len(all_traces[trace_index])
        update_infobox()
        plot_trace(active_trace_list[current_trace])
        status["text"] += "\nRestored " + str(len(frames_to_restore)) + " previously deleted frames to trace " + str(trace_index + 1)
        status["text"] += "\nThe number of trim actions this session is " + str(len(trim_undo_reference_stack)) + ". These actions can be undone."
        status.update()
    if len(trim_undo_reference_stack) == 0:
        undo_trim_button["state"] = tk.DISABLED


def auto_trim():
    noise_map = {
        "Cyanine 5": 10,
        "mCherry": 8,
        "GFP": 8,
    }

    global auto_trim_backup
    global auto_trim_fit_backup
    global trim_undo_frames_stack
    global trim_undo_reference_stack
    if len(auto_trim_backup) == 0:
        auto_trim_backup = list(all_traces)
        auto_trim_fit_backup = list(all_fits)

    for index in range(len(all_traces)):
        maximum_intensity = max(all_traces[index])
        trim_point = None
        noise_deviation = np.sqrt(maximum_intensity) * noise_map[trace_info[index][0]]
        threshold = maximum_intensity - noise_deviation
        for frame in range(0, int(len(all_traces[index]) / 6)):
            if all_traces[index][frame] > threshold:
                trim_point = frame
                break
        all_traces[index] = all_traces[index][trim_point:]
        all_fits[index] = all_fits[index][trim_point:]
        trace_info[index][2] = len(all_traces[index])
        undo_auto_trim_button["state"] = tk.NORMAL
        undo_trim_button["state"] = tk.DISABLED
        trim_undo_frames_stack = []
        trim_undo_reference_stack = []
    if ready_to_export:
        print(int(float(use_ai_var.get())))
        if int(float(use_ai_var.get())) == 1:
            ai_fit_trace()
        else:
            calculate_all_fits()
    status["text"] = "Used Smart trim function to automatically trim traces which begin below maximum. Previous trim" \
                     " actions can no longer be undone. Smart trim can be undone however any subsequent trim actions" \
                     " will also be discarded."
    status.update()
    plot_trace(active_trace_list[current_trace])
    update_infobox()


def undo_auto_trim():
    undo_auto_trim_button["state"] = tk.DISABLED
    global auto_trim_backup
    global auto_trim_fit_backup
    global trim_undo_reference_stack
    global trim_undo_frames_stack
    global all_traces, all_fits
    all_traces = list(auto_trim_backup)
    all_fits = list(auto_trim_fit_backup)
    for index in range(len(all_traces)):
        trace_info[index][2] = len(all_traces[index])
    auto_trim_backup = []
    auto_trim_fit_backup = []
    trim_undo_frames_stack = []
    trim_undo_reference_stack = []
    undo_trim_button["state"] = tk.DISABLED
    time.sleep(0.2)
    if ready_to_export:
        print(int(float(use_ai_var.get())))
        if int(float(use_ai_var.get())) == 1:
            ai_fit_trace()
        else:
            calculate_all_fits()
    status["text"] = "Smart trim has been undone."
    status.update()
    plot_trace(active_trace_list[current_trace])
    update_infobox()


def preference_update():
    setting = True
    while setting:
        set_GUI_state(tk.DISABLED)
        setting = easygui.choicebox(msg="Choose a preference to change.", title="Preferences", choices=list(preferences),
                                    preselect=0)

        if setting in preferences:
            preferences_function_map[setting]()
            save_preferences()

        if is_data_loaded:
            set_GUI_state(tk.NORMAL)
        load_pickle_button["state"] = tk.NORMAL
        load_button["state"] = tk.NORMAL
        preferences_button["state"] = tk.NORMAL
        open_tracking_button["state"] = tk.NORMAL


def next_trace():
    global current_trace
    global active_trace_list
    if current_trace < len(active_trace_list) - 1:
        current_trace += 1
        update_infobox()
        plot_trace(active_trace_list[current_trace])
    elif current_trace == len(active_trace_list) - 1:
        set_GUI_state(tk.DISABLED)
        easygui.msgbox(title="Warning!", msg="Last trace, can't continue")
        set_GUI_state(tk.NORMAL)


def previous_trace():
    global current_trace
    global active_trace_list
    if current_trace > 0:
        current_trace -= 1
        update_infobox()
        plot_trace(active_trace_list[current_trace])
    elif current_trace == 0:
        set_GUI_state(tk.DISABLED)
        easygui.msgbox(title="Warning!", msg="Earliest trace, can't go back further.")
        set_GUI_state(tk.NORMAL)


def skip_next():
    global current_trace
    global active_trace_list
    global preferences
    skip_amount = preferences["Trace skip number"]
    if len(active_trace_list) - 1 - current_trace >= skip_amount:
        current_trace += skip_amount
        update_infobox()
        plot_trace(active_trace_list[current_trace])
    else:
        current_trace = len(active_trace_list) - 1
        update_infobox()
        plot_trace(active_trace_list[current_trace])
        set_GUI_state(tk.DISABLED)
        easygui.msgbox(title="Warning!", msg="Gap between current trace and last trace in selection was less than skip"
                       " number. Skipped to last trace instead.")
        set_GUI_state(tk.NORMAL)


def skip_back():
    global current_trace
    global active_trace_list
    global preferences
    skip_amount = preferences["Trace skip number"]
    if current_trace >= skip_amount:
        current_trace -= skip_amount
        update_infobox()
        plot_trace(active_trace_list[current_trace])
    else:
        current_trace = active_trace_list[0]
        update_infobox()
        plot_trace(current_trace)
        set_GUI_state(tk.DISABLED)
        easygui.msgbox(title="Warning!", msg="Gap between current trace and first trace in selection was less than skip"
                       " number. Skipped to first trace instead.")
        set_GUI_state(tk.NORMAL)


def jump_to_trace():
    global active_trace_list
    global current_trace
    maximum_trace_in_active_list = max(active_trace_list)
    minimum_trace_in_active_list = min(active_trace_list)
    set_GUI_state(tk.DISABLED)
    selection = easygui.integerbox(title="Jump to trace", msg="Enter a trace number to jump to.",
                                   lowerbound=minimum_trace_in_active_list+1, upperbound=maximum_trace_in_active_list+1)
    set_GUI_state(tk.NORMAL)
    if selection:
        if (selection - 1) in active_trace_list:
            current_trace = active_trace_list.index(selection - 1)
            update_infobox()
            plot_trace(active_trace_list[current_trace])
        else:
            set_GUI_state(tk.DISABLED)
            easygui.msgbox(title="Warning!", msg="Invalid trace number. Trace may exist but is not selected by"
                           " current fluorophore filter.")
            set_GUI_state(tk.NORMAL)


def edit_trace_skip():
    global preferences
    global skip_forward_button
    global skip_back_button
    set_GUI_state(tk.DISABLED)
    new_amount = easygui.integerbox(msg="Enter number of traces to skip", title="Trace skip number",
                                    default=preferences["Trace skip number"], lowerbound=2, upperbound=100)
    set_GUI_state(tk.NORMAL)
    if new_amount:
        preferences["Trace skip number"] = new_amount
        skip_back_button["text"] = "<<-- "+str(preferences["Trace skip number"])
        skip_forward_button["text"] = str(preferences["Trace skip number"])+" -->>"
    status["text"] = "Updated trace skip amount to " + str(preferences["Trace skip number"])
    status.update()


def set_default_dir():
    global default_dir
    set_GUI_state(tk.DISABLED)
    new = easygui.diropenbox(msg="Select a default directory to load and save files from", default=default_dir)
    set_GUI_state(tk.NORMAL)
    if new:
        default_dir = new + "/"
    print(default_dir)
    status["text"] = "Changed default directory to '" + default_dir + "'"
    status.update()


def change_kernel():
    choice = easygui.choicebox(msg="Choose a convolution kernel", title="Kernels", choices=list(kernels),
                                preselect=1)
    if choice in kernels:
        preferences["kernel"] = choice
        plot_trace(active_trace_list[current_trace])


def set_gui_col():
    set_GUI_state(tk.DISABLED)
    global colours
    global col_window
    col_window = tk.Tk()
    col_window.title("Select GUI colour")
    inst_label = tk.Label(master=col_window, text="Select a colour", padx=10, pady=4)
    col1 = tk.Button(master=col_window, text="         ", padx=35, pady=20, bg=colours[0], command=lambda: setcol(0))
    col2 = tk.Button(master=col_window, text="         ", padx=35, pady=20, bg=colours[1], command=lambda: setcol(1))
    col3 = tk.Button(master=col_window, text="         ", padx=35, pady=20, bg=colours[2], command=lambda: setcol(2))
    col4 = tk.Button(master=col_window, text="         ", padx=35, pady=20, bg=colours[3], command=lambda: setcol(3))
    inst_label.grid(row=0, column=0, columnspan=2)
    col1.grid(row=1, column=0)
    col2.grid(row=1, column=1)
    col3.grid(row=2, column=0)
    col4.grid(row=2, column=1)
    col_window.protocol("WM_DELETE_WINDOW", on_close_col_win)
    col_window.mainloop()


def on_close_col_win():
    set_GUI_state(tk.NORMAL)
    col_window.quit()
    col_window.destroy()


def select_advanced():
    global advanced_window

    set_GUI_state(tk.DISABLED)
    fitting_button["state"] = tk.DISABLED
    current_fit_button["state"] = tk.DISABLED
    edit_fit_button["state"] = tk.DISABLED
    advanced_fit_button["state"] = tk.DISABLED
    edit_fit_button["relief"] = tk.RAISED
    edit_fit_button["bg"] = button_bg
    edit_fit_button["fg"] = button_fg
    trace_figure.canvas.mpl_disconnect(capture_mouse)

    advanced_window = tk.Tk()
    advanced_window.title("Advanced fit properties")
    advanced_window.geometry("350x250+400+400")
    f1 = tk.Frame(master=advanced_window, width=280, height=30)
    f2 = tk.Frame(master=advanced_window, width=280, height=100)
    f3 = tk.Frame(master=advanced_window, width=280, height=30)
    f1.pack(side=tk.TOP, padx=20, pady=0, fill="x")
    f2.pack(padx=20, pady=5, fill="x")
    f3.pack(side=tk.BOTTOM, padx=20, pady=10)

    fluoro_list = (preferences["Fluorophore config"][0], preferences["Fluorophore config"][0],
                   preferences["Fluorophore config"][1], preferences["Fluorophore config"][2])
    selected_fluoro = tk.StringVar(master=advanced_window)
    selected_fluoro.set(fluoro_list[0])
    f_menu = ttk.OptionMenu(f1, selected_fluoro, *fluoro_list)
    f_menu.pack(padx=10, pady=10, fill="x")

    tk.Label(master=f2, text="Window size").grid(row=1, column=0, padx=10, pady=1)
    tk.Label(master=f2, text="Threshold").grid(row=2, column=0, padx=10, pady=1)
    tk.Label(master=f2, text="Window stride").grid(row=3, column=0, padx=10, pady=1)
    size_entry = tk.Entry(master=f2)
    size_entry.grid(row=1, column=1, padx=10, pady=1)
    thresh_entry = tk.Entry(master=f2)
    thresh_entry.grid(row=2, column=1, padx=10, pady=1)
    stride_entry = tk.Entry(master=f2)
    stride_entry.grid(row=3, column=1, padx=10, pady=1)
    size_entry.insert(tk.END, str(preferences["Fitting tool window size"]))
    thresh_entry.insert(tk.END, str(preferences["Fitting tool threshold"]))
    stride_entry.insert(tk.END, str(preferences["Fitting tool window stride"]))

    conv_button = ttk.Checkbutton(master=f2, text="Pre-fit convolution")
    conv_button.grid(row=4, column=1, padx=10, pady=10)
    conv_button.state(["!alternate"])
    if preferences["Fit convolve"]:
        conv_button.state(["selected"])
    elif not preferences["Fit convolve"]:
        conv_button.state(["!selected"])

    fit_button = ttk.Button(master=f3, text="Calculate fits", command=lambda: check_param_valid(size_entry, thresh_entry, stride_entry, conv_button, selected_fluoro))
    cancel_button = ttk.Button(master=f3, text="Close", command=on_close_advanced)
    fit_button.pack(padx=20, pady=5, side=tk.LEFT)
    cancel_button.pack(padx=20, pady=5, side=tk.RIGHT)

    advanced_window.protocol("WM_DELETE_WINDOW", on_close_advanced)


def check_param_valid(size, threshold, stride, conv, selected):
    try:
        window_size = int(size.get())
        if window_size < 2:
            window_size = 2
            size.delete(0, tk.END)
            size.insert(tk.END, str(window_size))
        if window_size > 10:
            window_size = 10
            size.delete(0, tk.END)
            size.insert(tk.END, str(window_size))
    except:
        print("Invalid window size entered")
        easygui.msgbox(title="Warning!", msg="Invalid window size")
        return

    try:
        step_threshold = float(threshold.get())
        if step_threshold < 0.01:
            step_threshold = 0.01
            threshold.delete(0, tk.END)
            threshold.insert(tk.END, str(step_threshold))
        if step_threshold > 100:
            step_threshold = 100
            threshold.delete(0, tk.END)
            threshold.insert(tk.END, str(step_threshold))
    except:
        print("Invalid threshold entered")
        easygui.msgbox(title="Warning!", msg="Invalid threshold")
        return

    try:
        window_stride = int(stride.get())
        if window_stride < 1:
            window_stride = 1
            stride.delete(0, tk.END)
            stride.insert(tk.END, str(window_stride))
        if window_stride > 10:
            window_stride = 10
            stride.delete(0, tk.END)
            stride.insert(tk.END, str(window_stride))
    except:
        print("Invalid window stride entered")
        easygui.msgbox(title="Warning!", msg="Invalid window stride")
        return

    advanced_fit(window_size, step_threshold, window_stride, conv, selected)


def on_close_advanced():
    set_GUI_state(tk.NORMAL)
    fitting_button["state"] = tk.NORMAL
    current_fit_button["state"] = tk.NORMAL
    edit_fit_button["state"] = tk.NORMAL
    advanced_fit_button["state"] = tk.NORMAL
    try:
        advanced_window.destroy()
    except:
        """ Failed """


def setcol(col_index):
    global col_window
    preferences["GUI colours"] = colours[col_index]
    gui_space["bg"] = preferences["GUI colours"]
    gui_space.update()
    col_window.quit()
    col_window.destroy()
    set_GUI_state(tk.NORMAL)


def set_subsample():
    global preferences
    choices = ["ENABLE", "DISABLE"]
    set_GUI_state(tk.DISABLED)
    choice = easygui.choicebox(title="Enable / Disable subsampling", msg="Please select an option.", choices=choices,
                               preselect=1)
    set_GUI_state(tk.NORMAL)
    if choice:
        if choice == choices[0]:
            preferences["Subsampling"] = True
            # subsample_button["relief"] = tk.SUNKEN
            # subsample_button["bg"] = "#22ff55"
            # subsample_button["text"] = "Disable Subsampling"
        elif choice == choices[1]:
            preferences["Subsampling"] = False
            # subsample_button["relief"] = tk.RAISED
            # subsample_button["bg"] = "#eeeeee"
            # subsample_button["text"] = " Enable Subsampling"
        plot_trace(active_trace_list[current_trace])


def toggle_subsampling():
    global preferences
    if not preferences["Subsampling"]:
        preferences["Subsampling"] = True
        # subsample_button["relief"] = tk.SUNKEN
        # subsample_button["bg"] = "#22ff55"
        # subsample_button["text"] = "Disable Subsampling"
        plot_trace(active_trace_list[current_trace])
        save_preferences()
        return
    if preferences["Subsampling"]:
        preferences["Subsampling"] = False
        # subsample_button["relief"] = tk.RAISED
        # subsample_button["bg"] = "#eeeeee"
        # subsample_button["text"] = " Enable Subsampling"
        plot_trace(active_trace_list[current_trace])
        save_preferences()


def display_stats():
    global stats_window
    set_GUI_state(tk.DISABLED)
    stats_window = tk.Tk()
    stats_window.title("Visible step statistics for 488nm, 561nm, and 640nm")
    graph1_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph2_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph3_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph1_frame.pack(side=tk.LEFT, padx=20, pady=20)
    graph3_frame.pack(side=tk.RIGHT, padx=20, pady=20)
    graph2_frame.pack(side=tk.RIGHT, padx=20, pady=20)

    figure_488 = plt.Figure(figsize=(5, 5), dpi=100)
    figure_561 = plt.Figure(figsize=(5, 5), dpi=100)
    figure_640 = plt.Figure(figsize=(5, 5), dpi=100)
    canvas_488 = FigureCanvasTkAgg(figure_488, master=graph1_frame)
    canvas_561 = FigureCanvasTkAgg(figure_561, master=graph2_frame)
    canvas_640 = FigureCanvasTkAgg(figure_640, master=graph3_frame)
    canvas_488.draw()
    canvas_561.draw()
    canvas_640.draw()
    toolbar640 = NavigationToolbar2Tk(canvas_640, graph3_frame)
    toolbar561 = NavigationToolbar2Tk(canvas_561, graph2_frame)
    toolbar488 = NavigationToolbar2Tk(canvas_488, graph1_frame)
    toolbar640.update()
    toolbar561.update()
    toolbar488.update()
    canvas_488.get_tk_widget().pack(side=tk.TOP)
    canvas_561.get_tk_widget().pack(side=tk.TOP)
    canvas_640.get_tk_widget().pack(side=tk.TOP)

    plot_488 = figure_488.add_subplot(1, 1, 1)
    plot_488.set_title("488nm - "+preferences["Fluorophore config"][2]+" Visible steps")
    plot_561 = figure_561.add_subplot(1, 1, 1)
    plot_561.set_title("561nm - "+preferences["Fluorophore config"][1]+" Visible steps")
    plot_640 = figure_640.add_subplot(1, 1, 1)
    plot_640.set_title("Marker - "+preferences["Fluorophore config"][0]+" Visible steps")

    x_lab = ["P.B.", "0", "1", "2", "3", "4", "5+"]

    if GFP_trace_count > 0:
        GFP_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in GFP_sublist:
            if trace_info[trace][1] == "Partially bleached":
                GFP_steps[0] += 1
            elif trace_info[trace][1] == 0:
                GFP_steps[1] += 1
            elif trace_info[trace][1] == 1:
                GFP_steps[2] += 1
            elif trace_info[trace][1] == 2:
                GFP_steps[3] += 1
            elif trace_info[trace][1] == 3:
                GFP_steps[4] += 1
            elif trace_info[trace][1] == 4:
                GFP_steps[5] += 1
            elif trace_info[trace][1] >= 5:
                GFP_steps[6] += 1

        np_percent_GFP = np.array(GFP_steps)
        np_percent_GFP = np_percent_GFP / GFP_trace_count
        np_percent_GFP = np_percent_GFP * 100
        plot_488.bar(x_lab, np_percent_GFP, color="green")
        plot_488.set_xlabel("Number of steps")
        plot_488.set_ylabel("Frequency (%)")

    if mCherry_trace_count > 0:
        mCherry_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in mCherry_sublist:
            if trace_info[trace][1] == "Partially bleached":
                mCherry_steps[0] += 1
            elif trace_info[trace][1] == 0:
                mCherry_steps[1] += 1
            elif trace_info[trace][1] == 1:
                mCherry_steps[2] += 1
            elif trace_info[trace][1] == 2:
                mCherry_steps[3] += 1
            elif trace_info[trace][1] == 3:
                mCherry_steps[4] += 1
            elif trace_info[trace][1] == 4:
                mCherry_steps[5] += 1
            elif trace_info[trace][1] >= 5:
                mCherry_steps[6] += 1

        np_percent_mCherry = np.array(mCherry_steps)
        np_percent_mCherry = np_percent_mCherry / mCherry_trace_count
        np_percent_mCherry = np_percent_mCherry * 100
        plot_561.bar(x_lab, np_percent_mCherry, color="orange")
        plot_561.set_xlabel("Number of steps")
        plot_561.set_ylabel("Frequency (%)")

    if Cy5_trace_count > 0:
        Cy5_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in Cy5_sublist:
            if trace_info[trace][1] == "Partially bleached":
                Cy5_steps[0] += 1
            elif trace_info[trace][1] == 0:
                Cy5_steps[1] += 1
            elif trace_info[trace][1] == 1:
                Cy5_steps[2] += 1
            elif trace_info[trace][1] == 2:
                Cy5_steps[3] += 1
            elif trace_info[trace][1] == 3:
                Cy5_steps[4] += 1
            elif trace_info[trace][1] == 4:
                Cy5_steps[5] += 1
            elif trace_info[trace][1] >= 5:
                Cy5_steps[6] += 1

        np_percent_Cy5 = np.array(Cy5_steps)
        np_percent_Cy5 = np_percent_Cy5 / Cy5_trace_count
        np_percent_Cy5 = np_percent_Cy5 * 100
        plot_640.bar(x_lab, np_percent_Cy5, color="red")
        plot_640.set_xlabel("Number of steps")
        plot_640.set_ylabel("Frequency (%)")

    try:
        plot_488.set_title("488nm - " + preferences["Fluorophore config"][2] + " Visible steps, N = " + str(GFP_trace_count))
    except:
        """ Failed """
    try:
        plot_561.set_title("561nm - " + preferences["Fluorophore config"][1] + " Visible steps, N = " + str(mCherry_trace_count))
    except:
        """ Failed """
    try:
        plot_640.set_title("Marker - " + preferences["Fluorophore config"][0] + " Visible steps, N = " + str(Cy5_trace_count))
    except:
        """ Failed """

    stats_window.protocol("WM_DELETE_WINDOW", close_stats_window)


def display_stats_filtered():
    global stats_window
    set_GUI_state(tk.DISABLED)
    stats_window = tk.Tk()
    stats_window.title("Visible step statistics for 488nm, 561nm, and 640nm")
    graph1_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph2_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph3_frame = tk.Frame(master=stats_window, width=500, height=500, bg=colours[3])
    graph1_frame.pack(side=tk.LEFT, padx=20, pady=20)
    graph3_frame.pack(side=tk.RIGHT, padx=20, pady=20)
    graph2_frame.pack(side=tk.RIGHT, padx=20, pady=20)

    figure_488 = plt.Figure(figsize=(5, 5), dpi=100)
    figure_561 = plt.Figure(figsize=(5, 5), dpi=100)
    figure_640 = plt.Figure(figsize=(5, 5), dpi=100)
    canvas_488 = FigureCanvasTkAgg(figure_488, master=graph1_frame)
    canvas_561 = FigureCanvasTkAgg(figure_561, master=graph2_frame)
    canvas_640 = FigureCanvasTkAgg(figure_640, master=graph3_frame)
    canvas_488.draw()
    canvas_561.draw()
    canvas_640.draw()
    toolbar640 = NavigationToolbar2Tk(canvas_640, graph3_frame)
    toolbar561 = NavigationToolbar2Tk(canvas_561, graph2_frame)
    toolbar488 = NavigationToolbar2Tk(canvas_488, graph1_frame)
    toolbar640.update()
    toolbar561.update()
    toolbar488.update()
    canvas_488.get_tk_widget().pack(side=tk.TOP)
    canvas_561.get_tk_widget().pack(side=tk.TOP)
    canvas_640.get_tk_widget().pack(side=tk.TOP)

    plot_488 = figure_488.add_subplot(1, 1, 1)
    plot_488.set_title("488nm - "+preferences["Fluorophore config"][2]+" Visible steps")
    plot_561 = figure_561.add_subplot(1, 1, 1)
    plot_561.set_title("561nm - "+preferences["Fluorophore config"][1]+" Visible steps")
    plot_640 = figure_640.add_subplot(1, 1, 1)
    plot_640.set_title("Marker - "+preferences["Fluorophore config"][0]+" Visible steps")

    x_lab = ["P.B.", "0", "1", "2", "3", "4", "5+"]

    print(len(GFP_sublist))
    if GFP_trace_count > 0:
        GFP_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in GFP_sublist:
            if trace - 1 >= 0:
                if trace_info[trace - 1][0] == "Cyanine 5" and trace_info[trace - 1][1] == 1:
                    if trace_info[trace][1] == "Partially bleached":
                        GFP_steps[0] += 1
                    elif trace_info[trace][1] == 0:
                        GFP_steps[1] += 1
                    elif trace_info[trace][1] == 1:
                        GFP_steps[2] += 1
                    elif trace_info[trace][1] == 2:
                        GFP_steps[3] += 1
                    elif trace_info[trace][1] == 3:
                        GFP_steps[4] += 1
                    elif trace_info[trace][1] == 4:
                        GFP_steps[5] += 1
                    elif trace_info[trace][1] >= 5:
                        GFP_steps[6] += 1
        NUM_GFP = np.sum(GFP_steps)
        np_percent_GFP = np.array(GFP_steps)
        np_percent_GFP = np_percent_GFP / NUM_GFP
        np_percent_GFP = np_percent_GFP * 100
        plot_488.bar(x_lab, np_percent_GFP, color="green")
        plot_488.set_xlabel("Number of steps")
        plot_488.set_ylabel("Frequency (%)")

    if mCherry_trace_count > 0:
        mCherry_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in mCherry_sublist:
            if trace - 1 > 0:
                if trace_info[trace - 1][0] == "Cyanine 5" and trace_info[trace - 1][1] == 1:
                    if trace_info[trace][1] == "Partially bleached":
                        mCherry_steps[0] += 1
                    elif trace_info[trace][1] == 0:
                        mCherry_steps[1] += 1
                    elif trace_info[trace][1] == 1:
                        mCherry_steps[2] += 1
                    elif trace_info[trace][1] == 2:
                        mCherry_steps[3] += 1
                    elif trace_info[trace][1] == 3:
                        mCherry_steps[4] += 1
                    elif trace_info[trace][1] == 4:
                        mCherry_steps[5] += 1
                    elif trace_info[trace][1] >= 5:
                        mCherry_steps[6] += 1
            if trace - 2 >= 0:
                if trace_info[trace - 2][0] == "Cyanine 5" and trace_info[trace - 2][1] == 1:
                    if trace_info[trace][1] == "Partially bleached":
                        mCherry_steps[0] += 1
                    elif trace_info[trace][1] == 0:
                        mCherry_steps[1] += 1
                    elif trace_info[trace][1] == 1:
                        mCherry_steps[2] += 1
                    elif trace_info[trace][1] == 2:
                        mCherry_steps[3] += 1
                    elif trace_info[trace][1] == 3:
                        mCherry_steps[4] += 1
                    elif trace_info[trace][1] == 4:
                        mCherry_steps[5] += 1
                    elif trace_info[trace][1] >= 5:
                        mCherry_steps[6] += 1
        NUM_mCherry = np.sum(mCherry_steps)
        np_percent_mCherry = np.array(mCherry_steps)
        np_percent_mCherry = np_percent_mCherry / NUM_mCherry
        np_percent_mCherry = np_percent_mCherry * 100
        plot_561.bar(x_lab, np_percent_mCherry, color="orange")
        plot_561.set_xlabel("Number of steps")
        plot_561.set_ylabel("Frequency (%)")

    if Cy5_trace_count > 0:
        Cy5_steps = [0, 0, 0, 0, 0, 0, 0]
        for trace in Cy5_sublist:
            if trace_info[trace][1] == "Partially bleached":
                Cy5_steps[0] += 1
            elif trace_info[trace][1] == 0:
                Cy5_steps[1] += 1
            elif trace_info[trace][1] == 1:
                Cy5_steps[2] += 1
            elif trace_info[trace][1] == 2:
                Cy5_steps[3] += 1
            elif trace_info[trace][1] == 3:
                Cy5_steps[4] += 1
            elif trace_info[trace][1] == 4:
                Cy5_steps[5] += 1
            elif trace_info[trace][1] >= 5:
                Cy5_steps[6] += 1

        np_percent_Cy5 = np.array(Cy5_steps)
        np_percent_Cy5 = np_percent_Cy5 / Cy5_trace_count
        np_percent_Cy5 = np_percent_Cy5 * 100
        plot_640.bar(x_lab, np_percent_Cy5, color="red")
        plot_640.set_xlabel("Number of steps")
        plot_640.set_ylabel("Frequency (%)")

    try:
        plot_488.set_title("488nm - " + preferences["Fluorophore config"][2] + " Visible steps, N = " + str(NUM_GFP))
    except:
        """ Failed """
    try:
        plot_561.set_title("561nm - " + preferences["Fluorophore config"][1] + " Visible steps, N = " + str(NUM_mCherry))
    except:
        """ Failed """
    try:
        plot_640.set_title("Marker - " + preferences["Fluorophore config"][0] + " Visible steps, N = " + str(Cy5_trace_count))
    except:
        """ Failed """

    stats_window.protocol("WM_DELETE_WINDOW", close_stats_window)


def close_stats_window():
    global stats_window
    set_GUI_state(tk.NORMAL)
    stats_window.destroy()


def box_plot(data, edge_color, fill_color, labels, pl):
    global plot_area
    bp = pl.boxplot(data, patch_artist=True, labels=labels)

    for element in ['boxes', 'whiskers', 'fliers', 'means', 'medians', 'caps']:
        plt.setp(bp[element], color=edge_color)

    for patch in bp['boxes']:
        patch.set(facecolor=fill_color)

    return bp


def plot_snr():

    minimum_points = easygui.integerbox(msg="Enter minimum plateau length", lowerbound=3, upperbound=500)
    if not minimum_points:
        return

    traces640 = [all_traces[idx] for idx in Cy5_sublist]
    traces561 = [all_traces[idx] for idx in mCherry_sublist]
    traces488 = [all_traces[idx] for idx in GFP_sublist]

    info640 = [trace_info[idx] for idx in Cy5_sublist]
    info561 = [trace_info[idx] for idx in mCherry_sublist]
    info488 = [trace_info[idx] for idx in GFP_sublist]

    show_invalids = False
    aSNRs_640 = []
    for index in range(len(traces640)):
        val = asnr.calculate_asnr(traces640[index], info640[index], min_length=minimum_points)
        if show_invalids:
            aSNRs_640.append(val)
        else:
            if val:
                aSNRs_640.append(val)
        print(f"                    aSNR: {val}")
        print(f"Trace Number (index + 1): {index + 1}")
        print()

    show_invalids = False
    aSNRs_561 = []
    ref_561 = []
    for index in range(len(traces561)):
        val = asnr.calculate_asnr(traces561[index], info561[index], min_length=minimum_points)
        if show_invalids:
            aSNRs_561.append(val)
        else:
            if val:
                aSNRs_561.append(val)
                ref_561.append(index)
        print(f"                    aSNR: {val}")
        print(f"Trace Number (index + 1): {index + 1}")
        print()

    show_invalids = False
    aSNRs_488 = []
    for index in range(len(traces488)):
        val = asnr.calculate_asnr(traces488[index], info488[index], min_length=minimum_points)
        if show_invalids:
            aSNRs_488.append(val)
        else:
            if val:
                aSNRs_488.append(val)
        print(f"                    aSNR: {val}")
        print(f"Trace Number (index + 1): {index + 1}")
        print()

    fig = plt.figure(figsize=(8, 6), dpi=100)
    fig.subplots_adjust(top=0.95, bottom=0.07, left=0.09, right=0.98, wspace=0.3, hspace=0.3)

    plot_area = fig.add_subplot(1, 1, 1)
    plot_area.tick_params(axis='x', colors='black', labelsize=14)
    plot_area.tick_params(axis='y', colors='black', labelsize=12)
    plot_area.set_ylabel("aSNR", color="black", size=14)
    plot_area.set_title("aSNR distribution of traces", color="black", size=16)
    box_plot([aSNRs_640, aSNRs_561, aSNRs_488], "black", "#ccccff", labels=[f"Marker: {preferences['Fluorophore config'][0]}", f"561nm: {preferences['Fluorophore config'][1]}", f"488nm: {preferences['Fluorophore config'][2]}"], pl=plot_area)
    plt.show()


# def auto_correlation():
#     global auto_cor
#     if not auto_cor:
#         auto_cor = True
#         correlation_button["relief"] = tk.SUNKEN
#         correlation_button["bg"] = "#22ff55"
#         trace = all_traces[active_trace_list[current_trace]]
#         trace_index = active_trace_list[current_trace]
#         trace_label = trace_info[active_trace_list[current_trace]]
#         auto_correl = []
#         trace_square_mean = (sum(trace) / len(trace)) ** 2
#         for tau in range(len(trace) - 1):
#             count = 0
#             tau_sum = 0
#             for frame in range(len(trace)):
#                 if frame + tau < len(trace):
#                     count += 1
#                     tau_sum += trace[frame] * (trace[frame + tau])
#             auto_correl.append((tau_sum / count) / trace_square_mean)
#         trace_figure.clf()
#         trace_canvas.draw()
#         plot_area = trace_figure.add_subplot(1, 1, 1)
#         plot_area.plot(auto_correl)
#         plot_area.set_title("Autocorrelation "+str(trace_index + 1)+", "+str(trace_label[0]) + ", Step count: "+str(trace_label[1]))
#         plot_area.set_xlabel("Tau (frames)")
#         plot_area.set_ylabel("Correlation (A.U.)")
#         trace_canvas.draw()
#         return
#     else:
#         auto_cor = False
#         correlation_button["relief"] = tk.RAISED
#         correlation_button["bg"] = "#eeeeee"
#         plot_trace(active_trace_list[current_trace])


def integration():
    global integral
    global subtraction
    if not integral:
        integral = True
        integration_button["relief"] = tk.SUNKEN
        integration_button["bg"] = "#22ff55"
        integration_button["fg"] = "black"
        trace = all_traces[active_trace_list[current_trace]]
        trace_index = active_trace_list[current_trace]
        trace_label = trace_info[active_trace_list[current_trace]]
        integ = []
        integ.append(trace[0])
        for add in range(len(trace)):
            integ.append(integ[-1]+(trace[add] - subtraction))
        trace_figure.clf()
        trace_canvas.draw()
        plot_area = trace_figure.add_subplot(1, 1, 1)
        plot_area.set_facecolor("#222222")
        plot_area.spines['bottom'].set_color('blue')
        plot_area.spines['top'].set_color('blue')
        plot_area.spines['left'].set_color('blue')
        plot_area.spines['right'].set_color('blue')
        plot_area.xaxis.label.set_color('white')
        plot_area.yaxis.label.set_color('white')
        plot_area.tick_params(axis='x', colors='white')
        plot_area.tick_params(axis='y', colors='white')
        plot_area.grid(color="#333333")
        plot_area.plot(integ)
        plot_area.set_title("Integration "+str(trace_index + 1)+", "+internal_conversion[str(trace_label[0])]() +
                            ", Step count: "+str(trace_label[1])+" Subtraction = " +str(subtraction), color="white")
        plot_area.set_xlabel("Trace coordinate (frames)")
        plot_area.set_ylabel("Integrated trace (A.U.)")
        trace_canvas.draw()
        integ_inc_button["state"] = tk.NORMAL
        integ_dec_button["state"] = tk.NORMAL
        integ_sub_reset_button["state"] = tk.NORMAL
        return
    else:
        integral = False
        integration_button["relief"] = tk.RAISED
        integration_button["bg"] = button_bg
        integration_button["fg"] = button_fg
        plot_trace(active_trace_list[current_trace])
        integ_inc_button["state"] = tk.DISABLED
        integ_dec_button["state"] = tk.DISABLED
        integ_sub_reset_button["state"] = tk.DISABLED


def integration_inc():
    global subtraction
    subtraction += preferences["Subtraction amount"]
    integration()
    integration()


def integration_dec():
    global subtraction
    subtraction -= preferences["Subtraction amount"]
    integration()
    integration()


def integ_subtract_reset():
    global subtraction
    subtraction = 0
    integration()
    integration()


def change_subtraction():
    sub_amt = easygui.integerbox(msg="Change subtraction amount for trace integration", title="Change subtraction",
                                 default=preferences["Subtraction amount"], lowerbound=1, upperbound=100)
    preferences["Subtraction amount"] = sub_amt


def integral_scan():
    trace = all_traces[active_trace_list[current_trace]]
    trace_index = active_trace_list[current_trace]
    trace_label = trace_info[active_trace_list[current_trace]]
    max_bin = []
    for bin in range(len(trace) + 1):
        max_bin.append(0)
    for sub in range(20, int(max(trace)/1.75), 1):
        integ = []
        integ.append(trace[0])
        for add in range(len(trace)):
            integ.append(integ[-1] + (trace[add] - sub))
        try:
            max_idx = integ.index(max(integ))
            diff = integ[max_idx] - integ[max_idx + 7]

            test_idx = -8
            diff_ref = integ[test_idx] - integ[test_idx + 7]
            if diff > diff_ref / 4:
                max_bin[integ.index(max(integ))] += 1

        except:
            pass

    trace_figure.clf()
    trace_canvas.draw()
    plot_area = trace_figure.add_subplot(1, 1, 1)
    plot_area.plot(max_bin)
    plot_area.set_title("Integral maxima scan " + str(trace_index + 1) + ", " + str(trace_label[0]) + ", Step count: " +
                        str(trace_label[1]))
    plot_area.set_xlabel("Trace coordinate (frames)")
    plot_area.set_ylabel("Frequency of maxima over scan range(A.U.)")
    trace_canvas.draw()


def sigmoid(value):
    _sigmoid = 1 / (1 + np.exp(-value))
    return _sigmoid


def convolution():
    global kernel
    if kernel:
        kernel_button["relief"] = tk.RAISED
        kernel_button["bg"] = button_bg
        kernel_button["fg"] = button_fg
        kernel = False
        plot_trace(active_trace_list[current_trace])
        return
    else:
        kernel = True
        kernel_button["relief"] = tk.SUNKEN
        kernel_button["bg"] = "#22ff55"
        kernel_button["fg"] = "black"
        plot_trace(active_trace_list[current_trace])


def pregauss():
    choice = easygui.choicebox(msg="Enable or disable pre-gauss convolution kernel", title="Preferences",
                               choices=["Enable", "Disable"])
    if choice == "Enable":
        preferences["pre-gauss"] = True
        plot_trace(active_trace_list[current_trace])
    if choice == "Disable":
        preferences["pre-gauss"] = False
        plot_trace(active_trace_list[current_trace])


def neural_detection():
    global used_neural_network
    set_GUI_state(tk.DISABLED)
    confirmation = easygui.ccbox(msg="Use neural network to assign steps? Warning this action will replace all labels with CNN predictions and cannot be undone.",
                                 choices=('Proceed', 'Abort'), title="Deep convolutional neural network step detection V6.6")
    if confirmation:
        status["text"] = "Building tensorflow datasets from traces on per fluorophore basis..."
        status.update()

        dir_488 = preferences["488nm model dir"]
        dir_561 = preferences["561nm model dir"]
        dir_640 = preferences["640nm model dir"]

        dir_488 = dir_488[:11]
        dir_561 = dir_561[:11]
        dir_640 = dir_640[:11]

        model_resampling_dict = {
            "488nm_300fr": 300,
            "561nm_300fr": 300,
            "640nm_100fr": 100,
            "640nm_300fr": 300,
            "640nm_500fr": 500,
            "640nm_800fr": 800,
            "100ms_300fr": 300,
        }

        if not preferences["Custom NeuralNet integration"]:
            try:
                domain_size_488 = model_resampling_dict[dir_488]
                domain_size_561 = model_resampling_dict[dir_561]
                domain_size_640 = model_resampling_dict[dir_640]
            except KeyError:
                domain_size_488 = preferences["Custom NeuralNet settings"][2]
                domain_size_561 = preferences["Custom NeuralNet settings"][1]
                domain_size_640 = preferences["Custom NeuralNet settings"][0]
                easygui.msgbox(title="Warning!", msg="Neural network not recogized as internal model. Defaulting to "
                                                     "Resampling lengths provided in Custom NeuralNet settings.")
        else:
            domain_size_488 = preferences["Custom NeuralNet settings"][2]
            domain_size_561 = preferences["Custom NeuralNet settings"][1]
            domain_size_640 = preferences["Custom NeuralNet settings"][0]

        status["text"] += "\nTraces in "+preferences["Fluorophore config"][0]+" channel will be resampled to " + str(domain_size_640) + " frames."
        status["text"] += "\nTraces in "+preferences["Fluorophore config"][1]+" channel will be resampled to " + str(domain_size_561) + " frames."
        status["text"] += "\nTraces in "+preferences["Fluorophore config"][2]+" channel will be resampled to " + str(domain_size_488) + " frames.\n"

        BATCH = 100
        if len(Cy5_sublist) > 0:
            Cy5_pre_dataset = []
            for trace in Cy5_sublist:
                Cy5_pre_dataset.append(all_traces[trace])

            domain_size = domain_size_640
            for index in range(len(Cy5_pre_dataset)):
                tr_y = Cy5_pre_dataset[index]
                tr_x = np.arange(len(tr_y))
                interpolator = interp1d(tr_x, tr_y, kind="linear")
                scale_factor = domain_size / (len(tr_y) - 1)
                resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
                if len(resampled) > domain_size:
                    resampled = resampled[0:domain_size]
                Cy5_pre_dataset[index] = resampled

            for norm in range(len(Cy5_pre_dataset)):
                scaling = max(Cy5_pre_dataset[norm])
                for frame in range(len(Cy5_pre_dataset[norm])):
                    Cy5_pre_dataset[norm][frame] = Cy5_pre_dataset[norm][frame] / scaling

            if "CLDNN" in dir_640:
                for norm in range(len(Cy5_pre_dataset)):
                    minimum = min(Cy5_pre_dataset[norm])
                    for frame in range(len(Cy5_pre_dataset[norm])):
                        Cy5_pre_dataset[norm][frame] = Cy5_pre_dataset[norm][frame] - minimum
                for norm in range(len(Cy5_pre_dataset)):
                    scaling_factor = max(Cy5_pre_dataset[norm])
                    for frame in range(len(Cy5_pre_dataset[norm])):
                        Cy5_pre_dataset[norm][frame] = (Cy5_pre_dataset[norm][frame] / scaling_factor) * 999 + 1

            tf_predict_640 = tf.data.Dataset.from_tensor_slices((Cy5_pre_dataset))
            tf_predict_640 = tf_predict_640.batch(BATCH)
            try:
                predictions_640 = model_640.predict(x=tf_predict_640)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                status["text"] = "Neural network step detection failed. Check command line. Error is likely due to" + \
                                 " mismatch between model input and dataset tensor shape. Please make sure the" + \
                                 " correct neural network model has been assigned to the right fluorophore. Cy 5" + \
                                 " model expects input tensor with 100 features."
                status.update()
                set_GUI_state(tk.NORMAL)

            for i in range(len(predictions_640)):
                activations = list(predictions_640[i])
                int_pred = activations.index(max(activations))
                if len(activations) > 5:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4 + \
                             activations[5] * 5
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4] + \
                            activations[5]
                else:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4]
                weighted_sum = weight / w_sum
                e_sum = 0
                a_sum = 0
                for add in range(len(activations)):
                    if add != int_pred:
                        e_sum += activations[add] * abs(add - int_pred)
                for add in range(len(activations)):
                    a_sum += activations[add]
                confidence = activations[int_pred] / e_sum
                mean_dev = e_sum / a_sum
                normal_confidence = (sigmoid(np.log(confidence)))
                if int_pred == 6:
                    int_pred = "Partially bleached"

                trace_info[Cy5_sublist[i]][5] = round(weighted_sum, 3)
                trace_info[Cy5_sublist[i]][6] = round(normal_confidence, 2)
                trace_info[Cy5_sublist[i]][7] = round(mean_dev, 4)
                trace_info[Cy5_sublist[i]][1] = int_pred

            status["text"] += "\nUsed model '" + preferences["640nm model dir"] + "' to predict "+preferences["Fluorophore config"][0]+" steps"
            status.update()

        if len(mCherry_sublist) > 0:
            mCherry_pre_dataset = []
            for trace in mCherry_sublist:
                mCherry_pre_dataset.append(all_traces[trace])

            domain_size = domain_size_561
            for index in range(len(mCherry_pre_dataset)):
                tr_y = mCherry_pre_dataset[index]
                tr_x = np.arange(len(tr_y))
                interpolator = interp1d(tr_x, tr_y, kind="linear")
                scale_factor = domain_size / (len(tr_y) - 1)
                resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
                if len(resampled) > domain_size:
                    resampled = resampled[0:domain_size]
                mCherry_pre_dataset[index] = resampled

            for norm in range(len(mCherry_pre_dataset)):
                scaling = max(mCherry_pre_dataset[norm])
                for frame in range(len(mCherry_pre_dataset[norm])):
                    mCherry_pre_dataset[norm][frame] = mCherry_pre_dataset[norm][frame] / scaling

            if "CLDNN" in dir_561:
                for norm in range(len(mCherry_pre_dataset)):
                    minimum = min(mCherry_pre_dataset[norm])
                    for frame in range(len(mCherry_pre_dataset[norm])):
                        mCherry_pre_dataset[norm][frame] = mCherry_pre_dataset[norm][frame] - minimum
                for norm in range(len(mCherry_pre_dataset)):
                    scaling_factor = max(mCherry_pre_dataset[norm])
                    for frame in range(len(mCherry_pre_dataset[norm])):
                        mCherry_pre_dataset[norm][frame] = (mCherry_pre_dataset[norm][frame] / scaling_factor) * 999 + 1


            tf_predict_561 = tf.data.Dataset.from_tensor_slices((mCherry_pre_dataset))
            tf_predict_561 = tf_predict_561.batch(BATCH)
            try:
                predictions_561 = model_561.predict(x=tf_predict_561)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                status["text"] = "Neural network step detection failed. Check command line. Error is likely due to" + \
                                 " mismatch between model input and dataset tensor shape. Please make sure the" + \
                                 " correct neural network model has been assigned to the right fluorophore. mCherry" + \
                                 " model expects input tensor with 300 features."
                status.update()
                set_GUI_state(tk.NORMAL)

            for i in range(len(predictions_561)):
                activations = list(predictions_561[i])
                int_pred = activations.index(max(activations))
                if len(activations) > 5:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4 + \
                             activations[5] * 5
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4] + \
                            activations[5]
                else:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4]
                weighted_sum = weight / w_sum
                e_sum = 0
                a_sum = 0
                for add in range(len(activations)):
                    if add != int_pred:
                        e_sum += activations[add] * abs(add - int_pred)
                for add in range(len(activations)):
                    a_sum += activations[add]
                confidence = activations[int_pred] / e_sum
                mean_dev = e_sum / a_sum
                normal_confidence = (sigmoid(np.log(confidence)))
                if int_pred == 6:
                    int_pred = "Partially bleached"

                trace_info[mCherry_sublist[i]][5] = round(weighted_sum, 3)
                trace_info[mCherry_sublist[i]][6] = round(normal_confidence, 2)
                trace_info[mCherry_sublist[i]][7] = round(mean_dev, 4)
                trace_info[mCherry_sublist[i]][1] = int_pred

            status["text"] += "\nUsed model '" + preferences["561nm model dir"] + "' to predict "+preferences["Fluorophore config"][1]+" steps"
            status.update()

        if len(GFP_sublist) > 0:
            GFP_pre_dataset = []
            for trace in GFP_sublist:
                GFP_pre_dataset.append(all_traces[trace])

            domain_size = domain_size_488
            for index in range(len(GFP_pre_dataset)):
                tr_y = GFP_pre_dataset[index]
                tr_x = np.arange(len(tr_y))
                interpolator = interp1d(tr_x, tr_y, kind="linear")
                scale_factor = domain_size / (len(tr_y) - 1)
                resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
                if len(resampled) > domain_size:
                    resampled = resampled[0:domain_size]
                GFP_pre_dataset[index] = resampled

            for norm in range(len(GFP_pre_dataset)):
                scaling = max(GFP_pre_dataset[norm])
                for frame in range(len(GFP_pre_dataset[norm])):
                    GFP_pre_dataset[norm][frame] = GFP_pre_dataset[norm][frame] / scaling

            if "CLDNN" in dir_488:
                for norm in range(len(GFP_pre_dataset)):
                    minimum = min(GFP_pre_dataset[norm])
                    for frame in range(len(GFP_pre_dataset[norm])):
                        GFP_pre_dataset[norm][frame] = GFP_pre_dataset[norm][frame] - minimum
                for norm in range(len(GFP_pre_dataset)):
                    scaling_factor = max(GFP_pre_dataset[norm])
                    for frame in range(len(GFP_pre_dataset[norm])):
                        GFP_pre_dataset[norm][frame] = (GFP_pre_dataset[norm][frame] / scaling_factor) * 999 + 1


            tf_predict_488 = tf.data.Dataset.from_tensor_slices((GFP_pre_dataset))
            tf_predict_488 = tf_predict_488.batch(BATCH)
            try:
                predictions_488 = model_488.predict(x=tf_predict_488)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                status["text"] = "Neural network step detection failed. Check command line. Error is likely due to" + \
                                 " mismatch between model input and dataset tensor shape. Please make sure the" + \
                                 " correct neural network model has been assigned to the right fluorophore. GFP" + \
                                 " model expects input tensor with 300 features."
                status.update()
                set_GUI_state(tk.NORMAL)

            for i in range(len(predictions_488)):
                activations = list(predictions_488[i])
                int_pred = activations.index(max(activations))
                if len(activations) > 5:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4 + \
                             activations[5] * 5
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4] + \
                            activations[5]
                else:
                    weight = activations[1] * 1 + activations[2] * 2 + activations[3] * 3 + activations[4] * 4
                    w_sum = activations[0] + activations[1] + activations[2] + activations[3] + activations[4]
                weighted_sum = weight / w_sum
                e_sum = 0
                a_sum = 0
                for add in range(len(activations)):
                    if add != int_pred:
                        e_sum += activations[add] * abs(add - int_pred)
                for add in range(len(activations)):
                    a_sum += activations[add]
                confidence = activations[int_pred] / e_sum
                mean_dev = e_sum / a_sum
                normal_confidence = (sigmoid(np.log(confidence)))
                if int_pred == 6:
                    int_pred = "Partially bleached"

                trace_info[GFP_sublist[i]][5] = round(weighted_sum, 3)
                trace_info[GFP_sublist[i]][6] = round(normal_confidence, 2)
                trace_info[GFP_sublist[i]][7] = round(mean_dev, 4)
                trace_info[GFP_sublist[i]][1] = int_pred

            status["text"] += "\nUsed model '" + preferences["488nm model dir"] + "' to predict "+preferences["Fluorophore config"][2]+" steps"
            status.update()

        update_infobox()
        plot_trace(active_trace_list[current_trace])
    used_neural_network = True
    fitting_button["state"] = tk.NORMAL
    set_GUI_state(tk.NORMAL)


def fit_trace(trace, positions):
    trace = np.array(trace, dtype=np.float16)
    positions = sorted(positions)
    end_points = [0] + positions + [len(trace)]
    fit = np.zeros(shape=(len(trace)))
    for plateau in range(len(end_points) - 1):
        mean = np.mean(trace[end_points[plateau]:end_points[plateau+1]])
        fit[end_points[plateau]:end_points[plateau+1]] = mean
    return fit


def top_k_sample(steps, encoding):
    encoding = list(encoding)
    index = range(len(encoding))
    s = sorted(index, reverse=True, key=lambda i: encoding[i])
    actual = s[:8]
    consec = sorted(s[:8])
    print()
    print(consec)
    exclusion_list = []
    for idx in range(len(consec)-1):
        if consec[idx+1] - consec[idx] == 1 or consec[idx+1] - consec[idx] == -1:
            if encoding[consec[idx+1]] > encoding[consec[idx]]:
                exclusion_list.append(consec[idx])
            else:
                exclusion_list.append(consec[idx+1])
    s = []
    for idx in range(len(actual)):
        if actual[idx] not in exclusion_list:
            s.append(actual[idx])
    print(exclusion_list)
    print(s[:steps])
    return s[:steps]


def show_ai_activations():
    trace = list([all_traces[active_trace_list[current_trace]]])

    domain_size = 300
    tr_y = trace[0]
    tr_x = np.arange(len(tr_y))
    interpolator = interp1d(tr_x, tr_y, kind="linear")
    scale_factor = domain_size / (len(tr_y) - 1)
    resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
    if len(resampled) > domain_size:
        resampled = resampled[0:domain_size]
    trace[0] = resampled

    trace = np.array(trace, dtype=np.float16)
    trace = trace / np.max(trace)

    activations = position_model.predict(trace)
    plt.plot(trace.ravel())
    plt.plot(np.ravel(activations), linewidth=1)
    plt.title(f"Trace {current_trace + 1}, Position activations")
    plt.ylabel("Normalized Intensity / Activation")
    plt.xlabel("Resampled Trace / Neuron Index")
    plt.show()


def ai_fit_current_trace():
    trace = list([all_traces[active_trace_list[current_trace]]])

    domain_size = 300
    tr_y = trace[0]
    tr_x = np.arange(len(tr_y))
    interpolator = interp1d(tr_x, tr_y, kind="linear")
    scale_factor = domain_size / (len(tr_y) - 1)
    resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
    if len(resampled) > domain_size:
        resampled = resampled[0:domain_size]
    trace[0] = resampled

    trace = np.array(trace, dtype=np.float16)
    trace = trace / np.max(trace)

    predictions = position_model.predict(trace)
    predictions = np.array(predictions).ravel()

    valid_steps = [1, 2, 3, 4]
    if trace_info[active_trace_list[current_trace]][1] in valid_steps:
        count = trace_info[active_trace_list[current_trace]][1]
        posits = top_k_sample(count, predictions)
        posits = np.array(posits, dtype=np.float32)
        posits = posits * len(all_traces[active_trace_list[current_trace]]) / 300
        posits = list(posits.astype(dtype=np.int16))
        plateau_fit = fit_trace(all_traces[active_trace_list[current_trace]], posits)
        all_fits[active_trace_list[current_trace]] = plateau_fit
        trace_info[active_trace_list[current_trace]][4] = sorted(posits)
        heights, indices = np.unique(plateau_fit, return_index=True)
        heights = heights[np.argsort(indices)]
        step_heights = []
        for idx2 in range(len(heights) - 1):
            step_heights.append(heights[idx2] - heights[idx2 + 1])
        trace_info[active_trace_list[current_trace]][8] = step_heights

    plot_trace(active_trace_list[current_trace])
    update_infobox()


def ai_fit_trace():
    global ready_to_export
    set_GUI_state(tk.DISABLED)
    confirmation = easygui.ccbox(msg="Use neural network to predict step positions and fit traces. Traces with step counts of 1 - 4 will be fitted.",
                                 choices=('Proceed', 'Abort'), title="Deep convolutional neural network position detection V1.0a")
    if confirmation:
        trace_dataset = list(all_traces)
        domain_size = 300
        for index in range(len(trace_dataset)):
            tr_y = trace_dataset[index]
            tr_x = np.arange(len(tr_y))
            interpolator = interp1d(tr_x, tr_y, kind="linear")
            scale_factor = domain_size / (len(tr_y) - 1)
            resampled = interpolator(np.arange(0, len(tr_y) - 1, 1 / scale_factor))
            if len(resampled) > domain_size:
                resampled = resampled[0:domain_size]
            trace_dataset[index] = resampled
        trace_dataset = np.array(trace_dataset, dtype=np.float16)
        for norm in range(len(trace_dataset)):
            trace_dataset[norm] = trace_dataset[norm] / np.max(trace_dataset[norm])

        predictions = position_model.predict(trace_dataset)
        valid_steps = [1, 2, 3, 4]
        for idx in range(len(trace_info)):
            if trace_info[idx][1] in valid_steps:
                count = trace_info[idx][1]
                posits = top_k_sample(count, predictions[idx])
                posits = np.array(posits, dtype=np.float32)
                posits = posits * len(all_traces[idx]) / 300
                posits = list(posits.astype(dtype=np.int16))
                plateau_fit = fit_trace(all_traces[idx], posits)
                all_fits[idx] = plateau_fit
                trace_info[idx][4] = sorted(posits)
                heights, indices = np.unique(plateau_fit, return_index=True)
                heights = heights[np.argsort(indices)]
                step_heights = []
                for idx2 in range(len(heights)-1):
                    step_heights.append(heights[idx2] - heights[idx2+1])
                trace_info[idx][8] = step_heights


        ready_to_export = True
        plot_trace(active_trace_list[current_trace])
        update_infobox()

        status["text"] += "\nUsed position detection network to fit traces."
        status.update()

    set_GUI_state(tk.NORMAL)


def detect_steps(trace, steps, use_step_limit=True, strides=2):
    positions = [0]
    threshold = preferences["Fitting tool threshold"]
    accuracy = preferences["Fitting tool window size"]
    average_size = accuracy
    reference = 0
    old_average = sum(trace[0:average_size]) / average_size

    for scan in range(0, len(trace), strides):
        integration = 0
        av_size = 0
        for avrg in range(average_size):
            if scan + avrg < len(trace):
                integration += trace[scan + avrg]
                av_size += 1
        moving_average = integration / av_size
        if scan - reference >= accuracy:
            if old_average - moving_average >= threshold:
                positions.append(scan)
                reference = scan - 1
        # if scan % int(accuracy / 2) == 0:
        old_average = moving_average

    positions.append(len(trace))

    average_plateaus = []

    if len(positions) > 1:

        for generate_fit in range(1, len(positions)):
            integ = 0
            for find_av in range(positions[generate_fit] - positions[generate_fit - 1]):
                integ += trace[find_av + positions[generate_fit - 1]]
            av = integ / (positions[generate_fit] - positions[generate_fit - 1])
            average_plateaus.append(av)
        max_steps = []
        if steps != "Partially bleached" and steps != 5 and steps != 6 and use_step_limit:
            step_sizes = []
            for plats in range(len(average_plateaus) - 1):
                step_sizes.append(average_plateaus[plats] - average_plateaus[plats + 1])

            max_steps = []
            if len(step_sizes) > 0:
                for extract_max_steps in range(steps):
                    current_max = max(step_sizes)
                    max_index = step_sizes.index(current_max) + 1
                    max_steps.append(positions[max_index])
                    step_sizes[max_index - 1] = 0
                max_steps.sort()
                # print("Step positions:", max_steps)
                positions = [0] + max_steps + [len(trace)]
            else:
                positions = [0, len(trace)]

        average_plateaus = []

        for generate_fit in range(1, len(positions)):
            integ = 0
            for find_av in range(positions[generate_fit] - positions[generate_fit - 1]):
                integ += trace[find_av + positions[generate_fit - 1]]
            if positions[generate_fit] - positions[generate_fit - 1] != 0:
                av = integ / (positions[generate_fit] - positions[generate_fit - 1])
            else:
                av = integ / positions[generate_fit]
            average_plateaus.append(av)

        fit = []
        for generate_fit in range(1, len(positions)):
            for frame in range(positions[generate_fit] - positions[generate_fit - 1]):
                fit.append(average_plateaus[generate_fit - 1])

        plateau_heights = []
        for heights in range(len(average_plateaus) - 1):
            plateau_heights.append(round(average_plateaus[heights] - average_plateaus[heights + 1], 1))

    if not use_step_limit or steps == 5 or steps == 0 or steps == "Partially bleached":
        max_steps = []
        plateau_heights = []
    return fit, max_steps, plateau_heights


def advanced_steps(trace, steps, use_step_limit=True, strides=4, thresh=5, win_size=4):
    positions = [0]
    threshold = thresh
    accuracy = win_size
    average_size = accuracy
    reference = 0
    old_average = sum(trace[0:average_size]) / average_size

    for scan in range(0, len(trace), strides):
        integration = 0
        av_size = 0
        for avrg in range(average_size):
            if scan + avrg < len(trace):
                integration += trace[scan + avrg]
                av_size += 1
        moving_average = integration / av_size
        if scan - reference >= accuracy:
            if old_average - moving_average >= threshold:
                positions.append(scan)
                reference = scan - 1
        # if scan % int(accuracy / 2) == 0:
        old_average = moving_average

    positions.append(len(trace))

    average_plateaus = []

    if len(positions) > 1:

        for generate_fit in range(1, len(positions)):
            integ = 0
            for find_av in range(positions[generate_fit] - positions[generate_fit - 1]):
                integ += trace[find_av + positions[generate_fit - 1]]
            av = integ / (positions[generate_fit] - positions[generate_fit - 1])
            average_plateaus.append(av)
        max_steps = []
        if steps != "Partially bleached" and steps != 5 and steps != 6 and use_step_limit:
            step_sizes = []
            for plats in range(len(average_plateaus) - 1):
                step_sizes.append(average_plateaus[plats] - average_plateaus[plats + 1])

            max_steps = []
            if len(step_sizes) > 0:
                for extract_max_steps in range(steps):
                    current_max = max(step_sizes)
                    max_index = step_sizes.index(current_max) + 1
                    max_steps.append(positions[max_index])
                    step_sizes[max_index - 1] = 0
                max_steps.sort()
                # print("Step positions:", max_steps)
                positions = [0] + max_steps + [len(trace)]
            else:
                positions = [0, len(trace)]

        average_plateaus = []

        for generate_fit in range(1, len(positions)):
            integ = 0
            for find_av in range(positions[generate_fit] - positions[generate_fit - 1]):
                integ += trace[find_av + positions[generate_fit - 1]]
            if positions[generate_fit] - positions[generate_fit - 1] != 0:
                av = integ / (positions[generate_fit] - positions[generate_fit - 1])
            else:
                av = integ / positions[generate_fit]
            average_plateaus.append(av)

        fit = []
        for generate_fit in range(1, len(positions)):
            for frame in range(positions[generate_fit] - positions[generate_fit - 1]):
                fit.append(average_plateaus[generate_fit - 1])

        plateau_heights = []
        for heights in range(len(average_plateaus) - 1):
            plateau_heights.append(round(average_plateaus[heights] - average_plateaus[heights + 1], 1))

    if not use_step_limit or steps == 5 or steps == 0 or steps == "Partially bleached":
        max_steps = []
        plateau_heights = []
    return fit, max_steps, plateau_heights


def do_manual_assign(trace, max_steps, xpos=None, mxpos=None):
    try:
        if xpos >= len(trace) - 1:
            xpos = len(trace) - 1
        if xpos <= 0:
            xpos = 1
        if len(max_steps) > 0:
            diff = []
            for check in range(len(max_steps)):
                diff.append(abs(max_steps[check] - mxpos))
            index_to_amend = diff.index(min(diff))
            max_steps[index_to_amend] = int(xpos)
            max_steps.sort()

            positions = [0] + max_steps + [len(trace)]

            average_plateaus = []

            for generate_fit in range(1, len(positions)):
                integ = 0
                for find_av in range(positions[generate_fit] - positions[generate_fit - 1]):
                    integ += trace[find_av + positions[generate_fit - 1]]
                av = integ / (positions[generate_fit] - positions[generate_fit - 1])
                average_plateaus.append(av)

            fit = []
            for generate_fit in range(1, len(positions)):
                for frame in range(positions[generate_fit] - positions[generate_fit - 1]):
                    fit.append(average_plateaus[generate_fit - 1])

            plateau_heights = []
            for heights in range(len(average_plateaus) - 1):
                plateau_heights.append(round(average_plateaus[heights] - average_plateaus[heights + 1], 1))

            return fit, max_steps, plateau_heights

        else:
            fit, null, null2 = detect_steps(trace, 0, use_step_limit=False, strides=preferences["Fitting tool window stride"])
            return fit, [], []

    except:
        easygui.msgbox(title="Error!", msg="An Error occurred during manual reassignment. Trace may have to be refitted")
        fit, null, null2 = detect_steps(trace, 0, use_step_limit=False,
                                        strides=preferences["Fitting tool window stride"])
        return fit, [], []


def fit_convolution(filt, trace_in):
    kernel = kernels[filt]
    convolved_trace = []
    for convolve in range(2, len(trace_in) - 2):
        c1 = trace_in[convolve - 2] * kernel[0]
        c2 = trace_in[convolve - 1] * kernel[1]
        c3 = trace_in[convolve] * kernel[2]
        c4 = trace_in[convolve + 1] * kernel[3]
        c5 = trace_in[convolve + 2] * kernel[4]
        if filt != "trig":
            c_total = (c1 + c2 + c3 + c4 + c5) / sum(kernel)
        else:
            c_total = (c1 + c2 + c3 + c4 + c5)
        convolved_trace.append(c_total)
    convolved_trace.append(trace_in[-2])
    convolved_trace.append(trace_in[-1])
    convolved_trace.insert(0, trace_in[1])
    convolved_trace.insert(0, trace_in[0])
    return convolved_trace


def calculate_all_fits():
    print(int(float(use_ai_var.get())))
    if int(float(use_ai_var.get())) == 1:
        ai_fit_trace()
        return
    global trace_info
    global all_fits
    global has_manual_fit_happened, ready_to_export
    confirmation = True
    if has_manual_fit_happened:
        confirmation = easygui.ccbox(msg="Warning. Some traces have been manually refitted. Recalculating all fits will overwrite manually assigned"
                                     " fits. Do you wish to proceed?", title="Warning!", choices=('proceed', 'Abort'), default_choice='Abort')
    if confirmation:
        for calc_fit in range(len(trace_info)):
            trace = all_traces[calc_fit]
            if preferences["Fit convolve"]:
                trace = fit_convolution("gauss", trace)
            fit_line, step_pos, heights = detect_steps(trace, trace_info[calc_fit][1], use_step_limit=True, strides=preferences["Fitting tool window stride"])
            all_fits[calc_fit] = fit_line
            trace_info[calc_fit][4] = step_pos
            trace_info[calc_fit][8] = heights
        plot_trace(active_trace_list[current_trace])
        update_infobox()
        current_fit_button["state"] = tk.NORMAL
        advanced_fit_button["state"] = tk.NORMAL
        edit_fit_button["state"] = tk.NORMAL
        status["text"] += "\n\nUsed moving average fitting tool to apply fits to all traces."
        status.update()
        has_manual_fit_happened = False
        ready_to_export = True
        export_button["state"] = tk.NORMAL


def advanced_fit(win_size, thresh, win_strd, conv, selected_fluorophore):
    internal_names = ["Cyanine 5", "mCherry", "GFP"]
    f_name = selected_fluorophore.get()
    names = preferences["Fluorophore config"]
    index = names.index(f_name)

    global trace_info
    global all_fits
    global has_manual_fit_happened, ready_to_export
    confirmation = True
    if has_manual_fit_happened:
        confirmation = easygui.ccbox(
            msg="Warning. Some traces have been manually refitted. Recalculating all fits will overwrite manually assigned"
                " fits. Do you wish to proceed?", title="Warning!", choices=('proceed', 'Abort'),
            default_choice='Abort')
    if confirmation:
        for calc_fit in range(len(trace_info)):
            if trace_info[calc_fit][0] != internal_names[index]:
                continue
            trace = all_traces[calc_fit]
            if conv.instate(["selected"]):
                trace = fit_convolution("gauss", trace)
            fit_line, step_pos, heights = advanced_steps(trace, trace_info[calc_fit][1], use_step_limit=True,
                                                         strides=win_strd, thresh=thresh, win_size=win_size)
            all_fits[calc_fit] = fit_line
            trace_info[calc_fit][4] = step_pos
            trace_info[calc_fit][8] = heights
        plot_trace(active_trace_list[current_trace])
        update_infobox()
        current_fit_button["state"] = tk.NORMAL
        advanced_fit_button["state"] = tk.NORMAL
        edit_fit_button["state"] = tk.NORMAL
        status["text"] += "\n\nUsed moving average fitting tool to apply fits to " + str(selected_fluorophore.get()) + " traces."
        status.update()
        has_manual_fit_happened = False
        ready_to_export = True
        export_button["state"] = tk.NORMAL


def calculate_current_fit():
    print(int(float(use_ai_var.get())))
    if int(float(use_ai_var.get())) == 1:
        ai_fit_current_trace()
        return
    global trace_info
    global all_fits
    trace = all_traces[active_trace_list[current_trace]]
    if preferences["Fit convolve"]:
        trace = fit_convolution("gauss", trace)
    fit_line, step_pos, heights = detect_steps(trace, trace_info[active_trace_list[current_trace]][1], use_step_limit=True, strides=preferences["Fitting tool window stride"])
    all_fits[active_trace_list[current_trace]] = fit_line
    trace_info[active_trace_list[current_trace]][4] = step_pos
    trace_info[active_trace_list[current_trace]][8] = heights
    plot_trace(active_trace_list[current_trace])
    update_infobox()


def toggle_editing_mode():
    global capture_mouse, waiting_for_fit, show_fit
    if edit_fit_button["relief"] == tk.RAISED:
        edit_fit_button["relief"] = tk.SUNKEN
        edit_fit_button["bg"] = "#22aacc"
        edit_fit_button["fg"] = "black"
        capture_mouse = trace_figure.canvas.mpl_connect('button_press_event', select_step_to_refit)
        show_fit = True

    else:
        edit_fit_button["relief"] = tk.RAISED
        edit_fit_button["bg"] = button_bg
        edit_fit_button["fg"] = button_fg
        trace_figure.canvas.mpl_disconnect(capture_mouse)
        set_GUI_state(tk.NORMAL)
        export_button["state"] = tk.NORMAL
        fitting_button["state"] = tk.NORMAL
        current_fit_button["state"] = tk.NORMAL
        advanced_fit_button["state"] = tk.NORMAL
        waiting_for_fit = False
        plot_trace(active_trace_list[current_trace])


def select_step_to_refit(event):
    global capture_mouse, mouse_xpos, waiting_for_fit, vertical_position
    posits = trace_info[active_trace_list[current_trace]][4]
    mouse_xpos = event.xdata
    if len(posits) > 0:
        diff = []
        for check in range(len(posits)):
            diff.append(abs(posits[check] - mouse_xpos))
        s_position = posits[diff.index(min(diff))]
        vertical_position = s_position
        waiting_for_fit = True
        plot_trace(active_trace_list[current_trace])
    trace_figure.canvas.mpl_disconnect(capture_mouse)
    capture_mouse = trace_figure.canvas.mpl_connect('button_press_event', manual_assignment)
    status["text"] = "Selected nearest step to frame " + str(round(mouse_xpos)) + " to reposition. Waiting for user to click new step position..."
    status.update()
    set_GUI_state(tk.DISABLED)
    export_button["state"] = tk.DISABLED
    fitting_button["state"] = tk.DISABLED
    current_fit_button["state"] = tk.DISABLED
    advanced_fit_button["state"] = tk.DISABLED
    waiting_for_fit = False


def manual_assignment(event):
    global all_fits
    global trace_info
    global has_manual_fit_happened
    global mouse_xpos, capture_mouse
    trace = all_traces[active_trace_list[current_trace]]
    if preferences["Fit convolve"]:
        trace = fit_convolution("gauss", trace)
    posits = trace_info[active_trace_list[current_trace]][4]
    fit_line, step_pos, heights = do_manual_assign(trace, posits, xpos=event.xdata, mxpos=mouse_xpos)
    all_fits[active_trace_list[current_trace]] = fit_line
    trace_info[active_trace_list[current_trace]][4] = step_pos
    trace_info[active_trace_list[current_trace]][8] = heights
    plot_trace(active_trace_list[current_trace])
    update_infobox()
    has_manual_fit_happened = True
    trace_figure.canvas.mpl_disconnect(capture_mouse)
    capture_mouse = trace_figure.canvas.mpl_connect('button_press_event', select_step_to_refit)
    set_GUI_state(tk.NORMAL)
    export_button["state"] = tk.NORMAL
    fitting_button["state"] = tk.NORMAL
    current_fit_button["state"] = tk.NORMAL
    advanced_fit_button["state"] = tk.NORMAL
    status["text"] += "\nUser changed step position to frame " + str(round(event.xdata))
    status.update()


def pick_green_dir():
    global model_488, model_561, model_640
    find_dir = easygui.diropenbox(title="Select model folder for 488nm assignments", default=cwd)
    if find_dir:
        path = os.path.basename(find_dir)
        preferences["488nm model dir"] = path
        status["text"] = "updated neural network model directory for 488nm (GFP) to " + find_dir + "\nWarning: Please make sure base path '" + \
                         path + "' is in current working directory. (same folder as the program itself)"
        status["text"] += "\nReloading models according to new model directory..."
        status.update()
        model_488, model_561, model_640 = load_neural_network()
        status["text"] += "\nDone!"
        status["text"] += "\nThe following neural networks are assigned:\n\n" + \
                          "488nm predictor: " + str(preferences["488nm model dir"]) + \
                          "\n561nm predictor: " + str(preferences["561nm model dir"]) + \
                          "\n640nm predictor: " + str(preferences["640nm model dir"])
        status.update()


def pick_yellow_dir():
    global model_488, model_561, model_640
    find_dir = easygui.diropenbox(title="Select model folder for 561nm assignments", default=cwd)
    if find_dir:
        path = os.path.basename(find_dir)
        preferences["561nm model dir"] = path
        status["text"] = "updated neural network model directory for 561nm (mCherry) to " + find_dir + "\nWarning: Please make sure base path '" + \
                         path + "' is in current working directory. (same folder as the program itself)"
        status["text"] += "\nReloading models according to new model directory..."
        status.update()
        model_488, model_561, model_640 = load_neural_network()
        status["text"] += "\nDone!"
        status["text"] += "\nThe following neural networks are assigned:\n\n" + \
                          "488nm predictor: " + str(preferences["488nm model dir"]) + \
                          "\n561nm predictor: " + str(preferences["561nm model dir"]) + \
                          "\n640nm predictor: " + str(preferences["640nm model dir"])
        status.update()


def pick_red_dir():
    global model_488, model_561, model_640
    find_dir = easygui.diropenbox(title="Select model folder for 640nm assignments", default=cwd)
    if find_dir:
        path = os.path.basename(find_dir)
        preferences["640nm model dir"] = path
        status["text"] = "updated neural network model directory for 640nm (Cy 5) to " + find_dir + "\nWarning: Please make sure base path '" + \
                         path + "' is in current working directory. (same folder as the program itself)"
        status["text"] += "\nReloading models according to new model directory..."
        status.update()
        model_488, model_561, model_640 = load_neural_network()
        status["text"] += "\nDone!"
        status["text"] += "\nThe following neural networks are assigned:\n\n" + \
                          "488nm predictor: " + str(preferences["488nm model dir"]) + \
                          "\n561nm predictor: " + str(preferences["561nm model dir"]) + \
                          "\n640nm predictor: " + str(preferences["640nm model dir"])
        status.update()


def change_fitting_window():
    wnd_size = easygui.integerbox(title="Change Fitting parameter", msg="Change the size of the moving average window"
                                  " used by fitting algorithm. Default = 4 frames. Minimum = 2, Maximum = 10."
                                  " Warning: changing this setting may have an unpredictable effect on fitting. For advanced"
                                  " / expert users only.", lowerbound=2, upperbound=10, default=preferences["Fitting tool window size"]
                                  )
    if wnd_size:
        preferences["Fitting tool window size"] = wnd_size


def change_fitting_threshold():
    thresh = easygui.enterbox(title="Change Fitting parameter", msg="Change the threshold above which a change in the"
                                " height of the moving average window will be recorded as a step. This value is"
                                " not directly proportional to / does not represent actual step heights. Default = 5."
                                " Default value has only been tested when window size is default. Changing window size"
                                " may require a different threshold however this remains as yet untested."
                                " Warning: changing this setting may have unpredictable effects on fitting. Setting too"
                                " low may result in steps being missed. For advanced / expert users only.",
                                default=preferences["Fitting tool threshold"]
                                )
    if thresh:
        try:
            thresh = float(thresh)
            preferences["Fitting tool threshold"] = thresh
        except:
            easygui.msgbox(title="Error!", msg="Error! Invalid quanitity!")


def change_fitting_stride():
    stride = easygui.integerbox(title="Change fitting parameter", msg="Change the moving average window stride"
                                " (number of frames by which the window scans along the trace). This value should"
                                " be less than the frame width (size) of the window. Default = 4 frames"
                                " Warning: changing this setting"
                                " may have an unpredictable effect on fitting. For advanced / exper users only.",
                                lowerbound=1, upperbound=8, default=preferences["Fitting tool window stride"]
                                )
    if stride:
        preferences["Fitting tool window stride"] = stride


def change_scaling():
    global preferences
    choices = ["488nm scaling", "561nm scaling", "640nm scaling"]
    opts = easygui.choicebox(title="Preprocessing parameters", msg="Change the amount by which traces are downscaled"
                             " before entering neural network. Higher = more downscaling.",
                             choices=choices
                             )
    if opts:
        if opts == choices[0]:
            scale488 = easygui.integerbox(title="488nm Scaling", msg="Set downscaling for 488nm. Default = 200",
                                          lowerbound=100, upperbound=800, default=preferences["Preprocessing parameters"][0])
            if scale488:
                preferences["Preprocessing parameters"][0] = scale488

        elif opts == choices[1]:
            scale561 = easygui.integerbox(title="561nm Scaling", msg="Set downscaling for 561nm. Default = 200",
                                          lowerbound=100, upperbound=800,
                                          default=preferences["Preprocessing parameters"][1])
            if scale561:
                preferences["Preprocessing parameters"][1] = scale561

        elif opts == choices[2]:
            scale640 = easygui.integerbox(title="640nm Scaling", msg="Set downscaling for 640nm. Default = 800",
                                          lowerbound=200, upperbound=2000,
                                          default=preferences["Preprocessing parameters"][2])
            if scale640:
                preferences["Preprocessing parameters"][2] = scale640


def enable_fit_conv():
    choice = easygui.choicebox(msg="Enable or disable fitting convolution. This increases signal to noise but may result in failure to fit short plateaus.", title="Fit convolution",
                               choices=["Enable", "Disable"])
    if choice == "Enable":
        preferences["Fit convolve"] = True
    if choice == "Disable":
        preferences["Fit convolve"] = False


def set_fluoro_names():
    global preferences
    choices = ["marker fluorophore", "561nm fluorophore", "488nm fluorophore"]
    choice = easygui.choicebox(title="Configure fluorophores", msg="Select a laser wavelength, then enter the name of"
                                                                   " the fluorophore for that wavelength.",
                               choices=choices)
    if choice:
        index = choices.index(choice)
        name = easygui.enterbox(title=str(choice), msg="Enter name of " + str(choice), default=preferences["Fluorophore config"][index])
        if name:
            preferences["Fluorophore config"][index] = name
            status["text"] = "Updated fluorophore configuration."
            status.update()
        if is_data_loaded:
            plot_trace(active_trace_list[current_trace])
            update_infobox()


def set_frame_fraction():
    choice = easygui.choicebox(title="Set Frame Fraction", msg="Set the percentage fraction of frames used to generate sum views.",
                               choices=["Marker", "561nm", "488nm"])
    if choice:
        if choice == "Marker":
            frac = easygui.integerbox(title="Marker frames", msg="Set the percentage of marker frames.", lowerbound=1,
                                      upperbound=100, default=int(preferences["Sum view frame fraction"][0]*100))
            if frac:
                preferences["Sum view frame fraction"][0] = frac / 100

        if choice == "561nm":
            frac = easygui.integerbox(title="561nm frames", msg="Set the percentage of 561nm frames.", lowerbound=1,
                                      upperbound=100, default=int(preferences["Sum view frame fraction"][1]*100))
            if frac:
                preferences["Sum view frame fraction"][1] = frac / 100

        if choice == "488nm":
            frac = easygui.integerbox(title="488nm frames", msg="Set the percentage of 488nm frames.", lowerbound=1,
                                      upperbound=100, default=int(preferences["Sum view frame fraction"][2]*100))
            if frac:
                preferences["Sum view frame fraction"][2] = frac / 100


def set_intensity_target():
    while True:
        choice = easygui.choicebox(title="Set View Intensity Target", msg="Set view auto-brightness intensity target."
                                         " This quantity is not the same as the brightness shown in the user interface"
                                         ", it is the target mean view integration for the view optimizer.",
                                   choices=["Marker", "561nm", "488nm"])
        if choice:
            if choice == "Marker":
                targ = easygui.integerbox(title="Marker target", msg="Set the intensity target of marker view", lowerbound=1,
                                          upperbound=2000, default=int(preferences["Intensity target"][0] * 100))
                if targ:
                    preferences["Intensity target"][0] = targ / 100

            if choice == "561nm":
                targ = easygui.integerbox(title="561nm target", msg="Set the intensity target of 561nm view", lowerbound=1,
                                          upperbound=2000, default=int(preferences["Intensity target"][1] * 100))
                if targ:
                    preferences["Intensity target"][1] = targ / 100

            if choice == "488nm":
                targ = easygui.integerbox(title="488nm target", msg="Set the intensity target of 488nm view", lowerbound=1,
                                          upperbound=2000, default=int(preferences["Intensity target"][2] * 100))
                if targ:
                    preferences["Intensity target"][2] = targ / 100

        else:
            break
    return


def create_circle(x, y, r, canvasName, outline="white", width=1):
    x0 = x - r
    y0 = y - r
    x1 = x + r
    y1 = y + r
    return canvasName.create_oval(x0, y0, x1, y1, outline=outline, width=width)


def create_circle_dashed(x, y, r, canvasName, outline="white", width=1):
    x0 = x - r
    y0 = y - r
    x1 = x + r
    y1 = y + r
    return canvasName.create_oval(x0, y0, x1, y1, outline=outline, width=width, dash=(2, 2))


def create_mask(sigma_x, sigma_y, sx=11, sy=11):
    """ Mask spot in terms of its x and y standard deviation, by setting the radii of an ellipse to 3 x sigma
        We can capture at least 80% of the spot. """

    mask = np.ones((sy, sx))
    centre_x = int(sx / 2)
    centre_y = int(sy / 2)


    ellipse_radius_x = sigma_x * 1.5
    ellipse_radius_y = sigma_y * 1.5

    for x in range(sx):
        for y in range(sy):
            ax = x
            ay = y
            boundary_condition = (ax - centre_x)**2 / ellipse_radius_x**2 + (ay - centre_y)**2 / ellipse_radius_y**2
            if boundary_condition <= 1:
                mask[y][x] = 0

    return mask


def plot_histogram(data, linecol):
    if linecol == "red":
        raw_gui.hist_figM.clf()
        raw_gui.histM.draw()
        ax = raw_gui.hist_figM.add_subplot(111)
    elif linecol == "yellow":
        raw_gui.hist_fig5.clf()
        raw_gui.hist5.draw()
        ax = raw_gui.hist_fig5.add_subplot(111)
    else:
        raw_gui.hist_fig4.clf()
        raw_gui.hist4.draw()
        ax = raw_gui.hist_fig4.add_subplot(111)

    ax.set_facecolor("#222222")
    ax.spines['bottom'].set_color('blue')
    ax.spines['top'].set_color('blue')
    ax.spines['left'].set_color('blue')
    ax.spines['right'].set_color('blue')

    ax.plot(data, color=linecol, linewidth=1)
    raw_gui.histM.draw()
    raw_gui.hist5.draw()
    raw_gui.hist4.draw()


def calculate_spot_stats():
    coloc561 = 0
    coloc488 = 0
    if len(raw_gui.colocalizations) > 0:
        for index in range(len(raw_gui.colocalizations)):
            if raw_gui.colocalizations[index][1] is not None:
                coloc488 += 1
            if raw_gui.colocalizations[index][2] is not None:
                coloc561 += 1
    return len(raw_gui.raw_marker_spots), len(raw_gui.raw_488_spots), len(raw_gui.raw_561_spots), coloc488, coloc561


def create_view_marker(first_time=False):
    global raw_gui
    global imgM, imgMs
    global sum_array_marker
    global orgsum_marker
    if not raw_gui.is_data_loaded:
        return
    update_coloc()
    raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.use_marker.get() == 1:
        brightness = float(raw_gui.brightness_marker.get())
        if first_time and not raw_gui.high_pass:
            try:
                power, brightness = tirf.calculate_power(orgsum_marker, raw_setup["convolutions"][0],
                                                         int(raw_gui.bg_entry_marker.get()) / 100,
                                                         preferences["Intensity target"][0])
            except:
                easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: Marker.\n"
                                                   "Please make sure the view has valid frame limits.")
                raw_gui.analyse_marker.invoke()
                return
            raw_gui.power_marker.set(round(power, 2))
            raw_gui.brightness_marker.set(round(brightness*2, 1)/2)
        power = float(raw_gui.power_marker.get())
        brightness = float(raw_gui.brightness_marker.get())
        try:
            if not raw_gui.high_pass:
                subtracted, bins = tirf.sum_view(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, brightness, power, raw_setup["convolutions"][0])
                plot_histogram(bins, "red")
            else:
                subtracted, bins = tirf.wavelet_transform(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, power, raw_gui.use_deconv_M)
                plot_histogram(bins, "red")
        except:
            raw_gui.analyse_marker.invoke()
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: Marker.\n\n"
                                               + str(traceback.format_exc()))
            return

        sum_array_marker = np.copy(subtracted)
        h, w = np.shape(subtracted)[0], np.shape(subtracted)[1]
        mean_intensity = round(np.sum(np.clip(subtracted * brightness, 0, 255)) / (h*w), 3)
        raw_gui.view1_label["text"] = "  Marker - " + preferences["Fluorophore config"][0] + "       Mean Intensity: " \
            + str(mean_intensity)
        raw_gui.view1_label.update()
        sub_disp = np.zeros((h, w, 3))
        if raw_gui.marker_view_colour == "488":
            sub_disp[:, :, 1] = subtracted
        elif raw_gui.marker_view_colour == "561":
            sub_disp[:, :, 0] = subtracted
            sub_disp[:, :, 1] = subtracted
        else:
            sub_disp[:, :, 0] = subtracted
        sub_disp = sub_disp * brightness
        sub_disp = np.clip(sub_disp, 0, 255)
        sub_image = tirf.create_image(sub_disp, 512, 512)
        raw_gui.image_canvas_marker = sub_image
        imgM = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=sub_image)
        raw_gui.canvas_marker.delete("all")
        raw_gui.canvas_marker.create_image(2, 2, anchor="nw", image=imgM)

        org_disp = np.zeros((h, w, 3))
        subtracted2, null = tirf.wavelet_transform(orgsum_marker, 0.2, 0.3775, False)
        if not raw_gui.use_stokes:
            if raw_gui.marker_view_colour == "488":
                org_disp[:, :, 1] = orgsum_marker
            elif raw_gui.marker_view_colour == "561":
                org_disp[:, :, 0] = orgsum_marker
                org_disp[:, :, 1] = orgsum_marker
            else:
                org_disp[:, :, 0] = orgsum_marker
        else:
            if raw_gui.marker_view_colour == "488":
                background = tirf.true_low_pass(orgsum_marker)
                org_disp[:, :, 1] = (background / 255) * 246
                org_disp[:, :, 2] = (background / 255) * 255
                # org_disp[:, :, 2] = np.clip(org_disp[:, :, 2] - subtracted2/2, 0, 255)
                org_disp[:, :, 1] += subtracted2
                org_max = np.max(org_disp)
                org_disp = (org_disp / org_max) * 255
            elif raw_gui.marker_view_colour == "561":
                background = tirf.true_low_pass(orgsum_marker)
                org_disp[:, :, 0] = (background / 255) * 197
                org_disp[:, :, 1] = (background / 255) * 255
                org_disp = org_disp / 2
                org_disp[:, :, 0] += (subtracted2 / 255) * 255
                org_disp[:, :, 1] += (subtracted2 / 255) * 155
                org_max = np.max(org_disp)
                org_disp = (org_disp / org_max) * 255
            elif raw_gui.marker_view_colour == "640":
                background = tirf.true_low_pass(orgsum_marker)
                org_disp[:, :, 0] = (background / 255) * 255
                org_disp[:, :, 1] = (background / 255) * 50
                org_disp[:, :, 1] = np.clip(org_disp[:, :, 1] - subtracted2, 0, 255)
                org_disp[:, :, 0] += (subtracted2 / 255) * 255
                org_disp[:, :, 2] += (subtracted2 / 255) * 20
                org_max = np.max(org_disp)
                org_disp = (org_disp / org_max) * 255




        org_image = tirf.create_image(org_disp, 256, 256)
        imgMs = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=org_image)
        raw_gui.canvas_marker_mini.create_image(2, 2, anchor="nw", image=imgMs)

        if len(raw_gui.raw_marker_spots) > 0:
            for spot in raw_gui.raw_marker_spots:
                ring_colour = "#999999"
                if len(raw_gui.colocalizations) > 0:
                    coloc_markers = [channel[0] for channel in raw_gui.colocalizations]
                    if raw_gui.raw_marker_spots.index(spot) in coloc_markers:
                        coloc_index = coloc_markers.index(raw_gui.raw_marker_spots.index(spot))
                        if raw_gui.colocalizations[coloc_index][1] is not None and raw_gui.colocalizations[coloc_index][2] is None:
                            ring_colour = "#00ff77"
                        elif raw_gui.colocalizations[coloc_index][2] is not None and raw_gui.colocalizations[coloc_index][1] is None:
                            ring_colour = "#ffaa00"
                        elif raw_gui.colocalizations[coloc_index][1] is not None and raw_gui.colocalizations[coloc_index][2] is not None:
                            ring_colour = "#bb00ff"
                label = raw_gui.raw_marker_spots.index(spot) + 1
                create_circle(round(spot[0] * raw_gui.scale_factor) + 2, round(spot[1] * raw_gui.scale_factor) + 2, 5*raw_gui.scale_factor,
                              raw_gui.canvas_marker, outline=ring_colour, width=1)
                if raw_gui.scale_factor == 1:
                    f_size="arial 6"
                else:
                    f_size="arial 9"
                raw_gui.canvas_marker.create_text(round(spot[0] * raw_gui.scale_factor + 6.5*raw_gui.scale_factor),
                                                  round(spot[1] * raw_gui.scale_factor - 6.5*raw_gui.scale_factor), fill="white",
                                                  font=f_size, text=str(label))
        if raw_gui.region_initial[0] is not None and raw_gui.region_final[0] is not None:
            raw_gui.rect_640 = raw_gui.canvas_marker.create_rectangle(raw_gui.region_initial[0][0]*raw_gui.scale_factor, raw_gui.region_initial[0][1]*raw_gui.scale_factor,
                                                                  raw_gui.region_final[0][0]*raw_gui.scale_factor, raw_gui.region_final[0][1]*raw_gui.scale_factor,
                                                                  outline="#bbbbbb", width=1, dash=(2, 1))

        if raw_gui.display_stats:
            sp640, sp488, sp561, c488, c561 = calculate_spot_stats()
            raw_gui.canvas_marker.create_rectangle(2, 2, 188, 136, outline="#bbbbbb", width=2, fill="#aaaaaa", stipple="gray75")
            raw_gui.canvas_marker.create_text(8, 6, fill="#000000", anchor=tk.NW, font="TkDefaultFont 12 bold",
                                              text=f"Marker Spots: {sp640}\n561nm Spots: {sp561}\n488nm Spots: {sp488}\n\n561nm colocalized: {c561}\n488nm colocalized: {c488}",)

        raw_gui.raw_window.update()

    else:
        raw_gui.canvas_marker.delete("all")
        raw_gui.canvas_marker_mini.delete("all")
        # raw_gui.context_marker.entryconfig("Remove Spot", state=tk.DISABLED)
        # raw_gui.context_marker.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        raw_gui.raw_marker_spots = []
        plot_histogram([0], "red")

    raw_gui.set_rawgui_state(tk.NORMAL)


def create_view_561(first_time=False):
    global raw_gui
    global img561, img561s
    global sum_array_561
    global orgsum_561
    if not raw_gui.is_data_loaded:
        return
    update_coloc()
    if raw_gui.display_stats:
        create_view_marker()
    raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.use_561.get() == 1:
        brightness = float(raw_gui.brightness_561.get())
        if first_time and not raw_gui.high_pass:
            try:
                power, brightness = tirf.calculate_power(orgsum_561, raw_setup["convolutions"][1],
                                                         int(raw_gui.bg_entry_561.get()) / 100,
                                                         preferences["Intensity target"][1])
            except:
                easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: 561nm.\n"
                                                   "Please make sure the view has valid frame limits.")
                raw_gui.analyse_561.invoke()
                return
            raw_gui.power_561.set(round(power, 2))
            raw_gui.brightness_561.set(round(brightness*2, 1)/2)
        power = float(raw_gui.power_561.get())
        brightness = float(raw_gui.brightness_561.get())
        try:
            if not raw_gui.high_pass:
                subtracted, bins = tirf.sum_view(orgsum_561, int(raw_gui.bg_entry_561.get()) / 100, brightness, power, raw_setup["convolutions"][1])
                plot_histogram(bins, "yellow")
            else:
                subtracted, bins = tirf.wavelet_transform(orgsum_561, int(raw_gui.bg_entry_561.get()) / 100, power, raw_gui.use_deconv_561)
                plot_histogram(bins, "yellow")
        except:
            raw_gui.analyse_561.invoke()
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: 561.\n\n"
                                               + str(traceback.format_exc()))
            return
        sum_array_561 = np.copy(subtracted)
        h, w = np.shape(subtracted)[0], np.shape(subtracted)[1]
        mean_intensity = round(np.sum(np.clip(subtracted * brightness, 0, 255)) / (h * w), 3)
        raw_gui.view2_label["text"] = "  561nm - " + preferences["Fluorophore config"][1] + "       Mean Intensity: " \
                                      + str(mean_intensity)
        raw_gui.view2_label.update()
        sub_disp = np.zeros((h, w, 3))
        sub_disp[:, :, 0] = subtracted
        sub_disp[:, :, 1] = subtracted
        if raw_gui.overlay_561:
            brightness_marker = float(raw_gui.brightness_marker.get())
            power_marker = float(raw_gui.power_marker.get())
            if not raw_gui.high_pass:
                get_marker, null = tirf.sum_view(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, brightness_marker,
                                                 power_marker, raw_setup["convolutions"][0])
            else:
                get_marker, null = tirf.wavelet_transform(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, power_marker, raw_gui.use_deconv_M)
                sub_max = np.max(get_marker)
                get_marker = np.clip((get_marker/sub_max)*255, 0, 255)
            print(np.shape(get_marker))
            sub_disp[:, :, 0] += get_marker
            sub_disp = np.clip(sub_disp, 0, 255)
        sub_disp = sub_disp * brightness
        sub_disp = np.clip(sub_disp, 0, 255)
        sub_image = tirf.create_image(sub_disp, 512, 512)
        raw_gui.image_canvas_561 = sub_image
        img561 = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=sub_image)
        raw_gui.canvas_561.delete("all")
        raw_gui.canvas_561.create_image(2, 2, anchor="nw", image=img561)

        org_disp = np.zeros((h, w, 3))
        subtracted2, null = tirf.wavelet_transform(orgsum_561, 0.2, 0.3775, False)
        if not raw_gui.use_stokes:
            org_disp[:, :, 0] = orgsum_561
            org_disp[:, :, 1] = orgsum_561
        else:
            background = tirf.true_low_pass(orgsum_561)
            org_disp[:, :, 0] = (background / 255) * 197
            org_disp[:, :, 1] = (background / 255) * 255
            org_disp = org_disp / 2
            org_disp[:, :, 0] += (subtracted2 / 255) * 255
            org_disp[:, :, 1] += (subtracted2 / 255) * 155
            org_max = np.max(org_disp)
            org_disp = (org_disp / org_max) * 255
        org_image = tirf.create_image(org_disp, 256, 256)
        img561s = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=org_image)
        raw_gui.canvas_561_mini.create_image(2, 2, anchor="nw", image=img561s)
        if raw_gui.scale_factor == 1:
            f_size = "arial 6"
            f_size2 = "arial 6"
        else:
            f_size = "arial 9"
            f_size2 = "arial 8"
        if len(raw_gui.raw_marker_spots) > 0 and raw_gui.use_marker.get() == 1 and raw_gui.show_markers_561:
            for spot in raw_gui.raw_marker_spots:
                label = raw_gui.raw_marker_spots.index(spot) + 1
                create_circle(round(spot[0] * raw_gui.scale_factor) + 2, round(spot[1] * raw_gui.scale_factor) + 2,
                              4*raw_gui.scale_factor, raw_gui.canvas_561, outline="#881111", width=1)
                raw_gui.canvas_561.create_text(round(spot[0] * raw_gui.scale_factor - 5*raw_gui.scale_factor),
                                               round(spot[1] * raw_gui.scale_factor - 5*raw_gui.scale_factor), fill="#999999",
                                               font=f_size2, text=str(label))

        if len(raw_gui.raw_561_spots) > 0:
            for spot in raw_gui.raw_561_spots:
                ring_colour = "#999999"
                if len(raw_gui.colocalizations) > 0:
                    coloc_561 = [channel[2] for channel in raw_gui.colocalizations]
                    if raw_gui.raw_561_spots.index(spot) in coloc_561:
                        ring_colour = "#0055ff"
                label = raw_gui.raw_561_spots.index(spot) + 1
                show_all = True
                if raw_gui.display_stats:
                    show_all = False
                if show_all:
                    create_circle(round(spot[0] * raw_gui.scale_factor) + 2, round(spot[1] * raw_gui.scale_factor) + 2,
                                  5*raw_gui.scale_factor, raw_gui.canvas_561, outline=ring_colour, width=1)
                    raw_gui.canvas_561.create_text(round(spot[0] * raw_gui.scale_factor + 6.5*raw_gui.scale_factor),
                                                   round(spot[1] * raw_gui.scale_factor - 6.5*raw_gui.scale_factor), fill="white",
                                                   font=f_size, text=str(label))
                else:
                    if ring_colour == "#0055ff":
                        create_circle(round(spot[0] * raw_gui.scale_factor) + 2,
                                      round(spot[1] * raw_gui.scale_factor) + 2,
                                      5 * raw_gui.scale_factor, raw_gui.canvas_561, outline=ring_colour, width=1)
                        raw_gui.canvas_561.create_text(
                            round(spot[0] * raw_gui.scale_factor + 6.5 * raw_gui.scale_factor),
                            round(spot[1] * raw_gui.scale_factor - 6.5 * raw_gui.scale_factor), fill="white",
                            font=f_size, text=str(label))
                    else:
                        create_circle(round(spot[0] * raw_gui.scale_factor) + 2,
                                      round(spot[1] * raw_gui.scale_factor) + 2,
                                      5 * raw_gui.scale_factor, raw_gui.canvas_561, outline="#333333", width=1)

        if raw_gui.vectors_561 and raw_gui.marker_view_colour != "561":
            factor_dict = {
                "640": 0.52,
                "488": -0.48,
            }
            for x in range(0, w, 10):
                for y in range(0, h, 10):
                    newx, newy = coordinate_transform(x, y, factor_dict[raw_gui.marker_view_colour])
                    newx2, newy2 = coordinate_transform(x, y, factor_dict[raw_gui.marker_view_colour] * 1.5)
                    raw_gui.canvas_561.create_line(newx * raw_gui.scale_factor, newy * raw_gui.scale_factor,
                                                   newx2 * raw_gui.scale_factor, newy2 * raw_gui.scale_factor,
                                                   fill="#dd33ff", arrow=tk.LAST, arrowshape=(4, 5, 2))
                    raw_gui.canvas_561.create_line(x * raw_gui.scale_factor, y * raw_gui.scale_factor,
                                                   newx * raw_gui.scale_factor, newy * raw_gui.scale_factor,
                                                   fill="white")
        if raw_gui.region_initial[1] is not None and raw_gui.region_final[1] is not None:
            raw_gui.rect_561 = raw_gui.canvas_561.create_rectangle(raw_gui.region_initial[1][0]*raw_gui.scale_factor, raw_gui.region_initial[1][1]*raw_gui.scale_factor,
                                                                  raw_gui.region_final[1][0]*raw_gui.scale_factor, raw_gui.region_final[1][1]*raw_gui.scale_factor,
                                                                  outline="#bbbbbb", width=1, dash=(2, 1))

        raw_gui.raw_window.update()

    else:
        raw_gui.canvas_561.delete("all")
        raw_gui.canvas_561_mini.delete("all")
        # raw_gui.context_561.entryconfig("Remove Spot", state=tk.DISABLED)
        # raw_gui.context_561.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        raw_gui.raw_561_spots = []
        raw_gui.final_561_spots = []
        plot_histogram([0], "yellow")

    raw_gui.set_rawgui_state(tk.NORMAL)


def create_view_488(first_time=False):
    global raw_gui
    global img488, img488s
    global sum_array_488
    global orgsum_488
    if not raw_gui.is_data_loaded:
        return
    update_coloc()
    if raw_gui.display_stats:
        create_view_marker()
    raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.use_488.get() == 1:
        brightness = float(raw_gui.brightness_488.get())
        if first_time and not raw_gui.high_pass:
            try:
                power, brightness = tirf.calculate_power(orgsum_488, raw_setup["convolutions"][2],
                                                         int(raw_gui.bg_entry_488.get()) / 100,
                                                         preferences["Intensity target"][2])
            except:
                raw_gui.analyse_488.invoke()
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: 488.\n\n"
                                                   + str(traceback.format_exc()))
                return
            raw_gui.power_488.set(round(power, 2))
            raw_gui.brightness_488.set(round(brightness*2, 1)/2)
        power = float(raw_gui.power_488.get())
        brightness = float(raw_gui.brightness_488.get())
        try:
            if not raw_gui.high_pass:
                subtracted, bins = tirf.sum_view(orgsum_488, int(raw_gui.bg_entry_488.get()) / 100, brightness, power, raw_setup["convolutions"][2])
                plot_histogram(bins, "green")
            else:
                subtracted, bins = tirf.wavelet_transform(orgsum_488, int(raw_gui.bg_entry_488.get()) / 100, power, raw_gui.use_deconv_488)
                plot_histogram(bins, "green")
        except:
            raw_gui.analyse_488.invoke()
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="An error occurred while attempting to create view: 488.\n\n"
                                               + str(traceback.format_exc()))
            return
        sum_array_488 = np.copy(subtracted)
        h, w = np.shape(subtracted)[0], np.shape(subtracted)[1]
        mean_intensity = round(np.sum(np.clip(subtracted * brightness, 0, 255)) / (h * w), 3)
        raw_gui.view3_label["text"] = "  488nm - " + preferences["Fluorophore config"][2] + "       Mean Intensity: " \
                                      + str(mean_intensity)
        raw_gui.view3_label.update()
        sub_disp = np.zeros((h, w, 3))
        sub_disp[:, :, 1] = subtracted
        if raw_gui.overlay_488:
            brightness_marker = float(raw_gui.brightness_marker.get())
            power_marker = float(raw_gui.power_marker.get())
            if not raw_gui.high_pass:
                get_marker, null = tirf.sum_view(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, brightness_marker,
                                                 power_marker, raw_setup["convolutions"][0])
            else:
                get_marker, null = tirf.wavelet_transform(orgsum_marker, int(raw_gui.bg_entry_marker.get()) / 100, power_marker, raw_gui.use_deconv_M)
                sub_max = np.max(get_marker)
                get_marker = np.clip((get_marker/sub_max)*255, 0, 255)
            sub_disp[:, :, 0] += get_marker
        sub_disp = sub_disp * brightness
        sub_disp = np.clip(sub_disp, 0, 255)
        sub_image = tirf.create_image(sub_disp, 512, 512)
        raw_gui.image_canvas_488 = sub_image
        img488 = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=sub_image)
        raw_gui.canvas_488.delete("all")
        raw_gui.canvas_488.create_image(2, 2, anchor="nw", image=img488)

        org_disp = np.zeros((h, w, 3))
        subtracted2, null = tirf.wavelet_transform(orgsum_488, 0.2, 0.377, False)
        if not raw_gui.use_stokes:
            org_disp[:, :, 1] = orgsum_488
        else:
            background = tirf.true_low_pass(orgsum_488)
            org_disp[:, :, 1] = (background / 255) * 246
            org_disp[:, :, 2] = (background / 255) * 255
            # org_disp[:, :, 2] = np.clip(org_disp[:, :, 2] - subtracted2/2, 0, 255)
            org_disp[:, :, 1] += subtracted2
            org_max = np.max(org_disp)
            org_disp = (org_disp / org_max) * 255
        org_image = tirf.create_image(org_disp, 256, 256)
        img488s = tirf.ImageTk.PhotoImage(master=raw_gui.raw_window, image=org_image)
        raw_gui.canvas_488_mini.create_image(2, 2, anchor="nw", image=img488s)
        if raw_gui.scale_factor == 1:
            f_size = "arial 6"
            f_size2 = "arial 6"
        else:
            f_size = "arial 9"
            f_size2 = "arial 8"
        if len(raw_gui.raw_marker_spots) > 0 and raw_gui.use_marker.get() == 1 and raw_gui.show_markers_488:
            for spot in raw_gui.raw_marker_spots:
                label = raw_gui.raw_marker_spots.index(spot) + 1
                create_circle(round(spot[0] * raw_gui.scale_factor) + 2, round(spot[1] * raw_gui.scale_factor) + 2,
                              4*raw_gui.scale_factor, raw_gui.canvas_488, outline="#881111", width=1)
                raw_gui.canvas_488.create_text(round(spot[0] * raw_gui.scale_factor - 5*raw_gui.scale_factor),
                                               round(spot[1] * raw_gui.scale_factor - 5*raw_gui.scale_factor), fill="#999999",
                                               font=f_size2, text=str(label))

        if len(raw_gui.raw_488_spots) > 0:
            for spot in raw_gui.raw_488_spots:
                ring_colour = "#999999"
                if len(raw_gui.colocalizations) > 0:
                    coloc_488 = [channel[1] for channel in raw_gui.colocalizations]
                    if raw_gui.raw_488_spots.index(spot) in coloc_488:
                        ring_colour = "#0055ff"
                label = raw_gui.raw_488_spots.index(spot) + 1
                show_all = True
                if raw_gui.display_stats:
                    show_all = False
                if show_all:
                    create_circle(round(spot[0] * raw_gui.scale_factor) + 2, round(spot[1] * raw_gui.scale_factor) + 2,
                                  5*raw_gui.scale_factor, raw_gui.canvas_488,
                                  outline=ring_colour, width=1)
                    raw_gui.canvas_488.create_text(round(spot[0] * raw_gui.scale_factor + 6.5*raw_gui.scale_factor),
                                                   round(spot[1] * raw_gui.scale_factor - 6.5*raw_gui.scale_factor), fill="white",
                                                   font=f_size, text=str(label))
                else:
                    if ring_colour == "#0055ff":
                        create_circle(round(spot[0] * raw_gui.scale_factor) + 2,
                                      round(spot[1] * raw_gui.scale_factor) + 2,
                                      5 * raw_gui.scale_factor, raw_gui.canvas_488,
                                      outline=ring_colour, width=1)
                        raw_gui.canvas_488.create_text(
                            round(spot[0] * raw_gui.scale_factor + 6.5 * raw_gui.scale_factor),
                            round(spot[1] * raw_gui.scale_factor - 6.5 * raw_gui.scale_factor), fill="white",
                            font=f_size, text=str(label))
                    else:
                        create_circle(round(spot[0] * raw_gui.scale_factor) + 2,
                                      round(spot[1] * raw_gui.scale_factor) + 2,
                                      5 * raw_gui.scale_factor, raw_gui.canvas_488,
                                      outline="#333333", width=1)

        if raw_gui.vectors_488 and raw_gui.marker_view_colour != "488":
            factor_dict = {
                "640": 1,
                "561": 0.48,
            }
            for x in range(0, w, 10):
                for y in range(0, h, 10):
                    newx, newy = coordinate_transform(x, y, factor_dict[raw_gui.marker_view_colour])
                    newx2, newy2 = coordinate_transform(x, y, factor_dict[raw_gui.marker_view_colour]*1.5)
                    raw_gui.canvas_488.create_line(newx * raw_gui.scale_factor, newy * raw_gui.scale_factor,
                                                   newx2 * raw_gui.scale_factor, newy2 * raw_gui.scale_factor,
                                                   fill="#dd33ff", arrow=tk.LAST, arrowshape=(4, 5, 2))
                    raw_gui.canvas_488.create_line(x*raw_gui.scale_factor, y*raw_gui.scale_factor, newx*raw_gui.scale_factor, newy*raw_gui.scale_factor,
                                                   fill="white")
        if raw_gui.region_initial[2] is not None and raw_gui.region_final[2] is not None:
            raw_gui.rect_488 = raw_gui.canvas_488.create_rectangle(raw_gui.region_initial[2][0]*raw_gui.scale_factor, raw_gui.region_initial[2][1]*raw_gui.scale_factor,
                                                                  raw_gui.region_final[2][0]*raw_gui.scale_factor, raw_gui.region_final[2][1]*raw_gui.scale_factor,
                                                                  outline="#bbbbbb", width=1, dash=(2, 1))

        raw_gui.raw_window.update()

    else:
        raw_gui.canvas_488.delete("all")
        raw_gui.canvas_488_mini.delete("all")
        # raw_gui.context_488.entryconfig("Remove Spot", state=tk.DISABLED)
        # raw_gui.context_488.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        raw_gui.raw_488_spots = []
        raw_gui.final_488_spots = []
        plot_histogram([0], "green")

    raw_gui.set_rawgui_state(tk.NORMAL)


def load_tiff(auto=False, file=None):

    global array_TIF, array_TIF_len
    global orgsum_marker, orgsum_561, orgsum_488
    raw_gui.set_rawgui_state(tk.DISABLED)

    if not auto:
        raw_gui.load_tiff_button["state"] = tk.DISABLED
        path = easygui.fileopenbox(msg="Open 256x256 TIF stack from objective TIRF only",
                                   filetypes=["*.tif"], default=default_dir+"*.tif")
    else:
        path = file

    if path:
        if path in raw_gui.session_paths:
            confirm = easygui.ccbox(title="Warning!", msg="This file has already been opened this session! Load anyway?",
                                    choices=["Proceed", "Abort"], default_choice="Abort", cancel_choice="Abort")
            if not confirm:
                raw_gui.load_tiff_button["state"] = tk.NORMAL
                raw_gui.set_rawgui_state(tk.NORMAL)
                return
        raw_gui.status["text"] = "Current File: '" + str(path) + "'"
        raw_gui.current_data_path = path
        raw_gui.status.update()
        raw_gui.progress.place(x=740, y=826)
        raw_gui.progress.step(10)
        raw_gui.load_label.place(x=744, y=852)
        raw_gui.load_label["text"] = "Loading TIF file..."
        raw_gui.load_label.update()

        try:
            raw_data, num_frames = tirf.load_tiff(path)
        except FileNotFoundError:
            easygui.msgbox(title="Error!", msg="Warning file could not be located!")
            raw_gui.set_rawgui_state(tk.NORMAL)
            array_TIF = None
            array_TIF_len = None
            try:
                orgsum_marker = tirf.create_sum(array_TIF, int(raw_gui.start_entry_marker.get()),
                                                int(int(raw_gui.length_entry_marker.get()) * float(
                                                    preferences["Sum view frame fraction"][0])),
                                                raw_gui.view_modes[0])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ An error occurred """
                orgsum_marker = None

            try:
                orgsum_561 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_561.get()),
                                             int(int(raw_gui.length_entry_561.get()) * float(
                                                 preferences["Sum view frame fraction"][1])),
                                             raw_gui.view_modes[1])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ An error occurred """
                orgsum_561 = None

            try:
                orgsum_488 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_488.get()),
                                             int(int(raw_gui.length_entry_488.get()) * float(
                                                 preferences["Sum view frame fraction"][2])),
                                             raw_gui.view_modes[2])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ An error occurred """
                orgsum_488 = None

            raw_gui.is_data_loaded = False

            create_view_marker(first_time=True)
            raw_gui.progress.step(10)
            create_view_561(first_time=True)
            raw_gui.progress.step(18)
            create_view_488(first_time=True)
            raw_gui.progress.step(2)

            raw_gui.progress.stop()
            raw_gui.progress.place_forget()
            raw_gui.load_label.place_forget()

            return
        array_TIF = raw_data
        raw_gui.resolution = (np.shape(array_TIF)[0], np.shape(array_TIF)[1])
        array_TIF_len = num_frames
        raw_gui.scale_factor = 512 / np.shape(array_TIF)[0]
        raw_gui.status["text"] += "\nFile length (frames): " + str(array_TIF_len)
        raw_gui.status.update()

        raw_gui.progress.step(40)
        raw_gui.load_label["text"] = "Optimising initial params..."
        raw_gui.load_label.update()

        try:
            orgsum_marker = tirf.create_sum(array_TIF, int(raw_gui.start_entry_marker.get()),
                int(int(raw_gui.length_entry_marker.get()) * float(preferences["Sum view frame fraction"][0])),
                raw_gui.view_modes[0])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ An error occurred """
            orgsum_marker = None

        try:
            orgsum_561 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_561.get()),
                                             int(int(raw_gui.length_entry_561.get()) * float(
                                                 preferences["Sum view frame fraction"][1])),
                                         raw_gui.view_modes[1])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ An error occurred """
            orgsum_561 = None

        try:
            orgsum_488 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_488.get()),
                                             int(int(raw_gui.length_entry_488.get()) * float(
                                                 preferences["Sum view frame fraction"][2])),
                                         raw_gui.view_modes[2])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ An error occurred """
            orgsum_488 = None

        raw_gui.progress.step(20)
        raw_gui.load_label["text"] = "Creating views..."
        raw_gui.load_label.update()

        raw_gui.is_data_loaded = True

        create_view_marker(first_time=True)
        raw_gui.progress.step(10)
        create_view_561(first_time=True)
        raw_gui.progress.step(18)
        create_view_488(first_time=True)
        raw_gui.progress.step(2)

        raw_gui.progress.stop()
        raw_gui.progress.place_forget()
        raw_gui.load_label.place_forget()


        raw_gui.session_paths.append(path)
        raw_gui.load_tiff_button["state"] = tk.NORMAL

        raw_gui.raw_marker_spots = []
        raw_gui.raw_561_spots = []
        raw_gui.raw_488_spots = []
        raw_gui.final_561_spots = []
        raw_gui.final_488_spots = []

        raw_gui.traces_calculated = False
        remove_spots_no_update()

    else:
        raw_gui.load_tiff_button["state"] = tk.NORMAL

    raw_gui.set_rawgui_state(tk.NORMAL)


def detect_spots_callback():
    raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.is_data_loaded:
        raw_gui.progress.place(x=740, y=826)
        raw_gui.progress.step(10)
        raw_gui.load_label.place(x=744, y=852)
        raw_gui.load_label["text"] = "Analysing view: Marker..."
        raw_gui.load_label.update()

        if raw_gui.use_marker.get() == 1:
            spots = tirf.detect_spots(0, sum_array_marker, raw_fit_criteria, raw_gui.detection_mode)
            if raw_gui.region_initial[0] is not None and raw_gui.region_final[0] is not None:
                final_spots = []
                for spot in spots:
                    if spot[0] > raw_gui.region_initial[0][0] and spot[0] < raw_gui.region_final[0][0] \
                    and spot[1] > raw_gui.region_initial[0][1] and spot[1] < raw_gui.region_final[0][1]:
                        final_spots.append(spot)
                spots = final_spots
            raw_gui.raw_marker_spots = spots
            # raw_gui.context_marker.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_marker.entryconfig("Spot fitting parameters", state=tk.NORMAL)

        raw_gui.progress.step(30)
        raw_gui.load_label["text"] = "Analysing view: 561nm..."
        raw_gui.load_label.update()

        if raw_gui.use_561.get() == 1:
            if not raw_gui.force_coloc:
                spots = tirf.detect_spots(1, sum_array_561, raw_fit_criteria, raw_gui.detection_mode)
                if raw_gui.region_initial[1] is not None and raw_gui.region_final[1] is not None:
                    final_spots = []
                    for spot in spots:
                        if spot[0] > raw_gui.region_initial[1][0] and spot[0] < raw_gui.region_final[1][0] \
                                and spot[1] > raw_gui.region_initial[1][1] and spot[1] < raw_gui.region_final[1][1]:
                            final_spots.append(spot)
                    spots = final_spots
                raw_gui.raw_561_spots = spots
            # raw_gui.context_561.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_561.entryconfig("Spot fitting parameters", state=tk.NORMAL)
            # raw_gui.context_561.entryconfig("Show Marker Spots", state=tk.NORMAL)

        raw_gui.progress.step(30)
        raw_gui.load_label["text"] = "Analysing view: 488nm..."
        raw_gui.load_label.update()

        if raw_gui.use_488.get() == 1:
            if not raw_gui.force_coloc:
                spots = tirf.detect_spots(2, sum_array_488, raw_fit_criteria, raw_gui.detection_mode)
                if raw_gui.region_initial[2] is not None and raw_gui.region_final[2] is not None:
                    final_spots = []
                    for spot in spots:
                        if spot[0] > raw_gui.region_initial[2][0] and spot[0] < raw_gui.region_final[2][0] \
                                and spot[1] > raw_gui.region_initial[2][1] and spot[1] < raw_gui.region_final[2][1]:
                            final_spots.append(spot)
                    spots = final_spots
                raw_gui.raw_488_spots = spots
            # raw_gui.context_488.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_488.entryconfig("Spot fitting parameters", state=tk.NORMAL)
            # raw_gui.context_488.entryconfig("Show Marker Spots", state=tk.NORMAL)



        if not raw_gui.use_auto_optimization:
            raw_gui.progress.step(29)
            raw_gui.load_label["text"] = "Done."
            raw_gui.load_label.update()
            update_final_spots()
            create_view_488()
            create_view_561()
            create_view_marker()
        else:
            raw_gui.progress.step(29)
            raw_gui.load_label["text"] = "Optimizing Transform..."
            raw_gui.load_label.update()

        raw_gui.progress.stop()
        raw_gui.progress.place_forget()
        raw_gui.load_label.place_forget()

        if raw_gui.use_auto_optimization:
            print("Optimising parameters for chromatic shift correction...")
            colocalization_optimizer()
            create_view_488()
            create_view_561()
            create_view_marker()

    raw_gui.set_rawgui_state(tk.NORMAL)


def colocalization_optimizer():
    global opt_trial_calib
    param_history = []
    loss_history = []

    for x in range(preferences["Calibration optimizer settings"]["xmin"], preferences["Calibration optimizer settings"]["xmax"], preferences["Calibration optimizer settings"]["xstep"]):
        for y in range(preferences["Calibration optimizer settings"]["ymin"], preferences["Calibration optimizer settings"]["ymax"], preferences["Calibration optimizer settings"]["ystep"]):
            for SF in range(preferences["Calibration optimizer settings"]["minscale"], preferences["Calibration optimizer settings"]["maxscale"], preferences["Calibration optimizer settings"]["scalestep"]):
                opt_trial_calib = [x, y, SF, SF]
                calibration["Xc"] = x
                calibration["Yc"] = y
                calibration["SFx"] = SF
                calibration["SFy"] = SF
                update_final_spots()
                # create_view_marker()
                # create_view_561()
                # create_view_488()
                loss_function = calibration_optimizer_error_function(opt_trial_calib)
                param_history.append(opt_trial_calib)
                loss_history.append(loss_function)

    minimum = np.min(loss_history)
    opt_trial_calib = param_history[loss_history.index(minimum)]
    calibration["Xc"] = opt_trial_calib[0]
    calibration["Yc"] = opt_trial_calib[1]
    calibration["SFx"] = opt_trial_calib[2]
    calibration["SFy"] = opt_trial_calib[3]
    print()

    params = tirf.optimize.minimize(calibration_optimizer_finetune_error_function, opt_trial_calib,
                                    options={"maxiter": preferences["Calibration optimizer settings"]["maxiter"]},
                                    method="Nelder-Mead", bounds=[(-512, 1024), (-512, 1024), (60, 400), (60, 400)])
    print(params)
    print()
    print(f"Optimised params: {opt_trial_calib}")
    raw_gui.calib_params.append(opt_trial_calib)
    try:
        calib_window.update_params()
    except:
        """ Window was destroyed """
    update_final_spots()


def redo_frames(view):
    global orgsum_marker, orgsum_561, orgsum_488
    if view == 0:
        try:
            orgsum_marker = tirf.create_sum(array_TIF, int(raw_gui.start_entry_marker.get()),
                                            int(int(raw_gui.length_entry_marker.get()) * float(
                                                preferences["Sum view frame fraction"][0])), raw_gui.view_modes[0])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ An error occurred """
            orgsum_marker = None
        create_view_marker(first_time=True)

    if view == 1:
        try:
            orgsum_561 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_561.get()),
                                        int(int(raw_gui.length_entry_561.get()) * float(
                                            preferences["Sum view frame fraction"][1])), raw_gui.view_modes[1])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ A error occurred """
            orgsum_561 = None
        create_view_561(first_time=True)

    if view == 2:
        try:
            orgsum_488 = tirf.create_sum(array_TIF, int(raw_gui.start_entry_488.get()),
                                        int(int(raw_gui.length_entry_488.get()) * float(
                                            preferences["Sum view frame fraction"][2])), raw_gui.view_modes[2])
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ An error occurred """
            orgsum_488 = None
        create_view_488(first_time=True)


def remove_spots():
    raw_gui.raw_marker_spots = []
    raw_gui.raw_561_spots = []
    raw_gui.raw_488_spots = []
    raw_gui.final_561_spots = []
    raw_gui.final_488_spots = []
    raw_gui.colocalizations = []
    # raw_gui.context_marker.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_marker.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Show Marker Spots", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Show Marker Spots", state=tk.DISABLED)
    if raw_gui.use_marker.get() == 1:
        raw_gui.canvas_marker.delete("all")
        create_view_marker()
    if raw_gui.use_561.get() == 1:
        raw_gui.canvas_561.delete("all")
        create_view_561()
    if raw_gui.use_488.get() == 1:
        raw_gui.canvas_488.delete("all")
        create_view_488()


def remove_spots_no_update():
    raw_gui.raw_marker_spots = []
    raw_gui.raw_561_spots = []
    raw_gui.raw_488_spots = []
    raw_gui.final_561_spots = []
    raw_gui.final_488_spots = []
    raw_gui.colocalizations = []
    # raw_gui.context_marker.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_marker.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_561.entryconfig("Show Marker Spots", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Remove Spot", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Spot fitting parameters", state=tk.DISABLED)
    # raw_gui.context_488.entryconfig("Show Marker Spots", state=tk.DISABLED)
    if raw_gui.use_marker.get() == 1:
        raw_gui.canvas_marker.delete("all")
        create_view_marker()
    if raw_gui.use_561.get() == 1:
        raw_gui.canvas_561.delete("all")
        create_view_561()
    if raw_gui.use_488.get() == 1:
        raw_gui.canvas_488.delete("all")
        create_view_488()


def view_raw_stack():
    global view_raw
    try:
        view_raw.handle_close()
    except:
        """ No window was open """

    view_raw = RawStack()


def calculate_all_traces(auto=False):
    raw_gui.analysis_mode = "all"
    if not auto:
        raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.traces_calculated:
        choice = easygui.ccbox(title="Traces Already Calculated", msg="Warning: Traces from current TIF file have"
                               " already been added to the FluoroTensor import queue. Recalculating will append"
                               " newly calculated traces to this queue. This could result in duplicates.\n\n"
                               "Are You Sure?", choices=["Proceed", "Abort"], default_choice="Abort")
        if not choice:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return
    error = False
    total_spot_count = len(raw_gui.raw_marker_spots) + len(raw_gui.raw_561_spots) + len(raw_gui.raw_488_spots)
    raw_gui.progress["maximum"] = total_spot_count + 2
    raw_gui.progress.place(x=740, y=826)
    raw_gui.progress.step(1)
    raw_gui.load_label.place(x=744, y=852)
    raw_gui.load_label["text"] = "Calculating all traces..."
    raw_gui.load_label.update()
    if raw_gui.is_data_loaded and len(raw_gui.raw_marker_spots) > 0:

        for spot in raw_gui.raw_marker_spots:

            coords = [spot[0], spot[1]]
            # spread = [spot[2][0], spot[2][1]]
            mask = create_mask(1.6, 1.6)


            trace = []
            frame_start = int(float(raw_gui.start_entry_marker.get()))
            frame_end = frame_start + int(float(raw_gui.length_entry_marker.get()))

            h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

            x_low = 5
            x_hi = 6
            y_low = 5
            y_hi = 6

            if coords[0] < 5:
                x_low = coords[0]
            if coords[0] > w - 6:
                x_hi = w - coords[0]
            if coords[1] < 5:
                y_low = coords[1]
            if coords[1] > h - 6:
                y_hi = h - coords[1]

            spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                            int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
            spot_grid = spot_chunk[:, :, 0]

            if coords[0] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[:, shape_diff[1]:]
            if coords[1] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[shape_diff[0]:, :]
            if coords[0] > w-10:
                mask = mask[:, :np.shape(spot_grid)[1]]
            if coords[1] > h-10:
                mask = mask[:np.shape(spot_grid)[0], :]

            trace_org = []

            noise_mask = mask
            spot_mask = 1 - mask

            # plt.imshow(spot_mask)
            # plt.title('Signal Mask')
            # plt.show()
            #
            # plt.imshow(noise_mask)
            # plt.title('Background Mask')
            # plt.show()

            for fr in range(np.shape(spot_chunk)[2]):
                spot_grid = spot_chunk[:, :, fr]
                spot_matrix = spot_grid * spot_mask
                noise_matrix = spot_grid * noise_mask


                # plt.imshow(spot_grid)
                # plt.title('Spot Bounding Box')
                # plt.show()
                #
                # plt.imshow(spot_matrix)
                # plt.title('Isolated Signal')
                # plt.show()
                #
                # plt.imshow(noise_matrix)
                # plt.title('Isolated Background')
                # plt.show()


                noise_pixel_count = np.sum(noise_mask)
                spot_pixel_count = np.sum(spot_mask)

                trace_original = np.sum(spot_matrix) / spot_pixel_count
                noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                trace_corrected = trace_original - noise_intensity

                trace.append(trace_corrected)
                trace_org.append(trace_original)


            raw_gui.all_traces.append(trace)
            raw_gui.raw_traces.append(trace_org)
            raw_gui.trace_info.append(["Cyanine 5", 0, int(float(raw_gui.length_entry_marker.get())), raw_gui.raw_marker_spots.index(spot) + 1, spot])
            raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_marker.get())))))
            raw_gui.progress.step(1)
            raw_gui.progress.update()

        raw_gui.traces_calculated = True

    if raw_gui.is_data_loaded and len(raw_gui.raw_561_spots) > 0:

        for spot in raw_gui.raw_561_spots:

            coords = [spot[0], spot[1]]
            # spread = [spot[2][0], spot[2][1]]
            mask = create_mask(1.6, 1.6)

            trace = []
            frame_start = int(float(raw_gui.start_entry_561.get()))
            frame_end = frame_start + int(float(raw_gui.length_entry_561.get()))

            h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

            x_low = 5
            x_hi = 6
            y_low = 5
            y_hi = 6

            if coords[0] < 5:
                x_low = coords[0]
            if coords[0] > w - 6:
                x_hi = w - coords[0]
            if coords[1] < 5:
                y_low = coords[1]
            if coords[1] > h - 6:
                y_hi = h - coords[1]

            spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                            int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
            spot_grid = spot_chunk[:, :, 0]

            if coords[0] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[:, shape_diff[1]:]
            if coords[1] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[shape_diff[0]:, :]
            if coords[0] > w-10:
                mask = mask[:, :np.shape(spot_grid)[1]]
            if coords[1] > h-10:
                mask = mask[:np.shape(spot_grid)[0], :]

            trace_org = []

            noise_mask = mask
            spot_mask = 1 - mask

            for fr in range(np.shape(spot_chunk)[2]):
                spot_grid = spot_chunk[:, :, fr]
                spot_matrix = spot_grid * spot_mask
                noise_matrix = spot_grid * noise_mask

                noise_pixel_count = np.sum(noise_mask)
                spot_pixel_count = np.sum(spot_mask)

                trace_original = np.sum(spot_matrix) / spot_pixel_count
                noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                trace_corrected = trace_original - noise_intensity

                trace.append(trace_corrected)
                trace_org.append(trace_original)

            raw_gui.all_traces.append(trace)
            raw_gui.raw_traces.append(trace_org)
            raw_gui.trace_info.append(["mCherry", 0, int(float(raw_gui.length_entry_561.get())), raw_gui.raw_561_spots.index(spot) + 1, spot])
            raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_561.get())))))
            raw_gui.progress.step(1)
            raw_gui.progress.update()

        raw_gui.traces_calculated = True

    if raw_gui.is_data_loaded and len(raw_gui.raw_488_spots) > 0:

        for spot in raw_gui.raw_488_spots:

            coords = [spot[0], spot[1]]
            # spread = [spot[2][0], spot[2][1]]
            mask = create_mask(1.6, 1.6)

            trace = []
            frame_start = int(float(raw_gui.start_entry_488.get()))
            frame_end = frame_start + int(float(raw_gui.length_entry_488.get()))

            h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

            x_low = 5
            x_hi = 6
            y_low = 5
            y_hi = 6

            if coords[0] < 5:
                x_low = coords[0]
            if coords[0] > w - 6:
                x_hi = w - coords[0]
            if coords[1] < 5:
                y_low = coords[1]
            if coords[1] > h - 6:
                y_hi = h - coords[1]

            spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                            int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
            spot_grid = spot_chunk[:, :, 0]

            if coords[0] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[:, shape_diff[1]:]
            if coords[1] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[shape_diff[0]:, :]
            if coords[0] > w-10:
                mask = mask[:, :np.shape(spot_grid)[1]]
            if coords[1] > h-10:
                mask = mask[:np.shape(spot_grid)[0], :]

            trace_org = []

            noise_mask = mask
            spot_mask = 1 - mask

            for fr in range(np.shape(spot_chunk)[2]):
                spot_grid = spot_chunk[:, :, fr]
                spot_matrix = spot_grid * spot_mask
                noise_matrix = spot_grid * noise_mask

                noise_pixel_count = np.sum(noise_mask)
                spot_pixel_count = np.sum(spot_mask)

                trace_original = np.sum(spot_matrix) / spot_pixel_count
                noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                trace_corrected = trace_original - noise_intensity

                trace.append(trace_corrected)
                trace_org.append(trace_original)

            raw_gui.all_traces.append(trace)
            raw_gui.raw_traces.append(trace_org)
            raw_gui.trace_info.append(["GFP", 0, int(float(raw_gui.length_entry_488.get())), raw_gui.raw_488_spots.index(spot) + 1, spot])
            raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_488.get())))))
            raw_gui.progress.step(1)
            raw_gui.progress.update()

        raw_gui.traces_calculated = True

    raw_gui.marker_spot_count.append(len(raw_gui.raw_marker_spots))
    raw_gui._488_count.append(len(raw_gui.raw_488_spots))
    raw_gui._561_count.append(len(raw_gui.raw_561_spots))
    raw_gui.background640.append(int(float(raw_gui.bg_entry_marker.get())))
    raw_gui.background561.append(int(float(raw_gui.bg_entry_561.get())))
    raw_gui.background488.append(int(float(raw_gui.bg_entry_488.get())))
    if raw_gui.high_pass:
        raw_gui.enhancements.append(["Wavelet", str(raw_gui.detection_mode)])
    else:
        raw_gui.enhancements.append(["Default", str(raw_gui.detection_mode)])
    raw_gui.file_history_for_export.append(raw_gui.current_data_path)

    raw_gui.progress.stop()
    raw_gui.progress["maximum"] = 100
    raw_gui.progress.place_forget()
    raw_gui.load_label.place_forget()

    if not auto:
        raw_gui.set_rawgui_state(tk.NORMAL)
        if error:
            easygui.msgbox(msg="An error occured while calculating traces. Please check to ensure that frames do not"
                               " exceed the length of the TIF stack.", title="Error!")

        raw_gui.open_traces()


def calculate_non_coloc_traces(auto=False):
    raw_gui.analysis_mode = "all"
    if not auto:
        raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.traces_calculated:
        choice = easygui.ccbox(title="Traces Already Calculated", msg="Warning: Traces from current TIF file have"
                               " already been added to the FluoroTensor import queue. Recalculating will append"
                               " newly calculated traces to this queue. This could result in duplicates.\n\n"
                               "Are You Sure?", choices=["Proceed", "Abort"], default_choice="Abort")
        if not choice:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return
    error = False
    total_spot_count = len(raw_gui.raw_marker_spots) + len(raw_gui.raw_561_spots) + len(raw_gui.raw_488_spots)
    raw_gui.progress["maximum"] = total_spot_count + 2
    raw_gui.progress.place(x=740, y=826)
    raw_gui.progress.step(1)
    raw_gui.load_label.place(x=744, y=852)
    raw_gui.load_label["text"] = "Calculating non-coloc. traces..."
    raw_gui.load_label.update()

    coloc_marker_list = [coloc[0] for coloc in raw_gui.colocalizations if coloc[0] is not None]
    coloc_561_list = [coloc[2] for coloc in raw_gui.colocalizations if coloc[2] is not None]
    coloc_488_list = [coloc[1] for coloc in raw_gui.colocalizations if coloc[1] is not None]

    if raw_gui.is_data_loaded and len(raw_gui.raw_marker_spots) > 0:

        for spot in raw_gui.raw_marker_spots:
            if raw_gui.raw_marker_spots.index(spot) not in coloc_marker_list:
                coords = [spot[0], spot[1]]
                # spread = [spot[2][0], spot[2][1]]
                mask = create_mask(1.6, 1.6)


                trace = []
                frame_start = int(float(raw_gui.start_entry_marker.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_marker.get()))

                h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                                int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
                spot_grid = spot_chunk[:, :, 0]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w-10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h-10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                trace_org = []

                noise_mask = mask
                spot_mask = 1 - mask

                # plt.imshow(spot_mask)
                # plt.title('Signal Mask')
                # plt.show()
                #
                # plt.imshow(noise_mask)
                # plt.title('Background Mask')
                # plt.show()

                for fr in range(np.shape(spot_chunk)[2]):
                    spot_grid = spot_chunk[:, :, fr]
                    spot_matrix = spot_grid * spot_mask
                    noise_matrix = spot_grid * noise_mask


                    # plt.imshow(spot_grid)
                    # plt.title('Spot Bounding Box')
                    # plt.show()
                    #
                    # plt.imshow(spot_matrix)
                    # plt.title('Isolated Signal')
                    # plt.show()
                    #
                    # plt.imshow(noise_matrix)
                    # plt.title('Isolated Background')
                    # plt.show()


                    noise_pixel_count = np.sum(noise_mask)
                    spot_pixel_count = np.sum(spot_mask)

                    trace_original = np.sum(spot_matrix) / spot_pixel_count
                    noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                    trace_corrected = trace_original - noise_intensity

                    trace.append(trace_corrected)
                    trace_org.append(trace_original)


                raw_gui.all_traces.append(trace)
                raw_gui.raw_traces.append(trace_org)
                raw_gui.trace_info.append(["Cyanine 5", 0, int(float(raw_gui.length_entry_marker.get())), raw_gui.raw_marker_spots.index(spot) + 1, spot])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_marker.get())))))
                raw_gui.progress.step(1)
                raw_gui.progress.update()

        raw_gui.traces_calculated = True

    if raw_gui.is_data_loaded and len(raw_gui.raw_561_spots) > 0:

        for spot in raw_gui.raw_561_spots:
            if raw_gui.raw_561_spots.index(spot) not in coloc_561_list:
                coords = [spot[0], spot[1]]
                # spread = [spot[2][0], spot[2][1]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_561.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_561.get()))

                h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                                int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
                spot_grid = spot_chunk[:, :, 0]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w-10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h-10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                trace_org = []

                noise_mask = mask
                spot_mask = 1 - mask

                for fr in range(np.shape(spot_chunk)[2]):
                    spot_grid = spot_chunk[:, :, fr]
                    spot_matrix = spot_grid * spot_mask
                    noise_matrix = spot_grid * noise_mask

                    noise_pixel_count = np.sum(noise_mask)
                    spot_pixel_count = np.sum(spot_mask)

                    trace_original = np.sum(spot_matrix) / spot_pixel_count
                    noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                    trace_corrected = trace_original - noise_intensity

                    trace.append(trace_corrected)
                    trace_org.append(trace_original)

                raw_gui.all_traces.append(trace)
                raw_gui.raw_traces.append(trace_org)
                raw_gui.trace_info.append(["mCherry", 0, int(float(raw_gui.length_entry_561.get())), raw_gui.raw_561_spots.index(spot) + 1, spot])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_561.get())))))
                raw_gui.progress.step(1)
                raw_gui.progress.update()

        raw_gui.traces_calculated = True

    if raw_gui.is_data_loaded and len(raw_gui.raw_488_spots) > 0:

        for spot in raw_gui.raw_488_spots:
            if raw_gui.raw_488_spots.index(spot) not in coloc_488_list:
                coords = [spot[0], spot[1]]
                # spread = [spot[2][0], spot[2][1]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_488.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_488.get()))

                h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                                int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
                spot_grid = spot_chunk[:, :, 0]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w-10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h-10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                trace_org = []

                noise_mask = mask
                spot_mask = 1 - mask

                for fr in range(np.shape(spot_chunk)[2]):
                    spot_grid = spot_chunk[:, :, fr]
                    spot_matrix = spot_grid * spot_mask
                    noise_matrix = spot_grid * noise_mask

                    noise_pixel_count = np.sum(noise_mask)
                    spot_pixel_count = np.sum(spot_mask)

                    trace_original = np.sum(spot_matrix) / spot_pixel_count
                    noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                    trace_corrected = trace_original - noise_intensity

                    trace.append(trace_corrected)
                    trace_org.append(trace_original)

                raw_gui.all_traces.append(trace)
                raw_gui.raw_traces.append(trace_org)
                raw_gui.trace_info.append(["GFP", 0, int(float(raw_gui.length_entry_488.get())), raw_gui.raw_488_spots.index(spot) + 1, spot])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_488.get())))))
                raw_gui.progress.step(1)
                raw_gui.progress.update()

        raw_gui.traces_calculated = True

    raw_gui.marker_spot_count.append(len(raw_gui.raw_marker_spots))
    raw_gui._488_count.append(len(raw_gui.raw_488_spots))
    raw_gui._561_count.append(len(raw_gui.raw_561_spots))
    raw_gui.background640.append(int(float(raw_gui.bg_entry_marker.get())))
    raw_gui.background561.append(int(float(raw_gui.bg_entry_561.get())))
    raw_gui.background488.append(int(float(raw_gui.bg_entry_488.get())))
    if raw_gui.high_pass:
        raw_gui.enhancements.append(["Wavelet", str(raw_gui.detection_mode)])
    else:
        raw_gui.enhancements.append(["Default", str(raw_gui.detection_mode)])
    raw_gui.file_history_for_export.append(raw_gui.current_data_path)

    raw_gui.progress.stop()
    raw_gui.progress["maximum"] = 100
    raw_gui.progress.place_forget()
    raw_gui.load_label.place_forget()

    if not auto:
        raw_gui.set_rawgui_state(tk.NORMAL)
        if error:
            easygui.msgbox(msg="An error occured while calculating traces. Please check to ensure that frames do not"
                               " exceed the length of the TIF stack.", title="Error!")

        raw_gui.open_traces()


# old trace calc. not used, but kept just in case
def calculate_coloc_traces_old(auto=False):
    if not auto:
        raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.traces_calculated:
        choice = easygui.ccbox(title="Traces Already Calculated", msg="Warning: Traces from current TIF file have"
                               " already been added to the FluoroTensor import queue. Recalculating will append"
                               " newly calculated traces to this queue. This could result in duplicates.\n\n"
                               "Are You Sure?", choices=["Proceed", "Abort"], default_choice="Abort")
        if not choice:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return
    error = False
    total_spot_count = len(raw_gui.colocalizations)
    raw_gui.progress["maximum"] = total_spot_count + 1
    raw_gui.progress.place(x=740, y=826)
    raw_gui.progress.step(1)
    raw_gui.load_label.place(x=744, y=852)
    raw_gui.load_label["text"] = "Calculating coloc. traces..."
    raw_gui.load_label.update()
    if len(raw_gui.colocalizations) > 0:
        for coloc in raw_gui.colocalizations:
            coords = [raw_gui.raw_marker_spots[coloc[0]][0], raw_gui.raw_marker_spots[coloc[0]][1]]
            mask = create_mask(1.6, 1.6)

            trace = []
            frame_start = int(float(raw_gui.start_entry_marker.get()))
            frame_end = frame_start + int(float(raw_gui.length_entry_marker.get()))
            for fr in range(frame_start, frame_end):
                try:
                    detection_field = np.copy(array_TIF[:, :, fr])
                except:
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                    error = True

                h, w = np.shape(detection_field)[0], np.shape(detection_field)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_grid = detection_field[int(coords[1] - y_low):int(coords[1] + y_hi),
                            int(coords[0] - x_low):int(coords[0] + x_hi)]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w-10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h-10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                residual_matrix = spot_grid * mask
                mask_size = np.shape(mask)[0] * np.shape(mask)[1]
                mask_ratio = mask_size / np.sum(mask)

                trace_intensity = np.sum(spot_grid)
                background = np.sum(residual_matrix) * mask_ratio
                trace_corrected = trace_intensity - background
                trace_corrected = trace_corrected / (mask_size - np.sum(mask))

                trace.append(trace_corrected)

            raw_gui.all_traces.append(trace)
            raw_gui.trace_info.append(
                ["Cyanine 5", 0, int(float(raw_gui.length_entry_marker.get())),
                 coloc[0] + 1])
            raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_marker.get())))))


            if coloc[1] is not None:
                coords = [raw_gui.raw_488_spots[coloc[1]][0], raw_gui.raw_488_spots[coloc[1]][1]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_488.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_488.get()))
                for fr in range(frame_start, frame_end):
                    try:
                        detection_field = np.copy(array_TIF[:, :, fr])
                    except:
                        print("Warning error occured: Full traceback shown below:")
                        print(traceback.format_exc())
                        error = True

                    h, w = np.shape(detection_field)[0], np.shape(detection_field)[1]

                    x_low = 5
                    x_hi = 6
                    y_low = 5
                    y_hi = 6

                    if coords[0] < 5:
                        x_low = coords[0]
                    if coords[0] > w - 6:
                        x_hi = w - coords[0]
                    if coords[1] < 5:
                        y_low = coords[1]
                    if coords[1] > h - 6:
                        y_hi = h - coords[1]

                    spot_grid = detection_field[int(coords[1] - y_low):int(coords[1] + y_hi),
                                int(coords[0] - x_low):int(coords[0] + x_hi)]

                    if coords[0] < 10:
                        shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                        mask = mask[:, shape_diff[1]:]
                    if coords[1] < 10:
                        shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                        mask = mask[shape_diff[0]:, :]
                    if coords[0] > w-10:
                        mask = mask[:, :np.shape(spot_grid)[1]]
                    if coords[1] > h-10:
                        mask = mask[:np.shape(spot_grid)[0], :]

                    residual_matrix = spot_grid * mask
                    mask_size = np.shape(mask)[0] * np.shape(mask)[1]
                    mask_ratio = mask_size / np.sum(mask)

                    trace_intensity = np.sum(spot_grid)
                    background = np.sum(residual_matrix) * mask_ratio
                    trace_corrected = trace_intensity - background
                    trace_corrected = trace_corrected / (mask_size - np.sum(mask))

                    trace.append(trace_corrected)

                raw_gui.all_traces.append(trace)
                raw_gui.trace_info.append(
                    ["GFP", 0, int(float(raw_gui.length_entry_488.get())),
                     coloc[1] + 1])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_488.get())))))


            if coloc[2] is not None:
                coords = [raw_gui.raw_561_spots[coloc[2]][0], raw_gui.raw_561_spots[coloc[2]][1]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_561.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_561.get()))
                for fr in range(frame_start, frame_end):
                    try:
                        detection_field = np.copy(array_TIF[:, :, fr])
                    except:
                        print("Warning error occured: Full traceback shown below:")
                        print(traceback.format_exc())
                        error = True

                    h, w = np.shape(detection_field)[0], np.shape(detection_field)[1]

                    x_low = 5
                    x_hi = 6
                    y_low = 5
                    y_hi = 6

                    if coords[0] < 5:
                        x_low = coords[0]
                    if coords[0] > w - 6:
                        x_hi = w - coords[0]
                    if coords[1] < 5:
                        y_low = coords[1]
                    if coords[1] > h - 6:
                        y_hi = h - coords[1]

                    spot_grid = detection_field[int(coords[1] - y_low):int(coords[1] + y_hi),
                                int(coords[0] - x_low):int(coords[0] + x_hi)]

                    if coords[0] < 10:
                        shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                        mask = mask[:, shape_diff[1]:]
                    if coords[1] < 10:
                        shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                        mask = mask[shape_diff[0]:, :]
                    if coords[0] > w-10:
                        mask = mask[:, :np.shape(spot_grid)[1]]
                    if coords[1] > h-10:
                        mask = mask[:np.shape(spot_grid)[0], :]

                    residual_matrix = spot_grid * mask
                    mask_size = np.shape(mask)[0] * np.shape(mask)[1]
                    mask_ratio = mask_size / np.sum(mask)

                    trace_intensity = np.sum(spot_grid)
                    background = np.sum(residual_matrix) * mask_ratio
                    trace_corrected = trace_intensity - background
                    trace_corrected = trace_corrected / (mask_size - np.sum(mask))

                    trace.append(trace_corrected)

                raw_gui.all_traces.append(trace)
                raw_gui.trace_info.append(
                    ["mCherry", 0, int(float(raw_gui.length_entry_561.get())),
                     coloc[2] + 1])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_561.get())))))

            raw_gui.progress.step(1)
            raw_gui.progress.update()

        raw_gui.traces_calculated = True

    raw_gui.marker_spot_count.append(len(raw_gui.raw_marker_spots))
    raw_gui._488_count.append(len(raw_gui.raw_488_spots))
    raw_gui._561_count.append(len(raw_gui.raw_561_spots))

    coloc561 = 0
    coloc488 = 0
    if len(raw_gui.colocalizations) > 0:
        for index in range(len(raw_gui.colocalizations)):
            if raw_gui.colocalizations[index][1] is not None:
                coloc488 += 1
            if raw_gui.colocalizations[index][2] is not None:
                coloc561 += 1

    raw_gui.coloc_561.append(coloc561)
    raw_gui.coloc_488.append(coloc488)

    raw_gui.progress.stop()
    raw_gui.progress["maximum"] = 100
    raw_gui.progress.place_forget()
    raw_gui.load_label.place_forget()

    if not auto:
        raw_gui.set_rawgui_state(tk.NORMAL)
        if error:
            easygui.msgbox(msg="An error occured while calculating traces. Please check to ensure that frames do not"
                               " exceed the length of the TIF stack.", title="Error!")

        raw_gui.open_traces()


def calculate_coloc_traces(auto=False):
    raw_gui.analysis_mode = "coloc"
    if not auto:
        raw_gui.set_rawgui_state(tk.DISABLED)
    if raw_gui.traces_calculated:
        choice = easygui.ccbox(title="Traces Already Calculated", msg="Warning: Traces from current TIF file have"
                               " already been added to the FluoroTensor import queue. Recalculating will append"
                               " newly calculated traces to this queue. This could result in duplicates.\n\n"
                               "Are You Sure?", choices=["Proceed", "Abort"], default_choice="Abort")
        if not choice:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return
    error = False
    total_spot_count = len(raw_gui.colocalizations)
    raw_gui.progress["maximum"] = total_spot_count + 2
    raw_gui.progress.place(x=740, y=826)
    raw_gui.progress.step(1)
    raw_gui.load_label.place(x=744, y=852)
    raw_gui.load_label["text"] = "Calculating coloc. traces..."
    raw_gui.load_label.update()
    if len(raw_gui.colocalizations) > 0:
        for coloc in raw_gui.colocalizations:
            coords = [raw_gui.raw_marker_spots[coloc[0]][0], raw_gui.raw_marker_spots[coloc[0]][1]]
            spot = raw_gui.raw_marker_spots[coloc[0]]
            mask = create_mask(1.6, 1.6)

            trace = []
            frame_start = int(float(raw_gui.start_entry_marker.get()))
            frame_end = frame_start + int(float(raw_gui.length_entry_marker.get()))

            h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

            x_low = 5
            x_hi = 6
            y_low = 5
            y_hi = 6

            if coords[0] < 5:
                x_low = coords[0]
            if coords[0] > w - 6:
                x_hi = w - coords[0]
            if coords[1] < 5:
                y_low = coords[1]
            if coords[1] > h - 6:
                y_hi = h - coords[1]

            spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                            int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
            spot_grid = spot_chunk[:, :, 0]

            if coords[0] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[:, shape_diff[1]:]
            if coords[1] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[shape_diff[0]:, :]
            if coords[0] > w-10:
                mask = mask[:, :np.shape(spot_grid)[1]]
            if coords[1] > h-10:
                mask = mask[:np.shape(spot_grid)[0], :]

            # print(np.shape(spot_chunk))
            trace_org = []
            noise_mask = mask
            spot_mask = 1 - mask

            for fr in range(np.shape(spot_chunk)[2]):
                spot_grid = spot_chunk[:, :, fr]
                spot_matrix = spot_grid * spot_mask
                noise_matrix = spot_grid * noise_mask

                noise_pixel_count = np.sum(noise_mask)
                spot_pixel_count = np.sum(spot_mask)

                trace_original = np.sum(spot_matrix) / spot_pixel_count
                noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                trace_corrected = trace_original - noise_intensity

                trace.append(trace_corrected)
                trace_org.append(trace_original)

            raw_gui.all_traces.append(trace)
            raw_gui.raw_traces.append(trace_org)
            raw_gui.trace_info.append(
                ["Cyanine 5", 0, int(float(raw_gui.length_entry_marker.get())),
                 coloc[0] + 1, spot])
            raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_marker.get())))))


            if coloc[1] is not None:
                coords = [raw_gui.raw_488_spots[coloc[1]][0], raw_gui.raw_488_spots[coloc[1]][1]]
                spot = raw_gui.raw_488_spots[coloc[1]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_488.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_488.get()))

                h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                             int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
                spot_grid = spot_chunk[:, :, 0]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w - 10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h - 10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                # print(np.shape(spot_chunk))
                trace_org = []
                noise_mask = mask
                spot_mask = 1 - mask

                for fr in range(np.shape(spot_chunk)[2]):
                    spot_grid = spot_chunk[:, :, fr]
                    spot_matrix = spot_grid * spot_mask
                    noise_matrix = spot_grid * noise_mask

                    noise_pixel_count = np.sum(noise_mask)
                    spot_pixel_count = np.sum(spot_mask)

                    trace_original = np.sum(spot_matrix) / spot_pixel_count
                    noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                    trace_corrected = trace_original - noise_intensity

                    trace.append(trace_corrected)
                    trace_org.append(trace_original)

                raw_gui.all_traces.append(trace)
                raw_gui.raw_traces.append(trace_org)
                raw_gui.trace_info.append(
                    ["GFP", 0, int(float(raw_gui.length_entry_488.get())),
                     coloc[1] + 1, spot])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_488.get())))))


            if coloc[2] is not None:
                coords = [raw_gui.raw_561_spots[coloc[2]][0], raw_gui.raw_561_spots[coloc[2]][1]]
                spot = raw_gui.raw_561_spots[coloc[2]]
                mask = create_mask(1.6, 1.6)

                trace = []
                frame_start = int(float(raw_gui.start_entry_561.get()))
                frame_end = frame_start + int(float(raw_gui.length_entry_561.get()))
                h, w = np.shape(array_TIF)[0], np.shape(array_TIF)[1]

                x_low = 5
                x_hi = 6
                y_low = 5
                y_hi = 6

                if coords[0] < 5:
                    x_low = coords[0]
                if coords[0] > w - 6:
                    x_hi = w - coords[0]
                if coords[1] < 5:
                    y_low = coords[1]
                if coords[1] > h - 6:
                    y_hi = h - coords[1]

                spot_chunk = array_TIF[int(coords[1] - y_low):int(coords[1] + y_hi),
                             int(coords[0] - x_low):int(coords[0] + x_hi), frame_start:frame_end]
                spot_grid = spot_chunk[:, :, 0]

                if coords[0] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[:, shape_diff[1]:]
                if coords[1] < 10:
                    shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                    mask = mask[shape_diff[0]:, :]
                if coords[0] > w - 10:
                    mask = mask[:, :np.shape(spot_grid)[1]]
                if coords[1] > h - 10:
                    mask = mask[:np.shape(spot_grid)[0], :]

                # print(np.shape(spot_chunk))
                trace_org = []
                noise_mask = mask
                spot_mask = 1 - mask

                for fr in range(np.shape(spot_chunk)[2]):
                    spot_grid = spot_chunk[:, :, fr]
                    spot_matrix = spot_grid * spot_mask
                    noise_matrix = spot_grid * noise_mask

                    noise_pixel_count = np.sum(noise_mask)
                    spot_pixel_count = np.sum(spot_mask)

                    trace_original = np.sum(spot_matrix) / spot_pixel_count
                    noise_intensity = np.sum(noise_matrix) / noise_pixel_count

                    trace_corrected = trace_original - noise_intensity

                    trace.append(trace_corrected)
                    trace_org.append(trace_original)

                raw_gui.all_traces.append(trace)
                raw_gui.raw_traces.append(trace_org)
                raw_gui.trace_info.append(
                    ["mCherry", 0, int(float(raw_gui.length_entry_561.get())),
                     coloc[2] + 1, spot])
                raw_gui.all_fits.append(list(np.zeros(int(float(raw_gui.length_entry_561.get())))))

            raw_gui.progress.step(1)
            raw_gui.progress.update()

        raw_gui.traces_calculated = True

    raw_gui.marker_spot_count.append(len(raw_gui.raw_marker_spots))
    raw_gui._488_count.append(len(raw_gui.raw_488_spots))
    raw_gui._561_count.append(len(raw_gui.raw_561_spots))

    coloc561 = 0
    coloc488 = 0
    if len(raw_gui.colocalizations) > 0:
        for index in range(len(raw_gui.colocalizations)):
            if raw_gui.colocalizations[index][1] is not None:
                coloc488 += 1
            if raw_gui.colocalizations[index][2] is not None:
                coloc561 += 1

    raw_gui.coloc_561.append(coloc561)
    raw_gui.coloc_488.append(coloc488)
    raw_gui.background640.append(int(float(raw_gui.bg_entry_marker.get())))
    raw_gui.background561.append(int(float(raw_gui.bg_entry_561.get())))
    raw_gui.background488.append(int(float(raw_gui.bg_entry_488.get())))
    if raw_gui.high_pass:
        raw_gui.enhancements.append(["Wavelet", str(raw_gui.detection_mode)])
    else:
        raw_gui.enhancements.append(["Default", str(raw_gui.detection_mode)])
    raw_gui.file_history_for_export.append(raw_gui.current_data_path)


    raw_gui.progress.stop()
    raw_gui.progress["maximum"] = 100
    raw_gui.progress.place_forget()
    raw_gui.load_label.place_forget()

    if not auto:
        raw_gui.set_rawgui_state(tk.NORMAL)
        if error:
            easygui.msgbox(msg="An error occured while calculating traces. Please check to ensure that frames do not"
                               " exceed the length of the TIF stack.", title="Error!")

        raw_gui.open_traces()


def quality_check():
    index_choice = easygui.indexbox(title="Quality Control Options", msg="Load automation index file or curate new folder?",
                                    choices=["Load Index File", "Curate New Folder"],
                                    default_choice="Curate New Folder")
    if index_choice is None:
        return
    elif index_choice == 0:
        path = easygui.fileopenbox(title="Load Automation Index File", default=default_dir + "*.ind",
                                   filetypes=["*.aif", "Automation Index Files"])
        if path:
            with open(path, "rb") as load_index:
                raw_gui.auto_movie_list = pickle.load(load_index)
            check_index_file(path)
        return
    elif index_choice == 1:
        """ Continue to quality control """

    global quality_window

    raw_gui.session_paths = []
    raw_gui.auto_movie_list = []
    folder = easygui.diropenbox(title="Open folder containing TIF stacks from 1 single molecule experiment",
                                default=default_dir)
    if not folder:
        return
    for file in os.listdir(folder):
        if file.endswith(".tif"):
            raw_gui.auto_movie_list.append(os.path.join(folder, file))

    print(raw_gui.auto_movie_list)

    quality_window = QualityWin()
    quality_window.progress["maximum"] = len(raw_gui.auto_movie_list)


def check_index_file(path):
    is_path_valid = os.path.exists(raw_gui.auto_movie_list[0])
    if is_path_valid:
        return
    choice = easygui.indexbox(title="Data Moved!", msg="Warning! TIF files no longer exist in the location specified "
                                                       "by the index file. You may have moved or renamed the folder. "
                                                       "You will now be prompted to browse the new location of the folder "
                                                       "containing the raw data. FluoroTensor will then overwrite the "
                                                       "index file with the correct file paths according to the new folder "
                                                       "location. Attempting to automate without correcting may cause "
                                                       "FluoroTensor to crash.",
                              choices=["     Browse Folder Location      ", "Use Folder Containing Index File", "                Cancel                "], default_choice=0)
    if choice == 2 or choice is None:
        raw_gui.auto_movie_list = []
        return
    if choice == 1:
        new_folder_path = os.path.split(path)[0]
    elif choice == 0:
        new_folder_path = easygui.diropenbox(title="Select folder containing data.", default=default_dir)

    for idx, file_path in enumerate(raw_gui.auto_movie_list):
        file_stub = os.path.split(file_path)[1]
        new_path = os.path.join(new_folder_path, file_stub)
        raw_gui.auto_movie_list[idx] = new_path

    with open(path, "wb") as save_indices:
        pickle.dump(raw_gui.auto_movie_list, save_indices)


def automate():
    global auto_window
    raw_gui.set_rawgui_state(tk.DISABLED)
    raw_gui.all_traces = []
    raw_gui.all_fits = []
    raw_gui.trace_info = []
    raw_gui.traces_calculated = False
    raw_gui.session_paths = []

    raw_gui.marker_spot_count = []
    raw_gui.coloc_488 = []
    raw_gui.coloc_561 = []
    raw_gui._488_count = []
    raw_gui._561_count = []
    raw_gui.calib_params = []
    raw_gui.background640 = []
    raw_gui.background561 = []
    raw_gui.background488 = []
    raw_gui.enhancements = []
    raw_gui.file_history_for_export = []

    tif_stack_list = []
    used_list = False

    if len(raw_gui.auto_movie_list) > 0:
        list_choice = easygui.indexbox(title="File Source", msg="Use previously created quality control list?",
                                       choices=["           Use List           ", "Open Different Folder"],
                                       default_choice="           Use List           ")
        if list_choice is None:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return

        if list_choice == 0:
            tif_stack_list = raw_gui.auto_movie_list
            used_list = True
        else:
            folder = easygui.diropenbox(title="Open folder containing TIF stacks from 1 single molecule experiment",
                                default=default_dir)
            if not folder:
                raw_gui.set_rawgui_state(tk.NORMAL)
                return
            for file in os.listdir(folder):
                if file.endswith(".tif"):
                    tif_stack_list.append(os.path.join(folder, file))
    else:
        folder = easygui.diropenbox(title="Open folder containing TIF stacks from 1 single molecule experiment",
                                    default=default_dir)
        if not folder:
            raw_gui.set_rawgui_state(tk.NORMAL)
            return
        for file in os.listdir(folder):
            if file.endswith(".tif"):
                tif_stack_list.append(os.path.join(folder, file))


    print(tif_stack_list)
    if len(tif_stack_list) == 0:
        raw_gui.set_rawgui_state(tk.NORMAL)
        return

    choice = easygui.indexbox(title="Automation Mode", msg="Calculate colocalized traces or all traces?",
                              choices=["  Colocalized  ", "          All          ", "Non Colocalized"], default_choice="Colocalized")

    if choice is None:
        raw_gui.set_rawgui_state(tk.NORMAL)
        return

    auto_window = AutoWin()
    auto_window.progress["maximum"] = len(tif_stack_list)

    for tif_file in tif_stack_list:
        auto_window.label["text"] = "Automated analysis is in progress... (" + str(
                                    tif_stack_list.index(tif_file) + 1) + " / " + str(len(tif_stack_list)) + ")"
        auto_window.label.update()

        if auto_window.CancelFlag:
            break
        load_tiff(auto=True, file=tif_file)
        raw_gui.set_rawgui_state(tk.DISABLED)
        detect_spots_callback()
        raw_gui.set_rawgui_state(tk.DISABLED)
        if choice == 0:
            calculate_coloc_traces(auto=True)
        elif choice == 1:
            calculate_all_traces(auto=True)
        elif choice == 2:
            calculate_non_coloc_traces(auto=True)
        auto_window.progress.step(1)
        auto_window.progress.update()

    raw_gui.set_rawgui_state(tk.NORMAL)
    auto_window.handle_close()
    raw_gui.open_traces()


def update_final_spots():
    raw_gui.final_561_spots = []
    raw_gui.final_488_spots = []

    if raw_gui.force_coloc:
        raw_gui.raw_561_spots = []
        raw_gui.raw_488_spots = []

        if raw_gui.use_561.get() == 1 and len(raw_gui.raw_marker_spots) > 0:
            for spot in raw_gui.raw_marker_spots:
                newx, newy = reverse_coordinate_transform(spot[0], spot[1], 0.35)
                spot2 = list(spot)
                spot2[0], spot2[1] = newx, newy
                raw_gui.raw_561_spots.append(spot2)
        if raw_gui.use_488.get() == 1 and len(raw_gui.raw_marker_spots) > 0:
            for spot in raw_gui.raw_marker_spots:
                newx, newy = reverse_coordinate_transform(spot[0], spot[1], 1)
                spot2 = list(spot)
                spot2[0], spot2[1] = newx, newy
                raw_gui.raw_488_spots.append(spot2)

    else:
        if raw_gui.marker_view_colour == "640":
            if raw_gui.use_561.get() == 1 and len(raw_gui.raw_561_spots) > 0:
                for spot in raw_gui.raw_561_spots:
                    newx, newy = coordinate_transform(spot[0], spot[1], 0.35)
                    raw_gui.final_561_spots.append([newx, newy])

            if raw_gui.use_488.get() == 1 and len(raw_gui.raw_488_spots) > 0:
                for spot in raw_gui.raw_488_spots:
                    newx, newy = coordinate_transform(spot[0], spot[1], 1)
                    raw_gui.final_488_spots.append([newx, newy])

        elif raw_gui.marker_view_colour == "561":
            if raw_gui.use_488.get() == 1 and len(raw_gui.raw_488_spots) > 0:
                for spot in raw_gui.raw_488_spots:
                    newx, newy = coordinate_transform(spot[0], spot[1], 0.65)
                    raw_gui.final_488_spots.append([newx, newy])

            if raw_gui.use_561.get() == 1 and len(raw_gui.raw_561_spots) > 0:
                for spot in raw_gui.raw_561_spots:
                    newx, newy = spot[0], spot[1]
                    raw_gui.final_561_spots.append([newx, newy])

        elif raw_gui.marker_view_colour == "488":
            if raw_gui.use_561.get() == 1 and len(raw_gui.raw_561_spots) > 0:
                for spot in raw_gui.raw_561_spots:
                    newx, newy = reverse_coordinate_transform(spot[0], spot[1], 0.65)
                    raw_gui.final_561_spots.append([newx, newy])

            if raw_gui.use_488.get() == 1 and len(raw_gui.raw_488_spots) > 0:
                for spot in raw_gui.raw_488_spots:
                    newx, newy = spot[0], spot[1]
                    raw_gui.final_488_spots.append([newx, newy])


    update_coloc()


def update_coloc():
    raw_gui.colocalizations = []

    if raw_gui.use_marker.get() == 1 and len(raw_gui.raw_marker_spots) > 0:
        for index_marker, spot_marker in enumerate(raw_gui.raw_marker_spots):
            coloc = [index_marker]

            if raw_gui.use_488.get() == 1 and len(raw_gui.final_488_spots) > 0:
                distances = []
                for spot_488 in raw_gui.final_488_spots:
                    distances.append(np.sqrt((spot_marker[0] - spot_488[0])**2 + (spot_marker[1] - spot_488[1])**2))
                minimum = min(distances)
                if minimum < calibration["criterion"]:
                    index_488 = distances.index(minimum)
                    coloc.append(index_488)
                else:
                    coloc.append(None)
            else:
                coloc.append(None)

            if raw_gui.use_561.get() == 1 and len(raw_gui.final_561_spots) > 0:
                distances = []
                for spot_561 in raw_gui.final_561_spots:
                    distances.append(np.sqrt((spot_marker[0] - spot_561[0]) ** 2 + (spot_marker[1] - spot_561[1]) ** 2))
                minimum = min(distances)
                if minimum < calibration["criterion"]:
                    index_561 = distances.index(minimum)
                    coloc.append(index_561)
                else:
                    coloc.append(None)
            else:
                coloc.append(None)

            if coloc[1] is not None or coloc[2] is not None:
                raw_gui.colocalizations.append(coloc)


def calibration_optimizer_error_function(params):
    count = 0
    for coloc in raw_gui.colocalizations:
        if coloc[1] is not None:
            count += 1
        if coloc[2] is not None:
            count += 1
    loss = 1 / (count + 0.1)
    print(f"Loss: {loss}       Spots colocalized: {count}       Parameters: {params}")
    return loss


def calibration_optimizer_finetune_error_function(params):
    global opt_trial_calib
    opt_trial_calib = params
    update_final_spots()
    squared_error = 0
    for coloc in raw_gui.colocalizations:
        error_561 = 0
        error_488 = 0
        if coloc[0] is not None:
            coords_marker = [raw_gui.raw_marker_spots[coloc[0]][0], raw_gui.raw_marker_spots[coloc[0]][1]]
        if coloc[0] is not None and coloc[1] is not None:
            coords_488_mapped = [raw_gui.final_488_spots[coloc[1]][0], raw_gui.final_488_spots[coloc[1]][1]]
            error_488 = (coords_488_mapped[0] - coords_marker[0]) ** 2 + (coords_488_mapped[1] - coords_marker[1]) ** 2
        if coloc[0] is not None and coloc[2] is not None:
            coords_561_mapped = [raw_gui.final_561_spots[coloc[2]][0], raw_gui.final_561_spots[coloc[2]][1]]
            error_561 = (coords_561_mapped[0] - coords_marker[0]) ** 2 + (coords_561_mapped[1] - coords_marker[1]) ** 2

        squared_error += (error_561 + error_488)
    print(f"Loss: {round(squared_error, 9)}        Parameters: {params}")
    return round(squared_error, 9)



def coordinate_transform(x, y, factor):
    """ Transform the coordinates of a secondary spot onto the marker """

    if raw_gui.use_auto_optimization:
        calibration["Xc"], calibration["Yc"] = round(opt_trial_calib[0], 3), round(opt_trial_calib[1], 3)
        calibration["SFx"], calibration["SFy"] = round(opt_trial_calib[2], 3), round(opt_trial_calib[3], 3)

    delta_x = factor * (x + calibration["X0"] - calibration["Xc"]) / calibration["SFx"]
    delta_y = factor * (y + calibration["Y0"] - calibration["Yc"]) / calibration["SFy"]

    return x - delta_x, y - delta_y


def reverse_coordinate_transform(x, y ,factor):
    """ Transform the coordinates of a marker spot onto the expected position of a colocalized secondary spot """

    if raw_gui.use_auto_optimization:
        calibration["Xc"], calibration["Yc"] = round(opt_trial_calib[0], 3), round(opt_trial_calib[1], 3)
        calibration["SFx"], calibration["SFy"] = round(opt_trial_calib[2], 3), round(opt_trial_calib[3], 3)

    delta_x = factor * (x + calibration["X0"] - calibration["Xc"]) / calibration["SFx"]
    delta_y = factor * (y + calibration["Y0"] - calibration["Yc"]) / calibration["SFy"]

    return x + delta_x, y + delta_y


def linear_regression(series_x, series_y):
    x_sqrd_sum = 0
    xy_sum = 0
    x_sum = 0
    y_sum = 0
    for ind in range(len(series_x)):
        x_sqrd_sum += series_x[ind] ** 2
        xy_sum += series_x[ind] * series_y[ind]
        x_sum += series_x[ind]
        y_sum += series_y[ind]

    gradient = (len(series_x) * xy_sum - x_sum * y_sum) / (len(series_x) * x_sqrd_sum - x_sum**2)
    intercept = (y_sum - gradient*x_sum) / len(series_x)

    return gradient, intercept


def forced_zero_intercept_linear_regression(series_x, series_y):
    xy_sum = 0
    x_squared_sum = 0
    for index in range(len(series_x)):
        x_squared_sum += series_x[index] ** 2
        xy_sum += series_x[index] * series_y[index]
    gradient = (xy_sum / len(series_x)) / (x_squared_sum / len(series_x))
    return gradient


def save_and_exit_raw():
    global all_fits
    global all_traces
    global all_sublist
    global active_trace_list
    global trace_info
    global Cy5_sublist
    global mCherry_sublist
    global GFP_sublist
    global total_trace_count
    global Cy5_trace_count
    global mCherry_trace_count
    global GFP_trace_count
    global file_path
    global current_trace
    global load_from_excel
    global is_data_loaded
    global has_manual_fit_happened, ready_to_export, used_neural_network
    global trim_undo_reference_stack, trim_undo_frames_stack, auto_trim_backup, auto_trim_fit_backup
    global waiting_for_fit
    global colocalization_data

    if not len(raw_gui.all_traces) > 0:
        return

    load_from_excel = False
    all_fits = []
    all_traces = []
    all_sublist = []
    active_trace_list = []
    trace_info = []
    Cy5_sublist = []
    mCherry_sublist = []
    GFP_sublist = []
    total_trace_count = 0
    Cy5_trace_count = 0
    mCherry_trace_count = 0
    GFP_trace_count = 0
    file_path = None
    colocalization_data = None
    current_trace = 0
    is_data_loaded = True
    has_manual_fit_happened = False
    ready_to_export = False
    used_neural_network = False
    trim_undo_reference_stack = []
    trim_undo_frames_stack = []
    auto_trim_backup = []
    auto_trim_fit_backup = []
    waiting_for_fit = False

    if capture_mouse:
        trace_figure.canvas.mpl_disconnect(capture_mouse)

    all_traces = raw_gui.all_traces
    all_fits = raw_gui.all_fits
    trace_info = raw_gui.trace_info

    for count in range(len(trace_info)):
        fitting_params = trace_info[count].pop(-1)
        if len(trace_info[count]) < 5:
            trace_info[count].append([])
        if len(trace_info[count]) < 6:
            trace_info[count].append(None)
            trace_info[count].append(None)
            trace_info[count].append(None)
            trace_info[count].append([])
            trace_info[count].append(fitting_params)

    total_trace_count = len(trace_info)

    for count in range(total_trace_count):
        if trace_info[count][0] == "Cyanine 5":
            Cy5_trace_count += 1
            Cy5_sublist.append(count)
        if trace_info[count][0] == "mCherry":
            mCherry_trace_count += 1
            mCherry_sublist.append(count)
        if trace_info[count][0] == "GFP":
            GFP_trace_count += 1
            GFP_sublist.append(count)
        all_sublist.append(count)
    active_trace_list = all_sublist
    Cy5_trace_count = len(Cy5_sublist)
    mCherry_trace_count = len(mCherry_sublist)
    GFP_trace_count = len(GFP_sublist)

    tr_mode = raw_gui.analysis_mode
    print(tr_mode)

    colocalization_data = ColocData(raw_gui.marker_spot_count, raw_gui._488_count, raw_gui._561_count,
                                    raw_gui.coloc_488, raw_gui.coloc_561, tr_mode, raw_gui.calib_params,
                                    raw_gui.background640, raw_gui.background561, raw_gui.background488,
                                    raw_gui.enhancements, raw_gui.file_history_for_export)

    raw_gui.traces_calculated = False
    raw_gui.handle_raw_close()

    plot_trace(active_trace_list[current_trace])
    update_infobox()

    easygui.msgbox(title="Trace count", msg="Found " + str(total_trace_count) + " traces, of which\n" +
                   str(Cy5_trace_count) + " were " + preferences["Fluorophore config"][0] + " marker spots,\n" +
                   str(mCherry_trace_count) + " were " + preferences["Fluorophore config"][1] + " spots, and \n" +
                   str(GFP_trace_count) + " were " + preferences["Fluorophore config"][2] + " spots.")


def set_raw_defaults():
    raw_gui.defaults_button["state"] = tk.DISABLED
    global raw_defaults
    raw_defaults = RawDefaults()


def set_spot_criteria(view):
    raw_gui.criteria_marker["state"] = tk.DISABLED
    raw_gui.criteria_561["state"] = tk.DISABLED
    raw_gui.criteria_488["state"] = tk.DISABLED

    global spot_criteria
    spot_criteria = SpotCriteria(view)


def set_calibration():
    global calib_window
    calib_window = CalibWin()


def set_view_colour():
    global view_colour_window
    view_colour_window = ColWin()


def load_calibration():
    global calibration
    cwd = os.getcwd()
    path = cwd + "/calibration.xlsx"

    try:
        calib_sheet = opxl.load_workbook(path)
        sheet = calib_sheet["calibrations"]
        date_list = []
        force_list = []
        for check_column in range(2, 500):
            date = str(sheet.cell(row=1, column=check_column).value)
            force = str(sheet.cell(row=2, column=check_column).value)
            if date != "None":
                date = int(date)
            else:
                date = 0
            date_list.append(date)
            force_list.append(force)

        most_recent_column = date_list.index(max(date_list)) + 2
        if "F" in force_list:
            most_recent_column = force_list.index("F") + 2

        calibration["date"] = int(str(sheet.cell(row=1, column=most_recent_column).value))
        calibration["X0"] = float(str(sheet.cell(row=3, column=most_recent_column).value))
        calibration["Y0"] = float(str(sheet.cell(row=4, column=most_recent_column).value))
        calibration["Xc"] = float(str(sheet.cell(row=5, column=most_recent_column).value))
        calibration["Yc"] = float(str(sheet.cell(row=6, column=most_recent_column).value))
        calibration["SFx"] = float(str(sheet.cell(row=7, column=most_recent_column).value))
        calibration["SFy"] = float(str(sheet.cell(row=8, column=most_recent_column).value))
        calibration["criterion"] = float(str(sheet.cell(row=10, column=most_recent_column).value))

    except:
        print("Warning error occured: Full traceback shown below:")
        print(traceback.format_exc())
        easygui.msgbox(title="Error!", msg="Error!\n\nCould not load calibration file 'calibration.xlsx'. Please make"
                                           " sure it exists in the FluoroTensor working directory.\n\n"
                                           "Defaulting to hard-coded calibration as of 24/03/2022.")
        return


def create_raw_gui():
    global raw_gui
    set_GUI_state(tk.DISABLED)
    load_button["state"] = tk.DISABLED
    load_pickle_button["state"] = tk.DISABLED
    preferences_button["state"] = tk.DISABLED
    import_raw_button["state"] = tk.DISABLED
    raw_gui = RawUI()
    plot_histogram([0], "red")
    plot_histogram([0], "yellow")
    plot_histogram([0], "green")



def validate_UI_entries(value, action):
    if action == "1":
        if not value.isdigit():
            return False
    return True


def validate_percentages(value, action):
    if action == "1":
        if not value.isdigit():
            return False
        if len(value) > 2:
            return False
    return True


def validate_convolution(value, action):
    if action == "1":
        if not value.isdigit():
            return False
        if int(value) > 100:
            return False
    return True


def prevent_entry(value):
    if value:
        try:
            float(value)
        except ValueError:
            return False
        return True


def calibration_optimizer_settings():
    global opt_settings_win
    opt_settings_win = OptSettingsWin()


def view_licence():
    global licence_win
    licence_win = LicenceWin()


class RawUI:

    # create window and user interface for RAW TIF data spot detection and trace calculation

    def __init__(self):
        self.raw_window = tk.Tk()
        self.raw_window.iconbitmap("icon.ico")
        self.raw_window.geometry("1600x900+200+20")
        self.raw_window.protocol("WM_DELETE_WINDOW", self.handle_raw_close)
        self.raw_window["bg"] = "#555555"
        self.raw_window.resizable(False, False)
        self.raw_window.title("Import raw data into FluoroTensor - Spot detection and Traces")

        self.scale_factor = None

        self.mouse_event_data = []
        self.mouse_event_abs = []
        self.show_markers_561 = False
        self.show_markers_488 = False
        self.overlay_561 = False
        self.overlay_488 = False
        self.vectors_561 = False
        self.vectors_488 = False
        self.high_pass = False
        self.is_data_loaded = False
        self.warnings = False
        self.traces_calculated = False
        self.image_canvas_marker = None
        self.image_canvas_561 = None
        self.image_canvas_488 = None

        self.marker_view_colour = "640"
        self.region_initial = [None, None, None]
        self.region_final = [None, None, None]
        self.rect_640, self.rect_561, self.rect_488 = None, None, None

        self.use_auto_optimization = False

        self.all_traces = []
        self.raw_traces = []
        self.all_fits = []
        self.trace_info = []

        self.current_data_path = ""

        self.view1_label = tk.Label(self.raw_window, text="  Marker - " + preferences["Fluorophore config"][0],
                                    width=512, bg="#444444",
                                    anchor=tk.NW, relief=tk.RIDGE, fg="white")
        self.view2_label = tk.Label(self.raw_window, text="  561nm - " + preferences["Fluorophore config"][1],
                                    width=512, bg="#444444",
                                    anchor=tk.NW, relief=tk.RIDGE, fg="white")
        self.view3_label = tk.Label(self.raw_window, text="  488nm - " + preferences["Fluorophore config"][2],
                                    width=512, bg="#444444",
                                    anchor=tk.NW, relief=tk.RIDGE, fg="white")

        self.canvas_marker = tk.Canvas(self.raw_window, width=512, height=512, bg="#111144", highlightthickness=2,
                                       highlightbackground="black")
        self.canvas_561 = tk.Canvas(self.raw_window, width=512, height=512, bg="#111144", highlightthickness=2,
                                    highlightbackground="black")
        self.canvas_488 = tk.Canvas(self.raw_window, width=512, height=512, bg="#111144", highlightthickness=2,
                                    highlightbackground="black")
        self.canvas_marker_mini = tk.Canvas(self.raw_window, width=256, height=256, bg="#111144", highlightthickness=2,
                                            highlightbackground="black")
        self.canvas_561_mini = tk.Canvas(self.raw_window, width=256, height=256, bg="#111144", highlightthickness=2,
                                         highlightbackground="black")
        self.canvas_488_mini = tk.Canvas(self.raw_window, width=256, height=256, bg="#111144", highlightthickness=2,
                                         highlightbackground="black")

        self.view1_label.place(x=2, y=10, width=532, height=512 + 256 + 40)
        self.view2_label.place(x=2 + 512 + 20, y=10, width=532, height=512 + 256 + 40)
        self.view3_label.place(x=2 + 1024 + 40, y=10, width=532, height=512 + 256 + 40)

        self.canvas_marker.place(x=10, y=30)
        self.canvas_561.place(x=542, y=30)
        self.canvas_488.place(x=1074, y=30)
        self.canvas_marker_mini.place(x=10 + 256, y=552)
        self.canvas_561_mini.place(x=542 + 256, y=552)
        self.canvas_488_mini.place(x=1074 + 256, y=552)

        self.hist_frameM = tk.Frame(master=self.raw_window)
        self.hist_figM = plt.Figure(figsize=(0.5, 0.6), dpi=100)
        self.histM = FigureCanvasTkAgg(self.hist_figM, master=self.hist_frameM)
        self.hist_figM.subplots_adjust(top=0.999, bottom=0.01, left=0.001, right=0.98)
        self.histM.get_tk_widget().pack(side=tk.TOP)
        self.hist_frameM.place(x=200, y=730)

        self.hist_frame5 = tk.Frame(master=self.raw_window)
        self.hist_fig5 = plt.Figure(figsize=(0.5, 0.6), dpi=100)
        self.hist5 = FigureCanvasTkAgg(self.hist_fig5, master=self.hist_frame5)
        self.hist_fig5.subplots_adjust(top=0.999, bottom=0.01, left=0.001, right=0.98)
        self.hist5.get_tk_widget().pack(side=tk.TOP)
        self.hist_frame5.place(x=200+512+20, y=730)

        self.hist_frame4 = tk.Frame(master=self.raw_window)
        self.hist_fig4 = plt.Figure(figsize=(0.5, 0.6), dpi=100)
        self.hist4 = FigureCanvasTkAgg(self.hist_fig4, master=self.hist_frame4)
        self.hist_fig4.subplots_adjust(top=0.999, bottom=0.01, left=0.001, right=0.98)
        self.hist4.get_tk_widget().pack(side=tk.TOP)
        self.hist_frame4.place(x=200+1024+40, y=730)

        self.deconvM = tk.IntVar()
        self.deconv561 = tk.IntVar()
        self.deconv488 = tk.IntVar()
        self.use_deconv_M = False
        self.use_deconv_561 = False
        self.use_deconv_488 = False

        self.tk_use_stats = tk.IntVar()
        self.display_stats = False

        self.spot_warnings = tk.IntVar()
        self.enhancement_type = tk.IntVar()
        self.spot_warnings.set(0)
        self.context_marker = tk.Menu(master=self.canvas_marker, tearoff=0, bg="#777777", fg="#0f1544", bd=1,
                                      activeborderwidth=4, activebackground="#333377", activeforeground="white",
                                      disabledforeground="#999999")
        self.context_marker.add_command(label="Remove Spot", command=lambda: self.remove_spot(0))
        self.context_marker.add_command(label="Spot fitting parameters",
                                        command=lambda: self.display_spot_fit_params(0))
        self.context_marker.add_command(label="Change View Wavelength", command=set_view_colour)
        self.context_marker.add_checkbutton(label="Display Statistics", onvalue=1, command=self.toggle_stats,
                                            offvalue=0, variable=self.display_stats)
        self.context_marker.add_checkbutton(label="Show Spot Warnings", onvalue=1, command=self.toggle_warning,
                                            offvalue=0, variable=self.spot_warnings)
        self.context_marker.invoke(4)
        # self.context_marker.entryconfig("Remove Spot", state=tk.DISABLED)
        # self.context_marker.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        self.context_marker.add_separator()
        self.context_marker.add_checkbutton(label="Use 2D Wavelet Transform Enhancement", onvalue=1, command=self.toggle_enhancement_type,
                                            offvalue=0, variable=self.enhancement_type)
        self.context_marker.add_checkbutton(label="Use Wavelet Deconvolution", onvalue=1, offvalue=0, variable=self.deconvM,
                                            command=lambda: self.toggle_deconv(0))
        self.context_marker.add_command(label="Edit Point Spread Function")
        self.context_marker.add_separator()
        self.context_marker.add_command(label="Select Analysis Region", command=self.select_region_640)
        self.context_marker.add_command(label="Delete Region", command=lambda: self.delete_region(0))
        self.context_marker.add_separator()
        self.context_marker.add_command(label="Save Enhanced view as .png", command=lambda: self.save_view(0))
        self.simulate_stokes = tk.IntVar()
        self.simulate_stokes.set(0)
        self.use_stokes = False
        self.context_marker.add_checkbutton(label="Simulate Stokes Shift", onvalue=1, offvalue=0,
                                            variable=self.simulate_stokes, command=self.toggle_stokes)
        self.canvas_marker.bind("<ButtonRelease-3>", self.menu_marker)

        self.context_561_marker = tk.IntVar()
        self.context_561_over = tk.IntVar()
        self.context_561_vect = tk.IntVar()
        self.context_561_marker.set(0);
        self.context_561_over.set(0)
        self.context_561_vect.set(0)
        self.context_561 = tk.Menu(master=self.canvas_561, tearoff=0, bg="#777777", fg="#0f1544", bd=1,
                                   activeborderwidth=4, activebackground="#333377", activeforeground="white",
                                   disabledforeground="#999999")
        self.context_561.add_command(label="Remove Spot", command=lambda: self.remove_spot(1))
        self.context_561.add_command(label="Spot fitting parameters", command=lambda: self.display_spot_fit_params(1))
        self.context_561.add_checkbutton(label="Display Statistics", onvalue=1, command=self.toggle_stats,
                                            offvalue=0, variable=self.display_stats)
        self.context_561.add_checkbutton(label="Show Spot Warnings", onvalue=1, command=self.toggle_warning,
                                         offvalue=0, variable=self.spot_warnings)
        self.context_561.invoke(3)
        self.context_561.add_separator()
        self.context_561.add_checkbutton(label="Show Marker Spots", command=lambda: self.toggle_marker_spots(1),
                                         onvalue=1, offvalue=0, variable=self.context_561_marker)
        self.context_561.add_checkbutton(label="Marker View Overlay", command=lambda: self.toggle_overlay(1),
                                         onvalue=1, offvalue=0, variable=self.context_561_over)
        self.context_561.add_checkbutton(label="Show Aberration Vector Field", command=lambda: self.toggle_vectors(1),
                                         onvalue=1, offvalue=0, variable=self.context_561_vect)
        # show overlay and spot positions of marker over secondary channel 561
        self.context_561.invoke(5)
        self.context_561.invoke(6)
        # self.context_561.entryconfig("Remove Spot", state=tk.DISABLED)
        # self.context_561.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        self.context_561.add_separator()
        self.context_561.add_checkbutton(label="Use 2D Wavelet Transform Enhancement", onvalue=1,
                                            command=self.toggle_enhancement_type,
                                            offvalue=0, variable=self.enhancement_type)
        self.context_561.add_checkbutton(label="Use Wavelet Deconvolution", onvalue=1, offvalue=0,
                                            variable=self.deconv561,
                                            command=lambda: self.toggle_deconv(1))
        self.context_561.add_command(label="Edit Point Spread Function")
        self.context_561.add_separator()
        self.context_561.add_command(label="Select Analysis Region", command=self.select_region_561)
        self.context_561.add_command(label="Delete Region", command=lambda: self.delete_region(1))
        self.context_561.add_separator()
        self.context_561.add_command(label="Save Enhanced view as .png", command=lambda: self.save_view(1))
        self.context_561.add_checkbutton(label="Simulate Stokes Shift", onvalue=1, offvalue=0,
                                            variable=self.simulate_stokes, command=self.toggle_stokes)
        self.canvas_561.bind("<ButtonRelease-3>", self.menu_561)

        self.context_488_marker = tk.IntVar()
        self.context_488_over = tk.IntVar()
        self.context_488_vect = tk.IntVar()
        self.context_488_marker.set(0);
        self.context_488_over.set(0)
        self.context_488_vect.set(0)
        self.context_488 = tk.Menu(master=self.canvas_488, tearoff=0, bg="#777777", fg="#0f1544", bd=1,
                                   activeborderwidth=4, activebackground="#333377", activeforeground="white",
                                   disabledforeground="#999999")
        self.context_488.add_command(label="Remove Spot", command=lambda: self.remove_spot(2))
        self.context_488.add_command(label="Spot fitting parameters", command=lambda: self.display_spot_fit_params(2))
        self.context_488.add_checkbutton(label="Display Statistics", onvalue=1, command=self.toggle_stats,
                                            offvalue=0, variable=self.display_stats)
        self.context_488.add_checkbutton(label="Show Spot Warnings", onvalue=1, command=self.toggle_warning,
                                         offvalue=0, variable=self.spot_warnings)
        self.context_488.invoke(3)
        self.context_488.add_separator()
        self.context_488.add_checkbutton(label="Show Marker Spots", command=lambda: self.toggle_marker_spots(2),
                                         onvalue=1, offvalue=0, variable=self.context_488_marker)
        self.context_488.add_checkbutton(label="Marker View Overlay", command=lambda: self.toggle_overlay(2),
                                         onvalue=1, offvalue=0, variable=self.context_488_over)
        self.context_488.add_checkbutton(label="Show Aberration Vector Field", command=lambda: self.toggle_vectors(2),
                                         onvalue=1, offvalue=0, variable=self.context_488_vect)
        # show overlay and spot positions of marker over secondary channel 488
        self.context_488.invoke(5)
        self.context_488.invoke(6)
        # self.context_488.entryconfig("Remove Spot", state=tk.DISABLED)
        # self.context_488.entryconfig("Spot fitting parameters", state=tk.DISABLED)
        self.context_488.add_separator()
        self.context_488.add_checkbutton(label="Use 2D Wavelet Transform Enhancement", onvalue=1,
                                         command=self.toggle_enhancement_type,
                                         offvalue=0, variable=self.enhancement_type)
        self.context_488.add_checkbutton(label="Use Wavelet Deconvolution", onvalue=1, offvalue=0,
                                            variable=self.deconv488,
                                            command=lambda: self.toggle_deconv(2))
        self.context_488.add_command(label="Edit Point Spread Function")
        self.context_488.add_separator()
        self.context_488.add_command(label="Select Analysis Region", command=self.select_region_488)
        self.context_488.add_command(label="Delete Region", command=lambda: self.delete_region(2))
        self.context_488.add_separator()
        self.context_488.add_command(label="Save Enhanced view as .png", command=lambda: self.save_view(2))
        self.context_488.add_checkbutton(label="Simulate Stokes Shift", onvalue=1, offvalue=0,
                                            variable=self.simulate_stokes, command=self.toggle_stokes)
        self.canvas_488.bind("<ButtonRelease-3>", self.menu_488)

        self.canvas_marker.bind("<ButtonRelease-1>", self.add_spot_marker)
        self.canvas_561.bind("<ButtonRelease-1>", self.add_spot_561)
        self.canvas_488.bind("<ButtonRelease-1>", self.add_spot_488)

        self.canvas_marker_mini.bind("<ButtonRelease-1>", lambda x: self.view_mode(0))
        self.canvas_561_mini.bind("<ButtonRelease-1>", lambda x: self.view_mode(1))
        self.canvas_488_mini.bind("<ButtonRelease-1>", lambda x: self.view_mode(2))

        self.canvas_marker.bind("<MouseWheel>", self.mouse_wheel_640)
        self.canvas_561.bind("<MouseWheel>", self.mouse_wheel_561)
        self.canvas_488.bind("<MouseWheel>", self.mouse_wheel_488)

        tk.Label(master=self.raw_window, text="Start frame", bg="#444444", fg="white").place(x=30, y=560)
        tk.Label(master=self.raw_window, text="Number of frames", bg="#444444", fg="white").place(x=30, y=580)
        tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=30, y=600)
        tk.Label(master=self.raw_window, text="Analyse", bg="#444444", fg="white").place(x=30, y=620)

        tk.Label(master=self.raw_window, text="Start frame", bg="#444444", fg="white").place(x=50 + 512, y=560)
        tk.Label(master=self.raw_window, text="Number of frames", bg="#444444", fg="white").place(x=50 + 512, y=580)
        tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=50 + 512, y=600)
        tk.Label(master=self.raw_window, text="Analyse", bg="#444444", fg="white").place(x=50 + 512, y=620)

        tk.Label(master=self.raw_window, text="Start frame", bg="#444444", fg="white").place(x=70 + 1024, y=560)
        tk.Label(master=self.raw_window, text="Number of frames", bg="#444444", fg="white").place(x=70 + 1024, y=580)
        tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=70 + 1024, y=600)
        tk.Label(master=self.raw_window, text="Analyse", bg="#444444", fg="white").place(x=70 + 1024, y=620)

        self.start_entry_marker = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.start_entry_marker["validatecommand"] = (self.start_entry_marker.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.start_entry_marker, button_hv, "#333333")
        self.length_entry_marker = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.length_entry_marker["validatecommand"] = (
        self.length_entry_marker.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.length_entry_marker, button_hv, "#333333")
        self.bg_entry_marker = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.bg_entry_marker["validatecommand"] = (self.bg_entry_marker.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bg_entry_marker, button_hv, "#333333")
        self.start_entry_561 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.start_entry_561["validatecommand"] = (self.start_entry_561.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.start_entry_561, button_hv, "#333333")
        self.length_entry_561 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.length_entry_561["validatecommand"] = (self.length_entry_561.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.length_entry_561, button_hv, "#333333")
        self.bg_entry_561 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.bg_entry_561["validatecommand"] = (self.bg_entry_561.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bg_entry_561, button_hv, "#333333")
        self.start_entry_488 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.start_entry_488["validatecommand"] = (self.start_entry_488.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.start_entry_488, button_hv, "#333333")
        self.length_entry_488 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.length_entry_488["validatecommand"] = (self.length_entry_488.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.length_entry_488, button_hv, "#333333")
        self.bg_entry_488 = tk.Entry(master=self.raw_window, bg="#333333", fg="white", validate="key")
        self.bg_entry_488["validatecommand"] = (self.bg_entry_488.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bg_entry_488, button_hv, "#333333")

        self.start_entry_marker.place(x=180, y=560, width=40)
        self.length_entry_marker.place(x=180, y=580, width=40)
        self.bg_entry_marker.place(x=180, y=600, width=40)
        self.start_entry_561.place(x=180 + 512 + 20, y=560, width=40)
        self.length_entry_561.place(x=180 + 512 + 20, y=580, width=40)
        self.bg_entry_561.place(x=180 + 512 + 20, y=600, width=40)
        self.start_entry_488.place(x=180 + 1024 + 40, y=560, width=40)
        self.length_entry_488.place(x=180 + 1024 + 40, y=580, width=40)
        self.bg_entry_488.place(x=180 + 1024 + 40, y=600, width=40)

        self.start_entry_marker.insert(tk.END, str(raw_setup["marker frames"][0]))
        self.length_entry_marker.insert(tk.END, str(raw_setup["marker frames"][1]))
        self.bg_entry_marker.insert(tk.END, str(raw_setup["background"][0]))

        self.start_entry_561.insert(tk.END, str(raw_setup["561nm frames"][0]))
        self.length_entry_561.insert(tk.END, str(raw_setup["561nm frames"][1]))
        self.bg_entry_561.insert(tk.END, str(raw_setup["background"][1]))

        self.start_entry_488.insert(tk.END, str(raw_setup["488nm frames"][0]))
        self.length_entry_488.insert(tk.END, str(raw_setup["488nm frames"][1]))
        self.bg_entry_488.insert(tk.END, str(raw_setup["background"][2]))

        self.use_marker = tk.IntVar(master=self.raw_window);
        self.use_marker.set(raw_setup["analyse"][0])
        self.use_561 = tk.IntVar(master=self.raw_window);
        self.use_561.set(raw_setup["analyse"][1])
        self.use_488 = tk.IntVar(master=self.raw_window);
        self.use_488.set(raw_setup["analyse"][2])
        self.analyse_marker = tk.Checkbutton(master=self.raw_window, bg="#444444", variable=self.use_marker, onvalue=1,
                                             offvalue=0, activebackground="#444444", command=create_view_marker)
        self.analyse_561 = tk.Checkbutton(master=self.raw_window, bg="#444444", variable=self.use_561, onvalue=1,
                                          offvalue=0, activebackground="#444444", command=create_view_561)
        self.analyse_488 = tk.Checkbutton(master=self.raw_window, bg="#444444", variable=self.use_488, onvalue=1,
                                          offvalue=0, activebackground="#444444", command=create_view_488)

        self.analyse_marker.place(x=190, y=620)
        self.analyse_561.place(x=190 + 512 + 20, y=620)
        self.analyse_488.place(x=190 + 1024 + 40, y=620)

        self.criteria_marker = tk.Button(master=self.raw_window, text="Spot Criteria", padx=10, pady=2, bg="#333333",
                                         fg="#bbbbbb",
                                         bd=2, command=lambda: set_spot_criteria(0))
        change_col_hover(self.criteria_marker, button_hv, "#333333")
        self.criteria_561 = tk.Button(master=self.raw_window, text="Spot Criteria", padx=10, pady=2, bg="#333333",
                                      fg="#bbbbbb", bd=2,
                                      command=lambda: set_spot_criteria(1))
        change_col_hover(self.criteria_561, button_hv, "#333333")
        self.criteria_488 = tk.Button(master=self.raw_window, text="Spot Criteria", padx=10, pady=2, bg="#333333",
                                      fg="#bbbbbb", bd=2,
                                      command=lambda: set_spot_criteria(2))
        change_col_hover(self.criteria_488, button_hv, "#333333")

        self.criteria_marker.place(x=30, y=650)
        self.criteria_561.place(x=50 + 512, y=650)
        self.criteria_488.place(x=70 + 1024, y=650)

        self.refresh_view_marker = tk.Button(master=self.raw_window, text="Refresh View", padx=9, pady=2, bg="#333333",
                                             fg="#bbbbbb",
                                             bd=2, command=create_view_marker)
        change_col_hover(self.refresh_view_marker, button_hv, "#333333")
        self.refresh_view_561 = tk.Button(master=self.raw_window, text="Refresh View", padx=9, pady=2, bg="#333333",
                                          fg="#bbbbbb",
                                          bd=2, command=create_view_561)
        change_col_hover(self.refresh_view_561, button_hv, "#333333")
        self.refresh_view_488 = tk.Button(master=self.raw_window, text="Refresh View", padx=9, pady=2, bg="#333333",
                                          fg="#bbbbbb",
                                          bd=2, command=create_view_488)
        change_col_hover(self.refresh_view_488, button_hv, "#333333")

        self.refresh_view_marker.place(x=30, y=684)
        self.refresh_view_561.place(x=50 + 512, y=684)
        self.refresh_view_488.place(x=70 + 1024, y=684)

        self.redo_marker = tk.Button(master=self.raw_window, text="Redo Frames", padx=2, pady=2, bg="#333333",
                                     fg="#bbbbbb", bd=2, command=lambda: redo_frames(0))
        change_col_hover(self.redo_marker, button_hv, "#333333")
        self.redo_561 = tk.Button(master=self.raw_window, text="Redo Frames", padx=2, pady=2, bg="#333333",
                                  fg="#bbbbbb", bd=2, command=lambda: redo_frames(1))
        change_col_hover(self.redo_561, button_hv, "#333333")
        self.redo_488 = tk.Button(master=self.raw_window, text="Redo Frames", padx=2, pady=2, bg="#333333",
                                  fg="#bbbbbb", bd=2, command=lambda: redo_frames(2))
        change_col_hover(self.redo_488, button_hv, "#333333")

        self.redo_marker.place(x=160, y=684)
        self.redo_561.place(x=180 + 512, y=684)
        self.redo_488.place(x=200 + 1024, y=684)

        tk.Label(master=self.raw_window, text="Brightness", bg="#444444", fg="white").place(x=30, y=730)
        tk.Label(master=self.raw_window, text="Brightness", bg="#444444", fg="white").place(x=50 + 512, y=730)
        tk.Label(master=self.raw_window, text="Brightness", bg="#444444", fg="white").place(x=70 + 1024, y=730)

        tk.Label(master=self.raw_window, text="inv. Power", bg="#444444", fg="white").place(x=30, y=770)
        tk.Label(master=self.raw_window, text="inv. Power", bg="#444444", fg="white").place(x=50 + 512, y=770)
        tk.Label(master=self.raw_window, text="inv. Power", bg="#444444", fg="white").place(x=70 + 1024, y=770)

        self.brightness_marker = tk.StringVar(master=self.raw_window)
        self.brightness_561 = tk.StringVar(master=self.raw_window)
        self.brightness_488 = tk.StringVar(master=self.raw_window)

        self.power_marker = tk.StringVar(master=self.raw_window)
        self.power_561 = tk.StringVar(master=self.raw_window)
        self.power_488 = tk.StringVar(master=self.raw_window)

        self.brightness_marker.set(raw_setup["brightness"][0])
        self.brightness_561.set(raw_setup["brightness"][1])
        self.brightness_488.set(raw_setup["brightness"][2])

        self.power_marker.set(raw_setup["inversion power"][0])
        self.power_561.set(raw_setup["inversion power"][1])
        self.power_488.set(raw_setup["inversion power"][2])

        self.spin_box_marker = tk.Spinbox(master=self.raw_window, from_=0.25, to=10, increment=0.05, repeatinterval=400,
                                          textvariable=self.brightness_marker, bg="#bbbbbb", fg="black",
                                          command=create_view_marker, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_marker["validatecommand"] = (self.spin_box_marker.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_marker, "#eeeeee", "#bbbbbb")

        self.spin_box_561 = tk.Spinbox(master=self.raw_window, from_=0.25, to=10, increment=0.05, repeatinterval=400,
                                       textvariable=self.brightness_561, bg="#bbbbbb", fg="black",
                                       command=create_view_561, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_561["validatecommand"] = (self.spin_box_561.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_561, "#eeeeee", "#bbbbbb")

        self.spin_box_488 = tk.Spinbox(master=self.raw_window, from_=0.25, to=10, increment=0.05, repeatinterval=400,
                                       textvariable=self.brightness_488, bg="#bbbbbb", fg="black",
                                       command=create_view_488, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_488["validatecommand"] = (self.spin_box_488.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_488, "#eeeeee", "#bbbbbb")

        self.spin_box_marker_pow = tk.Spinbox(master=self.raw_window, from_=0.25, to=4, increment=0.02,
                                              repeatinterval=400,
                                              textvariable=self.power_marker, bg="#bbbbbb", fg="black",
                                              command=create_view_marker, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_marker_pow["validatecommand"] = (self.spin_box_marker_pow.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_marker_pow, "#eeeeee", "#bbbbbb")

        self.spin_box_561_pow = tk.Spinbox(master=self.raw_window, from_=0.25, to=4, increment=0.02, repeatinterval=400,
                                           textvariable=self.power_561, bg="#bbbbbb", fg="black",
                                           command=create_view_561, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_561_pow["validatecommand"] = (self.spin_box_561_pow.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_561_pow, "#eeeeee", "#bbbbbb")

        self.spin_box_488_pow = tk.Spinbox(master=self.raw_window, from_=0.25, to=4, increment=0.02, repeatinterval=400,
                                           textvariable=self.power_488, bg="#bbbbbb", fg="black",
                                           command=create_view_488, bd=2, font="Arial 12 bold", validate="key")
        self.spin_box_488_pow["validatecommand"] = (self.spin_box_488_pow.register(prevent_entry), "%P")
        change_col_hover_enterbox(self.spin_box_488_pow, "#eeeeee", "#bbbbbb")

        self.spin_box_marker.place(x=100, y=726, width=80, height=30)
        self.spin_box_561.place(x=100 + 512 + 20, y=726, width=80, height=30)
        self.spin_box_488.place(x=100 + 1024 + 40, y=726, width=80, height=30)

        self.spin_box_marker_pow.place(x=100, y=766, width=80, height=30)
        self.spin_box_561_pow.place(x=100 + 512 + 20, y=766, width=80, height=30)
        self.spin_box_488_pow.place(x=100 + 1024 + 40, y=766, width=80, height=30)

        self.load_tiff_button = tk.Button(master=self.raw_window, text="Load TIF", padx=7, pady=20, bg="#333377",
                                          fg="#bbbbbb",
                                          bd=2, command=load_tiff)
        change_col_hover(self.load_tiff_button, button_hv, "#333377")

        self.auto_button = tk.Button(master=self.raw_window, text="Automate", padx=7, pady=2, bg="#337733",
                                     fg="#bbbbbb",
                                     bd=2, command=automate)
        change_col_hover(self.auto_button, "#449944", "#337733")

        self.detect_spots_button = tk.Button(master=self.raw_window, text="Detect Spots", padx=4, pady=2, bg="#333333",
                                             fg="#bbbbbb",
                                             bd=2, command=detect_spots_callback)
        change_col_hover(self.detect_spots_button, button_hv, "#333333")

        self.remove_spots_button = tk.Button(master=self.raw_window, text="Remove Spots", padx=36, pady=2, bg="#333333",
                                             fg="#bbbbbb",
                                             bd=2, command=remove_spots)
        change_col_hover(self.remove_spots_button, button_hv, "#333333")

        self.calculate_all_button = tk.Button(master=self.raw_window, text="All Traces", padx=7, pady=2,
                                              bg="#333333", fg="#bbbbbb",
                                              bd=2, command=calculate_all_traces)
        change_col_hover(self.calculate_all_button, button_hv, "#333333")

        self.calculate_non_coloc_button = tk.Button(master=self.raw_window, text="Non Coloc.", padx=6, pady=2,
                                                    bg="#333333", fg="#bbbbbb",
                                                    bd=2, command=calculate_non_coloc_traces)
        change_col_hover(self.calculate_non_coloc_button, button_hv, "#333333")

        self.calculate_coloc_button = tk.Button(master=self.raw_window, text="Calculate Colocalized Traces", padx=2,
                                                pady=2, bg="#333333", fg="#bbbbbb",
                                                bd=2, command=calculate_coloc_traces)
        change_col_hover(self.calculate_coloc_button, button_hv, "#333333")

        self.save_and_exit_button = tk.Button(master=self.raw_window, text="Import to FluoroTensor", padx=10, pady=2,
                                              bg="#773333", fg="#bbbbbb",
                                              bd=2, command=save_and_exit_raw)
        change_col_hover(self.save_and_exit_button, "#994444", "#773333")

        self.view_traces_button = tk.Button(master=self.raw_window, text="Open Traces", padx=38, pady=2, bg="#333377",
                                            fg="#bbbbbb", bd=2, command=self.open_traces)
        change_col_hover(self.view_traces_button, button_hv, "#333377")

        self.defaults_button = tk.Button(master=self.raw_window, text="Set Defaults", padx=10, pady=2,
                                         bg="#111111", fg="#bbbbbb", bd=2, command=set_raw_defaults)
        change_col_hover(self.defaults_button, button_hv, "#111111")

        self.open_raw_button = tk.Button(master=self.raw_window, text="Open RAW", padx=12, pady=2, bg="#111111",
                                         fg="#bbbbbb", bd=2, command=view_raw_stack)
        change_col_hover(self.open_raw_button, button_hv, "#111111")

        self.calibration_button = tk.Button(master=self.raw_window, text="Coloc. Calibration", padx=10, pady=2,
                                            bg="#111111", fg="#bbbbbb", bd=2, command=set_calibration)
        change_col_hover(self.calibration_button, button_hv, "#111111")

        self.force_coloc_button = tk.Button(master=self.raw_window, text="Force Colocalization", padx=4, pady=2,
                                            bg="#222266", fg="#bbbbbb", bd=2, command=self.toggle_forced_coloc)
        change_col_hover(self.force_coloc_button, button_hv, "#222266")

        self.detection_mode_button = tk.Button(master=self.raw_window, text="Av. Mode", padx=2, pady=2, bg="#222266",
                                               fg="#bbbbbb", bd=2, command=self.toggle_detection_mode)
        change_col_hover(self.detection_mode_button, button_hv, "#222266")

        self.quality_button = tk.Button(master=self.raw_window, text="Quality Chk", padx=2, pady=2, bg="#222222",
                                     fg="#bbbbbb", bd=2, command=quality_check)
        change_col_hover(self.quality_button, button_hv, "#222222")


        self.load_tiff_button.place(x=10, y=824)
        self.auto_button.place(x=85, y=824)
        self.quality_button.place(x=85, y=859)
        self.save_and_exit_button.place(x=1440, y=859)
        self.view_traces_button.place(x=1440, y=824)
        self.detect_spots_button.place(x=243, y=824)
        self.detection_mode_button.place(x=170, y=824)
        self.remove_spots_button.place(x=170, y=859)
        self.calculate_all_button.place(x=336, y=824)
        self.calculate_non_coloc_button.place(x=417, y=824)
        self.calculate_coloc_button.place(x=336, y=859)
        self.defaults_button.place(x=508, y=859)
        self.open_raw_button.place(x=508, y=824)
        self.calibration_button.place(x=610, y=824)
        self.force_coloc_button.place(x=610, y=859)

        self.status = tk.Label(self.raw_window, text="", bg="#444444", anchor=tk.NW, relief=tk.RIDGE, fg="white",
                               padx=8, pady=4, wraplength=484, justify="left", font="TkDefaultFont 8")
        self.status.place(x=930, y=824, height=64, width=500)

        self.session_paths = []

        self.progress = ttk.Progressbar(master=self.raw_window, orient="horizontal", mode="determinate",
                                        length=180)
        self.load_label = tk.Label(master=self.raw_window, bg="#555555", fg="white")

        self.raw_marker_spots = []
        self.raw_561_spots = []
        self.raw_488_spots = []
        self.final_561_spots = []
        self.final_488_spots = []
        self.colocalizations = []
        self.marker_spot_count = []
        self._488_count = []
        self._561_count = []
        self.coloc_488 = []
        self.coloc_561 = []
        self.calib_params = []
        self.background640 = []
        self.background561 = []
        self.background488 = []
        self.enhancements = []
        self.file_history_for_export = []
        self.analysis_mode = None

        self.detection_mode = False
        self.force_coloc = False
        self.view_modes = ["b", "b", "b"]

        self.auto_movie_list = []

        self.resolution = None

    def toggle_stokes(self):
        if not self.use_stokes:
            self.use_stokes = True
            create_view_marker()
            create_view_561()
            create_view_488()
            return
        else:
            self.use_stokes = False
            create_view_marker()
            create_view_561()
            create_view_488()

    def toggle_stats(self):
        if not self.display_stats:
            self.display_stats = True
            create_view_marker()
            create_view_561()
            create_view_488()
            return
        else:
            self.display_stats = False
            create_view_marker()
            create_view_561()
            create_view_488()

    def toggle_deconv(self, view):
        if view == 0:
            if not self.use_deconv_M:
                self.use_deconv_M = True
                create_view_marker()
            else:
                self.use_deconv_M = False
                create_view_marker()

        elif view == 1:
            if not self.use_deconv_561:
                self.use_deconv_561 = True
                create_view_561()
            else:
                self.use_deconv_561 = False
                create_view_561()

        elif view == 2:
            if not self.use_deconv_488:
                self.use_deconv_488 = True
                create_view_488()
            else:
                self.use_deconv_488 = False
                create_view_488()


    def mouse_wheel_640(self, event):
        self.canvas_marker.unbind("<MouseWheel>")
        if event.delta == 120:
            change = 0.5
        else:
            change = -0.5
        current_brightness = float(self.brightness_marker.get())
        self.brightness_marker.set(round(current_brightness + change, 2))
        if float(self.brightness_marker.get()) < 0:
            self.brightness_marker.set(0)
        create_view_marker()
        self.canvas_marker.bind("<MouseWheel>", self.mouse_wheel_640)

    def mouse_wheel_561(self, event):
        self.canvas_561.unbind("<MouseWheel>")
        if event.delta == 120:
            change = 0.5
        else:
            change = -0.5
        current_brightness = float(self.brightness_561.get())
        self.brightness_561.set(round(current_brightness + change, 2))
        if float(self.brightness_561.get()) < 0:
            self.brightness_561.set(0)
        create_view_561()
        self.canvas_561.bind("<MouseWheel>", self.mouse_wheel_561)

    def mouse_wheel_488(self, event):
        self.canvas_488.unbind("<MouseWheel>")
        if event.delta == 120:
            change = 0.5
        else:
            change = -0.5
        current_brightness = float(self.brightness_488.get())
        self.brightness_488.set(round(current_brightness + change, 2))
        if float(self.brightness_488.get()) < 0:
            self.brightness_488.set(0)
        create_view_488()
        self.canvas_488.bind("<MouseWheel>", self.mouse_wheel_488)

    def delete_region(self, canv):
        self.region_initial[canv] = None
        if canv == 0:
            self.canvas_marker.delete(self.rect_640)
        elif canv == 1:
            self.canvas_561.delete(self.rect_561)
        elif canv == 2:
            self.canvas_488.delete(self.rect_488)

    def select_region_640(self):
        self.set_rawgui_state(tk.DISABLED)
        self.canvas_marker.bind("<Button-1>", lambda event, canv=0: self.region_start(event, canv))
        self.canvas_marker.bind("<ButtonRelease-1>", lambda event, canv=0: self.region_end(event, canv))
        self.canvas_marker.bind("<B1-Motion>", lambda event, canv=0: self.mouse_motion(event, canv))

    def select_region_561(self):
        self.set_rawgui_state(tk.DISABLED)
        self.canvas_561.bind("<Button-1>", lambda event, canv=1: self.region_start(event, canv))
        self.canvas_561.bind("<ButtonRelease-1>", lambda event, canv=1: self.region_end(event, canv))
        self.canvas_561.bind("<B1-Motion>", lambda event, canv=1: self.mouse_motion(event, canv))

    def select_region_488(self):
        self.set_rawgui_state(tk.DISABLED)
        self.canvas_488.bind("<Button-1>", lambda event, canv=2: self.region_start(event, canv))
        self.canvas_488.bind("<ButtonRelease-1>", lambda event, canv=2: self.region_end(event, canv))
        self.canvas_488.bind("<B1-Motion>", lambda event, canv=2: self.mouse_motion(event, canv))

    def region_start(self, event, canvas_index):
        self.region_initial[canvas_index] = [event.x, event.y]

    def mouse_motion(self, event, canvas_index):
        if canvas_index == 0:
            try:
                self.canvas_marker.delete(self.rect_640)
            except:
                """ No rectangle to delete """
            self.rect_640 = self.canvas_marker.create_rectangle(self.region_initial[0][0], self.region_initial[0][1],
                                                                event.x, event.y, outline="#bbbbbb", width=1, dash=(2, 1))

        if canvas_index == 1:
            try:
                self.canvas_561.delete(self.rect_561)
            except:
                """ No rectangle to delete """
            self.rect_561 = self.canvas_561.create_rectangle(self.region_initial[1][0], self.region_initial[1][1],
                                                             event.x, event.y, outline="#bbbbbb", width=1, dash=(2, 1))

        if canvas_index == 2:
            try:
                self.canvas_488.delete(self.rect_488)
            except:
                """ No rectangle to delete """
            self.rect_488 = self.canvas_488.create_rectangle(self.region_initial[2][0], self.region_initial[2][1],
                                                               event.x, event.y, outline="#bbbbbb", width=1, dash=(2, 1))

    def region_end(self, event, canvas_index):
        self.region_final[canvas_index] = [event.x, event.y]
        if self.region_final[canvas_index] == self.region_initial[canvas_index]:
            self.region_initial[canvas_index], self.region_final[canvas_index] = None, None
        region = [self.region_initial[canvas_index], self.region_final[canvas_index]]
        if region is not None:
            if region[0][0] > region[1][0]:
                region[0][0], region[1][0] = region[1][0], region[0][0]
            if region[0][1] > region[1][1]:
                region[0][1], region[1][1] = region[1][1], region[0][1]
        self.region_initial[canvas_index], self.region_final[canvas_index] = region[0], region[1]
        print(self.region_initial, self.region_final)
        try:
            self.region_initial[canvas_index][0] = self.region_initial[canvas_index][0] / self.scale_factor
            self.region_initial[canvas_index][1] = self.region_initial[canvas_index][1] / self.scale_factor
            self.region_final[canvas_index][0] = self.region_final[canvas_index][0] / self.scale_factor
            self.region_final[canvas_index][1] = self.region_final[canvas_index][1] / self.scale_factor
        except:
            """ NoneType detected """
        print(self.region_initial, self.region_final)
        self.canvas_marker.unbind("<Button-1>")
        self.canvas_marker.unbind("<ButtonRelease-1>")
        self.canvas_marker.unbind("<B1-Motion>")
        self.canvas_561.unbind("<Button-1>")
        self.canvas_561.unbind("<ButtonRelease-1>")
        self.canvas_561.unbind("<B1-Motion>")
        self.canvas_488.unbind("<Button-1>")
        self.canvas_488.unbind("<ButtonRelease-1>")
        self.canvas_488.unbind("<B1-Motion>")
        self.set_rawgui_state(tk.NORMAL)

    def handle_raw_close(self, forced=False):
        if not forced:
            if self.traces_calculated:
                decision = easygui.ccbox(title="Warning!", msg="There are unimported traces in the import queue."
                                                               "\n\nDo you wish to close anyway?",
                                         choices=["Close", "Cancel"],
                                         default_choice="Cancel")
                if not decision:
                    return
        global array_TIF, array_TIF_len, sum_array_marker, sum_array_561, sum_array_488
        global orgsum_marker, orgsum_561, orgsum_488
        array_TIF, array_TIF_len = None, None
        sum_array_marker = None
        sum_array_561 = None
        sum_array_488 = None
        array_TIF = None
        array_TIF_len = None
        orgsum_marker = None
        orgsum_561 = None
        orgsum_488 = None

        try:
            self.raw_window.destroy()
        except:
            """ Window was closed """
        try:
            raw_defaults.handle_close()
        except:
            """ Window was closed """
        try:
            spot_criteria.handle_close()
        except:
            """ Window was closed """
        try:
            display_params.handle_close()
        except:
            """ Window was closed """
        try:
            spot_warnings.handle_close()
        except:
            """ Window was closed """
        try:
            view_raw.handle_close()
        except:
            """ Window was closed """
        try:
            view_traces.handle_close()
        except:
            """ Window was closed """
        try:
            auto_window.handle_close()
        except:
            """ Window was closed """
        try:
            calib_window.handle_close()
        except:
            """ Window was closed """
        try:
            view_colour_window.handle_close()
        except:
            """ Window was closed """
        try:
            view_mode_win.handle_close()
        except:
            """ window was closed """
        try:
            voice_window.handle_close()
        except:
            """ Window was closed """
        try:
            quality_window.handle_close(forced=True)
        except:
            """ Window was closed """
        try:
            calib_analysis_win.handle_close()
        except:
            """ Window was closed """
        try:
            opt_settings_win.handle_close()
        except:
            """ Window was closed """
        finally:
            set_GUI_state(tk.NORMAL)
            load_button["state"] = tk.NORMAL
            load_pickle_button["state"] = tk.NORMAL
            preferences_button["state"] = tk.NORMAL
            import_raw_button["state"] = tk.NORMAL

    def set_rawgui_state(self, state):
        if state == tk.DISABLED:
            self.canvas_marker.unbind("<ButtonRelease-3>")
            self.canvas_561.unbind("<ButtonRelease-3>")
            self.canvas_488.unbind("<ButtonRelease-3>")
            self.canvas_marker.unbind("<ButtonRelease-1>")
            self.canvas_561.unbind("<ButtonRelease-1>")
            self.canvas_488.unbind("<ButtonRelease-1>")
        elif state == tk.NORMAL:
            self.canvas_marker.unbind("<Button-1>")
            self.canvas_marker.unbind("<ButtonRelease-1>")
            self.canvas_marker.unbind("<B1-Motion>")
            self.canvas_561.unbind("<Button-1>")
            self.canvas_561.unbind("<ButtonRelease-1>")
            self.canvas_561.unbind("<B1-Motion>")
            self.canvas_488.unbind("<Button-1>")
            self.canvas_488.unbind("<ButtonRelease-1>")
            self.canvas_488.unbind("<B1-Motion>")
            self.canvas_marker.bind("<ButtonRelease-3>", self.menu_marker)
            self.canvas_561.bind("<ButtonRelease-3>", self.menu_561)
            self.canvas_488.bind("<ButtonRelease-3>", self.menu_488)
            self.canvas_marker.bind("<ButtonRelease-1>", self.add_spot_marker)
            self.canvas_561.bind("<ButtonRelease-1>", self.add_spot_561)
            self.canvas_488.bind("<ButtonRelease-1>", self.add_spot_488)
        self.load_tiff_button["state"] = state
        self.auto_button["state"] = state
        self.detect_spots_button["state"] = state
        self.remove_spots_button["state"] = state
        self.calculate_all_button["state"] = state
        self.calculate_non_coloc_button["state"] = state
        self.calculate_coloc_button["state"] = state
        self.save_and_exit_button["state"] = state
        self.view_traces_button["state"] = state
        self.defaults_button["state"] = state
        self.open_raw_button["state"] = state
        self.calibration_button["state"] = state
        self.criteria_marker["state"] = state
        self.criteria_561["state"] = state
        self.criteria_488["state"] = state
        self.refresh_view_marker["state"] = state
        self.refresh_view_561["state"] = state
        self.refresh_view_488["state"] = state
        self.redo_marker["state"] = state
        self.redo_561["state"] = state
        self.redo_488["state"] = state

    def view_binding(self, state):
        if state == tk.DISABLED:
            self.canvas_marker.unbind("<ButtonRelease-3>")
            self.canvas_561.unbind("<ButtonRelease-3>")
            self.canvas_488.unbind("<ButtonRelease-3>")
            self.canvas_marker.unbind("<ButtonRelease-1>")
            self.canvas_561.unbind("<ButtonRelease-1>")
            self.canvas_488.unbind("<ButtonRelease-1>")
        elif state == tk.NORMAL:
            self.canvas_marker.bind("<ButtonRelease-3>", self.menu_marker)
            self.canvas_561.bind("<ButtonRelease-3>", self.menu_561)
            self.canvas_488.bind("<ButtonRelease-3>", self.menu_488)
            self.canvas_marker.bind("<ButtonRelease-1>", self.add_spot_marker)
            self.canvas_561.bind("<ButtonRelease-1>", self.add_spot_561)
            self.canvas_488.bind("<ButtonRelease-1>", self.add_spot_488)

    def toggle_detection_mode(self):
        if not self.detection_mode:
            self.detection_mode_button["fg"] = "black"
            self.detection_mode_button["bg"] = "#33ff66"
            self.detection_mode_button["relief"] = tk.SUNKEN
            self.detection_mode = True
            return
        elif self.detection_mode:
            self.detection_mode_button["fg"] = "#bbbbbb"
            self.detection_mode_button["bg"] = "#222222"
            self.detection_mode_button["relief"] = tk.RAISED
            self.detection_mode = False

    def toggle_forced_coloc(self):
        if not self.force_coloc:
            self.force_coloc_button["fg"] = "black"
            self.force_coloc_button["bg"] = "#33ff66"
            self.force_coloc_button["relief"] = tk.SUNKEN
            self.force_coloc = True
            easygui.msgbox(title="Force Colocalization", msg="Marker spots will be detected. Colocalization will be"
                           " assumed to occurr for all markers and traces calculated for transformed coordinates in"
                           " secondary channels. Only non-background traces will count towards true colocaliztion.")
            return
        elif self.force_coloc:
            self.force_coloc_button["fg"] = "#bbbbbb"
            self.force_coloc_button["bg"] = "#222266"
            self.force_coloc_button["relief"] = tk.RAISED
            self.force_coloc = False

    def menu_marker(self, event):
        self.mouse_event_data = [event.x, event.y]
        self.mouse_event_abs = [event.x_root, event.y_root]
        try:
            self.context_marker.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_marker.grab_release()

    def menu_561(self, event):
        self.mouse_event_data = [event.x, event.y]
        self.mouse_event_abs = [event.x_root, event.y_root]
        try:
            self.context_561.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_561.grab_release()

    def menu_488(self, event):
        self.mouse_event_data = [event.x, event.y]
        self.mouse_event_abs = [event.x_root, event.y_root]
        try:
            self.context_488.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_488.grab_release()

    def toggle_overlay(self, view):
        if view == 1:
            if not self.overlay_561:
                self.overlay_561 = True
                create_view_561()
                return
            else:
                self.overlay_561 = False
                create_view_561()
                return
        elif view == 2:
            if not self.overlay_488:
                self.overlay_488 = True
                create_view_488()
                return
            else:
                self.overlay_488 = False
                create_view_488()

    def toggle_marker_spots(self, view):
        if view == 1:
            if not self.show_markers_561:
                self.show_markers_561 = True
                create_view_561()
                return
            else:
                self.show_markers_561 = False
                create_view_561()
                return
        elif view == 2:
            if not self.show_markers_488:
                self.show_markers_488 = True
                create_view_488()
                return
            else:
                self.show_markers_488 = False
                create_view_488()

    def toggle_vectors(self, view):
        if view == 1:
            if not self.vectors_561:
                self.vectors_561 = True
                create_view_561()
                return
            else:
                self.vectors_561 = False
                create_view_561()
                return
        elif view == 2:
            if not self.vectors_488:
                self.vectors_488 = True
                create_view_488()
                return
            else:
                self.vectors_488 = False
                create_view_488()

    def toggle_enhancement_type(self):
        if not self.high_pass:
            self.high_pass = True
            self.power_marker.set(0.3775)
            self.power_561.set(0.3775)
            self.power_488.set(0.377)
            self.brightness_marker.set(1)
            self.brightness_561.set(1)
            self.brightness_488.set(1)
            create_view_marker()
            create_view_561()
            create_view_488()
            tk.Label(master=self.raw_window, text="Contrast            ", bg="#444444", fg="white").place(x=30, y=600)
            tk.Label(master=self.raw_window, text="Contrast            ", bg="#444444", fg="white").place(x=50 + 512, y=600)
            tk.Label(master=self.raw_window, text="Contrast            ", bg="#444444", fg="white").place(x=70 + 1024, y=600)
            return
        else:
            self.high_pass = False
            create_view_marker(first_time=True)
            create_view_561(first_time=True)
            create_view_488(first_time=True)
            tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=30, y=600)
            tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=50 + 512, y=600)
            tk.Label(master=self.raw_window, text="Background %", bg="#444444", fg="white").place(x=70 + 1024, y=600)


    def toggle_warning(self):
        if self.warnings:
            self.warnings = False
            return
        else:
            self.warnings = True

    def save_view(self, view):
        if self.is_data_loaded and view == 0:
            path = easygui.filesavebox(title="Save view - Marker", filetypes=["*.png", "PNG Files"], default="*.png")
            if path:
                if not path[-4:] == ".png":
                    path = path + ".png"
                self.image_canvas_marker.save(path)

        if self.is_data_loaded and view == 1:
            path = easygui.filesavebox(title="Save view - 561nm", filetypes=["*.png", "PNG Files"], default="*.png")
            if path:
                if not path[-4:] == ".png":
                    path = path + ".png"
                self.image_canvas_561.save(path)

        if self.is_data_loaded and view == 2:
            path = easygui.filesavebox(title="Save view - 488nm", filetypes=["*.png", "PNG Files"], default="*.png")
            if path:
                if not path[-4:] == ".png":
                    path = path + ".png"
                self.image_canvas_488.save(path)

    def remove_spot(self, view):
        coordinates = self.mouse_event_data
        coordinates[0] = (coordinates[0] - 2) / raw_gui.scale_factor
        coordinates[1] = (coordinates[1] - 2) / raw_gui.scale_factor

        if view == 0 and self.use_marker.get() == 1:
            residuals = []
            for spot in self.raw_marker_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)
            self.raw_marker_spots.pop(distances.index(minimum))
            update_final_spots()
            create_view_marker()
            create_view_561()
            create_view_488()

        if view == 1 and self.use_561.get() == 1:
            residuals = []
            for spot in self.raw_561_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)
            self.raw_561_spots.pop(distances.index(minimum))
            update_final_spots()
            create_view_marker()
            create_view_561()

        if view == 2 and self.use_488.get() == 1:
            residuals = []
            for spot in self.raw_488_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)
            self.raw_488_spots.pop(distances.index(minimum))
            update_final_spots()
            create_view_marker()
            create_view_488()

    def display_spot_fit_params(self, view):
        global display_params
        try:
            display_params.window.destroy()
        except:
            pass
        coordinates = self.mouse_event_data
        popup_coords = self.mouse_event_abs
        coordinates[0] = (coordinates[0] - 2) / raw_gui.scale_factor
        coordinates[1] = (coordinates[1] - 2) / raw_gui.scale_factor

        if view == 0 and self.use_marker.get() == 1:
            residuals = []
            for spot in self.raw_marker_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)

            # extract fitting information from spot [spotx, spoty, [sigmax, sigmay, amplitude, gaussian residual]]
            nearest_spot = self.raw_marker_spots[distances.index(minimum)]
            parameters = ["Marker", distances.index(minimum) + 1, nearest_spot[0], nearest_spot[1], nearest_spot[2][0],
                          nearest_spot[2][1], nearest_spot[2][2], nearest_spot[2][3]]

            display_params = ParamWin(parameters, popup_coords)

        if view == 1 and self.use_561.get() == 1:
            residuals = []
            for spot in self.raw_561_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)

            # extract fitting information from spot [spotx, spoty, [sigmax, sigmay, amplitude, gaussian residual]]
            nearest_spot = self.raw_561_spots[distances.index(minimum)]
            parameters = ["561nm", distances.index(minimum) + 1, nearest_spot[0], nearest_spot[1], nearest_spot[2][0],
                          nearest_spot[2][1], nearest_spot[2][2], nearest_spot[2][3]]

            display_params = ParamWin(parameters, popup_coords)

        if view == 2 and self.use_488.get() == 1:
            residuals = []
            for spot in self.raw_488_spots:
                residuals.append([np.abs(coordinates[0] - spot[0]), np.abs(coordinates[1] - spot[1])])
            distances = []
            for r in range(len(residuals)):
                distances.append(np.sqrt(residuals[r][0] ** 2 + residuals[r][1] ** 2))
            minimum = min(distances)

            # extract fitting information from spot [spotx, spoty, [sigmax, sigmay, amplitude, gaussian residual]]
            nearest_spot = self.raw_488_spots[distances.index(minimum)]
            parameters = ["488nm", distances.index(minimum) + 1, nearest_spot[0], nearest_spot[1], nearest_spot[2][0],
                          nearest_spot[2][1], nearest_spot[2][2], nearest_spot[2][3]]

            display_params = ParamWin(parameters, popup_coords)

    @staticmethod
    def generate_warning(spot, view):

        warning_message = ""
        if spot[2][4] < raw_fit_criteria["detection threshold"][view]:
            warning_message += "Spot Detection Threshold: " +str(round(spot[2][4], 3)) + " is less than minimum allowed: " \
                               + str(raw_fit_criteria["detection threshold"][view]) + "\n\n"
        if spot[2][0] < raw_fit_criteria["minimum sigma"][view]:
            warning_message += "Spot Sigma(x): " + str(round(spot[2][0], 3)) + " is less than minimum allowed:" \
                                                                               " " + str(
                raw_fit_criteria["minimum sigma"][view]) + "\n\n"
        if spot[2][0] > raw_fit_criteria["maximum sigma"][view]:
            warning_message += "Spot Sigma(x): " + str(round(spot[2][0], 3)) + " is greater than maximum allowed:" \
                                                                               " " + str(
                raw_fit_criteria["maximum sigma"][view]) + "\n\n"
        if spot[2][1] < raw_fit_criteria["minimum sigma"][view]:
            warning_message += "Spot Sigma(y): " + str(round(spot[2][1], 3)) + " is less than minimum allowed:" \
                                                                               " " + str(
                raw_fit_criteria["minimum sigma"][view]) + "\n\n"
        if spot[2][1] > raw_fit_criteria["maximum sigma"][view]:
            warning_message += "Spot Sigma(y): " + str(round(spot[2][1], 3)) + " is greater than maximum allowed:" \
                                                                               " " + str(
                raw_fit_criteria["maximum sigma"][view]) + "\n\n"
        if spot[2][2] < raw_fit_criteria["minimum gauss amplitude"][view]:
            warning_message += "Gaussian Amplitude: " + str(round(spot[2][2], 3)) + \
                               " is less than minimum allowed: " + str(
                raw_fit_criteria["minimum gauss amplitude"][view]) + "\n\n"
        if spot[2][3] > raw_fit_criteria["minimum gauss residual"][view]:
            warning_message += "Gaussian Fit Residual: " + str(round(spot[2][3], 3)) + \
                               " is greater than maximum allowed: " + str(
                raw_fit_criteria["minimum gauss residual"][view]) + "\n\n"
        eccentric = spot[2][1] / spot[2][0]
        if eccentric > 1:
            eccentric = 1 / eccentric
        if eccentric < raw_fit_criteria["eccentricity threshold"][view]:
            warning_message += "Eccentricity (min/maj): " + str(round(eccentric, 3)) + \
                               " is less than minimum allowed: " + str(
                raw_fit_criteria["eccentricity threshold"][view]) + "\n\n"

        return warning_message

    def add_spot_marker(self, event):
        global spot_warnings
        view = 0
        if self.is_data_loaded and self.use_marker.get() == 1:
            coords = [event.x, event.y]
            coords[0], coords[1] = (coords[0] - 2) / raw_gui.scale_factor, (coords[1] - 2) / raw_gui.scale_factor
            spot = tirf.add_spot(sum_array_marker, coords, raw_gui.detection_mode)
            if np.sqrt(np.abs(event.x - (spot[0] * raw_gui.scale_factor + 2)) ** 2 + np.abs(event.y - (spot[1] * raw_gui.scale_factor + 2)) ** 2) > 15:
                spot[0] = (event.x - 2) / raw_gui.scale_factor
                spot[1] = (event.y - 2) / raw_gui.scale_factor

            if self.warnings:
                warning_message = self.generate_warning(spot, view)
                comparison_pass = True
                for compare_spot in self.raw_marker_spots:
                    distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                    if distance < 6:
                        comparison_pass = False
                if not comparison_pass:
                    return

                if not warning_message == "":
                    try:
                        spot_warnings.handle_close()
                    except:
                        """ No previous warning was open """
                    finally:
                        spot_warnings = SpotWarn(spot, warning_message, view, event)
                    return

            comparison_pass = True
            for compare_spot in self.raw_marker_spots:
                distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                if distance < 6:
                    comparison_pass = False

            if comparison_pass:
                self.raw_marker_spots.append(spot)
                # self.context_marker.entryconfig("Remove Spot", state=tk.NORMAL)
                # self.context_marker.entryconfig("Spot fitting parameters", state=tk.NORMAL)
                # self.context_561.entryconfig("Show Marker Spots", state=tk.NORMAL)
                # self.context_488.entryconfig("Show Marker Spots", state=tk.NORMAL)
                update_final_spots()
                create_view_marker()
                create_view_561()
                create_view_488()
                try:
                    spot_warnings.handle_close()
                except:
                    """ window was closed """

    def add_spot_561(self, event):
        global spot_warnings
        view = 1
        if self.is_data_loaded and self.use_561.get() == 1:
            coords = [event.x, event.y]
            coords[0], coords[1] = (coords[0] - 2) / raw_gui.scale_factor, (coords[1] - 2) / raw_gui.scale_factor
            spot = tirf.add_spot(sum_array_561, coords, raw_gui.detection_mode)
            if np.sqrt(np.abs(event.x - (spot[0] * raw_gui.scale_factor + 2)) ** 2 + np.abs(event.y - (spot[1] * raw_gui.scale_factor + 2)) ** 2) > 15:
                spot[0] = (event.x - 2) / raw_gui.scale_factor
                spot[1] = (event.y - 2) / raw_gui.scale_factor

            if self.warnings:
                warning_message = self.generate_warning(spot, view)
                comparison_pass = True
                for compare_spot in self.raw_561_spots:
                    distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                    if distance < 6:
                        comparison_pass = False
                if not comparison_pass:
                    return

                if not warning_message == "":
                    try:
                        spot_warnings.handle_close()
                    except:
                        """ No previous warning was open """
                    finally:
                        spot_warnings = SpotWarn(spot, warning_message, view, event)
                    return

            comparison_pass = True
            for compare_spot in self.raw_561_spots:
                distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                if distance < 6:
                    comparison_pass = False

            if comparison_pass:
                self.raw_561_spots.append(spot)
                # self.context_561.entryconfig("Remove Spot", state=tk.NORMAL)
                # self.context_561.entryconfig("Spot fitting parameters", state=tk.NORMAL)
                update_final_spots()
                create_view_561()
                create_view_marker()
                try:
                    spot_warnings.handle_close()
                except:
                    """ window was closed """

    def add_spot_488(self, event):
        global spot_warnings
        view = 2
        if self.is_data_loaded and self.use_488.get() == 1:
            coords = [event.x, event.y]
            coords[0], coords[1] = (coords[0] - 2) / raw_gui.scale_factor, (coords[1] - 2) / raw_gui.scale_factor
            spot = tirf.add_spot(sum_array_488, coords, raw_gui.detection_mode)
            if np.sqrt(np.abs(event.x - (spot[0] * raw_gui.scale_factor + 2)) ** 2 + np.abs(event.y - (spot[1] * raw_gui.scale_factor + 2)) ** 2) > 15:
                spot[0] = (event.x - 2) / raw_gui.scale_factor
                spot[1] = (event.y - 2) / raw_gui.scale_factor

            if self.warnings:
                warning_message = self.generate_warning(spot, view)
                comparison_pass = True
                for compare_spot in self.raw_488_spots:
                    distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                    if distance < 6:
                        comparison_pass = False
                if not comparison_pass:
                    return

                if not warning_message == "":
                    try:
                        spot_warnings.handle_close()
                    except:
                        """ No previous warning was open """
                    finally:
                        spot_warnings = SpotWarn(spot, warning_message, view, event)
                    return

            comparison_pass = True
            for compare_spot in self.raw_488_spots:
                distance = np.sqrt(np.abs(spot[0] - compare_spot[0]) ** 2 + np.abs(spot[1] - compare_spot[1]) ** 2)
                if distance < 6:
                    comparison_pass = False

            if comparison_pass:
                self.raw_488_spots.append(spot)
                # self.context_488.entryconfig("Remove Spot", state=tk.NORMAL)
                # self.context_488.entryconfig("Spot fitting parameters", state=tk.NORMAL)
                update_final_spots()
                create_view_488()
                create_view_marker()
                try:
                    spot_warnings.handle_close()
                except:
                    """ window was closed """

    @staticmethod
    def open_traces():
        global view_traces
        view_traces = TraceWin()

    @staticmethod
    def view_mode(view):
        global view_mode_win
        try:
            view_mode_win.handle_close()
        except:
            """ window was closed """
        view_mode_win = ViewModeWin(view)


class SpotWarn:
    def __init__(self, spot, warnmsg, view, event):
        self.spot = spot
        self.window = tk.Tk()
        self.window["bg"] = "#444444"
        self.window.title("Warning: Spot does not fit criteria!")
        if event.x_root < 1400:
            self.geometry_string = "+" + str(event.x_root + 50) + "+" + str(event.y_root)
        else:
            self.geometry_string = "+" + str(event.x_root - 480) + "+" + str(event.y_root)
        self.window.geometry(self.geometry_string)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes("-topmost", True)
        self.window.focus_force()
        self.view = view
        if view == 0:
            self.matrix = raw_gui.canvas_marker.create_rectangle(event.x - 5*raw_gui.scale_factor, event.y - 5*raw_gui.scale_factor,
                                                                 event.x + 4.5*raw_gui.scale_factor, event.y + 4.5*raw_gui.scale_factor,
                                                                 outline="white", width=1)
            self.decide = create_circle_dashed(spot[0] * raw_gui.scale_factor + 2, spot[1] * raw_gui.scale_factor + 2,
                                               5*raw_gui.scale_factor, raw_gui.canvas_marker,
                                               outline="red", width=1)
        if view == 1:
            self.matrix = raw_gui.canvas_561.create_rectangle(event.x - 5*raw_gui.scale_factor, event.y - 5*raw_gui.scale_factor,
                                                              event.x + 4.5*raw_gui.scale_factor, event.y + 4.5*raw_gui.scale_factor,
                                                              outline="white", width=1)
            self.decide = create_circle_dashed(spot[0] * raw_gui.scale_factor + 2, spot[1] * raw_gui.scale_factor + 2,
                                               5*raw_gui.scale_factor, raw_gui.canvas_561, outline="red",
                                               width=1)
        if view == 2:
            self.matrix = raw_gui.canvas_488.create_rectangle(event.x - 5*raw_gui.scale_factor, event.y - 5*raw_gui.scale_factor,
                                                              event.x + 4.5*raw_gui.scale_factor, event.y + 4.5*raw_gui.scale_factor,
                                                              outline="white", width=1)
            self.decide = create_circle_dashed(spot[0] * raw_gui.scale_factor + 2, spot[1] * raw_gui.scale_factor + 2,
                                               5*raw_gui.scale_factor, raw_gui.canvas_488, outline="red",
                                               width=1)
        self.frame = tk.Frame(master=self.window, bg="#444444", padx=10, pady=6)
        self.frame.pack(padx=2, pady=4)
        tk.Label(master=self.frame, text=warnmsg, bg="#444444", fg="#ff8800", justify="left", font="calibri 11 bold",
                 padx=10).grid(row=0, column=0, columnspan=2)

        self.keep_spot = tk.Button(master=self.frame, text="Add Spot Anyway", bg="#222266", fg="#cccccc",
                                   padx=4, pady=1, command=self.add_spot)
        change_col_hover(self.keep_spot, button_hv, "#222266")
        self.discard = tk.Button(master=self.frame, text="Reject Spot", bg="#222222", fg="#cccccc",
                                 padx=20, pady=1, command=self.handle_close)
        change_col_hover(self.discard, button_hv, "#222222")
        self.keep_spot.grid(row=1, column=0, pady=12)
        self.discard.grid(row=1, column=1, pady=12)

        self.window.resizable(False, False)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
            if self.view == 0:
                raw_gui.canvas_marker.delete(self.decide)
                raw_gui.canvas_marker.delete(self.matrix)
            if self.view == 1:
                raw_gui.canvas_561.delete(self.decide)
                raw_gui.canvas_561.delete(self.matrix)
            if self.view == 2:
                raw_gui.canvas_488.delete(self.decide)
                raw_gui.canvas_488.delete(self.matrix)
        except:
            """ Window was closed """

    def add_spot(self):
        if self.view == 0:
            raw_gui.raw_marker_spots.append(self.spot)
            update_final_spots()
            create_view_marker()
            create_view_561()
            create_view_488()
            # raw_gui.context_marker.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_marker.entryconfig("Spot fitting parameters", state=tk.NORMAL)
            # raw_gui.context_561.entryconfig("Show Marker Spots", state=tk.NORMAL)
            # raw_gui.context_488.entryconfig("Show Marker Spots", state=tk.NORMAL)
        if self.view == 1:
            raw_gui.raw_561_spots.append(self.spot)
            update_final_spots()
            create_view_561()
            create_view_marker()
            # raw_gui.context_561.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_561.entryconfig("Spot fitting parameters", state=tk.NORMAL)
        if self.view == 2:
            raw_gui.raw_488_spots.append(self.spot)
            update_final_spots()
            create_view_488()
            create_view_marker()
            # raw_gui.context_488.entryconfig("Remove Spot", state=tk.NORMAL)
            # raw_gui.context_488.entryconfig("Spot fitting parameters", state=tk.NORMAL)

        try:
            self.handle_close()
        except:
            """ Window was closed """


class ParamWin:
    def __init__(self, params, coords):

        self.window = tk.Tk()
        self.window["bg"] = "#444444"
        self.window.title(params[0] + " Spot " + str(params[1]))
        if coords[0] < 1600:
            self.geometry_string = "+" + str(coords[0] + 50) + "+" + str(coords[1])
        else:
            self.geometry_string = "+" + str(coords[0] - 320) + "+" + str(coords[1])
        self.window.geometry(self.geometry_string)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.eccentric = params[4] / params[5]
        if self.eccentric > 1:
            self.eccentric = 1 / self.eccentric

        self.label_text1 = "Coordinate Origin: Top Left\n\n" \
                           "Coordinates (x, y):\n\n" \
                           "2D Gaussian Sigma(x):\n\n" \
                           "2D Gaussian Sigma(y):\n\n" \
                           "Gaussian Amplitude:\n\n" \
                           "GaussFit Residual:\n\n" \
                           "Eccentricity (min / maj):"

        self.label_text2 = "\n\n" + str(round(params[2], 1)) + " ,  " + str(round(params[3], 1)) + "\n\n" + \
                           str(round(params[4], 4)) + "\n\n" + str(round(params[5], 4)) + "\n\n" + \
                           str(round(params[6], 4)) + "\n\n" + str(round(params[7], 4)) + "\n\n" + \
                           str(round(self.eccentric, 4))

        tk.Label(master=self.window, text=self.label_text1, bg="#444444", fg="white", justify="left",
                 padx=10, pady=10).pack(side="left")
        tk.Label(master=self.window, text=self.label_text2, bg="#444444", fg="white", justify="left",
                 padx=20, pady=10).pack(side="right")

        self.window.resizable(False, False)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ window has been destroyed """


class RawDefaults:

    def __init__(self):

        self.window = tk.Tk()
        self.window.title("Default Parameters Set-Up")
        self.window.geometry("400x300+600+200")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        tk.Label(master=self.window, text="Start frame", bg="#444444", fg="white").place(x=30, y=70)
        tk.Label(master=self.window, text="Number of frames", bg="#444444", fg="white").place(x=30, y=100)
        tk.Label(master=self.window, text="Background %", bg="#444444", fg="white").place(x=30, y=130)
        tk.Label(master=self.window, text="Convolutions", bg="#444444", fg="white").place(x=30, y=180)

        tk.Label(master=self.window, text="Marker", bg="#444444", fg="white").place(x=190, y=30)
        tk.Label(master=self.window, text="561nm", bg="#444444", fg="white").place(x=260, y=30)
        tk.Label(master=self.window, text="488nm", bg="#444444", fg="white").place(x=330, y=30)

        self.startM = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.startM["validatecommand"] = (self.startM.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.startM, button_hv, "#333333")
        self.lenM = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.lenM["validatecommand"] = (self.lenM.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.lenM, button_hv, "#333333")
        self.bgM = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.bgM["validatecommand"] = (self.bgM.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bgM, button_hv, "#333333")
        self.convM = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.convM["validatecommand"] = (self.convM.register(validate_convolution), "%P", "%d")
        change_col_hover_enterbox(self.convM, button_hv, "#333333")
        self.startM.place(x=190, y=70, width=48)
        self.lenM.place(x=190, y=100, width=48)
        self.bgM.place(x=190, y=130, width=48)
        self.convM.place(x=190, y=180, width=48)
        self.startM.insert(tk.END, str(raw_setup["marker frames"][0]))
        self.lenM.insert(tk.END, str(raw_setup["marker frames"][1]))
        self.bgM.insert(tk.END, str(raw_setup["background"][0]))
        self.convM.insert(tk.END, str(raw_setup["convolutions"][0]))

        self.start561 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.start561["validatecommand"] = (self.start561.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.start561, button_hv, "#333333")
        self.len561 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.len561["validatecommand"] = (self.len561.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.len561, button_hv, "#333333")
        self.bg561 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.bg561["validatecommand"] = (self.bg561.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bg561, button_hv, "#333333")
        self.conv561 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.conv561["validatecommand"] = (self.conv561.register(validate_convolution), "%P", "%d")
        change_col_hover_enterbox(self.conv561, button_hv, "#333333")
        self.start561.place(x=260, y=70, width=48)
        self.len561.place(x=260, y=100, width=48)
        self.bg561.place(x=260, y=130, width=48)
        self.conv561.place(x=260, y=180, width=48)
        self.start561.insert(tk.END, str(raw_setup["561nm frames"][0]))
        self.len561.insert(tk.END, str(raw_setup["561nm frames"][1]))
        self.bg561.insert(tk.END, str(raw_setup["background"][1]))
        self.conv561.insert(tk.END, str(raw_setup["convolutions"][1]))

        self.start488 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.start488["validatecommand"] = (self.start488.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.start488, button_hv, "#333333")
        self.len488 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.len488["validatecommand"] = (self.len488.register(validate_UI_entries), "%P", "%d")
        change_col_hover_enterbox(self.len488, button_hv, "#333333")
        self.bg488 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.bg488["validatecommand"] = (self.bg488.register(validate_percentages), "%P", "%d")
        change_col_hover_enterbox(self.bg488, button_hv, "#333333")
        self.conv488 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.conv488["validatecommand"] = (self.conv488.register(validate_convolution), "%P", "%d")
        change_col_hover_enterbox(self.conv488, button_hv, "#333333")
        self.start488.place(x=330, y=70, width=48)
        self.len488.place(x=330, y=100, width=48)
        self.bg488.place(x=330, y=130, width=48)
        self.conv488.place(x=330, y=180, width=48)
        self.start488.insert(tk.END, str(raw_setup["488nm frames"][0]))
        self.len488.insert(tk.END, str(raw_setup["488nm frames"][1]))
        self.bg488.insert(tk.END, str(raw_setup["background"][2]))
        self.conv488.insert(tk.END, str(raw_setup["convolutions"][2]))

        self.apply = tk.Button(master=self.window, text="Apply", padx=30, pady=1, command=self.apply_defaults,
                               bg="#222266", fg="#cccccc")
        change_col_hover(self.apply, button_hv, "#222266")
        self.cancel = tk.Button(master=self.window, text="Close", padx=31, pady=1, command=self.handle_close,
                                bg="#222222", fg="#cccccc")
        change_col_hover(self.cancel, button_hv, "#222222")
        self.reset = tk.Button(master=self.window, text="Restore Defaults", padx=1, pady=1, command=self.reset_defaults,
                               bg="#662222", fg="#cccccc")
        change_col_hover(self.reset, "#994444", "#662222")
        self.use = tk.Button(master=self.window, text="Apply & Use", padx=12, pady=1, command=self.use_now,
                             bg="#222266", fg="#cccccc")
        change_col_hover(self.use, button_hv, "#222266")
        self.current = tk.Button(master=self.window, text="Import Current", padx=4, pady=1, command=self.import_current,
                                 bg="#662222", fg="#cccccc")
        change_col_hover(self.current, "#994444", "#662222")
        self.target = tk.Button(master=self.window, text="Intensity Targets", padx=12, pady=2, command=self.set_targets,
                                bg="#aa8800", fg="#222222")
        change_col_hover(self.target, "#bbaa44", "#aa8800")
        self.config = tk.Button(master=self.window, text="Fluorophore Config", padx=3, pady=1, command=self.fl_config,
                                bg="#aa8800", fg="#222222")
        change_col_hover(self.config, "#bbaa44", "#aa8800")
        self.apply.place(x=20, y=230)
        self.cancel.place(x=20, y=260)
        self.reset.place(x=284, y=230)
        self.use.place(x=130, y=230)
        self.current.place(x=284, y=260)
        self.target.place(x=20, y=10)
        self.config.place(x=20, y=40)
        self.window.update()

        raw_gui.view_binding(tk.DISABLED)

    def handle_close(self):
        try:
            self.window.destroy()
            raw_gui.defaults_button["state"] = tk.NORMAL
        except:
            """ Window was closed """

        raw_gui.view_binding(tk.NORMAL)

    def apply_defaults(self):
        raw_setup["marker frames"] = [int(self.startM.get()), int(self.lenM.get())]
        raw_setup["561nm frames"] = [int(self.start561.get()), int(self.len561.get())]
        raw_setup["488nm frames"] = [int(self.start488.get()), int(self.len488.get())]
        raw_setup["background"] = [int(self.bgM.get()), int(self.bg561.get()), int(self.bg488.get())]
        raw_setup["convolutions"] = [int(self.convM.get()), int(self.conv561.get()), int(self.conv488.get())]

        print(raw_setup["convolutions"])

        save_defaults()

    def reset_defaults(self):
        global raw_setup
        raw_setup = {
            "marker frames": [10, 300],
            "561nm frames": [320, 300],
            "488nm frames": [630, 300],
            "background": [96, 96, 96],
            "analyse": [1, 1, 1],
            "brightness": [4, 4, 4],
            "inversion power": [2, 3, 1],
            "convolutions": [8, 12, 6],
        }

        self.startM.delete(0, tk.END)
        self.lenM.delete(0, tk.END)
        self.bgM.delete(0, tk.END)
        self.convM.delete(0, tk.END)
        self.start561.delete(0, tk.END)
        self.len561.delete(0, tk.END)
        self.bg561.delete(0, tk.END)
        self.conv561.delete(0, tk.END)
        self.start488.delete(0, tk.END)
        self.len488.delete(0, tk.END)
        self.bg488.delete(0, tk.END)
        self.conv488.delete(0, tk.END)
        self.startM.insert(tk.END, str(raw_setup["marker frames"][0]))
        self.lenM.insert(tk.END, str(raw_setup["marker frames"][1]))
        self.bgM.insert(tk.END, str(raw_setup["background"][0]))
        self.convM.insert(tk.END, str(raw_setup["convolutions"][0]))
        self.start561.insert(tk.END, str(raw_setup["561nm frames"][0]))
        self.len561.insert(tk.END, str(raw_setup["561nm frames"][1]))
        self.bg561.insert(tk.END, str(raw_setup["background"][1]))
        self.conv561.insert(tk.END, str(raw_setup["convolutions"][1]))
        self.start488.insert(tk.END, str(raw_setup["488nm frames"][0]))
        self.len488.insert(tk.END, str(raw_setup["488nm frames"][1]))
        self.bg488.insert(tk.END, str(raw_setup["background"][2]))
        self.conv488.insert(tk.END, str(raw_setup["convolutions"][2]))

    def use_now(self):
        self.apply_defaults()

        raw_gui.start_entry_marker.delete(0, tk.END)
        raw_gui.length_entry_marker.delete(0, tk.END)
        raw_gui.bg_entry_marker.delete(0, tk.END)
        raw_gui.start_entry_561.delete(0, tk.END)
        raw_gui.length_entry_561.delete(0, tk.END)
        raw_gui.bg_entry_561.delete(0, tk.END)
        raw_gui.start_entry_488.delete(0, tk.END)
        raw_gui.length_entry_488.delete(0, tk.END)
        raw_gui.bg_entry_488.delete(0, tk.END)

        raw_gui.start_entry_marker.insert(tk.END, str(raw_setup["marker frames"][0]))
        raw_gui.length_entry_marker.insert(tk.END, str(raw_setup["marker frames"][1]))
        raw_gui.bg_entry_marker.insert(tk.END, str(raw_setup["background"][0]))

        raw_gui.start_entry_561.insert(tk.END, str(raw_setup["561nm frames"][0]))
        raw_gui.length_entry_561.insert(tk.END, str(raw_setup["561nm frames"][1]))
        raw_gui.bg_entry_561.insert(tk.END, str(raw_setup["background"][1]))

        raw_gui.start_entry_488.insert(tk.END, str(raw_setup["488nm frames"][0]))
        raw_gui.length_entry_488.insert(tk.END, str(raw_setup["488nm frames"][1]))
        raw_gui.bg_entry_488.insert(tk.END, str(raw_setup["background"][2]))

        save_defaults()
        self.handle_close()

    def import_current(self):
        self.startM.delete(0, tk.END);
        self.lenM.delete(0, tk.END);
        self.bgM.delete(0, tk.END)
        self.start561.delete(0, tk.END);
        self.len561.delete(0, tk.END);
        self.bg561.delete(0, tk.END)
        self.start488.delete(0, tk.END);
        self.len488.delete(0, tk.END);
        self.bg488.delete(0, tk.END)
        self.startM.insert(tk.END, str(raw_gui.start_entry_marker.get()))
        self.lenM.insert(tk.END, str(raw_gui.length_entry_marker.get()))
        self.bgM.insert(tk.END, str(raw_gui.bg_entry_marker.get()))
        self.start561.insert(tk.END, str(raw_gui.start_entry_561.get()))
        self.len561.insert(tk.END, str(raw_gui.length_entry_561.get()))
        self.bg561.insert(tk.END, str(raw_gui.bg_entry_561.get()))
        self.start488.insert(tk.END, str(raw_gui.start_entry_488.get()))
        self.len488.insert(tk.END, str(raw_gui.length_entry_488.get()))
        self.bg488.insert(tk.END, str(raw_gui.bg_entry_488.get()))

    @staticmethod
    def set_targets():
        set_intensity_target()
        save_preferences()

    @staticmethod
    def fl_config():
        set_fluoro_names()
        save_preferences()


class SpotCriteria:
    def __init__(self, view):
        self.view_map = {
            0: "Marker",
            1: "561nm",
            2: "488nm",
        }
        self.view = view
        self.window = tk.Tk()
        self.window.title("Spot Criteria - " + self.view_map[view])
        self.window.geometry("350x430+240+200")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.help_window = None
        self.help_text = ""
        self.err_label = tk.Label(master=self.window, text="", bg="#444444", fg="#ff9900", font="calibri 10 bold")
        self.err_label.place(x=90, y=12)

        tk.Label(master=self.window, text="Detection threshold", bg="#444444", fg="white").place(x=20, y=50)
        tk.Label(master=self.window, text="Averaging threshold", bg="#444444", fg="white").place(x=20, y=80)
        tk.Label(master=self.window, text="Kernel residual threshold", bg="#444444", fg="white").place(x=20, y=110)
        tk.Label(master=self.window, text="Minimum sigma (Gauss2D)", bg="#444444", fg="white").place(x=20, y=140)
        tk.Label(master=self.window, text="Maximum sigma (Gauss2D)", bg="#444444", fg="white").place(x=20, y=170)
        tk.Label(master=self.window, text="Absolute Intensity (0 - 255)", bg="#444444", fg="white").place(x=20, y=200)
        tk.Label(master=self.window, text="Gaussian Amplitude Threshold", bg="#444444", fg="white").place(x=20, y=230)
        tk.Label(master=self.window, text="Eccentricity threshold", bg="#444444", fg="white").place(x=20, y=260)
        tk.Label(master=self.window, text="Gaussian residual threshold", bg="#444444", fg="white").place(x=20, y=290)

        self.detection = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.detection["validatecommand"] = (self.detection.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.detection, button_hv, "#333333")
        self.averaging = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.averaging["validatecommand"] = (self.averaging.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.averaging, button_hv, "#333333")
        self.residual = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.residual["validatecommand"] = (self.residual.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.residual, button_hv, "#333333")
        self.minisig = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.minisig["validatecommand"] = (self.minisig.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.minisig, button_hv, "#333333")
        self.maxisig = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.maxisig["validatecommand"] = (self.maxisig.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.maxisig, button_hv, "#333333")
        self.intensity = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.intensity["validatecommand"] = (self.intensity.register(self.validate_int), "%P", "%d", "%s")
        change_col_hover_enterbox(self.intensity, button_hv, "#333333")
        self.amplitude = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.amplitude["validatecommand"] = (self.amplitude.register(self.validate_int), "%P", "%d", "%s")
        change_col_hover_enterbox(self.amplitude, button_hv, "#333333")
        self.eccentric = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.eccentric["validatecommand"] = (self.eccentric.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.eccentric, button_hv, "#333333")
        self.gaussresid = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.gaussresid["validatecommand"] = (self.gaussresid.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.gaussresid, button_hv, "#333333")

        self.detection.place(x=250, y=50, width=48)
        self.averaging.place(x=250, y=80, width=48)
        self.residual.place(x=250, y=110, width=48)
        self.minisig.place(x=250, y=140, width=48)
        self.maxisig.place(x=250, y=170, width=48)
        self.intensity.place(x=250, y=200, width=48)
        self.amplitude.place(x=250, y=230, width=48)
        self.eccentric.place(x=250, y=260, width=48)
        self.gaussresid.place(x=250, y=290, width=48)

        self.detection.insert(tk.END, raw_fit_criteria["detection threshold"][self.view])
        self.averaging.insert(tk.END, raw_fit_criteria["averaging distance"][self.view])
        self.residual.insert(tk.END, raw_fit_criteria["minimum kernel residual"][self.view])
        self.minisig.insert(tk.END, raw_fit_criteria["minimum sigma"][self.view])
        self.maxisig.insert(tk.END, raw_fit_criteria["maximum sigma"][self.view])
        self.intensity.insert(tk.END, raw_fit_criteria["minimum intensity"][self.view])
        self.amplitude.insert(tk.END, raw_fit_criteria["minimum gauss amplitude"][self.view])
        self.eccentric.insert(tk.END, raw_fit_criteria["eccentricity threshold"][self.view])
        self.gaussresid.insert(tk.END, raw_fit_criteria["minimum gauss residual"][self.view])

        self.save = tk.Button(master=self.window, text="Apply & Close", padx=8, pady=1, bg="#222266", fg="#cccccc",
                              command=self.apply)
        change_col_hover(self.save, button_hv, "#222266")
        self.close = tk.Button(master=self.window, text="Cancel", padx=27, pady=1, bg="#222222", fg="#cccccc",
                               command=self.handle_close)
        change_col_hover(self.close, button_hv, "#222222")
        self.reset = tk.Button(master=self.window, text="Restore Defaults", padx=3, pady=1, bg="#662222", fg="#cccccc",
                               command=self.reset_defaults)
        change_col_hover(self.reset, "#994444", "#662222")
        self.help = tk.Button(master=self.window, text="Help", padx=10, pady=1, bg="#333333", fg="#cccccc",
                              command=self.open_help_window)
        change_col_hover(self.help, button_hv, "#333333")

        self.save.place(x=40, y=380)
        self.close.place(x=200, y=380)
        self.reset.place(x=40, y=340)
        self.help.place(x=24, y=10)
        self.window.update()

        self.check_empty()

    def check_empty(self):
        if self.detection.get() and self.averaging.get() and self.residual.get() and self.minisig.get() and \
                self.maxisig.get() and self.intensity.get() and self.amplitude.get() and self.eccentric.get() and \
                self.gaussresid.get():
            self.err_label["text"] = ""
            self.err_label.update()
        else:
            self.err_label["text"] = "Blank fields will maintain previous values."
            self.err_label.update()

    def validate_float(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label["text"] = "Blank fields will maintain previous values."
            self.err_label.update()
        if action == "1":
            valid = True
            try:
                float(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label["text"] = ""
                self.err_label.update()
            else:
                self.check_empty()
        if value:
            try:
                float(value)
            except ValueError:
                return False
            if len(value) > 4:
                return False
        return True

    def validate_int(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label["text"] = "Blank fields will maintain previous values."
            self.err_label.update()
        if action == "1":
            valid = True
            try:
                int(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label["text"] = ""
                self.err_label.update()
            else:
                self.check_empty()
            if not value.isdigit():
                return False
            if int(float(value)) > 255:
                return False
        return True

    def handle_close(self):
        try:
            self.window.destroy()
            raw_gui.criteria_marker["state"] = tk.NORMAL
            raw_gui.criteria_561["state"] = tk.NORMAL
            raw_gui.criteria_488["state"] = tk.NORMAL
        except:
            """ Window was closed """
        try:
            self.help_window.destroy()
        except:
            """ Window was closed """

    def apply(self):
        if self.detection.get():
            raw_fit_criteria["detection threshold"][self.view] = float(self.detection.get())
        if self.averaging.get():
            raw_fit_criteria["averaging distance"][self.view] = float(self.averaging.get())
        if self.residual.get():
            raw_fit_criteria["minimum kernel residual"][self.view] = float(self.residual.get())
        if self.minisig.get():
            raw_fit_criteria["minimum sigma"][self.view] = float(self.minisig.get())
        if self.maxisig.get():
            raw_fit_criteria["maximum sigma"][self.view] = float(self.maxisig.get())
        if self.intensity.get():
            raw_fit_criteria["minimum intensity"][self.view] = int(float(self.intensity.get()))
        if self.amplitude.get():
            raw_fit_criteria["minimum gauss amplitude"][self.view] = int(float(self.amplitude.get()))
        if self.eccentric.get():
            raw_fit_criteria["eccentricity threshold"][self.view] = float(self.eccentric.get())
        if self.gaussresid.get():
            raw_fit_criteria["minimum gauss residual"][self.view] = float(self.gaussresid.get())

        save_criteria()
        self.handle_close()

    def reset_defaults(self):
        defaults = {
            "detection threshold": [3.0, 3.0, 3.0],
            "averaging distance": [2.0, 2.0, 2.0],
            "minimum kernel residual": [22.0, 25.0, 25.0],
            "minimum sigma": [0.8, 0.8, 0.8],
            "maximum sigma": [3.5, 3.5, 3.5],
            "minimum intensity": [20, 20, 20],
            "minimum gauss amplitude": [20, 20, 20],
            "eccentricity threshold": [0.6, 0.6, 0.6],
            "minimum gauss residual": [6.0, 6.0, 6.0],
        }

        raw_fit_criteria["detection threshold"][self.view] = defaults["detection threshold"][self.view]
        raw_fit_criteria["averaging distance"][self.view] = defaults["averaging distance"][self.view]
        raw_fit_criteria["minimum kernel residual"][self.view] = defaults["minimum kernel residual"][self.view]
        raw_fit_criteria["minimum sigma"][self.view] = defaults["minimum sigma"][self.view]
        raw_fit_criteria["maximum sigma"][self.view] = defaults["maximum sigma"][self.view]
        raw_fit_criteria["minimum intensity"][self.view] = defaults["minimum intensity"][self.view]
        raw_fit_criteria["minimum gauss amplitude"][self.view] = defaults["minimum gauss amplitude"][self.view]
        raw_fit_criteria["eccentricity threshold"][self.view] = defaults["eccentricity threshold"][self.view]
        raw_fit_criteria["minimum gauss residual"][self.view] = defaults["minimum gauss residual"][self.view]

        self.detection.delete(0, tk.END);
        self.averaging.delete(0, tk.END);
        self.residual.delete(0, tk.END)
        self.minisig.delete(0, tk.END);
        self.maxisig.delete(0, tk.END);
        self.intensity.delete(0, tk.END)
        self.amplitude.delete(0, tk.END);
        self.eccentric.delete(0, tk.END);
        self.gaussresid.delete(0, tk.END)

        self.detection.insert(tk.END, raw_fit_criteria["detection threshold"][self.view])
        self.averaging.insert(tk.END, raw_fit_criteria["averaging distance"][self.view])
        self.residual.insert(tk.END, raw_fit_criteria["minimum kernel residual"][self.view])
        self.minisig.insert(tk.END, raw_fit_criteria["minimum sigma"][self.view])
        self.maxisig.insert(tk.END, raw_fit_criteria["maximum sigma"][self.view])
        self.intensity.insert(tk.END, raw_fit_criteria["minimum intensity"][self.view])
        self.amplitude.insert(tk.END, raw_fit_criteria["minimum gauss amplitude"][self.view])
        self.eccentric.insert(tk.END, raw_fit_criteria["eccentricity threshold"][self.view])
        self.gaussresid.insert(tk.END, raw_fit_criteria["minimum gauss residual"][self.view])

    def open_help_window(self):
        self.help["state"] = tk.DISABLED
        self.help_window = tk.Tk()
        self.help_window.title("Spot Criteria - Help")
        self.help_window.attributes('-topmost', True)
        self.help_window.geometry("+1010+80")

        self.help_text = """
        An 8x8 pixel matrix is scanned across the background subtracted view. The location of any region
        of the image is recorded if the MAXIMUM intensity within the matrix is greater than the average
        intensity of the whole enhanced view by a factor of the 'Detection Threshold'. If 'Av. Mode' is
        enabled, the AVERAGE intensity within the matrix is used instead. 'Av. Mode' can help eliminate
        false positives but may result in an increase in false negatives. 

        Given that spots may lie on the borders between the matrices, a decision must be made whether the
        independent instances detected by different matrices are the same spot or not. This is achieved
        by calculating the weighted average maxima and thus position of each spot; Then the pythagorean
        distance is computed. If the spots are within a distance, in pixels, as defined by the 'Averaging
        threshold', the independent maxima are regarded as the same spot and the positions are averaged.

        Next, a Gaussian matrix is fitted over the the position of the spots. The modulus residual
        matrix is computed and summed to obtain a rough estimate of how close to a true symmetrical 
        Gaussian the spot is. The matrices are normalised for consistency and the threshold above which
        spots are rejected is defined by the 'Kernel residual threshold' This is to reduce the number of
        non-spots which need to be fitted by a true Gaussian resulting in drastically faster execution
        time of the algorithm.

        At this point, a true Gaussian fit is optimised for each spot individually, returning fitting
        parameters of (x, y, sigma x, sigma y, amplitude) which are then used to determine whether the 
        spot should be rejected or kept. If the latter, the coordinates returned by the fit are the sub-
        pixel accurate coordinates of the spot to be used for colocalization checks. 'Minimum sigma' 
        defines the lowest spread in x or y that the spot can have. In essence, a lower bound for spot
        size that makes sure dead pixels and abberations are rejected. 'Maximum sigma' is precisely the
        opposite and marks an upper bound for spot size to prevent large aggregates being kept.

        'Absolute intensity threshold' defines a lower bound on the normalized absolute intensity a spot
        can have to be kept in a range from 0 to 255 as to be consistent with the 'uint8' data type of
        8 bits per channel RGB images. This range is used only for spot detection, fitting and display
        and will not affect trace intensities, the raw data of which are stored in a float32 array.
        'Gaussian amplitude threshold' is a second fall-back that uses the fitted Gaussian amplitude
        as a minimum threshold. 'Eccentricity threshold' is the maximum factor by which the spot can
        be distorted in x or y before it is rejected under the presumption of being two spots within the
        diffraction limit. This Threshold should be less than 1.0 since distortions or aberrations can
        be introduced to single spots by optical effects in the microscope.

        The final decision is made when a kernel is generated using the fitted parameters and the residual
        sum computed above which the spot is rejected as defined by the 'Gaussian residual threshold'.
        """

        tk.Label(master=self.help_window, text=self.help_text, padx=10, pady=10, justify="left",
                 bg="#333333", fg="white", font="calibri 12").pack()
        self.help_window.resizable(False, False)
        self.help_window.protocol("WM_DELETE_WINDOW", self.del_win)

    def del_win(self):
        try:
            self.help_window.destroy()
        except:
            pass
        finally:
            self.help["state"] = tk.NORMAL


class RawStack:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Raw TIF Stack")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.canvas = tk.Canvas(master=self.window, bg="#222222", width=512, height=512)
        self.canvas.grid(row=0, column=0, columnspan=3)
        self.current_frame = tk.IntVar(master=self.window)
        if array_TIF_len is not None:
            self.slider = tk.Scale(master=self.window, from_=0, to=array_TIF_len - 1, bg="#333333", fg="#cccccc",
                                   length=436,
                                   width=24, orient=tk.HORIZONTAL, command=self.display_frame,
                                   variable=self.current_frame)
            change_col_hover_enterbox(self.slider, button_hv, "#333333")
            self.slider.grid(row=2, column=1, pady=4)
            self.back = tk.Button(master=self.window, text="<", bg="#222222", fg="white", padx=6, pady=12,
                                  command=self.dec)
            change_col_hover(self.back, button_hv, "#222222")
            self.next = tk.Button(master=self.window, text=">", bg="#222222", fg="white", padx=6, pady=12,
                                  command=self.inc)
            change_col_hover(self.next, button_hv, "#222222")
            self.normalize = tk.Button(master=self.window, text="Use Global Normalization", padx=10, pady=4,
                                       command=self.toggle_norm, bg="#222266", fg="#cccccc")
            change_col_hover(self.normalize, button_hv, "#222266")
            self.normalize.grid(row=1, column=1, pady=4)
            self.back.grid(row=2, column=0, padx=8)
            self.next.grid(row=2, column=2, padx=8)

            self.colour_maps = ["Viridis", "Cividis", "Magma", "Inferno", "Plasma", "Greys", "Original (Non Uniform)"]
            self.map_dict = {
                "Viridis": mplib.cm.viridis,
                "Cividis": mplib.cm.cividis,
                "Magma": mplib.cm.magma,
                "Inferno": mplib.cm.inferno,
                "Plasma": mplib.cm.plasma,
                "Greys": mplib.cm.gray,
            }
            self.cm_selected = tk.StringVar(master=self.window)
            self.cm_selected.set(self.colour_maps[3])
            self.select_map = tk.OptionMenu(self.window, self.cm_selected, *self.colour_maps,
                                            command=lambda x: self.display_frame(self.current_frame.get()))
            self.select_map["bg"] = "#444f55"
            self.select_map["fg"] = "#cccccc"
            self.select_map.config(highlightbackground="#444444")
            self.select_map.place(x=370, y=520, width=140)

            self.global_normalize = False

            if raw_gui.use_marker.get() == 1:
                self.startM = int(float(raw_gui.start_entry_marker.get()))
                self.lengthM = int(float(raw_gui.length_entry_marker.get()))

            if raw_gui.use_561.get() == 1:
                self.start5 = int(float(raw_gui.start_entry_561.get()))
                self.length5 = int(float(raw_gui.length_entry_561.get()))

            if raw_gui.use_488.get() == 1:
                self.start4 = int(float(raw_gui.start_entry_488.get()))
                self.length4 = int(float(raw_gui.length_entry_488.get()))

            self.global_maximum_m = np.max(array_TIF)
            self.global_maximum_5 = np.max(array_TIF)
            self.global_maximum_4 = np.max(array_TIF)

            try:
                self.global_maximum_m = np.max(array_TIF[:, :, self.startM:self.startM+self.lengthM-1])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ Frames out of range """
            try:
                self.global_maximum_5 = np.max(array_TIF[:, :, self.start5:self.start5+self.length5-1])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ Frames out of range """
            try:
                self.global_maximum_4 = np.max(array_TIF[:, :, self.start4:self.start4+self.length4-1])
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ Frames out of range """

            self.display_frame(0)

        self.window.resizable(False, False)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def inc(self):
        frame = int(float(self.current_frame.get()))
        if frame < array_TIF_len - 1:
            self.slider.set(frame + 1)

    def dec(self):
        frame = int(float(self.current_frame.get()))
        if frame > 0:
            self.slider.set(frame - 1)

    def display_frame(self, frame):
        global raw_img
        minimum = 0
        frame = int(float(frame))
        if raw_gui.is_data_loaded:
            frame_array = array_TIF[:, :, frame]
            if not self.global_normalize:
                maximum = np.max(frame_array)
                minimum = np.min(frame_array)
            elif self.global_normalize:
                try:
                    maximum = self.global_maximum_m
                    if frame < self.startM + self.lengthM:
                        maximum = self.global_maximum_m
                except:
                    """ no marker """
                try:
                    if frame > self.start5 and frame < self.start5 + self.length5:
                        maximum = self.global_maximum_5
                except:
                    """ no 561 """
                try:
                    if frame > self.start4 and frame < self.start4 + self.length4:
                        maximum = self.global_maximum_4
                except:
                    """ no 488 """

            if self.cm_selected.get() == "Original (Non Uniform)":
                frame_array = (frame_array / maximum) * 765
                frame_rgb = np.zeros((np.shape(frame_array)[0], np.shape(frame_array)[1], 3))
                blue = np.clip(np.copy(frame_array), 0, 255)
                green = np.clip(np.copy(frame_array - 255), 0, 255)
                red = np.clip(np.copy(frame_array - 510), 0, 255)
                frame_rgb[:, :, 2] = blue - (green + red) / 2
                frame_rgb[:, :, 1] = green - red / 1.01
                frame_rgb[:, :, 0] = red
            else:
                norm = mplib.colors.Normalize(vmin=minimum, vmax=maximum)
                frame_rgb = self.map_dict[self.cm_selected.get()](norm(frame_array))
                frame_rgb = frame_rgb[:, :, :3] * 255


            image = tirf.create_image(frame_rgb, 512, 512)
            raw_img = tirf.ImageTk.PhotoImage(master=self.canvas, image=image)
            self.canvas.create_image(2, 2, anchor="nw", image=raw_img)

    def toggle_norm(self):
        frame = int(float(self.current_frame.get()))
        if self.global_normalize:
            self.normalize["bg"] = "#222266"
            self.normalize["fg"] = "#cccccc"
            self.normalize["relief"] = tk.RAISED
            self.global_normalize = False
            self.display_frame(frame)
            return
        elif not self.global_normalize:
            self.normalize["bg"] = "#33ff66"
            self.normalize["fg"] = "black"
            self.normalize["relief"] = tk.SUNKEN
            self.global_normalize = True
            self.display_frame(frame)


class TraceWin:
    def __init__(self):
        self.handle_close()
        self.window = tk.Tk()
        self.window.title("FluoroTensor Import Queue - Traces")
        self.window["bg"] = "#666666"
        self.window.geometry("800x820+500+50")
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.frame = tk.Frame(master=self.window, width=780, height=600, bg="#222222")
        self.frame2 = tk.Frame(master=self.window, width=780, height=290, bg="#222222")
        self.frame.place(x=10, y=310)
        self.frame2.place(x=10, y=10)
        self.fig = plt.Figure(figsize=(7.8, 4), dpi=100)
        self.fig2 = plt.Figure(figsize=(7.8, 2.8), dpi=100)
        self.fig.set_facecolor("#333333")
        self.fig2.set_facecolor("#333333")
        self.fig.subplots_adjust(bottom=0.14, left=0.1, right=0.88)
        self.fig2.subplots_adjust(bottom=0.18, left=0.1, right=0.96)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.frame)
        self.canvas2 = FigureCanvasTkAgg(self.fig2, master=self.frame2)
        self.canvas.get_tk_widget().pack(side=tk.TOP)
        self.canvas2.get_tk_widget().pack()

        self.toolbar = NavigationToolbar2Tk(self.canvas, self.frame)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack()

        self.index = 0

        self.next = tk.Button(master=self.window, text=">", bg="#111111", fg="#bbbbbb", padx=10, pady=4,
                              command=self.increment)
        change_col_hover(self.next, button_hv, "#111111")
        self.back = tk.Button(master=self.window, text="<", bg="#111111", fg="#bbbbbb", padx=10, pady=4,
                              command=self.decrement)
        change_col_hover(self.back, button_hv, "#111111")
        self.start = tk.Button(master=self.window, text="<< Start", bg="#222266", fg="#bbbbbb", padx=4, pady=4,
                               command=self.first_trace)
        change_col_hover(self.start, button_hv, "#222266")
        self.end = tk.Button(master=self.window, text="End >>", bg="#222266", fg="#bbbbbb", padx=6, pady=4,
                             command=self.last_trace)
        change_col_hover(self.end, button_hv, "#222266")
        self.delete = tk.Button(master=self.window, text="Delete All Traces", bg="#662222", fg="#cccccc",
                                padx=8, pady=4, command=self.delete_all)
        change_col_hover(self.delete, "#994444", "#662222")
        self.done = tk.Button(master=self.window, text="Close", bg="#111111", fg="#bbbbbb", padx=35, pady=4,
                              command=self.handle_close)
        change_col_hover(self.done, button_hv, "#111111")

        self.start.place(x=280, y=770)
        self.back.place(x=360, y=770)
        self.next.place(x=410, y=770)
        self.end.place(x=470, y=770)
        self.delete.place(x=50, y=770)
        self.done.place(x=650, y=770)

        self.window.update()

        self.conversion_table = {
            "Cyanine 5": preferences["Fluorophore config"][0],
            "mCherry": preferences["Fluorophore config"][1],
            "GFP": preferences["Fluorophore config"][2],
        }

        if len(raw_gui.all_traces) > 0:
            self.display_trace()

        coloc_trend = self.fig2.add_subplot(111)
        coloc_trend.set_facecolor("#222222")
        coloc_trend.spines['bottom'].set_color('blue')
        coloc_trend.spines['top'].set_color('blue')
        coloc_trend.spines['left'].set_color('blue')
        coloc_trend.spines['right'].set_color('blue')
        coloc_trend.xaxis.label.set_color('white')
        coloc_trend.yaxis.label.set_color('white')
        coloc_trend.tick_params(axis='x', colors='white')
        coloc_trend.tick_params(axis='y', colors='white')
        coloc_trend.grid(color="#333333")

        coloc_percent_561 = []
        coloc_percent_488 = []
        for index, count in enumerate(raw_gui.marker_spot_count):
            coloc_percent_561.append(round(raw_gui.coloc_561[index] / count * 100, 1))
            coloc_percent_488.append(round(raw_gui.coloc_488[index] / count * 100, 1))

        coloc_trend.plot(coloc_percent_561, color="#886600", linewidth=1)
        coloc_trend.plot(coloc_percent_488, color="#007744", linewidth=1)

        x_Ser = [x for x, v in enumerate(raw_gui.marker_spot_count)]
        grad561, int561 = linear_regression(x_Ser, coloc_percent_561)
        grad488, int488 = linear_regression(x_Ser, coloc_percent_488)
        trendline561 = []
        trendline488 = []
        for x in x_Ser:
            trendline561.append(grad561 * x + int561)
            trendline488.append(grad488 * x + int488)

        coloc_trend.plot(trendline561, color="#ddaa00", linestyle="--")
        coloc_trend.plot(trendline488, color="#00ee88", linestyle="--")

        # for ind, col_561 in enumerate(coloc_percent_561):
        #     coloc_trend.annotate(str(col_561), xy=(ind, col_561), textcoords="data", color="lightgrey", size=6)
        # for ind, col_488 in enumerate(coloc_percent_488):
        #     coloc_trend.annotate(str(col_488), xy=(ind, col_488), textcoords="data", color="lightgrey", size=6)
        coloc_trend.set_xlabel("TIF file (in order they were loaded.)")
        coloc_trend.set_ylabel("Coloc. %")
        coloc_trend.set_title("Colocalization Percentage: Trend Over Time", color="white")

        self.canvas2.draw()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def display_trace(self):
        self.fig.clf()

        graph = self.fig.add_subplot(111)
        graph2 = graph.twinx()

        graph2.set_facecolor("#222222")
        graph.set_facecolor("#222222")
        graph.spines['bottom'].set_color('blue')
        graph.spines['top'].set_color('blue')
        graph.spines['left'].set_color('blue')
        graph.spines['right'].set_color('blue')
        graph.xaxis.label.set_color('white')
        graph.yaxis.label.set_color('white')
        graph.tick_params(axis='x', colors='white')
        graph.tick_params(axis='y', colors='white')

        graph2.yaxis.label.set_color('white')
        graph2.tick_params(axis='y', colors='white')


        graph.grid(color="#333333")
        if len(raw_gui.raw_traces) > 0:
            graph2.plot(raw_gui.raw_traces[self.index], color="green", linewidth=1)
        graph.plot(raw_gui.all_traces[self.index])
        graph.plot(raw_gui.all_fits[self.index])
        graph.set_xlabel("Trace coordinate (frames)")
        graph.set_ylabel("Background subtracted photon count (A.U.)")
        graph.set_title("Trace " + str(self.index + 1) + ", " + self.conversion_table[raw_gui.trace_info[self.index][0]]
                        + ", Spot " + str(raw_gui.trace_info[self.index][3]), color="white")

        graph2.set_ylabel("Uncorrected photon count (A.U.)")
        if len(raw_gui.raw_traces) > 0:
            graph2.set_ylim([0, np.max(raw_gui.raw_traces[self.index])*1.05])


        self.canvas.draw()

    def increment(self):
        if self.index < len(raw_gui.all_traces) - 1:
            self.index += 1
            self.display_trace()

    def decrement(self):
        if self.index > 0:
            self.index -= 1
            self.display_trace()

    def first_trace(self):
        self.index = 0
        self.display_trace()

    def last_trace(self):
        self.index = len(raw_gui.all_traces) - 1
        self.display_trace()

    def delete_all(self):
        confirm = easygui.ccbox(title="Delete All Traces?", msg="Delete all traces from FluoroTensor import queue."
                                                                "\n\nAre you sure?", choices=["Delete", "Cancel"],
                                default_choice="Cancel")
        if not confirm:
            return

        if len(raw_gui.all_traces) > 0:
            try:
                raw_gui.marker_spot_count.pop(-1)
                raw_gui._488_count.pop(-1)
                raw_gui._561_count.pop(-1)
                raw_gui.coloc_488.pop(-1)
                raw_gui.coloc_561.pop(-1)
                raw_gui.calib_params.pop(-1)
                raw_gui.background640.pop(-1)
                raw_gui.background561.pop(-1)
                raw_gui.background488.pop(-1)
                raw_gui.enhancements.pop(-1)
                raw_gui.file_history_for_export.pop(-1)
            except:
                "Could not delete most recent export data"
                print("An error occurred:\n\n" + traceback.format_exc())

        raw_gui.traces_calculated = False
        raw_gui.all_traces = []
        raw_gui.raw_traces = []
        raw_gui.all_fits = []
        raw_gui.trace_info = []
        self.fig.clf()
        self.canvas.draw()


class AutoWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Automated Analysis")
        self.window["bg"] = "#777777"
        self.window.geometry("300x130+800+200")
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.dont_close)
        self.window.attributes('-topmost', True)
        # self.window.overrideredirect(True)

        self.progress = ttk.Progressbar(master=self.window, orient="horizontal", mode="determinate", length=280)

        self.label = tk.Label(master=self.window, text="Automated analysis is in progress...", bg="#777777",
                              fg="white", anchor=tk.NW, justify="left")

        self.cancel = ttk.Button(master=self.window, text="        Stop Analysis        ", command=self.stop_analysis)

        self.progress.place(x=10, y=10)
        self.label.place(x=10, y=33)
        self.cancel.place(x=90, y=90)

        self.CancelFlag = False

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def stop_analysis(self):
        self.CancelFlag = True
        self.label["text"] = "Automated analysis will be cancelled\nafter current file is finished."
        self.label.update()

    def dont_close(self):
        """ Do nothing when closed """
        pass



class QualityWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Quality Control")
        self.window["bg"] = "#777777"
        self.window.geometry("300x150+800+200")
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.curated_file_list = []

        self.progress = ttk.Progressbar(master=self.window, orient="horizontal", mode="determinate", length=280)
        self.progress.place(x=10, y=10)
        self.label = tk.Label(master=self.window, text="Loading...", bg="#666666", fg="white", anchor=tk.NW, justify="left",
                              relief=tk.SUNKEN, bd=2, wraplength=270)
        self.label.place(x=10, y=38, width=280, height=70)

        self.keep_and_next = tk.Button(master=self.window, text="Keep File & Continue", bd=2, padx=5, pady=2,
                                       bg="#337733", fg="white", command=self.add_and_continue)
        change_col_hover(self.keep_and_next, "#449944", "#337733")
        self.keep_and_next.place(x=15, y=114)
        self.skip_and_next = tk.Button(master=self.window, text="Skip File & Continue", bd=2, padx=5, pady=2,
                                       bg="#773333", fg="white", command=self.skip_and_continue)
        change_col_hover(self.skip_and_next, "#994444", "#773333")
        self.skip_and_next.place(x=160, y=114)

        self.index = 0
        load_tiff(auto=True, file=raw_gui.auto_movie_list[self.index])
        self.label["text"] = f"File {self.index + 1} of {len(raw_gui.auto_movie_list)}: '{raw_gui.auto_movie_list[self.index]}'"
        self.label.update()
        self.progress["maximum"] = len(raw_gui.auto_movie_list) + 1
        self.progress.step(1)


    def handle_close(self, forced=False):
        if not forced:
            if len(self.curated_file_list) > 0:
                raw_gui.auto_movie_list = self.curated_file_list
                choice = easygui.indexbox(title="Before you close...", msg="Curated files have been indexed. The next "
                                                                           "automated run will prompt you to use these files.",
                                          choices=["Save Indices", "Don't Save"], default_choice="Save Indices")
                if choice == 0:
                    self.save_indices()

        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def save_indices(self):
        path = easygui.filesavebox(title="Save Indices", msg="Save curated file list indices for future automation.",
                                   default=default_dir + "*.ind", filetypes=["*.ind", "Automation Index File"])
        if path:
            ext = path[-4:]
            if ext == ".ind":
                path = path[:-4]
            path += ".ind"

            with open(path, "wb") as save_indices:
                pickle.dump(self.curated_file_list, save_indices)

    def add_and_continue(self):
        self.curated_file_list.append(raw_gui.auto_movie_list[self.index])
        if self.index < len(raw_gui.auto_movie_list) - 1:
            self.index += 1
            self.label["text"] = "Loading..."
            self.label.update()
            load_tiff(auto=True, file=raw_gui.auto_movie_list[self.index])
            self.label["text"] = f"File {self.index + 1} of {len(raw_gui.auto_movie_list)}: '{raw_gui.auto_movie_list[self.index]}'"
            self.label.update()
            self.progress.step(1)
        else:
            self.handle_close()


    def skip_and_continue(self):
        if self.index < len(raw_gui.auto_movie_list) - 1:
            self.index += 1
            self.label["text"] = "Loading..."
            self.label.update()
            load_tiff(auto=True, file=raw_gui.auto_movie_list[self.index])
            self.label["text"] = f"File {self.index + 1} of {len(raw_gui.auto_movie_list)}: '{raw_gui.auto_movie_list[self.index]}'"
            self.label.update()
            self.progress.step(1)
        else:
            self.handle_close()


class CalibWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Calibration")
        self.window.geometry("420x310+800+200")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.date = str(calibration["date"])
        self.date_format = self.date[6:] + "/" + self.date[4:6] + "/" + self.date[:4]

        self.info = "To maintain integrity of data, the calibration cannot be changed from" \
                    "\nwithin the program. If the calibration has changed, Enter a new column" \
                    "\nin the 'calibration.xlsx' excel file in the program's directory. New" \
                    "\ncalibrations must begin with the date in YYYYMMDD order in row 1 of" \
                    "\nthe spreadsheet and the new calibration parameters entered in the same" \
                    "\ncolumn. The program will automatically load the most recent" \
                    "\ncalibration according to the dates in the first row. To overide this" \
                    "\nand use an older calibration the letter 'F' should be typed into the" \
                    "\n2nd row (named 'Force calibration') in the column of the calibration" \
                    "\nyou wish to use. Then press the 'Reload Calibration' button."

        self.info_label = tk.Label(master=self.window, text=self.info, bg="#444444", bd=2, relief=tk.RIDGE,
                                   anchor=tk.NW, justify="left", fg="#cccccc", padx=6)
        self.info_label.place(x=10, y=10, width=400, height=160)

        self.current = "Current calibration in use:  " + self.date_format + "\n\n" \
                                                                            "Xo:                                 Yo:\n" \
                                                                            "Xc:                                 Yc:\n" \
                                                                            "SFx:                               SFy:\n" \
                                                                            "\nColocalization Criterion: " + str(
            calibration["criterion"]) + " pixels"
        self.current_scafold = tk.Label(master=self.window, text=self.current, bg="#444444", bd=2, relief=tk.RIDGE,
                                        anchor=tk.NW, justify="left", fg="#cccccc", padx=6)
        self.current_scafold.place(x=10, y=180, width=250, height=120)
        self.paramsX = str(calibration["X0"]) + "\n" + str(calibration["Xc"]) + "\n" + str(calibration["SFx"])
        self.paramsY = str(calibration["Y0"]) + "\n" + str(calibration["Yc"]) + "\n" + str(calibration["SFy"])

        self.xlabel = tk.Label(master=self.window, text=self.paramsX, bg="#444444",
                               anchor=tk.NW, justify="left", fg="#eeeeee", padx=6)
        self.ylabel = tk.Label(master=self.window, text=self.paramsY, bg="#444444",
                               anchor=tk.NW, justify="left", fg="#eeeeee", padx=6)

        self.xlabel.place(x=50, y=210)
        self.ylabel.place(x=160, y=210)

        self.reload = tk.Button(master=self.window, text="Reload Calibration", padx=6, pady=2, bg="#222266",
                                fg="#cccccc", command=self.reload)
        change_col_hover(self.reload, button_hv, "#222266")
        self.close = tk.Button(master=self.window, text="Close", padx=41, pady=2, bg="#222222",
                               fg="#cccccc", command=self.handle_close)
        change_col_hover(self.close, button_hv, "#222222")
        self.start_calib_button = tk.Button(master=self.window, text="Start Calibration", padx=6, pady=2, bg="#ff7722",
                                            fg="black", command=self.start_calibration)
        change_col_hover(self.start_calib_button, "#ffaa33", "#ff7722")
        self.auto_optimise_button = tk.Button(master=self.window, text="Auto-Optimize", padx=6, pady=2, bg="#3377ff",
                                              fg="black", command=self.toggle_auto_optimization)
        change_col_hover(self.auto_optimise_button, "#4499ff", "#3377ff")
        if raw_gui.use_auto_optimization:
            self.auto_optimise_button["bg"] = "#22ff55"
            self.auto_optimise_button["relief"] = tk.SUNKEN
        self.auto_optimise_button.place(x=270, y=240, width=100, height=25)
        self.settings_icon = tk.PhotoImage(master=self.window, file=cwd + "/icons/gear.png")
        self.auto_optimise_settings_button = tk.Button(master=self.window, pady=0, bg="#3377ff",
                                                       image=self.settings_icon, command=calibration_optimizer_settings)
        change_col_hover(self.auto_optimise_settings_button, "#4499ff", "#3377ff")
        self.auto_optimise_settings_button.place(x=380, y=240, width=30, height=25)

        self.start_calib_button.place(x=270, y=180, width=140, height=25)
        self.reload.place(x=270, y=210, width=140, height=25)
        self.close.place(x=270, y=275, width=140, height=25)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def update_params(self):
        self.paramsX = str(calibration["X0"]) + "\n" + str(calibration["Xc"]) + "\n" + str(calibration["SFx"])
        self.paramsY = str(calibration["Y0"]) + "\n" + str(calibration["Yc"]) + "\n" + str(calibration["SFy"])
        self.xlabel["text"] = self.paramsX
        self.ylabel["text"] = self.paramsY
        self.window.update()

    def reload(self):
        load_calibration()
        update_final_spots()
        self.date = str(calibration["date"])
        self.date_format = self.date[6:] + "/" + self.date[4:6] + "/" + self.date[:4]
        self.current = "Current calibration in use:  " + self.date_format + "\n\n" \
                                                                            "Xo:                                 Yo:\n" \
                                                                            "Xc:                                 Yc:\n" \
                                                                            "SFx:                               SFy:\n" \
                                                                            "\nColocalization Criterion: " + str(
            calibration["criterion"])
        self.current_scafold["text"] = self.current
        self.current_scafold.update()

        self.paramsX = str(calibration["X0"]) + "\n" + str(calibration["Xc"]) + "\n" + str(calibration["SFx"])
        self.paramsY = str(calibration["Y0"]) + "\n" + str(calibration["Yc"]) + "\n" + str(calibration["SFy"])
        self.xlabel["text"] = self.paramsX
        self.ylabel["text"] = self.paramsY

        self.xlabel.update()
        self.ylabel.update()
        create_view_marker()
        create_view_561()
        create_view_488()

    def start_calibration(self):
        global calib_analysis_win
        try:
            calib_analysis_win.handle_close()
        except:
            """ Failed to close window """
        calib_analysis_win = CalibAnalysisWin()

    def toggle_auto_optimization(self):
        if not raw_gui.use_auto_optimization:
            raw_gui.use_auto_optimization = True
            self.auto_optimise_button["bg"] = "#22ff55"
            self.auto_optimise_button["relief"] = tk.SUNKEN
            return
        else:
            raw_gui.use_auto_optimization = False
            self.auto_optimise_button["bg"] = "#3377ff"
            self.auto_optimise_button["relief"] = tk.RAISED


class OptSettingsWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Optimizer Settings")
        self.window.geometry("300x230+600+200")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        tk.Label(master=self.window, text="x - min", bg="#444444", fg="white").place(x=10, y=20)
        tk.Label(master=self.window, text="y - min", bg="#444444", fg="white").place(x=10, y=50)
        tk.Label(master=self.window, text="x - step", bg="#444444", fg="white").place(x=10, y=80)
        tk.Label(master=self.window, text="min. scale", bg="#444444", fg="white").place(x=10, y=110)
        tk.Label(master=self.window, text="max. iter.", bg="#444444", fg="white").place(x=10, y=140)
        tk.Label(master=self.window, text="x - max", bg="#444444", fg="white").place(x=160, y=20)
        tk.Label(master=self.window, text="y - max", bg="#444444", fg="white").place(x=160, y=50)
        tk.Label(master=self.window, text="y - step", bg="#444444", fg="white").place(x=160, y=80)
        tk.Label(master=self.window, text="max. scale", bg="#444444", fg="white").place(x=160, y=110)
        tk.Label(master=self.window, text="scale step", bg="#444444", fg="white").place(x=160, y=140)

        self.minx = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.minx, button_hv, "#333333")
        self.minx.place(x=80, y=21, width=60)
        self.miny = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.miny, button_hv, "#333333")
        self.miny.place(x=80, y=51, width=60)
        self.stepx = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.stepx, button_hv, "#333333")
        self.stepx.place(x=80, y=81, width=60)
        self.minscale = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.minscale, button_hv, "#333333")
        self.minscale.place(x=80, y=111, width=60)
        self.maxiter = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.maxiter, button_hv, "#333333")
        self.maxiter.place(x=80, y=141, width=60)

        self.maxx = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.maxx, button_hv, "#333333")
        self.maxx.place(x=230, y=21, width=60)
        self.maxy = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.maxy, button_hv, "#333333")
        self.maxy.place(x=230, y=51, width=60)
        self.stepy = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.stepy, button_hv, "#333333")
        self.stepy.place(x=230, y=81, width=60)
        self.maxscale = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.maxscale, button_hv, "#333333")
        self.maxscale.place(x=230, y=111, width=60)
        self.scalestep = tk.Entry(master=self.window, bg="#333333", fg="white")
        change_col_hover_enterbox(self.scalestep, button_hv, "#333333")
        self.scalestep.place(x=230, y=141, width=60)

        self.minx.insert(tk.END, str(preferences["Calibration optimizer settings"]["xmin"]))
        self.miny.insert(tk.END, str(preferences["Calibration optimizer settings"]["ymin"]))
        self.maxx.insert(tk.END, str(preferences["Calibration optimizer settings"]["xmax"]))
        self.maxy.insert(tk.END, str(preferences["Calibration optimizer settings"]["ymax"]))
        self.stepx.insert(tk.END, str(preferences["Calibration optimizer settings"]["xstep"]))
        self.stepy.insert(tk.END, str(preferences["Calibration optimizer settings"]["ystep"]))
        self.minscale.insert(tk.END, str(preferences["Calibration optimizer settings"]["minscale"]))
        self.maxscale.insert(tk.END, str(preferences["Calibration optimizer settings"]["maxscale"]))
        self.maxiter.insert(tk.END, str(preferences["Calibration optimizer settings"]["maxiter"]))
        self.scalestep.insert(tk.END, str(preferences["Calibration optimizer settings"]["scalestep"]))

        self.save_changes = tk.Button(master=self.window, text="Save Changes", padx=10, pady=2, bg="#222266",
                                      fg="white", command=self.done)
        change_col_hover(self.save_changes, button_hv, "#222266")
        self.save_changes.place(x=30, y=180)

        self.cancel = tk.Button(master=self.window, text="Cancel", padx=30, pady=2, bg="#662222",
                                fg="white", command=self.handle_close)
        change_col_hover(self.cancel, "#993333", "#662222")
        self.cancel.place(x=165, y=180)

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def done(self):
        try:
            preferences["Calibration optimizer settings"]["xmin"] = int(float(self.minx.get()))
            preferences["Calibration optimizer settings"]["ymin"] = int(float(self.miny.get()))
            preferences["Calibration optimizer settings"]["xmax"] = int(float(self.maxx.get()))
            preferences["Calibration optimizer settings"]["ymax"] = int(float(self.maxy.get()))
            preferences["Calibration optimizer settings"]["xstep"] = int(float(self.stepx.get()))
            preferences["Calibration optimizer settings"]["ystep"] = int(float(self.stepy.get()))
            preferences["Calibration optimizer settings"]["minscale"] = int(float(self.minscale.get()))
            preferences["Calibration optimizer settings"]["maxscale"] = int(float(self.maxscale.get()))
            preferences["Calibration optimizer settings"]["maxiter"] = int(float(self.maxiter.get()))
            preferences["Calibration optimizer settings"]["scalestep"] = int(float(self.scalestep.get()))
            save_preferences()
            self.handle_close()
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="Warning! Invalid value encountered while saving settings. Please correct it and try again.")


class CalibAnalysisWin:
    def __init__(self):
        global calib_analysis_win
        easygui.msgbox(title="Calibration", msg="Calibrate chromatic shift correction.\n\n"
                                                "Image Fluororescent beads using 1 mW powers for 640nm and 488nm lasers "
                                                "with EM GAIN set to 300. Use 640nm in marker channel and 488nm in "
                                                "488nm secondary channel. Analyse data using the calibration window "
                                                "until no fewer than 20 colocalized spots have been saved. "
                                                "Spots should be well separated as colocalization threshold will be "
                                                "modified to 5 pixels to ensure shifted beads are identified. Then "
                                                "export to the calibration optimizer template and use the excel solver "
                                                "add-in to minimise SSD by changing Xc, Yc, SFx, and SFy. Set "
                                                "constraints such that SFx and SFy must be greater or equal to 1. "
                                                "Then enter the values into the calibration excel spreadsheet to 3 "
                                                "decimal places. Enter the date in YYYYMMDD order, X0 and Y0 should "
                                                "be set to 0. Then, making sure the template is not open, save the "
                                                "calibration spreadsheet and press 'Reload Calibration' in the "
                                                "FluoroTensor calibration window.")

        try:
            calib_analysis_win.handle_close()
        except:
            """ Failed to close window """

        self.window = tk.Tk()
        self.window.title("Calibration")
        self.window.geometry("316x520+1240+200")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.load_calib_file = tk.Button(master=self.window, text="Analyse Calib. TIF", padx=6, pady=8, bg="#337733",
                                         fg="#cccccc", command=self.analyse_file)
        change_col_hover(self.load_calib_file, "#449944", "#337733")
        self.load_calib_file.place(x=10, y=10)

        self.export_spots = tk.Button(master=self.window, text="Export to Optimizer Template", padx=6, pady=8, bg="#333377",
                                         fg="#cccccc", command=self.export_to_opt)
        change_col_hover(self.export_spots, "#444499", "#333377")
        self.export_spots.place(x=130, y=10)

        tk.Label(master=self.window, text="Oligo Mode", bg="#444444", fg="white").place(x=235, y=55)
        self.oligo_mode = tk.IntVar(master=self.window)
        self.oligo_mode.set(0)
        self.oligo_mode_button = tk.Checkbutton(master=self.window, activebackground="#555555", bg="#444444",
                                                onvalue=1, offvalue=0, variable=self.oligo_mode, command=self.toggle_oligo_mode)
        self.oligo_mode_button.place(x=205, y=54)

        self.file_ID = tk.Label(master=self.window, text="Current File: None", fg="#cccccc", bg="#444444",
                                justify="left", anchor=tk.NW, font="TkDefaultFont 14")
        self.file_ID.place(x=10, y=70)

        self.frame_base = tk.Frame(master=self.window, width=276, height=380, bg="#555555", bd=2,
                                           relief=tk.SUNKEN)
        self.frame_base.place(x=10, y=110)
        self.scrollframe = Sframe(master=self.frame_base, width=276, height=380, bg="#555555")
        self.scrollframe.pack(side="top", expand=1, fill="both")
        self.scrollframe.bind_arrow_keys(self.frame_base)
        self.scrollframe.bind_scroll_wheel(self.frame_base)
        self.link_widget = self.scrollframe.display_widget(tk.Frame)

        self.spot_list = tk.Text(master=self.link_widget, bg="#555555", fg="#dddddd", width=32, height=22,
                                 font="TkDefaultFont 12")
        self.spot_list.pack()
        self.spot_list["state"] = tk.DISABLED

        calibration["criterion"] = 5
        self.calibration_list = []

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window wasn't open """
        try:
            load_calibration()
        except:
            """ Could not load calibration """

    def toggle_oligo_mode(self):
        if self.oligo_mode.get() == 1:
            calibration["criterion"] = 2
            if not raw_gui.use_auto_optimization:
                try:
                    calib_window.toggle_auto_optimization()
                except:
                    """ window not open """
        elif self.oligo_mode.get() == 0:
            calibration["criterion"] = 5
            if raw_gui.use_auto_optimization:
                try:
                    calib_window.toggle_auto_optimization()
                except:
                    """ window not open """

    def analyse_file(self):
        self.load_calib_file["relief"] = tk.SUNKEN
        self.load_calib_file["state"] = tk.DISABLED
        self.load_calib_file["bg"] = "#993333"
        load_tiff()
        path = raw_gui.current_data_path
        try:
            root_filename = path.split("\\")[-1]
            root_id_num = root_filename.split(" ")[1]
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            """ Filename text format not compatible """
            root_id_num = "N / A"
        self.file_ID["text"] = "Current File: " + str(root_id_num)
        self.file_ID.update()
        try:
            if raw_gui.is_data_loaded:
                if self.oligo_mode.get() == 0:
                    raw_gui.power_marker.set(2.5)
                    raw_gui.power_488.set(5)
                create_view_marker()
                create_view_488()

                detect_spots_callback()

                if len(raw_gui.colocalizations) > 0:
                    for coloc in raw_gui.colocalizations:
                        if coloc[0] is not None and coloc[1] is not None:
                            coords_marker = [raw_gui.raw_marker_spots[coloc[0]][0], raw_gui.raw_marker_spots[coloc[0]][1]]
                            coords_488 = [raw_gui.raw_488_spots[coloc[1]][0], raw_gui.raw_488_spots[coloc[1]][1]]
                            if [coords_marker, coords_488] not in self.calibration_list:
                                self.calibration_list.append([coords_marker, coords_488])
                                text = ""
                                for index, coord_pair in enumerate(self.calibration_list):
                                    text += f"{round(coord_pair[0][0], 2)}, {round(coord_pair[0][1], 2)}, " \
                                                              f"{round(coord_pair[1][0], 2)}, {round(coord_pair[1][1], 2)}\n"
                                self.spot_list["state"] = tk.NORMAL
                                self.spot_list.delete(1.0, tk.END)
                                self.spot_list.insert(tk.END, text)
                                self.spot_list["state"] = tk.DISABLED
                                if len(self.calibration_list) >= 22:
                                    self.spot_list["height"] = len(self.calibration_list)
                                self.spot_list.update()

        except ZeroDivisionError:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="Warning! An error occurred while attempting to analyse calibration file.")

        print(self.calibration_list)

        self.load_calib_file["relief"] = tk.RAISED
        self.load_calib_file["state"] = tk.NORMAL
        self.load_calib_file["bg"] = "#337733"

    def export_to_opt(self):
        export_path = easygui.fileopenbox(msg="Open Template excel file for export", default=default_dir + "*.xlsx")
        if export_path:
            optimizer = opxl.load_workbook(export_path)
            sheet = optimizer["Optimizer"]

            for index, coord_pair in enumerate(self.calibration_list):
                row = index + 4
                marker_x, marker_y = coord_pair[0][0], coord_pair[0][1]
                target_x, target_y = coord_pair[1][0], coord_pair[1][1]

                # marker coordinates
                sheet.cell(row=row, column=2).value = round(marker_x, 2)
                sheet.cell(row=row, column=3).value = round(marker_y, 2)

                # 488nm target coordinates
                sheet.cell(row=row, column=5).value = round(target_x, 2)
                sheet.cell(row=row, column=6).value = round(target_y, 2)

                # Trial coordinates using params to map markers onto target
                sheet.cell(row=row, column=8).value = f"=B{row}+(B{row}-$U$4)/$U$6"
                sheet.cell(row=row, column=9).value = f"=C{row}+(C{row}-$U$5)/$U$7"

                # Absolute error between target coordinates and trial coordinates
                sheet.cell(row=row, column=11).value = f"=ABS(H{row}-E{row})"
                sheet.cell(row=row, column=12).value = f"=ABS(I{row}-F{row})"

                # Squared error between target coordinates and trial coordinates
                sheet.cell(row=row, column=14).value = f"=(H{row}-E{row})^2"
                sheet.cell(row=row, column=15).value = f"=(I{row}-F{row})^2"

                # Shift distances between marker and target in x and y
                sheet.cell(row=row, column=17).value = f"=ABS(E{row}-B{row})"
                sheet.cell(row=row, column=18).value = f"=ABS(F{row}-C{row})"

            try:
                optimizer.save(export_path)
                easygui.msgbox(title="Export complete", msg="Data export to excel optimizer has finished successfully.")
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                easygui.msgbox(title="Error!", msg="An error occurred while exporting template. Make sure the file "
                                                   "is not currently open in excel and try again. If this problem "
                                                   "persists, contact the developer.")


class ColWin:
    def __init__(self):
        try:
            self.handle_close()
        except:
            """ Window wasn't open """

        self.window = tk.Tk()
        self.window.title("Marker Colour")
        self.window.geometry("200x100+900+500")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.label = tk.Label(master=self.window, text="640nm         561nm         488nm", bg="#444444")
        self.label.place(x=10, y=10)

        self.red = tk.Button(master=self.window, padx=10, pady=5, bg="red", command=self.choose_red)
        change_col_hover(self.red, "#ff9999", "red")
        self.yellow = tk.Button(master=self.window, padx=10, pady=5, bg="yellow", command=self.choose_yellow)
        change_col_hover(self.yellow, "#ffff99", "yellow")
        self.green = tk.Button(master=self.window, padx=10, pady=5, bg="green", command=self.choose_green)
        change_col_hover(self.green, "#99ff99", "green")

        self.red.place(x=20, y=50)
        self.yellow.place(x=80, y=50)
        self.green.place(x=140, y=50)


    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def choose_red(self):
        raw_gui.marker_view_colour = "640"
        create_view_marker()
        self.handle_close()

    def choose_yellow(self):
        raw_gui.marker_view_colour = "561"
        create_view_marker()
        self.handle_close()

    def choose_green(self):
        raw_gui.marker_view_colour = "488"
        create_view_marker()
        self.handle_close()


class ColocData:
    def __init__(self, count_marker, count_488, count_561, coloc488, coloc561, mode, calibration_params,
                 bg640, bg561, bg488, enhance, history):
        self.count_marker_list = count_marker
        self.count_488_list = count_488
        self.count_561_list = count_561
        self.coloc_488_list = coloc488
        self.coloc_561_list = coloc561
        self.trace_mode = mode
        self.calib_params = calibration_params
        self.backgrounds640 = bg640
        self.backgrounds561 = bg561
        self.backgrounds488 = bg488
        self.enhancements = enhance
        self.file_history = history


class ViewModeWin:

    def __init__(self, view):
        self.window = tk.Tk()
        self.window.title("View mode settings")
        self.window.geometry("200x100+900+500")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.sum_button = tk.Button(master=self.window, text="Sum", bd=2, padx=10, pady=4, bg="#222266", fg="#cccccc",
                                    command=self.toggle_sum)
        change_col_hover(self.sum_button, button_hv, "#222266")

        self.max_button = tk.Button(master=self.window, text="Max", bd=2, padx=10, pady=4, bg="#222266", fg="#cccccc",
                                    command=self.toggle_max)
        change_col_hover(self.max_button, button_hv, "#222266")

        self.close_button = tk.Button(master=self.window, text="OK", bd=2, padx=10, pady=2, bg="#222222", fg="#cccccc",
                                      command=self.apply_settings)
        change_col_hover(self.close_button, button_hv, "#222222")

        self.sum_button.place(x=30, y=15)
        self.max_button.place(x=115, y=15)
        self.close_button.place(x=76, y=60)

        self.view = view
        self.mode = raw_gui.view_modes[self.view]
        self.initialize()

    def initialize(self):
        if self.mode == "b":
            self.sum_on()
            self.max_on()
        elif self.mode == "s":
            self.sum_on()
            self.max_off()
        elif self.mode == "m":
            self.sum_off()
            self.max_on()

    def sum_on(self):
        self.sum_button["relief"] = tk.SUNKEN
        self.sum_button["bg"] = "#22ff77"
        self.sum_button["fg"] = "black"

    def sum_off(self):
        self.sum_button["relief"] = tk.RAISED
        self.sum_button["bg"] = "#222266"
        self.sum_button["fg"] = "#cccccc"

    def max_on(self):
        self.max_button["relief"] = tk.SUNKEN
        self.max_button["bg"] = "#22ff77"
        self.max_button["fg"] = "black"

    def max_off(self):
        self.max_button["relief"] = tk.RAISED
        self.max_button["bg"] = "#222266"
        self.max_button["fg"] = "#cccccc"

    def toggle_sum(self):
        if self.mode == "b":
            self.mode = "m"
            self.initialize()
            return
        if self.mode == "s":
            self.mode = "m"
            self.initialize()
            return
        if self.mode == "m":
            self.mode = "b"
            self.initialize()

    def toggle_max(self):
        if self.mode == "b":
            self.mode = "s"
            self.initialize()
            return
        if self.mode == "m":
            self.mode = "s"
            self.initialize()
            return
        if self.mode == "s":
            self.mode = "b"
            self.initialize()

    def apply_settings(self):
        raw_gui.view_modes[self.view] = self.mode
        redo_frames(self.view)
        self.handle_close()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ window was closed """


class ExpSelectWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.geometry("250x340+1000+300")
        self.window.title("Select markers")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.vars = []
        self.check_buttons = []
        self.available = [0, 1, 2, 3, 4, 5, "Partially bleached"]

        self.final_selection = False
        self.cancelled = False

        tk.Message(master=self.window, bg="#444444", text='Filter secondary distributions by marker steps. If none are '
                                                        'selected, unfiltered secondary distributions will be exported.',
                 fg='#cccccc', justify='left', anchor=tk.NW, width=240).place(x=4, y=4, width=250)

        for i in range(len(self.available)):
            self.vars.append(tk.IntVar(master=self.window))
            self.vars[-1].set(0)
        for i in range(len(self.available)):
            self.check_buttons.append(tk.Checkbutton(master=self.window, bg="#444444", variable=self.vars[i], onvalue=1,
                                          offvalue=0, activebackground="#444444", text=str(self.available[i])))
            self.check_buttons[-1].place(x=20, y=80 + i*30)

        self.select_all_button = tk.Button(master=self.window, text="Select All", padx=5, pady=1, bg="#222222", fg="#cccccc",
                                           command=self.select_all)
        change_col_hover(self.select_all_button, button_hv, "#222222")
        self.clear_all_button = tk.Button(master=self.window, text="Clear All", padx=5, pady=1, bg="#662222", fg="#cccccc",
                                          command=self.clear)
        change_col_hover(self.clear_all_button, "#993333", "#662222")
        self.export_button = tk.Button(master=self.window, text="Export", padx=16, pady=1, bg="#337733", fg="#cccccc",
                               command=self.return_to_export)
        change_col_hover(self.export_button, "#449944", "#337733")

        self.select_all_button.place(x=10, y=300)
        self.clear_all_button.place(x=85, y=300)
        self.export_button.place(x=160, y=300)
        self.window.update()

    def handle_close(self):
        self.cancelled = True
        self.final_selection = True
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def return_to_export(self):
        self.cancelled = False
        self.final_selection = []
        for i in range(len(self.vars)):
            if self.vars[i].get() == 1:
                self.final_selection.append(self.available[i])
        if len(self.final_selection) == 0:
            easygui.msgbox(title="No Markers!", msg="No markers selected. FluoroTensor will now export unfiltered distributions.")
            self.final_selection = ['unfiltered']
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def select_all(self):
        for i in range(len(self.vars)):
            if self.vars[i].get() == 0:
                self.check_buttons[i].invoke()

    def clear(self):
        for i in range(len(self.vars)):
            if self.vars[i].get() == 1:
                self.check_buttons[i].invoke()


class FilterWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.geometry("270x430+1000+300")
        self.window.title("Filter by Properties")
        self.window["bg"] = "#444444"
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.vars = []
        self.check_buttons = []
        self.available = [preferences["Fluorophore config"][0], preferences["Fluorophore config"][1], preferences["Fluorophore config"][2], 0, 1, 2, 3, 4, 5, "Partially bleached"]

        self.final_selection = False
        self.cancelled = False

        tk.Message(master=self.window, bg="#444444", text='Filter displayed traces by fluorophore and step count. Selecting a fluorophore will filter by all traces of that fluorophore present unless any steps are selected.',
                 fg='#cccccc', justify='left', anchor=tk.NW, width=240).place(x=4, y=4, width=250)

        for i in range(len(self.available)):
            self.vars.append(tk.IntVar(master=self.window))
            self.vars[-1].set(0)
        for i in range(len(self.available)):
            self.check_buttons.append(tk.Checkbutton(master=self.window, bg="#444444", variable=self.vars[i], onvalue=1,
                                          offvalue=0, activebackground="#444444", text=str(self.available[i])))
            self.check_buttons[-1].place(x=20, y=80 + i*30)

        if Cy5_trace_count == 0:
            self.check_buttons[0]["state"] = tk.DISABLED
        if mCherry_trace_count == 0:
            self.check_buttons[1]["state"] = tk.DISABLED
        if GFP_trace_count == 0:
            self.check_buttons[2]["state"] = tk.DISABLED

        steps = []
        for idx in range(len(trace_info)):
            steps.append(trace_info[idx][1])
        for idx in range(3, len(self.available)):
            if self.available[idx] not in steps:
                self.check_buttons[idx]["state"] = tk.DISABLED

        self.select_all_button = tk.Button(master=self.window, text="Select All", padx=5, pady=1, bg="#222222", fg="#cccccc",
                                           command=self.select_all)
        change_col_hover(self.select_all_button, button_hv, "#222222")
        self.clear_all_button = tk.Button(master=self.window, text="Clear All", padx=5, pady=1, bg="#662222", fg="#cccccc",
                                          command=self.clear)
        change_col_hover(self.clear_all_button, "#993333", "#662222")
        self.done = tk.Button(master=self.window, text="Done", padx=20, pady=1, bg="#337733", fg="#cccccc",
                               command=self.set_filter)
        change_col_hover(self.done, "#449944", "#337733")

        self.select_all_button.place(x=10, y=390)
        self.clear_all_button.place(x=85, y=390)
        self.done.place(x=160, y=390)
        self.window.update()

    def handle_close(self):
        active_trace_list = list(all_sublist)
        status["text"] = "No choices selected, filtering by all..."
        status.update()
        plot_trace(active_trace_list[current_trace])
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def set_filter(self):
        global active_trace_list, current_trace
        active_trace_list = []
        current_trace = 0
        final_selection = []
        valid_steps = []
        for idx in range(len(self.vars)):
            final_selection.append(int(float(self.vars[idx].get())))
            if final_selection[-1] == 1:
                valid_steps.append(self.available[idx])
        if np.sum(final_selection[3:]) == 0:
            if final_selection[0] == 1:
                active_trace_list += Cy5_sublist
            if final_selection[1] == 1:
                active_trace_list += mCherry_sublist
            if final_selection[2] == 1:
                active_trace_list += GFP_sublist
        else:
            if final_selection[0] == 1:
                for idx in range(len(Cy5_sublist)):
                    if trace_info[Cy5_sublist[idx]][1] in valid_steps:
                        active_trace_list.append(Cy5_sublist[idx])
            if final_selection[1] == 1:
                for idx in range(len(mCherry_sublist)):
                    if trace_info[mCherry_sublist[idx]][1] in valid_steps:
                        active_trace_list.append(mCherry_sublist[idx])
            if final_selection[2] == 1:
                for idx in range(len(GFP_sublist)):
                    if trace_info[GFP_sublist[idx]][1] in valid_steps:
                        active_trace_list.append(GFP_sublist[idx])

        if len(active_trace_list) > 0:
            status["text"] = f"Updated trace filter to {valid_steps}"
            status.update()
            plot_trace(active_trace_list[current_trace])
            update_infobox()

            try:
                self.window.destroy()
            except:
                """ Window was closed """
        else:
            easygui.msgbox(title="Warning!", msg="No traces could be found with the current filter. Please try again.")

    def select_all(self):
        for i in range(len(self.vars)):
            if self.vars[i].get() == 0:
                self.check_buttons[i].invoke()

    def clear(self):
        for i in range(len(self.vars)):
            if self.vars[i].get() == 1:
                self.check_buttons[i].invoke()


class CustomNetWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Custom Neural Network Setup")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.geometry("300x200+500+200")
        self.window.resizable(False, False)

        tk.Label(master=self.window, text="Neural Networks                       Domain Size (Frames)\n\n488nm\n\n561nm\n\n640nm",
                 bg="#444444", fg="white", anchor=tk.NW, justify="left").place(x=20, y=10)

        self.domain_488 = tk.Entry(master=self.window, bg="#333333", fg="white", relief=tk.SUNKEN, bd=2)
        self.domain_488.place(x=180, y=40, width=100)
        self.domain_488.insert(tk.END, preferences["Custom NeuralNet settings"][2])
        self.domain_561 = tk.Entry(master=self.window, bg="#333333", fg="white", relief=tk.SUNKEN, bd=2)
        self.domain_561.place(x=180, y=70, width=100)
        self.domain_561.insert(tk.END, preferences["Custom NeuralNet settings"][1])
        self.domain_640 = tk.Entry(master=self.window, bg="#333333", fg="white", relief=tk.SUNKEN, bd=2)
        self.domain_640.place(x=180, y=100, width=100)
        self.domain_640.insert(tk.END, preferences["Custom NeuralNet settings"][0])

        self.save_changes = tk.Button(master=self.window, text="Save Changes", padx=10, pady=2, bg="#222266", fg="white",
                                      command=self.done)
        change_col_hover(self.save_changes, button_hv, "#222266")
        self.save_changes.place(x=30, y=150)

        self.cancel = tk.Button(master=self.window, text="Cancel", padx=30, pady=2, bg="#662222",
                                fg="white", command=self.handle_close)
        change_col_hover(self.cancel, "#993333", "#662222")
        self.cancel.place(x=160, y=150)

    def handle_close(self):
        set_GUI_state(tk.NORMAL)
        try:
            self.window.destroy()
        except:
            """ Failed """

    def done(self):
        try:
            preferences["Custom NeuralNet settings"][0] = int(float(self.domain_640.get()))
            preferences["Custom NeuralNet settings"][1] = int(float(self.domain_561.get()))
            preferences["Custom NeuralNet settings"][2] = int(float(self.domain_488.get()))
            save_preferences()
            self.handle_close()
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="Warning! Invalid value encountered while saving settings. Please correct it and try again.")


class FRETWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title(f"FRET Analysis")
        self.window.geometry("1600x662+210+180")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.conversion_table = {
            "Cyanine 5": preferences["Fluorophore config"][0],
            "mCherry": preferences["Fluorophore config"][1],
            "GFP": preferences["Fluorophore config"][2],
        }

        if len(unique_ref_list) % 2 != 0:
            easygui.msgbox(title="Error!", msg="Uneven number of markers to secondary FRET traces! Cannot proceed.")
            self.handle_close

        self.frame_time = 0.05

        self.FRET_data_list = []

        self.frame = tk.Frame(master=self.window)
        self.frame.place(x=6, y=46)
        self.figure = plt.Figure(figsize=(12.88, 5.68), dpi=100)
        self.grid_spec = gridspec.GridSpec(2, 2, width_ratios=[3, 1])
        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.95, bottom=0.08, left=0.06, right=0.98, wspace=0.16, hspace=0.36)

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)

        self.toolbar = NavigationToolbar2Tk(self.canvas, self.window)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack()

        tk.Label(master=self.window, text="Frame Duration                      ms", bg="#444444", fg="#cccccc").place(x=10, y=10, height=20)
        self.frame_time_entry = tk.Entry(master=self.window, bg="#333333", fg="white")
        self.frame_time_entry.insert(tk.END, 50)
        change_col_hover_enterbox(self.frame_time_entry, button_hv, "#333333")
        self.frame_time_entry.place(x=110, y=8, width=40, height=23)
        self.frame_time_entry["state"] = tk.DISABLED

        self.back_button = tk.Button(master=self.window, text="<- Back", padx=10, pady=1, bg="#333377", fg="white",
                                     command=self.previous_set)
        change_col_hover(self.back_button, button_hv, "#333377")
        self.back_button.place(x=200, y=8, height=23)

        self.next_button = tk.Button(master=self.window, text="Next >", padx=10, pady=1, bg="#333377", fg="white",
                                     command=self.next_set)
        change_col_hover(self.next_button, button_hv, "#333377")
        self.next_button.place(x=280, y=8, height=23)

        self.stats_box = tk.Message(master=self.window, text="", anchor=tk.NW, bg="#555555", padx=10, pady=10,
                                       relief=tk.RIDGE, width=270, bd=2, fg="#dddddd", font="TkDefaultFont 14")
        self.stats_box.place(x=1300, y=80, width=290, height=200)

        # self.help_button = self.next_button = tk.Button(master=self.window, text="Info", padx=10, pady=1, bg="#ffbb22", fg="black",
        #                                                 command=self.help)
        # change_col_hover(self.help_button, "#ffcc55", "#ffbb22")
        # self.help_button.place(x=380, y=8, height=23)


        self.current_pair = 0

        self.window.update()

        self.get_data()
        self.fit_all_FRET_data()
        self.plot_current()

    def help(self):
        # global FRET_help_fig
        easygui.msgbox(title="Information", image=cwd+"/FRET_fig.png", msg="")

    def previous_set(self):
        if self.current_pair > 0:
            self.current_pair -= 1
            self.plot_current()

    def next_set(self):
        if self.current_pair <= len(self.FRET_data_list) - 1:
            self.current_pair += 1
            self.plot_current()

    def get_data(self):
        for idx in range(0, len(unique_ref_list), 2):
            fret_data = FRETData()
            fret_data.time_axis = np.linspace(0, len(unique_ref_list[idx][1]) * float(int(self.frame_time_entry.get())/1000) - float(int(self.frame_time_entry.get())/1000), len(unique_ref_list[idx][1]))
            fret_data.original_marker_trace = list(unique_ref_list[idx][1])
            fret_data.marker_info = list(unique_ref_list[idx][0])
            fret_data.marker_trace = list(unique_ref_list[idx][1])
            fret_data.original_secondary_trace = list(unique_ref_list[idx+1][1])
            fret_data.secondary_info = list(unique_ref_list[idx+1][0])
            fret_data.secondary_trace = list(unique_ref_list[idx+1][1])

            self.FRET_data_list.append(fret_data)

    def gauss(self, x, a, x0, sigma):
        return a * np.exp(-(x - x0) ** 2 / (2 * sigma ** 2))

    def fit_all_FRET_data(self):
        for idx in range(len(self.FRET_data_list)):
            self.retrieve_FRET_pair_gmm(idx)
            self.get_average_fits(idx)

    def get_average_fits(self, idx):
        data = self.FRET_data_list[idx]
        trace_m = data.marker_trace
        mean_off_m = data.gmm_features_marker[0][0]
        mean_on_m = data.gmm_features_marker[1][0]
        data.marker_fit, data.marker_boundary_points, data.marker_on_times  = self.moving_average_fit(trace_m, mean_off_m, mean_on_m)

        trace_s = data.secondary_trace
        mean_off_s = data.gmm_features_secondary[0][0]
        mean_on_s = data.gmm_features_secondary[1][0]
        data.secondary_fit, data.secondary_boundary_points, data.secondary_on_times = self.moving_average_fit(trace_s, mean_off_s, mean_on_s)


    def moving_average_fit(self, trace, mean_off, mean_on):
        delineator = (mean_on + mean_off) / 2
        fit = []
        transition_points = []
        initial_state = (trace[0] + trace[1] + trace[2] + trace[3] + trace[4]) / 5
        if initial_state < delineator:
            state = mean_off
        else:
            state = mean_on
        init_state = state
        for init in range(3):
            fit.append(state)
        for x in range(1, len(trace) - 4):
            moving_average = (trace[x] + trace[x+1] + trace[x+2] + trace[x+3] + trace[x+4]) / 5
            if moving_average < delineator:
                new_state = mean_off
            else:
                new_state = mean_on
            fit.append(new_state)
            if new_state != state:
                transition_points.append(x+3)
                state = new_state

        fit.append(new_state)
        fit.append(new_state)

        on_times = []
        try:
            if init_state == mean_off and len(transition_points)%2 == 0:
                for t in range(len(transition_points) - 1):
                    on_times.append(transition_points[t + 1] - transition_points[t])
            elif init_state == mean_off and len(transition_points)%2 != 0:
                for t in range(len(transition_points) - 2):
                    on_times.append(transition_points[t + 1] - transition_points[t])
                on_times.append(len(trace) - transition_points[-1])
            elif init_state == mean_on and len(transition_points)%2 != 0:
                on_times.append(transition_points[0])
                for t in range(1, len(transition_points) - 1):
                    on_times.append(transition_points[t + 1] - transition_points[t])
            elif init_state == mean_on and len(transition_points)%2 == 0:
                on_times.append(transition_points[0])
                for t in range(len(transition_points) - 2):
                    on_times.append(transition_points[t + 1] - transition_points[t])
                on_times.append(len(trace) - transition_points[-1])
        except:
            print(traceback.format_exc())
            """ Not enough data points """

        return fit, transition_points, on_times

    def retrieve_FRET_pair_gmm(self, index):
        data = self.FRET_data_list[index]
        marker_features, marker_x_axis, marker_gaussians = self.get_gmm_fit(data.marker_trace)
        secondary_features, secondary_x_axis, secondary_gaussians = self.get_gmm_fit(data.secondary_trace)

        data.gmm_features_marker = marker_features
        data.gmm_features_secondary = secondary_features
        data.gaussian_x_axis_marker = marker_x_axis
        data.gaussian_x_axis_secondary = secondary_x_axis
        data.gaussians_marker = marker_gaussians
        data.gaussians_secondary = secondary_gaussians

    def get_gmm_fit(self, distribution):
        fits = []
        np_dist = np.array(distribution)
        gmm_data = np_dist.reshape(-1, 1)
        np_dist.sort()
        gmm = mixture.GaussianMixture(n_components=2, covariance_type="full")
        gmm.fit(gmm_data)

        weights = gmm.weights_
        means = gmm.means_
        covars = gmm.covariances_
        features = []

        start = np.min(np_dist)
        end = np.max(np_dist)
        x_axis = np.linspace(start, end, 100)

        for fit_index in range(2):
            features.append([float(means[fit_index].ravel()), float(np.sqrt(covars[fit_index]).ravel()),
                             float(weights[fit_index].ravel())])
        features.sort()

        for fit_index in range(2):
            init_fit = features[fit_index][2] * scipy_stats.norm.pdf(np_dist, features[fit_index][0],
                                                                     features[fit_index][1]).ravel()
            scale = np.max(init_fit)
            features[fit_index][2] = scale
            fit = self.gauss(x_axis, scale, features[fit_index][0], features[fit_index][1]).ravel()
            fits.append(fit)

        # features ([[mean1, sigma1, weight1], [mean2, sigma2, weight2]])
        return features, x_axis, fits


    def plot_current(self):
        data = self.FRET_data_list[self.current_pair]

        titles = [
            f"FRET pair {self.current_pair + 1}, Marker: {self.conversion_table[data.marker_info[0]]}",
            "Intensity Histogram",
            f"FRET pair {self.current_pair + 1}, Secondary: {self.conversion_table[data.secondary_info[0]]}",
            "Intensity Histogram",
                  ]

        x_labels = ["Time (s)", "Intensity (A.U.)", "Time (s)", "Intensity (A.U.)"]
        y_labels = ["Intensity (A.U.)", "Frequency Density", "Intensity (A.U.)", "Frequency Density"]

        self.figure.clf()

        for subplot in range(4):

            bg_col = "#222222"
            box_col = "blue"
            text_col = "white"
            grid_col = "#333333"
            line_col = "white"
            fig_col = "#333333"

            self.figure.set_facecolor(fig_col)
            plot_area = self.figure.add_subplot(self.grid_spec[subplot])
            plot_area.set_facecolor(bg_col)
            plot_area.spines['bottom'].set_color(box_col)
            plot_area.spines['top'].set_color(box_col)
            plot_area.spines['left'].set_color(box_col)
            plot_area.spines['right'].set_color(box_col)
            plot_area.xaxis.label.set_color(text_col)
            plot_area.yaxis.label.set_color(text_col)
            plot_area.tick_params(axis='x', colors=text_col, labelsize=8)
            plot_area.tick_params(axis='y', colors=text_col, labelsize=8)
            plot_area.set_xlabel(x_labels[subplot], color=text_col, size=10)
            plot_area.set_ylabel(y_labels[subplot], color=text_col, size=10)
            plot_area.set_title(titles[subplot], color=text_col, size=11)
            if subplot == 0:
                plot_area.plot(data.time_axis, data.marker_trace, linewidth=1)
                plot_area.grid(color=grid_col)
                if len(data.marker_fit) > 0:
                    plot_area.plot(data.time_axis, data.marker_fit)
            elif subplot == 2:
                plot_area.plot(data.time_axis, data.secondary_trace, linewidth=1)
                plot_area.grid(color=grid_col)
                if len(data.secondary_fit) > 0:
                    plot_area.plot(data.time_axis, data.secondary_fit)
            elif subplot == 1:
                plot_area.hist(data.marker_trace, density=True, bins=100)
                try:
                    plot_area.plot(data.gaussian_x_axis_marker, data.gaussians_marker[0], color="red")
                    plot_area.plot(data.gaussian_x_axis_marker, data.gaussians_marker[1], color="green")
                except:
                    """ GMM fit data not present """
                    print(traceback.format_exc())
            elif subplot == 3:
                plot_area.hist(data.secondary_trace, density=True, bins=100)
                try:
                    plot_area.plot(data.gaussian_x_axis_secondary, data.gaussians_secondary[0], color="red")
                    plot_area.plot(data.gaussian_x_axis_secondary, data.gaussians_secondary[1], color="green")
                except:
                    """ GMM fit data not present """
                    print(traceback.format_exc())

        self.canvas.draw()
        try:
            self.display_info()
        except:
            print(traceback.format_exc())
            """ Calculations Failed """

    def display_info(self):
        data = self.FRET_data_list[self.current_pair]
        mean_on_marker = np.mean(data.marker_on_times)
        mean_on_marker_seconds = mean_on_marker * (int(self.frame_time_entry.get()) / 1000)
        mean_on_secondary = np.mean(data.secondary_on_times)
        mean_on_secondary_seconds = mean_on_secondary * (int(self.frame_time_entry.get()) / 1000)
        print(mean_on_marker)
        print(mean_on_marker_seconds)
        print(mean_on_secondary)
        print(mean_on_secondary_seconds)
        k_one = 1 / mean_on_secondary_seconds
        k_minus_one = 1 / mean_on_marker_seconds
        K_EQ = k_one / k_minus_one
        self.stats_box["text"] = f"k1 = {round(k_one, 5)} s^-1\nk-1 = {round(k_minus_one, 5)} s^-1\n\nK (EQ) = {round(K_EQ, 5)}"
        self.stats_box.update()

    def handle_close(self):
        set_GUI_state(tk.NORMAL)
        try:
            self.window.destroy()
        except:
            """ Failed """


class FRETData:
    def __init__(self):
        # Trace & Kinetic properties
        self.time_axis = []
        self.original_marker_trace = []
        self.original_secondary_trace = []
        self.marker_trace = []
        self.secondary_trace = []
        self.marker_info = []
        self.secondary_info = []
        self.marker_fit = []
        self.secondary_fit = []
        self.marker_mean_on = None
        self.marker_mean_off = None
        self.secondary_mean_off = None
        self.secondary_mean_on = None
        self.marker_on_times = []
        self.secondary_on_times = []

        #Gaussian Mixture Parameters
        self.marker_boundary_points = [None, None]
        self.secondary_boundary_points = [None, None]
        # gmm data - features ([[mean1, sigma1, weight1], [mean2, sigma2, weight2]])
        self.gmm_features_marker = []
        self.gmm_features_secondary = []
        # gaussians [[y1], [y2]]
        self.gaussian_x_axis_marker = []
        self.gaussian_x_axis_secondary = []
        self.gaussians_marker = []
        self.gaussians_secondary = []


class VoiceWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.geometry("500x900+800+20")
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window["bg"] = "#555555"
        self.window.resizable(False, False)
        self.window.title("Voice Control")

        self.log = tk.Label(master=self.window, bg="#444444", bd=2, fg="#cccccc", relief=tk.SUNKEN,
                            justify="left", anchor=tk.NW, text="Listening...\n", width=460, wraplength=460)
        self.log.place(x=10, y=10, width=480, height=880)
        self.terminate = False
        self.window.update()
        audio_stream.start_stream()
        # self.calculations("evaluate gas constant begin calculation")
        self.voice_loop()


    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to close """
        self.terminate = True
        audio_stream.stop_stream()


    def voice_loop(self):

        count = 1
        while not self.terminate:
            count += 1
            command_phrase = ""
            audio_data = audio_stream.read(int(frame_buffer / 2), exception_on_overflow=False)

            if recog.AcceptWaveform(audio_data):
                output = recog.PartialResult()

                phrase = output[14:-3]
                word_list = phrase.split(' ')
                if word_list[0] == 'the':
                    word_list.pop(0)
                processed_phrase = ''
                for i in range(len(word_list)):
                    processed_phrase = processed_phrase + word_list[i] + ' '
                processed_phrase = processed_phrase[:-1]
                print(processed_phrase)
                if "computer" in processed_phrase and "evaluate" not in processed_phrase and len(processed_phrase) > 100:
                    recog.Reset()
                if "computer" not in processed_phrase:
                    recog.Reset()
                if "cancel" in processed_phrase or "council" in processed_phrase or "counsel" in processed_phrase:
                    recog.Reset()
                if "computer" in processed_phrase:
                    command_phrase = processed_phrase[processed_phrase.index("computer")+8:]
                    if "terminate" in command_phrase and ("voice" in command_phrase or "control" in command_phrase):
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        self.terminate = True
                        Speech.speak("Terminating voice control")
                        recog.Reset()
                    if "detect" in command_phrase and "spot" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Detecting spots")
                        recog.Reset()
                        detect_spots_callback()
                    if "remove" in command_phrase and "spot" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Removing spots")
                        recog.Reset()
                        remove_spots()
                    if "calculate" in command_phrase and "all" in command_phrase and "traces" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Calculating all traces")
                        recog.Reset()
                        calculate_all_traces(auto=True)
                    if "calculate" in command_phrase and "co-localized" in command_phrase and "traces" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Calculating co-localized traces")
                        recog.Reset()
                        calculate_coloc_traces(auto=True)
                    if "calculate" in command_phrase and "co localized" in command_phrase and "traces" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Calculating co-localized traces")
                        recog.Reset()
                        calculate_coloc_traces(auto=True)
                    if "open" in command_phrase and ("raw data" in command_phrase or "rule data" in command_phrase
                                                     or "tiff file" in command_phrase):
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Opening raw data")
                        recog.Reset()
                        view_raw_stack()
                    if "import" in command_phrase:
                        self.log["text"] += command_phrase + "\n"
                        self.log.update()
                        Speech.speak("Importing analysis")
                        recog.Reset()
                        save_and_exit_raw()
                    if "evaluate" in command_phrase and "begin calculation" in command_phrase:
                        recog.Reset()
                        try:
                            self.calculations(command_phrase)
                        except:
                            print("Warning error occured: Full traceback shown below:")
                            print(traceback.format_exc())
                            Speech.speak("Failed to parse expression")
                            self.log["text"] += "Failed to parse expression" + "\n"
                            self.log.update()

            if count % 2 == 0:
                self.window.update()
                raw_gui.raw_window.update()

        if self.terminate:
            recog.Reset()
            self.handle_close()

    def calculations(self, expression):
        d = {
            "gas constant ": 8.3145,
            "guess constant": 8.3145,
            "pi ": 3.14159265358979,
            "pie ": 3.14159265358979,
            "avogadro constant ": 6.02e23,
        }
        expression_list = expression.split(" ")
        command_index = expression_list.index("evaluate")
        end_exp_index = expression_list.index("begin")
        pure_expression_list = expression_list[command_index + 1:end_exp_index]
        print(expression)
        print(expression_list)
        print(pure_expression_list)
        for index in range(len(pure_expression_list)):
            if pure_expression_list[index] == "times":
                pure_expression_list[index] = "*"
            if index < len(pure_expression_list) - 1:
                if pure_expression_list[index] == "divided" and pure_expression_list[index + 1] == "by":
                    pure_expression_list[index] = "/"
                    pure_expression_list[index + 1] = ""
            if pure_expression_list[index] == "minus":
                pure_expression_list[index] = "-"
            if pure_expression_list[index] == "plus":
                pure_expression_list[index] = "+"
            if pure_expression_list[index] == "squared":
                pure_expression_list[index] = "**2"
            if pure_expression_list[index] == "square":
                pure_expression_list[index] = "**2"
            if pure_expression_list[index] == "cubed":
                pure_expression_list[index] = "**3"
            if pure_expression_list[index] == "cube":
                pure_expression_list[index] = "**3"
            if pure_expression_list[index] == "to":
                pure_expression_list[index] = "two"
            if pure_expression_list[index] == "for":
                pure_expression_list[index] = "four"
            if pure_expression_list[index] == "a":
                pure_expression_list[index] = "eight"
            if pure_expression_list[index] == "the":
                pure_expression_list[index] = ""
        print(pure_expression_list)

        grouped_list = []
        concat = ""
        for index in range(len(pure_expression_list)):
            if pure_expression_list[index] != "*" and pure_expression_list[index] != "/" and \
                    pure_expression_list[index] != "-" and pure_expression_list[index] != "+" and \
                    pure_expression_list[index] != "**2" and pure_expression_list[index] != "**3" and \
                    pure_expression_list[index] != "":
                concat = concat + pure_expression_list[index] + " "
            else:
                grouped_list.append(concat)
                grouped_list.append(pure_expression_list[index])
                concat = ""
        grouped_list.append(concat)
        print(grouped_list)
        for index in range(len(grouped_list)):
            if grouped_list[index] not in d:
                if len(grouped_list[index]) > 3:
                    converted = w2n.word_to_num(grouped_list[index])
                    grouped_list[index] = str(converted)
            else:
                grouped_list[index] = str(d[grouped_list[index]])
        concat = ""
        for index in range(len(grouped_list)):
            if grouped_list[index] != "":
                concat = concat + grouped_list[index] + " "
        print(concat)

        result = eval(concat)

        if result < 0:
            result_phrase = "minus " + convert(-result)
        else:
            result_phrase = convert(result)
        print(result_phrase)
        self.log["text"] += concat + " = " + str(result) + "\n"
        self.log.update()
        Speech.speak(result_phrase)


def start_voice_window():
    voice_window = VoiceWin()


class Speech:

    @classmethod
    def speak(cls, text):
        mp3_file_object = BytesIO()
        tts = gTTS(text, lang='en')
        tts.write_to_fp(mp3_file_object)
        pygame.init()
        pygame.mixer.init()
        pygame.mixer.music.load(mp3_file_object, 'mp3')
        pygame.mixer.music.play()


class LicenceWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.geometry("950x640+200+100")
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window["bg"] = "#555555"
        self.window.resizable(False, False)
        self.window.title("FluoroTensor Licence - Creative Commons NonCommercial 4.0")

        self.frame_base = tk.Frame(master=self.window, width=930, height=580, bg="#555555", bd=2,
                                   relief=tk.SUNKEN)
        self.frame_base.place(x=10, y=10)
        self.scrollframe = Sframe(master=self.frame_base, width=910, height=580, bg="#555555")
        self.scrollframe.pack(side="top", expand=1, fill="both")
        self.scrollframe.bind_arrow_keys(self.frame_base)
        self.scrollframe.bind_scroll_wheel(self.frame_base)
        self.link_widget = self.scrollframe.display_widget(tk.Frame)

        self.licence_text = tk.Text(master=self.link_widget, bg="#ffffff", fg="#444444", width=100, height=338,
                                    font="consolas 12")
        self.licence_text.pack()
        self.scrollframe.bind_scroll_wheel(self.licence_text)
        self.licence_text.insert(tk.END,
                                 """
=======================================================================

Creative Commons Attribution-NonCommercial 4.0 International Public
License

By exercising the Licensed Rights (defined below), You accept and agree
to be bound by the terms and conditions of this Creative Commons
Attribution-NonCommercial 4.0 International Public License ("Public
License"). To the extent this Public License may be interpreted as a
contract, You are granted the Licensed Rights in consideration of Your
acceptance of these terms and conditions, and the Licensor grants You
such rights in consideration of benefits the Licensor receives from
making the Licensed Material available under these terms and
conditions.


Section 1 -- Definitions.

  a. Adapted Material means material subject to Copyright and Similar
     Rights that is derived from or based upon the Licensed Material
     and in which the Licensed Material is translated, altered,
     arranged, transformed, or otherwise modified in a manner requiring
     permission under the Copyright and Similar Rights held by the
     Licensor. For purposes of this Public License, where the Licensed
     Material is a musical work, performance, or sound recording,
     Adapted Material is always produced where the Licensed Material is
     synched in timed relation with a moving image.

  b. Adapter's License means the license You apply to Your Copyright
     and Similar Rights in Your contributions to Adapted Material in
     accordance with the terms and conditions of this Public License.

  c. Copyright and Similar Rights means copyright and/or similar rights
     closely related to copyright including, without limitation,
     performance, broadcast, sound recording, and Sui Generis Database
     Rights, without regard to how the rights are labeled or
     categorized. For purposes of this Public License, the rights
     specified in Section 2(b)(1)-(2) are not Copyright and Similar
     Rights.
  d. Effective Technological Measures means those measures that, in the
     absence of proper authority, may not be circumvented under laws
     fulfilling obligations under Article 11 of the WIPO Copyright
     Treaty adopted on December 20, 1996, and/or similar international
     agreements.

  e. Exceptions and Limitations means fair use, fair dealing, and/or
     any other exception or limitation to Copyright and Similar Rights
     that applies to Your use of the Licensed Material.

  f. Licensed Material means the artistic or literary work, database,
     or other material to which the Licensor applied this Public
     License.

  g. Licensed Rights means the rights granted to You subject to the
     terms and conditions of this Public License, which are limited to
     all Copyright and Similar Rights that apply to Your use of the
     Licensed Material and that the Licensor has authority to license.

  h. Licensor means the individual(s) or entity(ies) granting rights
     under this Public License.

  i. NonCommercial means not primarily intended for or directed towards
     commercial advantage or monetary compensation. For purposes of
     this Public License, the exchange of the Licensed Material for
     other material subject to Copyright and Similar Rights by digital
     file-sharing or similar means is NonCommercial provided there is
     no payment of monetary compensation in connection with the
     exchange.

  j. Share means to provide material to the public by any means or
     process that requires permission under the Licensed Rights, such
     as reproduction, public display, public performance, distribution,
     dissemination, communication, or importation, and to make material
     available to the public including in ways that members of the
     public may access the material from a place and at a time
     individually chosen by them.

  k. Sui Generis Database Rights means rights other than copyright
     resulting from Directive 96/9/EC of the European Parliament and of
     the Council of 11 March 1996 on the legal protection of databases,
     as amended and/or succeeded, as well as other essentially
     equivalent rights anywhere in the world.

  l. You means the individual or entity exercising the Licensed Rights
     under this Public License. Your has a corresponding meaning.


Section 2 -- Scope.

  a. License grant.

       1. Subject to the terms and conditions of this Public License,
          the Licensor hereby grants You a worldwide, royalty-free,
          non-sublicensable, non-exclusive, irrevocable license to
          exercise the Licensed Rights in the Licensed Material to:

            a. reproduce and Share the Licensed Material, in whole or
               in part, for NonCommercial purposes only; and

            b. produce, reproduce, and Share Adapted Material for
               NonCommercial purposes only.

       2. Exceptions and Limitations. For the avoidance of doubt, where
          Exceptions and Limitations apply to Your use, this Public
          License does not apply, and You do not need to comply with
          its terms and conditions.

       3. Term. The term of this Public License is specified in Section
          6(a).

       4. Media and formats; technical modifications allowed. The
          Licensor authorizes You to exercise the Licensed Rights in
          all media and formats whether now known or hereafter created,
          and to make technical modifications necessary to do so. The
          Licensor waives and/or agrees not to assert any right or
          authority to forbid You from making technical modifications
          necessary to exercise the Licensed Rights, including
          technical modifications necessary to circumvent Effective
          Technological Measures. For purposes of this Public License,
          simply making modifications authorized by this Section 2(a)
          (4) never produces Adapted Material.

       5. Downstream recipients.

            a. Offer from the Licensor -- Licensed Material. Every
               recipient of the Licensed Material automatically
               receives an offer from the Licensor to exercise the
               Licensed Rights under the terms and conditions of this
               Public License.

            b. No downstream restrictions. You may not offer or impose
               any additional or different terms or conditions on, or
               apply any Effective Technological Measures to, the
               Licensed Material if doing so restricts exercise of the
               Licensed Rights by any recipient of the Licensed
               Material.

       6. No endorsement. Nothing in this Public License constitutes or
          may be construed as permission to assert or imply that You
          are, or that Your use of the Licensed Material is, connected
          with, or sponsored, endorsed, or granted official status by,
          the Licensor or others designated to receive attribution as
          provided in Section 3(a)(1)(A)(i).

  b. Other rights.

       1. Moral rights, such as the right of integrity, are not
          licensed under this Public License, nor are publicity,
          privacy, and/or other similar personality rights; however, to
          the extent possible, the Licensor waives and/or agrees not to
          assert any such rights held by the Licensor to the limited
          extent necessary to allow You to exercise the Licensed
          Rights, but not otherwise.

       2. Patent and trademark rights are not licensed under this
          Public License.

       3. To the extent possible, the Licensor waives any right to
          collect royalties from You for the exercise of the Licensed
          Rights, whether directly or through a collecting society
          under any voluntary or waivable statutory or compulsory
          licensing scheme. In all other cases the Licensor expressly
          reserves any right to collect such royalties, including when
          the Licensed Material is used other than for NonCommercial
          purposes.


Section 3 -- License Conditions.

Your exercise of the Licensed Rights is expressly made subject to the
following conditions.

  a. Attribution.

       1. If You Share the Licensed Material (including in modified
          form), You must:

            a. retain the following if it is supplied by the Licensor
               with the Licensed Material:

                 i. identification of the creator(s) of the Licensed
                    Material and any others designated to receive
                    attribution, in any reasonable manner requested by
                    the Licensor (including by pseudonym if
                    designated);

                ii. a copyright notice;

               iii. a notice that refers to this Public License;

                iv. a notice that refers to the disclaimer of
                    warranties;

                 v. a URI or hyperlink to the Licensed Material to the
                    extent reasonably practicable;

            b. indicate if You modified the Licensed Material and
               retain an indication of any previous modifications; and

            c. indicate the Licensed Material is licensed under this
               Public License, and include the text of, or the URI or
               hyperlink to, this Public License.

       2. You may satisfy the conditions in Section 3(a)(1) in any
          reasonable manner based on the medium, means, and context in
          which You Share the Licensed Material. For example, it may be
          reasonable to satisfy the conditions by providing a URI or
          hyperlink to a resource that includes the required
          information.

       3. If requested by the Licensor, You must remove any of the
          information required by Section 3(a)(1)(A) to the extent
          reasonably practicable.

       4. If You Share Adapted Material You produce, the Adapter's
          License You apply must not prevent recipients of the Adapted
          Material from complying with this Public License.


Section 4 -- Sui Generis Database Rights.

Where the Licensed Rights include Sui Generis Database Rights that
apply to Your use of the Licensed Material:

  a. for the avoidance of doubt, Section 2(a)(1) grants You the right
     to extract, reuse, reproduce, and Share all or a substantial
     portion of the contents of the database for NonCommercial purposes
     only;

  b. if You include all or a substantial portion of the database
     contents in a database in which You have Sui Generis Database
     Rights, then the database in which You have Sui Generis Database
     Rights (but not its individual contents) is Adapted Material; and

  c. You must comply with the conditions in Section 3(a) if You Share
     all or a substantial portion of the contents of the database.

For the avoidance of doubt, this Section 4 supplements and does not
replace Your obligations under this Public License where the Licensed
Rights include other Copyright and Similar Rights.


Section 5 -- Disclaimer of Warranties and Limitation of Liability.

  a. UNLESS OTHERWISE SEPARATELY UNDERTAKEN BY THE LICENSOR, TO THE
     EXTENT POSSIBLE, THE LICENSOR OFFERS THE LICENSED MATERIAL AS-IS
     AND AS-AVAILABLE, AND MAKES NO REPRESENTATIONS OR WARRANTIES OF
     ANY KIND CONCERNING THE LICENSED MATERIAL, WHETHER EXPRESS,
     IMPLIED, STATUTORY, OR OTHER. THIS INCLUDES, WITHOUT LIMITATION,
     WARRANTIES OF TITLE, MERCHANTABILITY, FITNESS FOR A PARTICULAR
     PURPOSE, NON-INFRINGEMENT, ABSENCE OF LATENT OR OTHER DEFECTS,
     ACCURACY, OR THE PRESENCE OR ABSENCE OF ERRORS, WHETHER OR NOT
     KNOWN OR DISCOVERABLE. WHERE DISCLAIMERS OF WARRANTIES ARE NOT
     ALLOWED IN FULL OR IN PART, THIS DISCLAIMER MAY NOT APPLY TO YOU.

  b. TO THE EXTENT POSSIBLE, IN NO EVENT WILL THE LICENSOR BE LIABLE
     TO YOU ON ANY LEGAL THEORY (INCLUDING, WITHOUT LIMITATION,
     NEGLIGENCE) OR OTHERWISE FOR ANY DIRECT, SPECIAL, INDIRECT,
     INCIDENTAL, CONSEQUENTIAL, PUNITIVE, EXEMPLARY, OR OTHER LOSSES,
     COSTS, EXPENSES, OR DAMAGES ARISING OUT OF THIS PUBLIC LICENSE OR
     USE OF THE LICENSED MATERIAL, EVEN IF THE LICENSOR HAS BEEN
     ADVISED OF THE POSSIBILITY OF SUCH LOSSES, COSTS, EXPENSES, OR
     DAMAGES. WHERE A LIMITATION OF LIABILITY IS NOT ALLOWED IN FULL OR
     IN PART, THIS LIMITATION MAY NOT APPLY TO YOU.

  c. The disclaimer of warranties and limitation of liability provided
     above shall be interpreted in a manner that, to the extent
     possible, most closely approximates an absolute disclaimer and
     waiver of all liability.


Section 6 -- Term and Termination.

  a. This Public License applies for the term of the Copyright and
     Similar Rights licensed here. However, if You fail to comply with
     this Public License, then Your rights under this Public License
     terminate automatically.

  b. Where Your right to use the Licensed Material has terminated under
     Section 6(a), it reinstates:

       1. automatically as of the date the violation is cured, provided
          it is cured within 30 days of Your discovery of the
          violation; or

       2. upon express reinstatement by the Licensor.

     For the avoidance of doubt, this Section 6(b) does not affect any
     right the Licensor may have to seek remedies for Your violations
     of this Public License.

  c. For the avoidance of doubt, the Licensor may also offer the
     Licensed Material under separate terms or conditions or stop
     distributing the Licensed Material at any time; however, doing so
     will not terminate this Public License.

  d. Sections 1, 5, 6, 7, and 8 survive termination of this Public
     License.


Section 7 -- Other Terms and Conditions.

  a. The Licensor shall not be bound by any additional or different
     terms or conditions communicated by You unless expressly agreed.

  b. Any arrangements, understandings, or agreements regarding the
     Licensed Material not stated herein are separate from and
     independent of the terms and conditions of this Public License.


Section 8 -- Interpretation.

  a. For the avoidance of doubt, this Public License does not, and
     shall not be interpreted to, reduce, limit, restrict, or impose
     conditions on any use of the Licensed Material that could lawfully
     be made without permission under this Public License.

  b. To the extent possible, if any provision of this Public License is
     deemed unenforceable, it shall be automatically reformed to the
     minimum extent necessary to make it enforceable. If the provision
     cannot be reformed, it shall be severed from this Public License
     without affecting the enforceability of the remaining terms and
     conditions.

  c. No term or condition of this Public License will be waived and no
     failure to comply consented to unless expressly agreed to by the
     Licensor.

  d. Nothing in this Public License constitutes or may be interpreted
     as a limitation upon, or waiver of, any privileges and immunities
     that apply to the Licensor or You, including from the legal
     processes of any jurisdiction or authority.

=======================================================================
                                 """
                                 )
        self.licence_text["state"] = tk.DISABLED

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

# trackXpress extension package for FluoroTensor ---------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


class TrackingUI:
    def __init__(self):
        self.window = tk.Tk()
        self.window.iconbitmap("icon.ico")
        self.window.geometry("1850x900+40+20")
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window["bg"] = "#555555"
        self.window.resizable(False, False)
        self.window.title("TrackXpress single molecule tracking extension package")

        self.raw_data_tracking, self.proc_data_tracking, self.data_length_tracking = None, None, None
        self.raw_data_coloc1, self.proc_data_coloc1, self.data_length_coloc1 = None, None, None
        self.raw_data_coloc2, self.proc_data_coloc2, self.data_length_coloc2 = None, None, None

        self.current_field_length = 1

        self.display_frames = [tk.IntVar(master=self.window), tk.IntVar(master=self.window),
                               tk.IntVar(master=self.window)]
        self.display_brightness = [tk.IntVar(master=self.window), tk.IntVar(master=self.window),
                                   tk.IntVar(master=self.window)]
        self.display_contrast = [tk.IntVar(master=self.window), tk.IntVar(master=self.window),
                                 tk.IntVar(master=self.window)]
        self.display_frames[0].set(0)
        self.display_frames[1].set(0)
        self.display_frames[2].set(0)
        self.display_brightness[0].set(50)
        self.display_brightness[1].set(50)
        self.display_brightness[2].set(50)
        self.display_contrast[0].set(0)
        self.display_contrast[1].set(0)
        self.display_contrast[2].set(0)
        self.filenames = [None, None, None]
        self.auto_file = "C:/"

        self.ready_for_evolve = False

        self.initial_state = None
        self.tracking_data_trk = []
        self.tracking_data_c1 = []
        self.tracking_data_c2 = []

        self.current_trajectory_series = TrajectorySeries()
        self.cancel_flag = False
        self.saved_flag = True
        self.g_fonts = {
            "title": 11,
            "axis": 8,
            "labels": 10,
        }

        tk.Label(master=self.window, text="Tracking Control", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=6, y=6, width=600, height=888)
        tk.Label(master=self.window, text="Primary Tracking field", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=606, y=6, width=268, height=888)
        tk.Label(master=self.window, text="Colocalization field 1", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=874, y=6, width=268, height=888)
        tk.Label(master=self.window, text="Colocalization field 2", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=1142, y=6, width=268, height=888)
        tk.Label(master=self.window, text="Trajectory Analysis", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=1410, y=6, width=434, height=888)

        self.popout_trk = tk.Button(master=self.window, bg="#222222", fg="#cccccc", text="Float", font="TkDefaultFont 8",
                                    command=lambda: self.float_view(0))
        change_col_hover(self.popout_trk, button_hv, "#222222")
        self.popout_trk.place(x=838, y=6)
        self.popout_c1 = tk.Button(master=self.window, bg="#222222", fg="#cccccc", text="Float", font="TkDefaultFont 8",
                                   command=lambda: self.float_view(1))
        change_col_hover(self.popout_c1, button_hv, "#222222")
        self.popout_c1.place(x=1106, y=6)
        self.popout_c2 = tk.Button(master=self.window, bg="#222222", fg="#cccccc", text="Float", font="TkDefaultFont 8",
                                   command=lambda: self.float_view(2))
        change_col_hover(self.popout_c2, button_hv, "#222222")
        self.popout_c2.place(x=1374, y=6)

        self.boxes = [tk.IntVar(master=self.window), tk.IntVar(master=self.window), tk.IntVar(master=self.window)]
        self.boxes[0].set(1)
        self.boxes[1].set(1)
        self.boxes[2].set(1)
        self.boxes_trk = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                        onvalue=1, offvalue=0, variable=self.boxes[0],
                                        command=self.display_tracking)
        self.boxes_trk.place(x=738, y=8, height=20)
        tk.Label(master=self.window, text="Show boxes", bg="#333333", fg="#cccccc").place(x=758, y=8, height=20)
        self.boxes_c1 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                        onvalue=1, offvalue=0, variable=self.boxes[1],
                                        command=self.display_coloc1)
        self.boxes_c1.place(x=1006, y=8, height=20)
        tk.Label(master=self.window, text="Show boxes", bg="#333333", fg="#cccccc").place(x=1026, y=8, height=20)
        self.boxes_c2 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                        onvalue=1, offvalue=0, variable=self.boxes[2],
                                        command=self.display_coloc2)
        self.boxes_c2.place(x=1274, y=8, height=20)
        tk.Label(master=self.window, text="Show boxes", bg="#333333", fg="#cccccc").place(x=1294, y=8, height=20)


        self.popout_state = [False, False, False]

        self.canvas_tracking = tk.Canvas(self.window, width=256, height=512, bg="#111144", highlightthickness=2,
                                         highlightbackground="black")
        self.canvas_tracking.place(x=610, y=30)
        self.canvas_coloc1 = tk.Canvas(self.window, width=256, height=512, bg="#111144", highlightthickness=2,
                                       highlightbackground="black")
        self.canvas_coloc1.place(x=878, y=30)
        self.canvas_coloc2 = tk.Canvas(self.window, width=256, height=512, bg="#111144", highlightthickness=2,
                                       highlightbackground="black")
        self.canvas_coloc2.place(x=1146, y=30)

        self.canvas_tracking.bind("<ButtonRelease-1>", self.canvas_trk_clicked)
        self.canvas_coloc1.bind("<ButtonRelease-1>", self.canvas_c1_clicked)
        self.canvas_coloc2.bind("<ButtonRelease-1>", self.canvas_c2_clicked)

        tk.Label(master=self.window, text="Scroll Frames", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=610, y=550, height=20)
        tk.Label(master=self.window, text="Scroll Frames", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=878, y=550, height=20)
        tk.Label(master=self.window, text="Scroll Frames", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=1146, y=550, height=20)

        self.slider_trk = tk.Scale(master=self.window, from_=0, to=self.data_length_tracking, bg="#333333",
                                   fg="#cccccc",
                                   length=253, width=24, orient=tk.HORIZONTAL,
                                   command=lambda x: self.display_tracking(),
                                   variable=self.display_frames[0])
        self.slider_trk["state"] = tk.DISABLED
        change_col_hover_enterbox(self.slider_trk, button_hv, "#333333")
        self.slider_trk.place(x=610, y=574)

        self.slider_c1 = tk.Scale(master=self.window, from_=0, to=self.data_length_coloc1, bg="#333333",
                                  fg="#cccccc",
                                  length=253, width=24, orient=tk.HORIZONTAL, command=lambda x: self.display_coloc1(),
                                  variable=self.display_frames[1])
        self.slider_c1["state"] = tk.DISABLED
        change_col_hover_enterbox(self.slider_c1, button_hv, "#333333")
        self.slider_c1.place(x=878, y=574)

        self.slider_c2 = tk.Scale(master=self.window, from_=0, to=self.data_length_coloc2, bg="#333333",
                                  fg="#cccccc",
                                  length=253, width=24, orient=tk.HORIZONTAL, command=lambda x: self.display_coloc2(),
                                  variable=self.display_frames[2])
        self.slider_c2["state"] = tk.DISABLED
        change_col_hover_enterbox(self.slider_c2, button_hv, "#333333")
        self.slider_c2.place(x=1146, y=574)

        tk.Label(master=self.window, text="Brightness", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=610, y=630, height=20)
        tk.Label(master=self.window, text="Brightness", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=878, y=630, height=20)
        tk.Label(master=self.window, text="Brightness", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=1146, y=630, height=20)

        self.bright_trk = tk.Scale(master=self.window, from_=1, to=100, bg="#333333",
                                   fg="#cccccc", length=253, width=24, orient=tk.HORIZONTAL,
                                   command=lambda x: self.display_tracking(update=True),
                                   variable=self.display_brightness[0])
        self.bright_trk["state"] = tk.DISABLED
        change_col_hover_enterbox(self.bright_trk, button_hv, "#333333")
        self.bright_trk.place(x=610, y=654)

        self.bright_c1 = tk.Scale(master=self.window, from_=1, to=100, bg="#333333", fg="#cccccc",
                                  length=253, width=24, orient=tk.HORIZONTAL,
                                  command=lambda x: self.display_coloc1(update=True),
                                  variable=self.display_brightness[1])
        self.bright_c1["state"] = tk.DISABLED
        change_col_hover_enterbox(self.bright_c1, button_hv, "#333333")
        self.bright_c1.place(x=878, y=654)

        self.bright_c2 = tk.Scale(master=self.window, from_=1, to=100, bg="#333333", fg="#cccccc",
                                  length=253, width=24, orient=tk.HORIZONTAL,
                                  command=lambda x: self.display_coloc2(update=True),
                                  variable=self.display_brightness[2])
        self.bright_c2["state"] = tk.DISABLED
        change_col_hover_enterbox(self.bright_c2, button_hv, "#333333")
        self.bright_c2.place(x=1146, y=654)

        tk.Label(master=self.window, text="Contrast", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=610, y=710, height=20)
        tk.Label(master=self.window, text="Contrast", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=878, y=710, height=20)
        tk.Label(master=self.window, text="Constrast", anchor=tk.NW, padx=6, pady=2,
                 bg="#333333", fg="#dddddd").place(x=1146, y=710, height=20)

        self.cont_trk = tk.Scale(master=self.window, from_=-100, to=100, bg="#333333",
                                 fg="#cccccc", length=253, width=24, orient=tk.HORIZONTAL,
                                 command=lambda x: self.display_tracking(update=True),
                                 variable=self.display_contrast[0])
        self.cont_trk["state"] = tk.DISABLED
        change_col_hover_enterbox(self.cont_trk, button_hv, "#333333")
        self.cont_trk.place(x=610, y=734)

        self.cont_c1 = tk.Scale(master=self.window, from_=-100, to=100, bg="#333333", fg="#cccccc",
                                length=253, width=24, orient=tk.HORIZONTAL,
                                command=lambda x: self.display_coloc1(update=True),
                                variable=self.display_contrast[1])
        self.cont_c1["state"] = tk.DISABLED
        change_col_hover_enterbox(self.cont_c1, button_hv, "#333333")
        self.cont_c1.place(x=878, y=734)

        self.cont_c2 = tk.Scale(master=self.window, from_=-100, to=100, bg="#333333", fg="#cccccc",
                                length=253, width=24, orient=tk.HORIZONTAL,
                                command=lambda x: self.display_coloc2(update=True),
                                variable=self.display_contrast[2])
        self.cont_c2["state"] = tk.DISABLED
        change_col_hover_enterbox(self.cont_c2, button_hv, "#333333")
        self.cont_c2.place(x=1146, y=734)

        tk.Label(master=self.window, text="File", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=12, y=30, width=588, height=120)

        tk.Label(master=self.window, text="Select File", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd").place(x=14, y=113, height=20)

        self.file_info = tk.Label(master=self.window, text="File: None\n\nFrame count: None", anchor=tk.NW, padx=6,
                                  pady=2, font="TkDefaultFont 8", bg="#555555", fg="#dddddd", bd=2,
                                  relief=tk.SUNKEN, wraplength=290, justify="left")
        self.file_info.place(x=190, y=50, width=300, height=90)

        self.field_list = ["Tracking", "Coloc. 1", "Coloc. 2"]
        self.field_dict = {
            "Tracking": 0,
            "Coloc. 1": 1,
            "Coloc. 2": 2,
        }
        self.field_selected = tk.StringVar(master=self.window)
        self.field_selected.set(self.field_list[0])
        self.select_field = tk.OptionMenu(self.window, self.field_selected, *self.field_list,
                                          command=self.update_fileinfo)
        self.select_field["bg"] = "#444f55"
        self.select_field["fg"] = "#cccccc"
        self.select_field.config(highlightbackground="#444444")
        self.select_field.place(x=90, y=110, width=95)

        self.load_button = tk.Button(master=self.window, text="Load TIF", padx=10, pady=1, bd=2, bg="#222266",
                                     fg="#cccccc", command=self.load_file)
        change_col_hover(self.load_button, button_hv, "#222266")
        self.load_button.place(x=20, y=50)
        self.automate_button = tk.Button(master=self.window, text="Automate", padx=10, pady=1, bd=2, bg="#338833",
                                         fg="#cccccc", command=self.automate_tracking)
        change_col_hover(self.automate_button, "#44aa44", "#338833")
        self.automate_button.place(x=101, y=50)
        self.save_processed_button = tk.Button(master=self.window, text="Export Processed TIF", padx=23, pady=1, bd=2,
                                               bg="#222222", fg="#cccccc", command=self.export_tiff_file)
        change_col_hover(self.save_processed_button, button_hv, "#222222")
        self.save_processed_button.place(x=20, y=80)
        self.view_raw_button = tk.Button(master=self.window, text="Open Raw", padx=16, pady=8, bd=2, bg="#222266",
                                         fg="#cccccc", command=self.view_raw_data)
        change_col_hover(self.view_raw_button, button_hv, "#222266")
        self.view_raw_button.place(x=498, y=50, width=94)
        self.view_proc_button = tk.Button(master=self.window, text="Open Processed", padx=3, pady=8, bd=2, bg="#222266",
                                          fg="#cccccc", font="TkDefaultFont 8", command=self.view_proc_data)
        change_col_hover(self.view_proc_button, button_hv, "#222266")
        self.view_proc_button.place(x=498, y=101)

        tk.Label(master=self.window, text="Preprocess & Enhance", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=12, y=154, width=588, height=70)

        self.redo_proc_button = tk.Button(master=self.window, text="Redo Preprocessing", padx=25, pady=7, bd=2,
                                          bg="#222266",
                                          fg="#cccccc",
                                          command=lambda: self.process_raw_data(field=self.field_selected.get()))
        change_col_hover(self.redo_proc_button, button_hv, "#222266")
        self.redo_proc_button.place(x=20, y=174)

        tk.Label(master=self.window, text="Filter Power:", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd").place(x=190, y=181, height=20)
        self.filter_power = tk.StringVar(master=self.window)
        self.filter_power.set(1.10)
        self.spin_box_filter = tk.Spinbox(master=self.window, from_=0.25, to=10, increment=0.05, repeatinterval=40,
                                          textvariable=self.filter_power, bg="#bbbbbb", fg="black",
                                          bd=2, font="Arial 12 bold", validate="key")
        change_col_hover_enterbox(self.spin_box_filter, "#cccccc", "#bbbbbb")
        # self.spin_box_filter["validatecommand"] = (self.spin_box_filter.register(prevent_entry), "%P")
        self.spin_box_filter.place(x=280, y=177, width=80, height=32)

        tk.Label(master=self.window, text="Average Frames:", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd").place(x=400, y=181, height=20)
        self.frames_to_average = tk.StringVar(master=self.window)
        self.frames_to_average.set(1)
        self.spin_box_frames = tk.Spinbox(master=self.window, from_=1, to=50, increment=1, repeatinterval=40,
                                          textvariable=self.frames_to_average, bg="#bbbbbb", fg="black",
                                          bd=2, font="Arial 12 bold", validate="key")
        change_col_hover_enterbox(self.spin_box_frames, "#cccccc", "#bbbbbb")
        # self.spin_box_frames["validatecommand"] = (self.spin_box_frames.register(prevent_entry), "%P")
        self.spin_box_frames.place(x=515, y=177, width=75, height=32)

        tk.Label(master=self.window, text="Initial Selection & Particle Detection", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=12, y=228, width=588, height=300)

        self.err_label = tk.Label(master=self.window, text="", bg="#444444", fg="#ff9900", font="calibri 10 bold")
        self.err_label.place(x=240, y=229)

        tk.Label(master=self.window, text="Detection threshold", bg="#444444", fg="white").place(x=20, y=260)
        tk.Label(master=self.window, text="Averaging threshold", bg="#444444", fg="white").place(x=20, y=290)
        tk.Label(master=self.window, text="Kernel residual threshold", bg="#444444", fg="white").place(x=20, y=320)
        tk.Label(master=self.window, text="Minimum sigma (Gauss2D)", bg="#444444", fg="white").place(x=20, y=350)
        tk.Label(master=self.window, text="Maximum sigma (Gauss2D)", bg="#444444", fg="white").place(x=20, y=380)
        tk.Label(master=self.window, text="Absolute Intensity (0 - 255)", bg="#444444", fg="white").place(x=20, y=410)
        tk.Label(master=self.window, text="Gaussian Amplitude Threshold", bg="#444444", fg="white").place(x=20, y=440)
        tk.Label(master=self.window, text="Eccentricity threshold", bg="#444444", fg="white").place(x=20, y=470)
        tk.Label(master=self.window, text="Gaussian residual threshold", bg="#444444", fg="white").place(x=20, y=500)
        tk.Label(master=self.window, text="Start Frame", bg="#444444", fg="white").place(x=460, y=235)

        self.detection = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.detection["validatecommand"] = (self.detection.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.detection, button_hv, "#333333")
        self.averaging = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.averaging["validatecommand"] = (self.averaging.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.averaging, button_hv, "#333333")
        self.residual = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.residual["validatecommand"] = (self.residual.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.residual, button_hv, "#333333")
        self.minisig = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.minisig["validatecommand"] = (self.minisig.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.minisig, button_hv, "#333333")
        self.maxisig = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.maxisig["validatecommand"] = (self.maxisig.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.maxisig, button_hv, "#333333")
        self.intensity = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.intensity["validatecommand"] = (self.intensity.register(self.validate_int), "%P", "%d", "%s")
        change_col_hover_enterbox(self.intensity, button_hv, "#333333")
        self.amplitude = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.amplitude["validatecommand"] = (self.amplitude.register(self.validate_int), "%P", "%d", "%s")
        change_col_hover_enterbox(self.amplitude, button_hv, "#333333")
        self.eccentric = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.eccentric["validatecommand"] = (self.eccentric.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.eccentric, button_hv, "#333333")
        self.gaussresid = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.gaussresid["validatecommand"] = (self.gaussresid.register(self.validate_float), "%P", "%d", "%s")
        change_col_hover_enterbox(self.gaussresid, button_hv, "#333333")
        self.init_frame = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.init_frame["validatecommand"] = (self.init_frame.register(self.init_frame_trigger), "%P", "%d", "%s")
        change_col_hover_enterbox(self.init_frame, button_hv, "#333333")

        self.detection.place(x=200, y=263, width=48)
        self.averaging.place(x=200, y=293, width=48)
        self.residual.place(x=200, y=323, width=48)
        self.minisig.place(x=200, y=353, width=48)
        self.maxisig.place(x=200, y=383, width=48)
        self.intensity.place(x=200, y=413, width=48)
        self.amplitude.place(x=200, y=443, width=48)
        self.eccentric.place(x=200, y=473, width=48)
        self.gaussresid.place(x=200, y=503, width=48)
        self.init_frame.place(x=540, y=238, width=52)

        self.detection.insert(tk.END, 4.0)
        self.averaging.insert(tk.END, 2.0)
        self.residual.insert(tk.END, 23.0)
        self.minisig.insert(tk.END, 0.7)
        self.maxisig.insert(tk.END, 3.5)
        self.intensity.insert(tk.END, 10)
        self.amplitude.insert(tk.END, 10)
        self.eccentric.insert(tk.END, 0.65)
        self.gaussresid.insert(tk.END, 10.0)
        self.init_frame.insert(tk.END, 0)

        self.canvas_initial = tk.Canvas(self.window, width=128, height=256, bg="#111144", highlightthickness=2,
                                        highlightbackground="black")
        self.canvas_initial.place(x=460, y=263)

        self.detect_spots_button = tk.Button(master=self.window, text="Calculate Initial State", padx=7, pady=4, bd=2,
                                             bg="#225577", fg="#dddddd", command=self.get_initial_spots)
        change_col_hover(self.detect_spots_button, "#337799", "#225577")
        self.detect_spots_button["state"] = tk.DISABLED
        self.detect_spots_button.place(x=255, y=263)

        self.discard_spots_button = tk.Button(master=self.window, text="Discard", padx=5, pady=4, bd=2,
                                              bg="#662222", fg="#dddddd", command=self.discard_spots)
        change_col_hover(self.discard_spots_button, "#993333", "#662222")
        self.discard_spots_button.place(x=395, y=263)

        self.spotlistframe_base = tk.Frame(master=self.window, width=197, height=223, bg="#555555", bd=2,
                                           relief=tk.SUNKEN)
        self.spotlistframe_base.place(x=255, y=300)
        self.spot_scrollframe = Sframe(master=self.spotlistframe_base, width=175, height=199, bg="#555555")
        self.spot_scrollframe.pack(side="top", expand=1, fill="both")
        self.spot_scrollframe.bind_arrow_keys(self.spotlistframe_base)
        self.spot_scrollframe.bind_scroll_wheel(self.spotlistframe_base)
        self.spot_link_widget = self.spot_scrollframe.display_widget(tk.Frame)

        self.spot_list = tk.Text(master=self.spot_link_widget, bg="#555555", fg="#dddddd", width=82, height=300,
                                 font="TkDefaultFont 8")
        self.spot_list.pack()
        self.spot_list["state"] = tk.DISABLED

        tk.Label(master=self.window, text="Tracking Parameters", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=12, y=532, width=588, height=300)

        self.err_label2 = tk.Label(master=self.window, text="", bg="#444444", fg="#ff9900", font="calibri 10 bold")
        self.err_label2.place(x=160, y=533)

        self.start_tracking_button = tk.Button(master=self.window, text="Analyse", padx=29, pady=2, bd=2,
                                               bg="#338833", fg="#dddddd", command=self.analyse)
        self.start_tracking_button["state"] = tk.DISABLED
        change_col_hover(self.start_tracking_button, "#44aa44", "#338833")
        self.start_tracking_button.place(x=484, y=560)

        tk.Label(master=self.window, text="Tracking Mode", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd").place(x=240, y=561, height=20)
        self.analysis_mode = tk.StringVar(master=self.window)
        self.analysis_mode.set("Continuous")
        self.select_mode = tk.OptionMenu(self.window, self.analysis_mode, *["Continuous"],
                                         command=self.toggle_button_state)
        self.select_mode["bg"] = "#444f55"
        self.select_mode["fg"] = "#cccccc"
        self.select_mode.config(highlightbackground="#444444")
        self.select_mode.place(x=340, y=558, width=140, height=32)

        tk.Label(master=self.window, text="Calibration", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="#dddddd", bd=2, relief=tk.RAISED).place(x=240, y=600, width=352, height=100)
        tk.Label(master=self.window, text="Distance Units", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="#dddddd").place(x=242, y=630, height=20)
        self.units = tk.StringVar(master=self.window)
        self.units.set("Pixels")
        self.select_units = tk.OptionMenu(self.window, self.units, *["Pixels (px)", "Nanometres (nm)"],
                                          command=self.update_units)
        self.select_units["bg"] = "#444f55"
        self.select_units["fg"] = "#cccccc"
        self.select_units.config(highlightbackground="#555555")
        self.select_units.place(x=340, y=627, width=140, height=32)

        self.coloc_calibration_button = tk.Button(master=self.window, text="Coloc. Transform", padx=0, pady=2, bd=2,
                                                  bg="#222222", fg="#dddddd")
        change_col_hover(self.coloc_calibration_button, button_hv, "#222222")
        self.coloc_calibration_button.place(x=484, y=629)

        tk.Label(master=self.window, text="Pixel Size", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="white").place(x=268, y=664, height=20)
        tk.Label(master=self.window, text="nm", anchor=tk.NW, padx=2, pady=2,
                 bg="#555555", fg="white").place(x=393, y=664, height=20)
        self.pixel_size = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.pixel_size["validatecommand"] = (self.pixel_size.register(self.validate_float_track), "%P", "%d", "%s")
        self.pixel_size.insert(tk.END, 160)
        change_col_hover_enterbox(self.pixel_size, button_hv, "#333333")
        self.pixel_size.place(x=342, y=666, width=50)

        tk.Label(master=self.window, text="Frame Interval", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="white").place(x=430, y=664, height=20)
        tk.Label(master=self.window, text="s", anchor=tk.NW, padx=2, pady=2,
                 bg="#555555", fg="white").place(x=573, y=664, height=20)
        self.frame_interval = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.frame_interval["validatecommand"] = (
        self.frame_interval.register(self.validate_float_track), "%P", "%d", "%s")
        self.frame_interval.insert(tk.END, 0.1)
        change_col_hover_enterbox(self.frame_interval, button_hv, "#333333")
        self.frame_interval.place(x=530, y=666, width=40)

        tk.Label(master=self.window, text="Start at", bg="#444444", fg="white").place(x=20, y=561)
        tk.Label(master=self.window, text="Stop at", bg="#444444", fg="white").place(x=136, y=561)
        self.start_frame = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.start_frame["validatecommand"] = (self.start_frame.register(self.validate_int_track), "%P", "%d", "%s")
        self.start_frame.insert(tk.END, 0)
        change_col_hover_enterbox(self.start_frame, button_hv, "#333333")
        self.start_frame.place(x=74, y=564, width=40)
        self.stop_frame = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.stop_frame["validatecommand"] = (self.stop_frame.register(self.validate_int_track), "%P", "%d", "%s")
        self.stop_frame.insert(tk.END, 0)
        change_col_hover_enterbox(self.stop_frame, button_hv, "#333333")
        self.stop_frame.place(x=190, y=564, width=40)

        tk.Label(master=self.window, text="Lock End Frame", bg="#444444", fg="#cccccc").place(x=216, y=538)
        self.lock_stop = tk.IntVar(master=self.window)
        self.lock_stop.set(0)
        self.lock_stop_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                                   onvalue=1, offvalue=0, variable=self.lock_stop)
        self.lock_stop_button.place(x=190, y=536)

        tk.Label(master=self.window, text="Particle Discriminator", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="#dddddd", bd=2, relief=tk.RAISED).place(x=20, y=600, width=214, height=100)
        tk.Label(master=self.window, text="Max. Displacement", bg="#555555", fg="white").place(x=25, y=621)
        self.unitl1 = tk.Label(master=self.window, text="px", bg="#555555", fg="white")
        self.unitl1.place(x=184, y=621)
        self.max_displacement = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.max_displacement["validatecommand"] = (
        self.max_displacement.register(self.validate_float_track), "%P", "%d", "%s")
        self.max_displacement.insert(tk.END, 5)
        change_col_hover_enterbox(self.max_displacement, button_hv, "#333333")
        self.max_displacement.place(x=140, y=623, width=40)
        tk.Label(master=self.window, text="Min. Path Duration", bg="#555555", fg="white").place(x=25, y=643)
        tk.Label(master=self.window, text="frames", bg="#555555", fg="white").place(x=184, y=643)
        self.min_duration = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.min_duration["validatecommand"] = (self.min_duration.register(self.validate_int_track), "%P", "%d", "%s")
        self.min_duration.insert(tk.END, 20)
        change_col_hover_enterbox(self.min_duration, button_hv, "#333333")
        self.min_duration.place(x=140, y=645, width=40)
        tk.Label(master=self.window, text="Max. Dark Duration", bg="#555555", fg="white").place(x=25, y=665)
        tk.Label(master=self.window, text="frames", bg="#555555", fg="white").place(x=184, y=665)
        self.max_dark_time = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.max_dark_time["validatecommand"] = (self.max_dark_time.register(self.validate_int_track), "%P", "%d", "%s")
        self.max_dark_time.insert(tk.END, 1)
        change_col_hover_enterbox(self.max_dark_time, button_hv, "#333333")
        self.max_dark_time.place(x=140, y=667, width=40)


        tk.Label(master=self.window, text="Diffusion Analysis", anchor=tk.NW, padx=6, pady=2,
                 bg="#555555", fg="#dddddd", bd=2, relief=tk.RAISED).place(x=20, y=706, width=572, height=118)

        tk.Label(master=self.window, text="Min. Proportion", bg="#555555", fg="white").place(x=25, y=731)
        tk.Label(master=self.window, text="%", bg="#555555", fg="white").place(x=184, y=731)
        tk.Label(master=self.window, text="Min. Data Points", bg="#555555", fg="white").place(x=25, y=753)
        tk.Label(master=self.window, text="dy/dx Threshold", bg="#555555", fg="white").place(x=25, y=775)
        tk.Label(master=self.window, text="Min. Coeff. Determ.", bg="#555555", fg="white").place(x=25, y=797)

        tk.Label(master=self.window, text="Zero Diff.", bg="#555555", fg="white").place(x=230, y=731)
        tk.Label(master=self.window, text="um^2 /s", bg="#555555", fg="white").place(x=369, y=731)
        tk.Label(master=self.window, text="Uncertainty", bg="#555555", fg="white").place(x=230, y=753)
        tk.Label(master=self.window, text="St.D.", bg="#555555", fg="white").place(x=369, y=753)

        self.info_icon = tk.PhotoImage(master=self.window, file=cwd + "/icons/info.png")
        info_1 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                           command=lambda: self.info_message(title="Minimum Fitting Percentage",
                                                             message="The fitting tool will scan over a range of fitting proportions "
                                                                     "until an optimum is found. The minimum fitting percentage is the lowest "
                                                                     "proportion of the MSD plot on which a fit will be attempted."))
        info_1.place(x=202, y=732, width=19, height=19)
        change_col_hover(info_1, "#666666", "#555555")
        info_2 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Minimum Data Points",
                                                    message="After the fitting tool has scanned over a range of fitting proportions "
                                                            "it will test the R^2 values for each attempted linear regression and "
                                                            "select the optimal fit. Minimum data points is the fewest number of "
                                                            "time points in the MSD plot allowable for the final fit."))
        info_2.place(x=202, y=754, width=19, height=19)
        change_col_hover(info_2, "#666666", "#555555")
        info_3 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Gradient Threshold",
                                                    message="The gradient threshold (dy/dx threshold) is the gradient set point of R^2 "
                                                            "as a function of the fitting proportion at which the call will be made to fit "
                                                            "the MSD at that proportion. If set to zero, the fitting proportion will be taken "
                                                            "as the last fitting proportion before the gradient of R^2 becomes negative "
                                                            "which is the point where R^2 decreases as a greater proportion of the MSD is fitted."))
        info_3.place(x=202, y=776, width=19, height=19)
        change_col_hover(info_3, "#666666", "#555555")
        info_4 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Minimum Coefficient of Determination",
                                                    message="The fitting tool will attempt to determine if a spot is stationary. A stationary spot "
                                                            "will almost universally result in the R^2 function returning an extremely negative value. "
                                                            "If any R^2 value in R^2 as a function of the fitting proportion is less than this threshold, "
                                                            "and the mean diffusion coefficient of all the fits up to and including that negative value is "
                                                            "less than the the 'Zero Diffusion Threshold' then the spot will be classified as potentially "
                                                            "stationary. (i.e. D = 0)"))
        info_4.place(x=202, y=798, width=19, height=19)
        change_col_hover(info_4, "#666666", "#555555")
        info_5 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Zero Diffusion Threshold",
                                                    message="The diffusion coefficient threshold at which a track will be considered stationary if its "
                                                            "minimum R^2 value is less than the minimum threshold. (See info on 'Minimum Coefficient "
                                                            "of Determination)."))
        info_5.place(x=420, y=732, width=19, height=19)
        change_col_hover(info_5, "#666666", "#555555")
        info_6 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Degrees of Uncertainty",
                                                    message="The number of standard deviations used to caultulate the ± percentage error "
                                                            "of the diffusion coefficient."))
        info_6.place(x=420, y=754, width=19, height=19)
        change_col_hover(info_6, "#666666", "#555555")


        self.minimum_msd_prop = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.minimum_msd_prop["validatecommand"] = (
            self.minimum_msd_prop.register(self.validate_int_track), "%P", "%d", "%s")
        self.minimum_msd_prop.insert(tk.END, 5)
        change_col_hover_enterbox(self.minimum_msd_prop, button_hv, "#333333")
        self.minimum_msd_prop.place(x=140, y=733, width=40)

        self.minimum_msd_timepoints = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.minimum_msd_timepoints["validatecommand"] = (
            self.minimum_msd_timepoints.register(self.validate_int_track), "%P", "%d", "%s")
        self.minimum_msd_timepoints.insert(tk.END, 5)
        change_col_hover_enterbox(self.minimum_msd_timepoints, button_hv, "#333333")
        self.minimum_msd_timepoints.place(x=140, y=755, width=40)

        self.msd_gradient_threshold_r2 = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.msd_gradient_threshold_r2["validatecommand"] = (
            self.msd_gradient_threshold_r2.register(self.validate_float_track), "%P", "%d", "%s")
        self.msd_gradient_threshold_r2.insert(tk.END, 0.0)
        change_col_hover_enterbox(self.msd_gradient_threshold_r2, button_hv, "#333333")
        self.msd_gradient_threshold_r2.place(x=140, y=777, width=40)

        self.msd_stationary_r2_threshold = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.msd_stationary_r2_threshold.insert(tk.END, -10)
        change_col_hover_enterbox(self.msd_stationary_r2_threshold, button_hv, "#333333")
        self.msd_stationary_r2_threshold.place(x=140, y=799, width=40)

        self.msd_min_diff = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.msd_min_diff["validatecommand"] = (
            self.msd_min_diff.register(self.validate_float_track), "%P", "%d", "%s")
        self.msd_min_diff.insert(tk.END, 0.0001)
        change_col_hover_enterbox(self.msd_min_diff, button_hv, "#333333")
        self.msd_min_diff.place(x=308, y=733, width=56)

        self.msd_uncertainty_degree = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.msd_uncertainty_degree["validatecommand"] = (
            self.msd_uncertainty_degree.register(self.validate_int_track), "%P", "%d", "%s")
        self.msd_uncertainty_degree.insert(tk.END, 1)
        change_col_hover_enterbox(self.msd_uncertainty_degree, button_hv, "#333333")
        self.msd_uncertainty_degree.place(x=308, y=755, width=56)


        self.calulate_optimal_msds = tk.Button(master=self.window, text="Apply Constraints", command=self.msd_autofit,
                                               bg="#338833", fg="white", padx=2, pady=2)
        change_col_hover(self.calulate_optimal_msds, "#44aa44", "#338833")
        self.calulate_optimal_msds.place(x=232, y=786)
        self.plot_msd_fitting_results = tk.Button(master=self.window, text="Plot Current", command=self.open_r2_win,
                                                  bg="#222266", fg="white", padx=2, pady=2)
        change_col_hover(self.plot_msd_fitting_results, button_hv, "#222266")
        self.plot_msd_fitting_results.place(x=346, y=786)

        tk.Frame(master=self.window, bg="#888888").place(x=442, y=710, width=1, height=70)
        tk.Label(master=self.window, text="Don't export track if:", bg="#555555", fg="white",
                 justify="left", anchor=tk.NW).place(x=470, y=708)

        info_7 = tk.Button(master=self.window, image=self.info_icon, bg="#555555", relief=tk.FLAT, activebackground='#777777',
                  command=lambda: self.info_message(title="Remove Tracks",
                                                    message="Remove a track from the export queue if it is either stationary or "
                                                            "the percentage error in the diffusion coefficient is greater than the threshold. "
                                                            "Tracks that are too short and cause errors in the auto-fit routine will be "
                                                            "automatically rejected."))
        info_7.place(x=450, y=710, width=19, height=19)
        change_col_hover(info_7, "#666666", "#555555")

        self.exclude_stationary = tk.IntVar(master=self.window)
        self.exclude_stationary.set(1)
        self.exclude_stationary_button = tk.Checkbutton(master=self.window, activebackground="#555555", bg="#555555",
                                                        onvalue=1, offvalue=0, variable=self.exclude_stationary)
        self.exclude_stationary_button.place(x=450, y=731)

        self.exclude_based_on_error = tk.IntVar(master=self.window)
        self.exclude_based_on_error.set(1)
        self.exclude_based_on_error_button = tk.Checkbutton(master=self.window, activebackground="#555555", bg="#555555",
                                                        onvalue=1, offvalue=0, variable=self.exclude_based_on_error)
        self.exclude_based_on_error_button.place(x=450, y=753)

        tk.Label(master=self.window, text="Spot is stationary", bg="#555555", fg="#cccccc").place(x=476, y=731)
        tk.Label(master=self.window, text="Error >                   %", bg="#555555", fg="#cccccc").place(x=476, y=753)

        self.msd_maximum_error = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.msd_maximum_error["validatecommand"] = (
            self.msd_maximum_error.register(self.validate_int_track), "%P", "%d", "%s")
        self.msd_maximum_error.insert(tk.END, 10)
        change_col_hover_enterbox(self.msd_maximum_error, button_hv, "#333333")
        self.msd_maximum_error.place(x=520, y=755, width=48)

        self.plot_all_msds = tk.Button(master=self.window, text="View All MSDs", command=self.open_all_msds,
                                                  bg="#222266", fg="white", padx=3, pady=2)
        change_col_hover(self.plot_all_msds, button_hv, "#222266")
        self.plot_all_msds.place(x=450, y=786)

        self.hist_icon = tk.PhotoImage(master=self.window, file=cwd + "/icons/hist.png")
        self.dataset_histogram_button = tk.Button(master=self.window, image=self.hist_icon, command=self.open_hist_win,
                                                  bg="#446688")
        change_col_hover(self.dataset_histogram_button, "#6688aa", "#446688")
        self.dataset_histogram_button.place(x=548, y=786, height=28)


        tk.Label(master=self.window, text="Save / Load Profiles", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=12, y=836, width=438, height=52)

        self.save_experiment = tk.Button(master=self.window, text="Save Experiment", padx=10, pady=1, bd=2,
                                         bg="#446688", fg="#dddddd", command=self.save_document)
        change_col_hover(self.save_experiment, "#6688aa", "#446688")
        self.save_experiment.place(x=20, y=856)
        self.load_experiment = tk.Button(master=self.window, text="Open Experiment", padx=10, pady=1, bd=2,
                                         bg="#446688", fg="#dddddd", command=self.open_document)
        change_col_hover(self.load_experiment, "#6688aa", "#446688")
        self.load_experiment.place(x=140, y=856)
        self.new_experiment_button = tk.Button(master=self.window, text="New Experiment", padx=10, pady=1, bd=2,
                                               bg="#446688", fg="#dddddd", command=self.new_exp)
        change_col_hover(self.new_experiment_button, "#6688aa", "#446688")
        self.new_experiment_button.place(x=265, y=856)


        tk.Label(master=self.window, text="Plots", anchor=tk.NW, padx=6, pady=2,
                 bg="#444444", fg="#dddddd", bd=2, relief=tk.RIDGE).place(x=454, y=836, width=146, height=52)
        self.dark_mode = tk.IntVar(master=self.window)
        self.dark_mode.set(1)
        self.dark_mode_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                               onvalue=1, offvalue=0, variable=self.dark_mode,
                                               command=self.display_trajectory)
        self.dark_mode_button.place(x=500, y=861, height=20)
        tk.Label(master=self.window, text="Dark Theme", bg="#444444", fg="#cccccc").place(x=524, y=862, height=20)

        self.plot_grids = tk.IntVar(master=self.window)
        self.plot_grids.set(1)
        self.plot_grids_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                               onvalue=1, offvalue=0, variable=self.plot_grids,
                                               command=self.display_trajectory)
        self.plot_grids_button.place(x=500, y=840, height=20)
        tk.Label(master=self.window, text="Grid Guides", bg="#444444", fg="#cccccc").place(x=524, y=841, height=20)

        self.font_icon = tk.PhotoImage(master=self.window, file=cwd + "/icons/font.png")
        self.font_settings_button = tk.Button(master=self.window, image=self.font_icon, command=self.open_font_win,
                                              bg="#777777")
        change_col_hover(self.font_settings_button, "#999999", "#777777")
        self.font_settings_button.place(x=462, y=856, height=27)


        self.plot_trajectory_frame = tk.Frame(master=self.window, width=424, height=394)
        self.plot_trajectory_frame.place(x=1414, y=30)
        self.plot_trajectory_figure = plt.Figure(figsize=(4.24, 3.94), dpi=100)
        self.plot_trajectory_canvas = FigureCanvasTkAgg(self.plot_trajectory_figure, master=self.plot_trajectory_frame)
        self.plot_trajectory_canvas.draw()
        self.plot_trajectory_canvas.get_tk_widget().pack(side=tk.TOP)
        self.plot_trajectory_figure.set_facecolor("#222222")
        self.traj_toolbar = NavigationToolbar2Tk(self.plot_trajectory_canvas, self.plot_trajectory_frame)
        self.traj_toolbar.config(background="#25252f")
        for w in self.traj_toolbar.winfo_children():
            w.config(background="#25252f")
        self.traj_toolbar._message_label.config(background="#25252f")
        self.traj_toolbar._message_label.config(foreground="#cccccc")
        self.traj_toolbar.update()
        self.plot_trajectory_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.X)
        self.plot_trajectory_figure.subplots_adjust(bottom=0.08, left=0.18, right=0.94, top=0.95)

        self.show_dataset = tk.Label(master=self.window, text="Field:                 Tracking", bg="#333333",
                                     fg="#dddddd")
        self.show_dataset.place(x=1420, y=470)
        tk.Label(master=self.window, text="Select Path", bg="#333333", fg="#dddddd").place(x=1420, y=500)
        self.current_trajectory = tk.StringVar(master=self.window)
        self.trajectory_list = []
        self.select_trajectory = tk.Spinbox(master=self.window, repeatinterval=400,
                                            textvariable=self.current_trajectory, bg="#bbbbbb", fg="black",
                                            bd=2, font="Calibri 12 bold", state="readonly", value=self.trajectory_list,
                                            command=self.display_trajectory)
        self.select_trajectory.place(x=1500, y=496, height=30, width=130)
        tk.Label(master=self.window, text="Rapid Scroll", bg="#333333", fg="#dddddd").place(x=1420, y=530)
        self.quick_current = tk.StringVar(master=self.window)
        self.quick_trajectory = tk.Spinbox(master=self.window, repeatinterval=5,
                                           textvariable=self.quick_current, bg="#bbbbbb", fg="black",
                                           bd=2, font="Calibri 12 bold", state="readonly", value=self.trajectory_list,
                                           command=self.quick_scroll)
        self.quick_trajectory.place(x=1500, y=526, height=30, width=130)

        self.compute_and_show_button = tk.Button(master=self.window, text="Analyse and Display", padx=15, pady=1, bd=2,
                                                 bg="#338833", fg="#dddddd", command=open_time_series)
        change_col_hover(self.compute_and_show_button, "#44aa44", "#338833")
        self.compute_and_show_button.place(x=1690, y=498, width=146)

        self.export_button = tk.Button(master=self.window, text="Export to Excel", padx=28, pady=1, bd=2,
                                       bg="#446688", fg="#dddddd", command=self.export_to_xl)
        change_col_hover(self.export_button, "#6688aa", "#446688")
        self.export_button.place(x=1690, y=528, width=146)

        tk.Label(master=self.window, text="Trajectory Information", anchor=tk.NW, padx=6,
                 pady=2, bg="#555555", fg="#dddddd", bd=2, relief=tk.RIDGE,
                 wraplength=290, justify="left").place(x=1416, y=564, width=422, height=324)
        self.trajectory_info = tk.Label(master=self.window, text="", anchor=tk.NW, padx=6,
                                        pady=2, bg="#444444", fg="#dddddd", bd=2, relief=tk.SUNKEN,
                                        wraplength=390, justify="left", font="TkDefaultFont 8")
        self.trajectory_info.place(x=1422, y=600, width=410, height=280)
        self.msd_proportion = tk.StringVar(master=self.window)
        self.msd_proportion.set(20)
        self.msd_proportion_box = tk.Spinbox(master=self.window, from_=5, to=100, increment=5, repeatinterval=100,
                                             textvariable=self.msd_proportion, bg="#bbbbbb", fg="black",
                                             bd=2, font="Arial 11 bold", command=self.set_msd_prop)
        self.msd_proportion_box.bind("<Return>", self.set_msd_prop)
        self.msd_proportion_box.bind("<0>", self.set_msd_prop)
        self.msd_proportion_box.bind("<Key>", self.set_msd_prop)
        self.msd_proportion_box.place(x=1750, y=568, width=80, height=29)
        tk.Label(master=self.window, text="MSD Percentage", bg="#555555", fg="white").place(x=1650, y=572)

        tk.Label(master=self.window, text="Export", bg="#333333", fg="white").place(x=1730, y=472)
        self.include_track_var = tk.IntVar(master=self.window)
        self.include_track_var.set(1)
        self.include_track_button = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                   onvalue=1, offvalue=0, variable=self.include_track_var,
                                                   command=self.toggle_export_include)
        self.include_track_button["state"] = tk.DISABLED
        self.include_track_button.place(x=1700, y=470)

        self.refresh_track_info_button = tk.Button(master=self.window, text="Refresh", bg="#222222", fg="#cccccc",
                                                   padx=4, pady=0, command=self.display_trajectory)
        change_col_hover(self.refresh_track_info_button, button_hv, "#222222")
        self.refresh_track_info_button.place(x=1554, y=567, height=19)

        self.region_trk_button = tk.Button(master=self.window, text="Select Region", bg="#222266", padx=10, pady=2,
                                           fg="#cccccc", command=self.select_region_trk)
        change_col_hover(self.region_trk_button, button_hv, "#222266")
        self.region_trk_button.place(x=610, y=790, width=120)
        self.region_c1_button = tk.Button(master=self.window, text="Select Region", bg="#222266", padx=10, pady=2,
                                           fg="#cccccc", command=self.select_region_c1)
        change_col_hover(self.region_c1_button, button_hv, "#222266")
        self.region_c1_button.place(x=878, y=790, width=120)
        self.region_c2_button = tk.Button(master=self.window, text="Select Region", bg="#222266", padx=10, pady=2,
                                           fg="#cccccc", command=self.select_region_c2)
        change_col_hover(self.region_c2_button, button_hv, "#222266")
        self.region_c2_button.place(x=1146, y=790, width=120)

        self.del_region_trk_button = tk.Button(master=self.window, text="Delete Region", bg="#222266", padx=10, pady=2,
                                           fg="#cccccc", command=lambda: self.delete_region(0))
        change_col_hover(self.del_region_trk_button, button_hv, "#222266")
        self.del_region_trk_button.place(x=610, y=820, width=120)
        self.del_region_c1_button = tk.Button(master=self.window, text="Delete Region", bg="#222266", padx=10, pady=2,
                                          fg="#cccccc", command=lambda: self.delete_region(1))
        change_col_hover(self.del_region_c1_button, button_hv, "#222266")
        self.del_region_c1_button.place(x=878, y=820, width=120)
        self.del_region_c2_button = tk.Button(master=self.window, text="Delete Region", bg="#222266", padx=10, pady=2,
                                          fg="#cccccc", command=lambda: self.delete_region(2))
        change_col_hover(self.del_region_c2_button, button_hv, "#222266")
        self.del_region_c2_button.place(x=1146, y=820, width=120)

        self.restore_minimized_task_button = tk.Button(master=self.window, text="Restore Progress Bar", bg="#222222",
                                                       padx=10, pady=2, fg="#cccccc", command=self.restore_progress_win)
        change_col_hover(self.restore_minimized_task_button, button_hv, "#222222")

        tk.Label(master=self.window, text="Invert Selection", bg="#333333", fg="white").place(x=770, y=792)
        self.invert_selection_flag_trk = tk.IntVar(master=self.window)
        self.invert_selection_flag_trk.set(0)
        self.invert_selection_trk = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                   onvalue=1, offvalue=0, variable=self.invert_selection_flag_trk,
                                                   command=self.display_tracking)
        self.invert_selection_trk.place(x=740, y=790)
        tk.Label(master=self.window, text="Invert Selection", bg="#333333", fg="white").place(x=1038, y=792)
        self.invert_selection_flag_c1 = tk.IntVar(master=self.window)
        self.invert_selection_flag_c1.set(0)
        self.invert_selection_c1 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                  onvalue=1, offvalue=0, variable=self.invert_selection_flag_c1,
                                                  command=self.display_coloc1)
        self.invert_selection_c1.place(x=1008, y=790)
        tk.Label(master=self.window, text="Invert Selection", bg="#333333", fg="white").place(x=1306, y=792)
        self.invert_selection_flag_c2 = tk.IntVar(master=self.window)
        self.invert_selection_flag_c2.set(0)
        self.invert_selection_c2 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                  onvalue=1, offvalue=0, variable=self.invert_selection_flag_c2,
                                                  command=self.display_coloc2)
        self.invert_selection_c2.place(x=1276, y=790)

        tk.Label(master=self.window, text="Use Global Norm.", bg="#333333", fg="#cccccc", pady=2).place(x=770, y=550, height=20)
        self.global_norm_trk = tk.IntVar(master=self.window)
        self.global_norm_trk.set(1)
        self.set_global_norm_trk = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                  onvalue=1, offvalue=0, variable=self.global_norm_trk,
                                                  command=self.display_tracking)
        self.set_global_norm_trk.place(x=740, y=548)
        tk.Label(master=self.window, text="Use Global Norm.", bg="#333333", fg="#cccccc", pady=2).place(x=1038, y=550, height=20)
        self.global_norm_c1 = tk.IntVar(master=self.window)
        self.global_norm_c1.set(1)
        self.set_global_norm_c1 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                 onvalue=1, offvalue=0, variable=self.global_norm_c1,
                                                 command=self.display_coloc1)
        self.set_global_norm_c1.place(x=1008, y=548)
        tk.Label(master=self.window, text="Use Global Norm.", bg="#333333", fg="#cccccc", pady=2).place(x=1306, y=550, height=20)
        self.global_norm_c2 = tk.IntVar(master=self.window)
        self.global_norm_c2.set(1)
        self.set_global_norm_c2 = tk.Checkbutton(master=self.window, activebackground="#333333", bg="#333333",
                                                 onvalue=1, offvalue=0, variable=self.global_norm_c2,
                                                 command=self.display_coloc2)
        self.set_global_norm_c2.place(x=1276, y=548)

        self.region_initial = [None, None, None]
        self.region_final = [None, None, None]
        self.rect_trk, self.rect_c1, self.rect_c2 = None, None, None

    @staticmethod
    def restore_progress_win():
        try:
            progress_win.restore()
        except:
            """ Failed """

    def info_message(self, title, message):
        easygui.msgbox(title=title, msg=message)

    def delete_region(self, canv):
        self.region_initial[canv] = None
        if canv == 0:
            self.canvas_tracking.delete(self.rect_trk)
        elif canv == 1:
            self.canvas_coloc1.delete(self.rect_c1)
        elif canv == 2:
            self.canvas_coloc2.delete(self.rect_c2)

    def select_region_trk(self):
        if self.region_trk_button["relief"] == tk.RAISED:
            self.region_trk_button["relief"] = tk.SUNKEN
            self.canvas_tracking.unbind("<ButtonRelease-1>")
            self.canvas_tracking.bind("<Button-1>", lambda event, canv=0: self.region_start(event, canv))
            self.canvas_tracking.bind("<ButtonRelease-1>", lambda event, canv=0: self.region_end(event, canv))
            self.canvas_tracking.bind("<B1-Motion>", lambda event, canv=0: self.mouse_motion(event, canv))
        else:
            self.region_trk_button["relief"] = tk.RAISED
            self.canvas_tracking.unbind("<Button-1>")
            self.canvas_tracking.unbind("<ButtonRelease-1>")
            self.canvas_tracking.unbind("<B1-Motion>")
            self.canvas_tracking.bind("<ButtonRelease-1>", self.canvas_trk_clicked)

    def select_region_c1(self):
        if self.region_c1_button["relief"] == tk.RAISED:
            self.region_c1_button["relief"] = tk.SUNKEN
            self.canvas_coloc1.unbind("<ButtonRelease-1>")
            self.canvas_coloc1.bind("<Button-1>", lambda event, canv=1: self.region_start(event, canv))
            self.canvas_coloc1.bind("<ButtonRelease-1>", lambda event, canv=1: self.region_end(event, canv))
            self.canvas_coloc1.bind("<B1-Motion>", lambda event, canv=1: self.mouse_motion(event, canv))
        else:
            self.region_c1_button["relief"] = tk.RAISED
            self.canvas_coloc1.unbind("<Button-1>")
            self.canvas_coloc1.unbind("<ButtonRelease-1>")
            self.canvas_coloc1.unbind("<B1-Motion>")
            self.canvas_coloc1.bind("<ButtonRelease-1>", self.canvas_c1_clicked)

    def select_region_c2(self):
        if self.region_c2_button["relief"] == tk.RAISED:
            self.region_c2_button["relief"] = tk.SUNKEN
            self.canvas_coloc2.unbind("<ButtonRelease-1>")
            self.canvas_coloc2.bind("<Button-1>", lambda event, canv=2: self.region_start(event, canv))
            self.canvas_coloc2.bind("<ButtonRelease-1>", lambda event, canv=2: self.region_end(event, canv))
            self.canvas_coloc2.bind("<B1-Motion>", lambda event, canv=2: self.mouse_motion(event, canv))
        else:
            self.region_c2_button["relief"] = tk.RAISED
            self.canvas_coloc2.unbind("<Button-1>")
            self.canvas_coloc2.unbind("<ButtonRelease-1>")
            self.canvas_coloc2.unbind("<B1-Motion>")
            self.canvas_coloc2.bind("<ButtonRelease-1>", self.canvas_c2_clicked)

    def region_start(self, event, canvas_index):
        self.region_initial[canvas_index] = [event.x, event.y]
        print(self.region_initial)

    def mouse_motion(self, event, canvas_index):
        if canvas_index == 0:
            try:
                self.canvas_tracking.delete(self.rect_trk)
            except:
                """ No rectangle to delete """
            if self.invert_selection_flag_trk.get() == 1:
                col = "red"
            else:
                col = "green"
            self.rect_trk = self.canvas_tracking.create_rectangle(self.region_initial[0][0], self.region_initial[0][1],
                                                  event.x, event.y, outline=col, width=1, dash=(2, 1))

        if canvas_index == 1:
            try:
                self.canvas_coloc1.delete(self.rect_c1)
            except:
                """ No rectangle to delete """
            if self.invert_selection_flag_c1.get() == 1:
                col = "red"
            else:
                col = "green"
            self.rect_c1 = self.canvas_coloc1.create_rectangle(self.region_initial[1][0], self.region_initial[1][1],
                                                  event.x, event.y, outline=col, width=1, dash=(2, 1))

        if canvas_index == 2:
            try:
                self.canvas_coloc2.delete(self.rect_c2)
            except:
                """ No rectangle to delete """
            if self.invert_selection_flag_c2.get() == 1:
                col = "red"
            else:
                col = "green"
            self.rect_c2 = self.canvas_coloc2.create_rectangle(self.region_initial[2][0], self.region_initial[2][1],
                                                  event.x, event.y, outline=col, width=1, dash=(2, 1))

    def region_end(self, event, canvas_index):
        self.region_final[canvas_index] = [event.x, event.y]
        if self.region_final[canvas_index] == self.region_initial[canvas_index]:
            self.region_initial[canvas_index], self.region_final[canvas_index] = None, None
        print(self.region_final)

    def toggle_export_include(self):
        try:
            name = self.current_trajectory.get()
            index = self.trajectory_list.index(name)
            if self.field_selected.get() == self.field_list[0]:
                data = self.tracking_data_trk
            elif self.field_selected.get() == self.field_list[1]:
                data = self.tracking_data_c1
            elif self.field_selected.get() == self.field_list[2]:
                data = self.tracking_data_c2
            traj = data[index]

            if self.include_track_var.get() == 0:
                traj.include_export = False
            else:
                traj.include_export = True

        except:
            print("Failed to select\n")
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())

    def set_msd_prop(self, event=None):
        if event is not None:
            print(event.char)
        try:
            name = self.current_trajectory.get()
            index = self.trajectory_list.index(name)
            if self.field_selected.get() == self.field_list[0]:
                data = self.tracking_data_trk
            elif self.field_selected.get() == self.field_list[1]:
                data = self.tracking_data_c1
            elif self.field_selected.get() == self.field_list[2]:
                data = self.tracking_data_c2
            traj = data[index]

            traj.MSD_prop = int(self.msd_proportion.get())
            if event is not None:
                traj.MSD_prop = int(self.msd_proportion.get() + event.char)
        except ValueError:
            traj.MSD_prop = int(event.char)
        except:
            traj.MSD_prop = 20
            print("Failed to select\n")
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())

    def handle_close(self, forced=False):

        if not forced:
            if not self.saved_flag:
                decision = easygui.ccbox(title="Warning!", msg="There is an unsaved experiment open."
                                                               "\n\nDo you wish to close anyway?",
                                         choices=["Close", "Cancel"],
                                         default_choice="Cancel")
                if not decision:
                    return

        set_GUI_state(tk.NORMAL)
        open_tracking_button["state"] = tk.NORMAL
        load_button["state"] = tk.NORMAL
        load_pickle_button["state"] = tk.NORMAL
        preferences_button["state"] = tk.NORMAL
        import_raw_button["state"] = tk.NORMAL

        try:
            raw_stack_window.window.destroy()
        except:
            """ Failed to destroy window """
        try:
            proc_stack_window.window.destroy()
        except:
            """ Failed to destroy window """
        try:
            progress_win.window.destroy()
        except:
            """ Failed to destroy window """
        try:
            graph_window.window.destroy()
        except:
            """ Failed to destroy window """
        try:
            enlarge_win.handle_close()
        except:
            """ Failed to destroy window """
        try:
            msd_win.handle_close()
        except:
            """ Failed to destroy window """
        try:
            r2_win.handle_close()
        except:
            """ Failed to destroy window """
        try:
            histwin.handle_close()
        except:
            """ Failed to close window """
        try:
            auto_track_win.handle_close()
        except:
            """ Failed to close window """
        try:
            auto_select_win.handle_close()
        except:
            """ Failed to close window """
        self.cancel_flag = True
        try:
            histwin.cancel_flag = True
        except:
            """ Object does not exist """
        try:
            font_win.handle_close()
        except:
            """ Failed to close window """
        try:
            self.window.destroy()
        except:
            """ Failed to destroy window """

    def export_tiff_file(self):
        global progress_win
        data = None
        if self.field_selected.get() == self.field_list[0]:
            data = self.proc_data_tracking
        elif self.field_selected.get() == self.field_list[1]:
            data = self.proc_data_coloc1
        elif self.field_selected.get() == self.field_list[2]:
            data = self.proc_data_coloc2
        if data is None:
            return
        shape = np.shape(data)
        path = easygui.filesavebox(title="Export Processed TIF", default=default_dir+"*.tif", filetypes=["*.tif", "TIF files"])
        if not path:
            return
        ext = path[-4:]
        if ext == ".tif":
            path = path[:-4]
        elif ext == "tiff":
            path = path[:-5]
        path += ".tif"

        progress_win = ProgressWin("Preparing movie...", "Converting format...")
        progress_win.cancel_button.place_forget()
        progress_win.progress["maximum"] = shape[2] + 1

        movie = np.zeros((shape[2], shape[0], shape[1]))
        for frame in range(shape[2]):
            movie[frame, :, :] = data[:, :, frame]
            progress_win.progress.step(1)
            progress_win.progress.update()

        progress_win.label["text"] = "Saving movie..."
        progress_win.label.update()
        progress_win.window.title("Saving...")
        try:
            movie = movie.astype('uint16')
            with tifffile.TiffWriter(path) as tif:
                tif.write(movie)
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="An error occurred while attempting to export tif.\n" + str(traceback.format_exc()))

        try:
            progress_win.handle_close()
        except:
            """ Window already closed """


    def load_file(self, file=None, auto=False):
        if not auto:
            path = easygui.fileopenbox(title="Load TIF file", default=default_dir+"*.TIF", filetypes=["*.TIF", "*.tiff", "TIF files"])
        else:
            path = file
        if path:
            try:
                if self.field_selected.get() == self.field_list[0]:
                    self.raw_data_tracking, self.data_length_tracking = tirf.load_tiff(path)
                    self.filenames[0] = path
                    self.tracking_data_trk = []
                elif self.field_selected.get() == self.field_list[1]:
                    self.raw_data_coloc1, self.data_length_coloc1 = tirf.load_tiff(path)
                    self.filenames[1] = path
                    self.tracking_data_c1 = []
                elif self.field_selected.get() == self.field_list[2]:
                    self.raw_data_coloc2, self.data_length_coloc2 = tirf.load_tiff(path)
                    self.filenames[2] = path
                    self.tracking_data_c2 = []
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                easygui.msgbox(title="Error!", msg="An error occured while loading this file: '" + str(path) + "'.\n\n" + str(traceback.format_exc()))
            self.process_raw_data(self.field_selected.get())
            self.saved_flag = False

    def update_fileinfo(self, field):
        self.file_info["text"] = f"File: '{self.filenames[self.field_dict[field]]}'\n\n"
        if field == self.field_list[0]:
            self.file_info["text"] += f"Frame count: {self.data_length_tracking}"
            if self.data_length_tracking is not None:
                self.current_field_length = self.data_length_tracking
            else:
                self.current_field_length = 1
        elif field == self.field_list[1]:
            self.file_info["text"] += f"Frame count: {self.data_length_coloc1}"
            if self.data_length_coloc1 is not None:
                self.current_field_length = self.data_length_coloc1
            else:
                self.current_field_length = 1
        elif field == self.field_list[2]:
            self.file_info["text"] += f"Frame count: {self.data_length_coloc2}"
            if self.data_length_coloc2 is not None:
                self.current_field_length = self.data_length_coloc2
            else:
                self.current_field_length = 1
        if self.filenames[self.field_dict[field]] is None:
            self.file_info["text"] = "File: None\n\nFrame count: None"
        self.file_info.update()
        self.init_frame.delete(0, tk.END)
        self.init_frame.insert(tk.END, 0)
        if self.lock_stop.get() == 0:
            self.stop_frame.delete(0, tk.END)
            self.stop_frame.insert(tk.END, self.current_field_length)
        self.load_init_frame()
        self.update_trajectories()
        self.show_dataset["text"] = "Field:                 " + self.field_selected.get()
        self.show_dataset.update()
        self.toggle_button_state(self.analysis_mode.get())

    def update_trajectories(self):
        self.trajectory_list = []
        self.current_trajectory.set("")
        self.quick_current.set("")
        if self.field_selected.get() == self.field_list[0] and len(self.tracking_data_trk) > 0:
            for ind, path in enumerate(self.tracking_data_trk):
                self.trajectory_list.append("Trajectory " + str(ind + 1))
        if self.field_selected.get() == self.field_list[1] and len(self.tracking_data_c1) > 0:
            for ind, path in enumerate(self.tracking_data_c1):
                self.trajectory_list.append("Trajectory " + str(ind + 1))
        if self.field_selected.get() == self.field_list[2] and len(self.tracking_data_c2) > 0:
            for ind, path in enumerate(self.tracking_data_c2):
                self.trajectory_list.append("Trajectory " + str(ind + 1))
        self.select_trajectory["value"] = self.trajectory_list
        self.quick_trajectory["value"] = self.trajectory_list
        self.select_trajectory.update()
        self.quick_trajectory.update()
        self.display_trajectory()

    def process_raw_data(self, field):
        global progress_win
        if field == self.field_list[0]:
            data = self.raw_data_tracking
        elif field == self.field_list[1]:
            data = self.raw_data_coloc1
        elif field == self.field_list[2]:
            data = self.raw_data_coloc2

        if data is None:
            return

        shape = np.shape(data)
        length = shape[2]

        progress_win = ProgressWin("Processing TIF Stack", "Enhancing frames...")
        progress_win.cancel_button.place_forget()
        progress_win.progress["maximum"] = shape[2] + 1

        result = np.zeros(shape)
        for frame in range(length):
            processed = tirf.low_pass(data[:, :, frame], float(self.filter_power.get()))
            result[:, :, frame] = processed
            progress_win.progress.step(1)
            progress_win.progress.update()

        progress_win.progress.stop()
        progress_win.progress["maximum"] = shape[2] + 1 - int(float(self.frames_to_average.get()))
        progress_win.label["text"] = "Averaging frames..."
        progress_win.label.update()

        if int(float(self.frames_to_average.get())) > 1 and int(float(self.frames_to_average.get())) < length:
            copy = np.copy(result)
            result = np.zeros((shape[0], shape[1], shape[2] - int(float(self.frames_to_average.get())) + 1))
            for frame in range(length - int(float(self.frames_to_average.get())) + 1):
                averaged_frame = np.zeros((shape[0], shape[1]))
                for add in range(int(float(self.frames_to_average.get()))):
                    averaged_frame = averaged_frame + copy[:, :, frame + add]
                result[:, :, frame] = averaged_frame
                progress_win.progress.step(1)
                progress_win.progress.update()

        progress_win.handle_close()
        result = np.clip(result, 0, 65535)

        if self.lock_stop.get() == 0:
            self.stop_frame.delete(0, tk.END)
            self.stop_frame.insert(tk.END, np.shape(result)[2])

        if field == self.field_list[0]:
            self.proc_data_tracking = result
            self.data_length_tracking = np.shape(result)[2]
        elif field == self.field_list[1]:
            self.proc_data_coloc1 = result
            self.data_length_coloc1 = np.shape(result)[2]
        elif field == self.field_list[2]:
            self.proc_data_coloc2 = result
            self.data_length_coloc2 = np.shape(result)[2]

        if self.proc_data_tracking is not None:
            self.slider_trk["state"] = tk.NORMAL
            self.slider_trk["to"] = self.data_length_tracking - 1
            if self.display_frames[0].get() > self.data_length_tracking - 1:
                self.slider_trk.set(self.data_length_tracking - 1)
            self.bright_trk["state"] = tk.NORMAL
            self.cont_trk["state"] = tk.NORMAL
            self.display_tracking()
        else:
            self.slider_trk["state"] = tk.DISABLED
            self.bright_trk["state"] = tk.DISABLED
            self.cont_trk["state"] = tk.DISABLED
            self.display_tracking(erase=True)
        if self.proc_data_coloc1 is not None:
            self.slider_c1["state"] = tk.NORMAL
            self.slider_c1["to"] = self.data_length_coloc1 - 1
            if self.display_frames[1].get() > self.data_length_coloc1 - 1:
                self.slider_c1.set(self.data_length_coloc1 - 1)
            self.bright_c1["state"] = tk.NORMAL
            self.cont_c1["state"] = tk.NORMAL
            self.display_coloc1()
        else:
            self.slider_c1["state"] = tk.DISABLED
            self.bright_c1["state"] = tk.DISABLED
            self.cont_c1["state"] = tk.DISABLED
            self.display_coloc1(erase=True)
        if self.proc_data_coloc2 is not None:
            self.slider_c2["state"] = tk.NORMAL
            self.slider_c2["to"] = self.data_length_coloc2 - 1
            if self.display_frames[2].get() > self.data_length_coloc2 - 1:
                self.slider_c2.set(self.data_length_coloc2 - 1)
            self.bright_c2["state"] = tk.NORMAL
            self.cont_c2["state"] = tk.NORMAL
            self.display_coloc2()
        else:
            self.slider_c2["state"] = tk.DISABLED
            self.bright_c2["state"] = tk.DISABLED
            self.cont_c2["state"] = tk.DISABLED
            self.display_coloc2(erase=True)
        self.update_fileinfo(self.field_selected.get())

    def new_exp(self):
        self.handle_close()
        open_tracking_window()

    def save_document(self, auto=False, autopath=None):
        try:
            global progress_win
            if not auto:
                save_file_path = easygui.filesavebox(title="Save Experiment", default=default_dir + "untitled experiment.txp",
                                                     filetypes=["*.txp", "Tracking experiment files"])
            else:
                save_file_path = autopath
            if save_file_path:
                ext = save_file_path[-4:]
                if ext == ".txp":
                    save_file_path = save_file_path[:-4]

                progress_win = ProgressWin(title="Save", msg="Saving experiment...")
                progress_win.cancel_button.place_forget()
                progress_win.progress["maximum"] = 5
                progress_win.progress["mode"] = "determinate"

                save_file_path += ".txp"

                with open(save_file_path, "wb") as save_doc:
                    progress_win.progress.step(1)
                    progress_win.progress.update()
                    pickle.dump(self.raw_data_tracking, save_doc)
                    pickle.dump(self.proc_data_tracking, save_doc)
                    pickle.dump(self.data_length_tracking, save_doc)

                    progress_win.progress.step(1)
                    progress_win.progress.update()
                    pickle.dump(self.raw_data_coloc1, save_doc)
                    pickle.dump(self.proc_data_coloc1, save_doc)
                    pickle.dump(self.data_length_coloc1, save_doc)

                    progress_win.progress.step(1)
                    progress_win.progress.update()
                    pickle.dump(self.raw_data_coloc2, save_doc)
                    pickle.dump(self.proc_data_coloc2, save_doc)
                    pickle.dump(self.data_length_coloc2, save_doc)

                    progress_win.progress.step(1)
                    progress_win.progress.update()
                    pickle.dump(self.filenames, save_doc)

                    pickle.dump(self.tracking_data_trk, save_doc)
                    pickle.dump(self.tracking_data_c1, save_doc)
                    pickle.dump(self.tracking_data_c2, save_doc)

                    pickle.dump(self.display_frames[0].get(), save_doc)
                    pickle.dump(self.display_frames[1].get(), save_doc)
                    pickle.dump(self.display_frames[2].get(), save_doc)
                    pickle.dump(self.display_brightness[0].get(), save_doc)
                    pickle.dump(self.display_brightness[1].get(), save_doc)
                    pickle.dump(self.display_brightness[2].get(), save_doc)
                    pickle.dump(self.display_contrast[0].get(), save_doc)
                    pickle.dump(self.display_contrast[1].get(), save_doc)
                    pickle.dump(self.display_contrast[2].get(), save_doc)

                    pickle.dump(self.current_trajectory.get(), save_doc)
                    pickle.dump(self.field_selected.get(), save_doc)

                    pickle.dump(self.detection.get(), save_doc)
                    pickle.dump(self.averaging.get(), save_doc)
                    pickle.dump(self.residual.get(), save_doc)
                    pickle.dump(self.minisig.get(), save_doc)
                    pickle.dump(self.maxisig.get(), save_doc)
                    pickle.dump(self.intensity.get(), save_doc)
                    pickle.dump(self.amplitude.get(), save_doc)
                    pickle.dump(self.eccentric.get(), save_doc)
                    pickle.dump(self.gaussresid.get(), save_doc)
                    pickle.dump(self.init_frame.get(), save_doc)
                    pickle.dump(self.pixel_size.get(), save_doc)
                    pickle.dump(self.frame_interval.get(), save_doc)
                    pickle.dump(self.start_frame.get(), save_doc)
                    pickle.dump(self.stop_frame.get(), save_doc)
                    pickle.dump(self.max_displacement.get(), save_doc)
                    pickle.dump(self.min_duration.get(), save_doc)
                    pickle.dump(self.max_dark_time.get(), save_doc)

                    pickle.dump(self.region_initial, save_doc)
                    pickle.dump(self.region_final, save_doc)

                    pickle.dump(self.invert_selection_flag_trk.get(), save_doc)
                    pickle.dump(self.invert_selection_flag_c1.get(), save_doc)
                    pickle.dump(self.invert_selection_flag_c2.get(), save_doc)

                    pickle.dump(self.minimum_msd_prop.get(), save_doc)
                    pickle.dump(self.minimum_msd_timepoints.get(), save_doc)
                    pickle.dump(self.msd_gradient_threshold_r2.get(), save_doc)
                    pickle.dump(self.msd_stationary_r2_threshold.get(), save_doc)
                    pickle.dump(self.msd_min_diff.get(), save_doc)
                    pickle.dump(self.msd_uncertainty_degree.get(), save_doc)

                    pickle.dump(self.exclude_stationary.get(), save_doc)
                    pickle.dump(self.exclude_based_on_error.get(),save_doc)
                    pickle.dump(self.msd_maximum_error.get(), save_doc)

                progress_win.progress.step(1)
                progress_win.progress.stop()
                progress_win.handle_close()
                self.saved_flag = True
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="Warning! an error occurred while attempting to save experiment:\n\n" + traceback.format_exc())
            try:
                progress_win.progress.step(1)
                progress_win.progress.stop()
                progress_win.handle_close()
            except:
                """ No progress bar window open """


    def open_document(self):
        try:
            if not self.saved_flag:
                decision = easygui.ccbox(title="Warning!", msg="There is an unsaved experiment open."
                                                               "\n\nDo you wish to open a new experiment anyway?",
                                         choices=["Proceed", "Cancel"],
                                         default_choice="Cancel")
                if not decision:
                    return

            open_file_path = easygui.fileopenbox(title="Open Experiment", default=default_dir+"*.txp",
                                                 filetypes=["*.txp", "Tracking experiment files"])
            if open_file_path:

                progress_win = ProgressWin(title="Open", msg="Opening experiment...")
                progress_win.cancel_button.place_forget()
                progress_win.progress["maximum"] = 50
                progress_win.progress["mode"] = "determinate"
                try:
                    with open(open_file_path, "rb") as open_doc:
                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_tracking = pickle.load(open_doc)
                        self.proc_data_tracking = pickle.load(open_doc)
                        self.data_length_tracking = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_coloc1 = pickle.load(open_doc)
                        self.proc_data_coloc1 = pickle.load(open_doc)
                        self.data_length_coloc1 = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_coloc2 = pickle.load(open_doc)
                        self.proc_data_coloc2 = pickle.load(open_doc)
                        self.data_length_coloc2 = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.filenames = pickle.load(open_doc)

                        self.tracking_data_trk = pickle.load(open_doc)
                        self.tracking_data_c1 = pickle.load(open_doc)
                        self.tracking_data_c2 = pickle.load(open_doc)

                        display_frames0 = pickle.load(open_doc)
                        display_frames1 = pickle.load(open_doc)
                        display_frames2 = pickle.load(open_doc)
                        self.display_brightness[0].set(pickle.load(open_doc))
                        self.display_brightness[1].set(pickle.load(open_doc))
                        self.display_brightness[2].set(pickle.load(open_doc))
                        self.display_contrast[0].set(pickle.load(open_doc))
                        self.display_contrast[1].set(pickle.load(open_doc))
                        self.display_contrast[2].set(pickle.load(open_doc))

                        current_traj = pickle.load(open_doc)
                        field = pickle.load(open_doc)

                        self.detection.delete(0, tk.END)
                        self.averaging.delete(0, tk.END)
                        self.residual.delete(0, tk.END)
                        self.minisig.delete(0, tk.END)
                        self.maxisig.delete(0, tk.END)
                        self.intensity.delete(0, tk.END)
                        self.amplitude.delete(0, tk.END)
                        self.eccentric.delete(0, tk.END)
                        self.gaussresid.delete(0, tk.END)
                        self.init_frame.delete(0, tk.END)
                        self.pixel_size.delete(0, tk.END)
                        self.frame_interval.delete(0, tk.END)
                        self.start_frame.delete(0, tk.END)
                        self.stop_frame.delete(0, tk.END)
                        self.max_displacement.delete(0, tk.END)
                        self.min_duration.delete(0, tk.END)
                        self.max_dark_time.delete(0, tk.END)

                        self.detection.insert(0, pickle.load(open_doc))
                        self.averaging.insert(0, pickle.load(open_doc))
                        self.residual.insert(0, pickle.load(open_doc))
                        self.minisig.insert(0, pickle.load(open_doc))
                        self.maxisig.insert(0, pickle.load(open_doc))
                        self.intensity.insert(0, pickle.load(open_doc))
                        self.amplitude.insert(0, pickle.load(open_doc))
                        self.eccentric.insert(0, pickle.load(open_doc))
                        self.gaussresid.insert(0, pickle.load(open_doc))
                        init = pickle.load(open_doc)
                        self.pixel_size.insert(0, pickle.load(open_doc))
                        self.frame_interval.insert(0, pickle.load(open_doc))
                        start = pickle.load(open_doc)
                        stop = pickle.load(open_doc)
                        self.max_displacement.insert(0, pickle.load(open_doc))
                        self.min_duration.insert(0, pickle.load(open_doc))
                        self.max_dark_time.insert(0, pickle.load(open_doc))

                        self.region_initial = pickle.load(open_doc)
                        self.region_final = pickle.load(open_doc)

                        self.invert_selection_flag_trk.set(pickle.load(open_doc))
                        self.invert_selection_flag_c1.set(pickle.load(open_doc))
                        self.invert_selection_flag_c2.set(pickle.load(open_doc))

                        min_msd_prop = pickle.load(open_doc)
                        self.minimum_msd_prop.delete(0, tk.END)
                        self.minimum_msd_timepoints.delete(0, tk.END)
                        self.msd_gradient_threshold_r2.delete(0, tk.END)
                        self.msd_stationary_r2_threshold.delete(0, tk.END)
                        self.msd_min_diff.delete(0, tk.END)
                        self.msd_uncertainty_degree.delete(0, tk.END)
                        self.msd_maximum_error.delete(0, tk.END)

                        self.minimum_msd_prop.insert(0, min_msd_prop)
                        self.minimum_msd_timepoints.insert(0, pickle.load(open_doc))
                        self.msd_gradient_threshold_r2.insert(0, pickle.load(open_doc))
                        self.msd_stationary_r2_threshold.insert(0, pickle.load(open_doc))
                        self.msd_min_diff.insert(0, pickle.load(open_doc))
                        self.msd_uncertainty_degree.insert(0, pickle.load(open_doc))

                        self.exclude_stationary.set(pickle.load(open_doc))
                        self.exclude_based_on_error.set(pickle.load(open_doc))
                        self.exclude_stationary_button.update()
                        self.exclude_based_on_error_button.update()
                        self.msd_maximum_error.insert(0, pickle.load(open_doc))

                except EOFError:
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                    """ File was missing components """
                try:
                    for add_msd_prop in range(len(self.tracking_data_trk)):
                        try:
                            confirm_MSD_prop = self.tracking_data_trk[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_trk[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_trk[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_trk[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_trk[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_trk[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_trk[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_trk[add_msd_prop].end_frame
                            self.tracking_data_trk[add_msd_prop] = new_obj

                    for add_msd_prop in range(len(self.tracking_data_c1)):
                        try:
                            confirm_MSD_prop = self.tracking_data_c1[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_c1[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_c1[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_c1[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_c1[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_c1[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_c1[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_c1[add_msd_prop].end_frame
                            self.tracking_data_c1[add_msd_prop] = new_obj

                    for add_msd_prop in range(len(self.tracking_data_c2)):
                        try:
                            confirm_MSD_prop = self.tracking_data_c2[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_c2[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_c2[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_c2[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_c2[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_c2[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_c2[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_c2[add_msd_prop].end_frame
                            self.tracking_data_c2[add_msd_prop] = new_obj

                except:
                    print("An Error occurred. Full traceback shown below:\n")
                    print(traceback.format_exc())

                progress_win.progress.step(9)

                self.field_selected.set(field)

                self.include_track_button["state"] = tk.NORMAL
                if self.proc_data_tracking is not None:
                    self.slider_trk["state"] = tk.NORMAL
                    self.slider_trk["to"] = self.data_length_tracking - 1
                    if self.display_frames[0].get() > self.data_length_tracking - 1:
                        self.slider_trk.set(self.data_length_tracking - 1)
                    self.bright_trk["state"] = tk.NORMAL
                    self.cont_trk["state"] = tk.NORMAL
                    self.display_frames[0].set(display_frames0)
                    self.display_tracking()
                else:
                    self.slider_trk["state"] = tk.DISABLED
                    self.bright_trk["state"] = tk.DISABLED
                    self.cont_trk["state"] = tk.DISABLED
                    self.display_tracking(erase=True)
                if self.proc_data_coloc1 is not None:
                    self.slider_c1["state"] = tk.NORMAL
                    self.slider_c1["to"] = self.data_length_coloc1 - 1
                    if self.display_frames[1].get() > self.data_length_coloc1 - 1:
                        self.slider_c1.set(self.data_length_coloc1 - 1)
                    self.bright_c1["state"] = tk.NORMAL
                    self.cont_c1["state"] = tk.NORMAL
                    self.display_frames[1].set(display_frames1)
                    self.display_coloc1()
                else:
                    self.slider_c1["state"] = tk.DISABLED
                    self.bright_c1["state"] = tk.DISABLED
                    self.cont_c1["state"] = tk.DISABLED
                    self.display_coloc1(erase=True)
                if self.proc_data_coloc2 is not None:
                    self.slider_c2["state"] = tk.NORMAL
                    self.slider_c2["to"] = self.data_length_coloc2 - 1
                    if self.display_frames[2].get() > self.data_length_coloc2 - 1:
                        self.slider_c2.set(self.data_length_coloc2 - 1)
                    self.bright_c2["state"] = tk.NORMAL
                    self.cont_c2["state"] = tk.NORMAL
                    self.display_frames[2].set(display_frames2)
                    self.display_coloc2()
                else:
                    self.slider_c2["state"] = tk.DISABLED
                    self.bright_c2["state"] = tk.DISABLED
                    self.cont_c2["state"] = tk.DISABLED
                    self.display_coloc2(erase=True)
                self.update_fileinfo(self.field_selected.get())

                self.start_frame.insert(0, start)
                self.stop_frame.insert(0, stop)

                try:
                    self.update_trajectories()
                    self.display_trajectory()
                except:
                    """ Failed """

                self.current_trajectory.set(current_traj)
                self.select_trajectory.update()
                self.display_trajectory()

                progress_win.progress.stop()
                progress_win.handle_close()
                self.saved_flag = True

        except PermissionError:
            easygui.msgbox(title="Error!", msg="Warning! an error occurred while attempting to open experiment:\n"
                                               "PermissionError: Permission denied!")
            try:
                progress_win.progress.step(1)
                progress_win.progress.stop()
                progress_win.handle_close()
            except:
                """ No progress bar window open """

    def open_document_auto(self, filename):
        try:
            open_file_path = filename
            if open_file_path:

                progress_win = ProgressWin(title="Open", msg="Opening experiment...")
                progress_win.cancel_button.place_forget()
                progress_win.progress["maximum"] = 50
                progress_win.progress["mode"] = "determinate"
                try:
                    with open(open_file_path, "rb") as open_doc:
                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_tracking = pickle.load(open_doc)
                        self.proc_data_tracking = pickle.load(open_doc)
                        self.data_length_tracking = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_coloc1 = pickle.load(open_doc)
                        self.proc_data_coloc1 = pickle.load(open_doc)
                        self.data_length_coloc1 = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.raw_data_coloc2 = pickle.load(open_doc)
                        self.proc_data_coloc2 = pickle.load(open_doc)
                        self.data_length_coloc2 = pickle.load(open_doc)

                        progress_win.progress.step(10)
                        progress_win.progress.update()
                        self.filenames = pickle.load(open_doc)

                        self.tracking_data_trk = pickle.load(open_doc)
                        self.tracking_data_c1 = pickle.load(open_doc)
                        self.tracking_data_c2 = pickle.load(open_doc)

                        display_frames0 = pickle.load(open_doc)
                        display_frames1 = pickle.load(open_doc)
                        display_frames2 = pickle.load(open_doc)
                        self.display_brightness[0].set(pickle.load(open_doc))
                        self.display_brightness[1].set(pickle.load(open_doc))
                        self.display_brightness[2].set(pickle.load(open_doc))
                        self.display_contrast[0].set(pickle.load(open_doc))
                        self.display_contrast[1].set(pickle.load(open_doc))
                        self.display_contrast[2].set(pickle.load(open_doc))

                        current_traj = pickle.load(open_doc)
                        field = pickle.load(open_doc)

                        self.detection.delete(0, tk.END)
                        self.averaging.delete(0, tk.END)
                        self.residual.delete(0, tk.END)
                        self.minisig.delete(0, tk.END)
                        self.maxisig.delete(0, tk.END)
                        self.intensity.delete(0, tk.END)
                        self.amplitude.delete(0, tk.END)
                        self.eccentric.delete(0, tk.END)
                        self.gaussresid.delete(0, tk.END)
                        self.init_frame.delete(0, tk.END)
                        self.pixel_size.delete(0, tk.END)
                        self.frame_interval.delete(0, tk.END)
                        self.start_frame.delete(0, tk.END)
                        self.stop_frame.delete(0, tk.END)
                        self.max_displacement.delete(0, tk.END)
                        self.min_duration.delete(0, tk.END)
                        self.max_dark_time.delete(0, tk.END)

                        self.detection.insert(0, pickle.load(open_doc))
                        self.averaging.insert(0, pickle.load(open_doc))
                        self.residual.insert(0, pickle.load(open_doc))
                        self.minisig.insert(0, pickle.load(open_doc))
                        self.maxisig.insert(0, pickle.load(open_doc))
                        self.intensity.insert(0, pickle.load(open_doc))
                        self.amplitude.insert(0, pickle.load(open_doc))
                        self.eccentric.insert(0, pickle.load(open_doc))
                        self.gaussresid.insert(0, pickle.load(open_doc))
                        init = pickle.load(open_doc)
                        self.pixel_size.insert(0, pickle.load(open_doc))
                        self.frame_interval.insert(0, pickle.load(open_doc))
                        start = pickle.load(open_doc)
                        stop = pickle.load(open_doc)
                        self.max_displacement.insert(0, pickle.load(open_doc))
                        self.min_duration.insert(0, pickle.load(open_doc))
                        self.max_dark_time.insert(0, pickle.load(open_doc))

                        self.region_initial = pickle.load(open_doc)
                        self.region_final = pickle.load(open_doc)

                        self.invert_selection_flag_trk.set(pickle.load(open_doc))
                        self.invert_selection_flag_c1.set(pickle.load(open_doc))
                        self.invert_selection_flag_c2.set(pickle.load(open_doc))

                        lock_constraints = False
                        try:
                            if histwin.lock_constraints.get() == 1:
                                lock_constraints = True
                        except:
                            """ Window was closed """

                        if not lock_constraints:
                            min_msd_prop = pickle.load(open_doc)
                            self.minimum_msd_prop.delete(0, tk.END)
                            self.minimum_msd_timepoints.delete(0, tk.END)
                            self.msd_gradient_threshold_r2.delete(0, tk.END)
                            self.msd_stationary_r2_threshold.delete(0, tk.END)
                            self.msd_min_diff.delete(0, tk.END)
                            self.msd_uncertainty_degree.delete(0, tk.END)
                            self.msd_maximum_error.delete(0, tk.END)

                            self.minimum_msd_prop.insert(0, min_msd_prop)
                            self.minimum_msd_timepoints.insert(0, pickle.load(open_doc))
                            self.msd_gradient_threshold_r2.insert(0, pickle.load(open_doc))
                            self.msd_stationary_r2_threshold.insert(0, pickle.load(open_doc))
                            self.msd_min_diff.insert(0, pickle.load(open_doc))
                            self.msd_uncertainty_degree.insert(0, pickle.load(open_doc))

                            self.exclude_stationary.set(pickle.load(open_doc))
                            self.exclude_based_on_error.set(pickle.load(open_doc))
                            self.exclude_stationary_button.update()
                            self.exclude_based_on_error_button.update()
                            self.msd_maximum_error.insert(0, pickle.load(open_doc))

                except EOFError:
                    print("Warning error occured: Full traceback shown below:")
                    print(traceback.format_exc())
                    """ File was missing components """
                try:
                    for add_msd_prop in range(len(self.tracking_data_trk)):
                        try:
                            confirm_MSD_prop = self.tracking_data_trk[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_trk[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_trk[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_trk[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_trk[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_trk[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_trk[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_trk[add_msd_prop].end_frame
                            self.tracking_data_trk[add_msd_prop] = new_obj

                    for add_msd_prop in range(len(self.tracking_data_c1)):
                        try:
                            confirm_MSD_prop = self.tracking_data_c1[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_c1[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_c1[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_c1[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_c1[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_c1[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_c1[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_c1[add_msd_prop].end_frame
                            self.tracking_data_c1[add_msd_prop] = new_obj

                    for add_msd_prop in range(len(self.tracking_data_c2)):
                        try:
                            confirm_MSD_prop = self.tracking_data_c2[add_msd_prop].MSD_prop
                        except:
                            """ MSD_prop attribute not present in file. creating new object and transferring data """
                            new_obj = TrackingData()
                            new_obj.fitting_params = self.tracking_data_c2[add_msd_prop].fitting_params
                            new_obj.include_export = self.tracking_data_c2[add_msd_prop].include_export
                            new_obj.termination_reason = self.tracking_data_c2[add_msd_prop].termination_reason
                            new_obj.coordinates = self.tracking_data_c2[add_msd_prop].coordinates
                            new_obj.frame_list = self.tracking_data_c2[add_msd_prop].frame_list
                            new_obj.start_frame = self.tracking_data_c2[add_msd_prop].start_frame
                            new_obj.end_frame = self.tracking_data_c2[add_msd_prop].end_frame
                            self.tracking_data_c2[add_msd_prop] = new_obj

                except:
                    print("An Error occurred. Full traceback shown below:\n")
                    print(traceback.format_exc())

                progress_win.progress.step(9)

                self.field_selected.set(field)

                self.include_track_button["state"] = tk.NORMAL
                if self.proc_data_tracking is not None:
                    self.slider_trk["state"] = tk.NORMAL
                    self.slider_trk["to"] = self.data_length_tracking - 1
                    if self.display_frames[0].get() > self.data_length_tracking - 1:
                        self.slider_trk.set(self.data_length_tracking - 1)
                    self.bright_trk["state"] = tk.NORMAL
                    self.cont_trk["state"] = tk.NORMAL
                    self.display_frames[0].set(display_frames0)
                    self.display_tracking()
                else:
                    self.slider_trk["state"] = tk.DISABLED
                    self.bright_trk["state"] = tk.DISABLED
                    self.cont_trk["state"] = tk.DISABLED
                    self.display_tracking(erase=True)
                if self.proc_data_coloc1 is not None:
                    self.slider_c1["state"] = tk.NORMAL
                    self.slider_c1["to"] = self.data_length_coloc1 - 1
                    if self.display_frames[1].get() > self.data_length_coloc1 - 1:
                        self.slider_c1.set(self.data_length_coloc1 - 1)
                    self.bright_c1["state"] = tk.NORMAL
                    self.cont_c1["state"] = tk.NORMAL
                    self.display_frames[1].set(display_frames1)
                    self.display_coloc1()
                else:
                    self.slider_c1["state"] = tk.DISABLED
                    self.bright_c1["state"] = tk.DISABLED
                    self.cont_c1["state"] = tk.DISABLED
                    self.display_coloc1(erase=True)
                if self.proc_data_coloc2 is not None:
                    self.slider_c2["state"] = tk.NORMAL
                    self.slider_c2["to"] = self.data_length_coloc2 - 1
                    if self.display_frames[2].get() > self.data_length_coloc2 - 1:
                        self.slider_c2.set(self.data_length_coloc2 - 1)
                    self.bright_c2["state"] = tk.NORMAL
                    self.cont_c2["state"] = tk.NORMAL
                    self.display_frames[2].set(display_frames2)
                    self.display_coloc2()
                else:
                    self.slider_c2["state"] = tk.DISABLED
                    self.bright_c2["state"] = tk.DISABLED
                    self.cont_c2["state"] = tk.DISABLED
                    self.display_coloc2(erase=True)
                self.update_fileinfo(self.field_selected.get())
                self.init_frame.insert(0, 0)
                self.start_frame.insert(0, start)
                self.stop_frame.insert(0, stop)

                try:
                    self.update_trajectories()
                    self.display_trajectory()
                except:
                    """ Failed """

                self.current_trajectory.set(current_traj)
                self.select_trajectory.update()
                self.display_trajectory()

                progress_win.progress.stop()
                progress_win.handle_close()
                self.saved_flag = True

        except PermissionError:
            easygui.msgbox(title="Error!", msg="Warning! an error occurred while attempting to open experiment:\n"
                                               "PermissionError: Permission denied!")
            try:
                progress_win.progress.step(1)
                progress_win.progress.stop()
                progress_win.handle_close()
            except:
                """ No progress bar window open """

    def canvas_trk_clicked(self, event, callback=False):
        if self.field_selected.get() != self.field_list[0]:
            return
        if not callback:
            mouse = [event.x, event.y]
        else:
            mouse = event
        coords = []
        indices = []
        for ind, spot in enumerate(self.tracking_data_trk):
            try:
                frame_list_index = spot.frame_list.index(self.display_frames[0].get())
                coords.append(spot.coordinates[frame_list_index])
                indices.append(ind)
            except:
                """ No spots """
        dists = []
        for i in range(len(coords)):
            dists.append((coords[i][0] - mouse[0])**2 + (coords[i][1] - mouse[1])**2)
        if len(dists) > 0:
            index = dists.index(np.min(dists))
            name = "Trajectory " + str(indices[index] + 1)
            self.current_trajectory.set(name)
            self.select_trajectory.update()
            self.display_trajectory()

    def canvas_c1_clicked(self, event, callback=False):
        if self.field_selected.get() != self.field_list[1]:
            return
        if not callback:
            mouse = [event.x, event.y]
        else:
            mouse = event
        coords = []
        indices = []
        for ind, spot in enumerate(self.tracking_data_c1):
            try:
                frame_list_index = spot.frame_list.index(self.display_frames[1].get())
                coords.append(spot.coordinates[frame_list_index])
                indices.append(ind)
            except:
                """ No spots """
        dists = []
        for i in range(len(coords)):
            dists.append((coords[i][0] - mouse[0])**2 + (coords[i][1] - mouse[1])**2)
        if len(dists) > 0:
            index = dists.index(np.min(dists))
            name = "Trajectory " + str(indices[index] + 1)
            self.current_trajectory.set(name)
            self.select_trajectory.update()
            self.display_trajectory()

    def canvas_c2_clicked(self, event, callback=False):
        if self.field_selected.get() != self.field_list[2]:
            return
        if not callback:
            mouse = [event.x, event.y]
        else:
            mouse = event
        coords = []
        indices = []
        for ind, spot in enumerate(self.tracking_data_c2):
            try:
                frame_list_index = spot.frame_list.index(self.display_frames[2].get())
                coords.append(spot.coordinates[frame_list_index])
                indices.append(ind)
            except:
                """ No spots """
        dists = []
        for i in range(len(coords)):
            dists.append((coords[i][0] - mouse[0])**2 + (coords[i][1] - mouse[1])**2)
        if len(dists) > 0:
            index = dists.index(np.min(dists))
            name = "Trajectory " + str(indices[index] + 1)
            self.current_trajectory.set(name)
            self.select_trajectory.update()
            self.display_trajectory()

    def enlarged_trk_clicked(self, event):
        self.canvas_trk_clicked([event.x / 1.75, event.y / 1.75], callback=True)

    def enlarged_c1_clicked(self, event):
        self.canvas_c1_clicked([event.x / 1.75, event.y / 1.75], callback=True)

    def enlarged_c2_clicked(self, event):
        self.canvas_c2_clicked([event.x / 1.75, event.y / 1.75], callback=True)

    def float_view(self, view):
        global enlarge_win
        d = {
            0: "Primary Tracking Field",
            1: "Colocalization Field 1",
            2: "Colocalization Field 2",
        }
        d2 = {
            0: self.display_tracking,
            1: self.display_coloc1,
            2: self.display_coloc2,
        }
        d3 = {
            0: self.proc_data_tracking,
            1: self.proc_data_coloc1,
            2: self.proc_data_coloc2,
        }
        d4 = {
            0: self.enlarged_trk_clicked,
            1: self.enlarged_c1_clicked,
            2: self.enlarged_c2_clicked,
        }
        try:
            enlarge_win.window.destroy()
        except:
            """ Failed """
        if not self.popout_state[view]:
            d2[view](erase=True)
        self.popout_state[view] = True
        enlarge_win = EnlargedView(title=d[view], shape=(np.shape(d3[view])[0], np.shape(d3[view])[1]),
                                   function=d2[view], callback=d4[view])
        d2[view]()

    def display_tracking(self, update=False, erase=False):
        if self.proc_data_tracking is None:
            self.canvas_tracking.delete("all")
            return
        if self.popout_state[0]:
            canvas = enlarge_win.canvas
            scale = 1.75
        else:
            canvas = self.canvas_tracking
            scale = 1
        self.slider_trk["repeatinterval"] = int(float(self.frame_interval.get()) * 1000)
        global track_img
        shape = np.shape(self.proc_data_tracking)
        frame = np.zeros((shape[0], shape[1], 3))
        frame[:, :, 0] = self.proc_data_tracking[:, :, self.display_frames[0].get()]
        frame[:, :, 1] = self.proc_data_tracking[:, :, self.display_frames[0].get()]
        frame[:, :, 2] = self.proc_data_tracking[:, :, self.display_frames[0].get()]
        if self.global_norm_trk.get() == 1:
            maximum = np.max(self.proc_data_tracking)
        else:
            maximum = np.max(frame)
        frame = (frame / maximum) * 255 * self.display_brightness[0].get() / 50
        frame = np.clip(frame, 0, 255)
        frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                              frame[:, :, 2],
                                                                              self.display_contrast[0].get())
        frame = np.clip(frame * 2, 0, 255)
        image = tirf.create_image(frame, int(shape[0]*scale), int(shape[1]*scale))
        if self.popout_state[0]:
            track_img = tirf.ImageTk.PhotoImage(master=enlarge_win.window, image=image)
        else:
            track_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
        canvas.delete("all")
        canvas.create_image(2, 2, anchor="nw", image=track_img)
        if self.tracking_data_trk is not None:
            for ind, spot in enumerate(self.tracking_data_trk):
                try:
                    frame_list_index = spot.frame_list.index(self.display_frames[0].get())
                    coords = spot.coordinates[frame_list_index]
                    col = "yellow"
                    try:
                        if self.trajectory_list.index(self.current_trajectory.get()) == ind:
                            col = "red"
                    except:
                        """ No trajectory data """
                    try:
                        if self.boxes[0].get() == 1 or col == "red":
                            canvas.create_rectangle((coords[0] - 3)*scale, (coords[1] - 3)*scale, (coords[0] + 7)*scale,
                                                                  (coords[1] + 7)*scale,
                                                                  outline=col, width=1)
                            canvas.create_text(round(coords[0] + 10)*scale, round(coords[1] - 10)*scale, fill="white",
                                                             font="arial 8", text=str(ind + 1))
                    except:
                        """ Coordinates invalid """
                except ValueError:
                    """ trajectory not present in this frame """

        if self.region_initial[0] is not None and self.region_final[0] is not None:
            if self.invert_selection_flag_trk.get() == 1:
                col2 = "red"
            else:
                col2 = "green"
            self.rect_trk = self.canvas_tracking.create_rectangle(self.region_initial[0][0], self.region_initial[0][1],
                                                                  self.region_final[0][0], self.region_final[0][1],
                                                                  outline=col2, width=1, dash=(2, 1))

        if erase:
            self.canvas_tracking.delete("all")

        if update:
            self.load_init_frame()

    def display_coloc1(self, update=False, erase=False):
        if self.proc_data_coloc1 is None:
            self.canvas_coloc1.delete("all")
            return
        if self.popout_state[1]:
            canvas = enlarge_win.canvas
            scale = 1.75
        else:
            canvas = self.canvas_coloc1
            scale = 1
        self.slider_c1["repeatinterval"] = int(float(self.frame_interval.get()) * 1000)
        global col1_img
        shape = np.shape(self.proc_data_coloc1)
        frame = np.zeros((shape[0], shape[1], 3))
        frame[:, :, 0] = self.proc_data_coloc1[:, :, self.display_frames[1].get()]
        frame[:, :, 1] = self.proc_data_coloc1[:, :, self.display_frames[1].get()]
        frame[:, :, 2] = self.proc_data_coloc1[:, :, self.display_frames[1].get()]
        if self.global_norm_c1.get() == 1:
            maximum = np.max(self.proc_data_coloc1)
        else:
            maximum = np.max(frame)
        frame = (frame / maximum) * 255 * self.display_brightness[1].get() / 50
        frame = np.clip(frame, 0, 255)
        frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                              frame[:, :, 2],
                                                                              self.display_contrast[1].get())
        frame = np.clip(frame * 2, 0, 255)
        image = tirf.create_image(frame, int(shape[0]*scale), int(shape[1]*scale))
        if self.popout_state[1]:
            col1_img = tirf.ImageTk.PhotoImage(master=enlarge_win.window, image=image)
        else:
            col1_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
        canvas.delete("all")
        canvas.create_image(2, 2, anchor="nw", image=col1_img)
        if self.tracking_data_c1 is not None:
            for ind, spot in enumerate(self.tracking_data_c1):
                try:
                    frame_list_index = spot.frame_list.index(self.display_frames[1].get())
                    coords = spot.coordinates[frame_list_index]
                    col = "yellow"
                    try:
                        if self.trajectory_list.index(self.current_trajectory.get()) == ind:
                            col = "red"
                    except:
                        """ No trajectory data """
                    try:
                        if self.boxes[1].get() == 1 or col == "red":
                            canvas.create_rectangle((coords[0] - 3)*scale, (coords[1] - 3)*scale, (coords[0] + 7)*scale,
                                                                  (coords[1] + 7)*scale,
                                                                  outline=col, width=1)
                            canvas.create_text(round(coords[0] + 10)*scale, round(coords[1] - 10)*scale, fill="white",
                                                             font="arial 8", text=str(ind + 1))
                    except:
                        """ Coordinates invalid """
                except ValueError:
                    """ trajectory not present in this frame """

        if self.region_initial[1] is not None and self.region_final[1] is not None:
            if self.invert_selection_flag_c1.get() == 1:
                col2 = "red"
            else:
                col2 = "green"
            self.rect_c1 = self.canvas_coloc1.create_rectangle(self.region_initial[1][0], self.region_initial[1][1],
                                                               self.region_final[1][0], self.region_final[1][1],
                                                               outline=col2, width=1, dash=(2, 1))

        if erase:
            self.canvas_coloc1.delete("all")

        if update:
            self.load_init_frame()

    def display_coloc2(self, update=False, erase=False):
        if self.proc_data_coloc2 is None:
            self.canvas_coloc2.delete("all")
            return
        if self.popout_state[2]:
            canvas = enlarge_win.canvas
            scale = 1.75
        else:
            canvas = self.canvas_coloc2
            scale = 1
        self.slider_c2["repeatinterval"] = int(float(self.frame_interval.get()) * 1000)
        global col2_img
        shape = np.shape(self.proc_data_coloc2)
        frame = np.zeros((shape[0], shape[1], 3))
        frame[:, :, 0] = self.proc_data_coloc2[:, :, self.display_frames[2].get()]
        frame[:, :, 1] = self.proc_data_coloc2[:, :, self.display_frames[2].get()]
        frame[:, :, 2] = self.proc_data_coloc2[:, :, self.display_frames[2].get()]
        if self.global_norm_c2.get() == 1:
            maximum = np.max(self.proc_data_coloc2)
        else:
            maximum = np.max(frame)
        frame = (frame / maximum) * 255 * self.display_brightness[2].get() / 50
        frame = np.clip(frame, 0, 255)
        frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                              frame[:, :, 2],
                                                                              self.display_contrast[2].get())
        frame = np.clip(frame * 2, 0, 255)
        image = tirf.create_image(frame, int(shape[0]*scale), int(shape[1]*scale))
        if self.popout_state[2]:
            col2_img = tirf.ImageTk.PhotoImage(master=enlarge_win.window, image=image)
        else:
            col2_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
        canvas.delete("all")
        canvas.create_image(2, 2, anchor="nw", image=col2_img)
        if self.tracking_data_c2 is not None:
            for ind, spot in enumerate(self.tracking_data_c2):
                try:
                    frame_list_index = spot.frame_list.index(self.display_frames[2].get())
                    coords = spot.coordinates[frame_list_index]
                    col = "yellow"
                    try:
                        if self.trajectory_list.index(self.current_trajectory.get()) == ind:
                            col = "red"
                    except:
                        """ No trajectory data """
                    try:
                        if self.boxes[2].get() == 1 or col == "red":
                            canvas.create_rectangle((coords[0] - 3) * scale, (coords[1] - 3) * scale,
                                                    (coords[0] + 7) * scale,
                                                    (coords[1] + 7) * scale,
                                                    outline=col, width=1)
                            canvas.create_text(round(coords[0] + 10) * scale, round(coords[1] - 10) * scale,
                                               fill="white",
                                               font="arial 8", text=str(ind + 1))
                    except:
                        """ Coordinates invalid """
                except ValueError:
                    """ trajectory not present in this frame """

        if self.region_initial[2] is not None and self.region_final[2] is not None:
            if self.invert_selection_flag_c2.get() == 1:
                col2 = "red"
            else:
                col2 = "green"
            self.rect_c2 = self.canvas_coloc2.create_rectangle(self.region_initial[2][0], self.region_initial[2][1],
                                                               self.region_final[2][0], self.region_final[2][1],
                                                               outline=col2, width=1, dash=(2, 1))

        if erase:
            self.canvas_coloc2.delete("all")

        if update:
            self.load_init_frame()

    def load_init_frame(self, initial=None):
        global init_img
        if initial is None:
            initial = self.init_frame.get()
        if initial != "" and self.field_selected.get() == self.field_list[0]:
            if self.proc_data_tracking is not None:
                shape = np.shape(self.proc_data_tracking)
                frame = np.zeros((shape[0], shape[1], 3))
                frame[:, :, 0] = self.proc_data_tracking[:, :, int(initial)]
                frame[:, :, 1] = self.proc_data_tracking[:, :, int(initial)]
                frame[:, :, 2] = self.proc_data_tracking[:, :, int(initial)]
                maximum = np.max(self.proc_data_tracking)
                frame = (frame / maximum) * 255 * self.display_brightness[0].get() / 20
                frame = np.clip(frame, 0, 255)
                frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                                      frame[:, :, 2],
                                                                                      self.display_contrast[0].get())
                image = tirf.create_image(frame, int(shape[0] / 2), int(shape[1] / 2))
                init_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
                self.canvas_initial.create_image(2, 2, anchor="nw", image=init_img)
                self.detect_spots_button["state"] = tk.NORMAL

        elif initial != "" and self.field_selected.get() == self.field_list[1]:
            if self.proc_data_coloc1 is not None:
                shape = np.shape(self.proc_data_coloc1)
                frame = np.zeros((shape[0], shape[1], 3))
                frame[:, :, 0] = self.proc_data_coloc1[:, :, int(initial)]
                frame[:, :, 1] = self.proc_data_coloc1[:, :, int(initial)]
                frame[:, :, 2] = self.proc_data_coloc1[:, :, int(initial)]
                maximum = np.max(self.proc_data_coloc1)
                frame = (frame / maximum) * 255 * self.display_brightness[1].get() / 20
                frame = np.clip(frame, 0, 255)
                frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                                      frame[:, :, 2],
                                                                                      self.display_contrast[1].get())
                image = tirf.create_image(frame, int(shape[0] / 2), int(shape[1] / 2))
                init_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
                self.canvas_initial.create_image(2, 2, anchor="nw", image=init_img)
                self.detect_spots_button["state"] = tk.NORMAL

        elif initial != "" and self.field_selected.get() == self.field_list[2]:
            if self.proc_data_coloc2 is not None:
                shape = np.shape(self.proc_data_coloc2)
                frame = np.zeros((shape[0], shape[1], 3))
                frame[:, :, 0] = self.proc_data_coloc2[:, :, int(initial)]
                frame[:, :, 1] = self.proc_data_coloc2[:, :, int(initial)]
                frame[:, :, 2] = self.proc_data_coloc2[:, :, int(initial)]
                maximum = np.max(self.proc_data_coloc2)
                frame = (frame / maximum) * 255 * self.display_brightness[2].get() / 20
                frame = np.clip(frame, 0, 255)
                frame[:, :, 0], frame[:, :, 1], frame[:, :, 2] = self.adjust_contrast(frame[:, :, 0], frame[:, :, 1],
                                                                                      frame[:, :, 2],
                                                                                      self.display_contrast[2].get())
                image = tirf.create_image(frame, int(shape[0] / 2), int(shape[1] / 2))
                init_img = tirf.ImageTk.PhotoImage(master=self.window, image=image)
                self.canvas_initial.create_image(2, 2, anchor="nw", image=init_img)
                self.detect_spots_button["state"] = tk.NORMAL

        else:
            self.detect_spots_button["state"] = tk.DISABLED

    def check_empty(self, pos=None):
        if self.detection.get() and self.averaging.get() and self.residual.get() and self.minisig.get() and \
                self.maxisig.get() and self.intensity.get() and self.amplitude.get() and self.eccentric.get() and \
                self.gaussresid.get():
            self.err_label["text"] = ""
            self.err_label.update()
        else:
            self.err_label["text"] = "Warning! Required field left blank!"
            self.err_label.update()

    def validate_float(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label["text"] = "Warning! Required field left blank!"
            self.err_label.update()
        if action == "1":
            valid = True
            try:
                float(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label["text"] = ""
                self.err_label.update()
            else:
                self.check_empty()
        if value:
            try:
                float(value)
            except ValueError:
                return False
            if len(value) > 4:
                return False
        return True

    def validate_int(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label["text"] = "Warning! Required field left blank!"
            self.err_label.update()
        if action == "1":
            valid = True
            try:
                int(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label["text"] = ""
                self.err_label.update()
            else:
                self.check_empty()
            if not value.isdigit():
                return False
            if int(float(value)) > 255:
                return False
        return True

    def validate_float_track(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label2["text"] = "Warning! Required field left blank!"
            self.err_label2.update()
        if action == "1":
            valid = True
            try:
                float(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label2["text"] = ""
                self.err_label2.update()
            else:
                self.check_empty()
        if value:
            try:
                float(value)
            except ValueError:
                return False
            if len(value) > 7:
                return False
        return True

    def validate_int_track(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label2["text"] = "Warning! Required field left blank!"
            self.err_label2.update()
        if action == "1":
            valid = True
            try:
                int(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label2["text"] = ""
                self.err_label2.update()
            else:
                self.check_empty()
            if not value.isdigit():
                return False
            if int(float(value)) > 100000:
                return False
        return True

    def init_frame_trigger(self, value, action, prev):
        if action == "0" and len(value) == 0:
            self.err_label["text"] = "Warning! Required field left blank!"
            self.err_label.update()
        if action == "1":
            valid = True
            try:
                int(value)
            except:
                valid = False
            if len(prev) == 0 and valid:
                self.err_label["text"] = ""
                self.err_label.update()
            else:
                self.check_empty()
            if not value.isdigit():
                return False
            if int(float(value)) > self.current_field_length - 1:
                return False
        self.load_init_frame(initial=value)
        return True

    def get_initial_spots(self):
        initial_state_params = {
            "detection threshold": [float(self.detection.get())],
            "averaging distance": [float(self.averaging.get())],
            "minimum kernel residual": [float(self.residual.get())],
            "minimum sigma": [float(self.minisig.get())],
            "maximum sigma": [float(self.maxisig.get())],
            "minimum intensity": [int(self.intensity.get())],
            "minimum gauss amplitude": [int(self.amplitude.get())],
            "eccentricity threshold": [float(self.eccentric.get())],
            "minimum gauss residual": [float(self.gaussresid.get())],
        }

        region = None
        region_type = None

        if self.field_selected.get() == self.field_list[0]:
            frame = self.proc_data_tracking[:, :, int(self.init_frame.get())]
            norm = self.global_norm_trk.get()
            full_max = np.max(self.proc_data_tracking)
            maximum_frame_index = np.unravel_index(np.argmax(self.proc_data_tracking), self.proc_data_tracking.shape)[2]
            max_frame = self.proc_data_tracking[:, :, maximum_frame_index]
            if self.region_initial[0] is not None and self.region_final[0] is not None:
                region = (self.region_initial[0], self.region_final[0])
                region_type = self.invert_selection_flag_trk.get()

        elif self.field_selected.get() == self.field_list[1]:
            frame = self.proc_data_coloc1[:, :, int(self.init_frame.get())]
            norm = self.global_norm_c1.get()
            full_max = np.max(self.proc_data_coloc1)
            maximum_frame_index = np.unravel_index(np.argmax(self.proc_data_coloc1), self.proc_data_coloc1.shape)[2]
            max_frame = self.proc_data_coloc1[:, :, maximum_frame_index]
            if self.region_initial[1] is not None and self.region_final[1] is not None:
                region = (self.region_initial[1], self.region_final[1])
                region_type = self.invert_selection_flag_c1.get()

        elif self.field_selected.get() == self.field_list[2]:
            frame = self.proc_data_coloc2[:, :, int(self.init_frame.get())]
            norm = self.global_norm_c2.get()
            full_max = np.max(self.proc_data_coloc2)
            maximum_frame_index = np.unravel_index(np.argmax(self.proc_data_coloc2), self.proc_data_coloc2.shape)[2]
            max_frame = self.proc_data_coloc2[:, :, maximum_frame_index]
            if self.region_initial[2] is not None and self.region_final[2] is not None:
                region = (self.region_initial[2], self.region_final[2])
                region_type = self.invert_selection_flag_c2.get()

        print(f"Maximum frame index: {maximum_frame_index}")
        print(f"Max. array int: {full_max}")
        print(f"Max. pixel int: {self.proc_data_tracking[(np.unravel_index(np.argmax(self.proc_data_tracking), self.proc_data_tracking.shape)[0])][(np.unravel_index(np.argmax(self.proc_data_tracking), self.proc_data_tracking.shape)[1])][(np.unravel_index(np.argmax(self.proc_data_tracking), self.proc_data_tracking.shape)[2])]}")
        print(f"Max. array int normalized: {np.max(max_frame)}")
        print(f"Max. frame int normalized: {np.max(frame)}")
        if region is not None:
            if region[0][0] > region[1][0]:
                region[0][0], region[1][0] = region[1][0], region[0][0]
            if region[0][1] > region[1][1]:
                region[0][1], region[1][1] = region[1][1], region[0][1]


        shape = np.shape(frame)

        if norm == 0:
            maximum = np.max(frame)
        elif norm == 1:
            maximum = full_max

        frame = (frame / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
        frame = np.clip(frame, 0, 255)
        frame[:, :] = self.adjust_contrast(0, 0, frame[:, :],
                                           self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                           single_channel=True)

        max_frame = (max_frame / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
        max_frame = np.clip(max_frame, 0, 255)
        max_frame[:, :] = self.adjust_contrast(0, 0, max_frame[:, :],
                                           self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                           single_channel=True)

        print(np.max(max_frame))
        print(np.max(frame))

        if norm == 0:
            maximum = np.max(frame)
        elif norm == 1:
            maximum = np.max(max_frame)

        frame = (frame / maximum) * 255

        spots = tirf.continuous_track(0, frame, initial_state_params, 0, region=region, inverted=region_type)
        self.initial_state = spots

        text = ""
        for i in range(len(spots)):
            text += f"Spot {i + 1}, Coordinates: {round(spots[i][0], 1)}, {round(spots[i][1], 1)}, Sigma(x): " \
                    f"{round(spots[i][2][0], 2)}, Sigma(y): {round(spots[i][2][1], 2)}, Amplitude: " \
                    f"{round(spots[i][2][2], 1)}, Residual: {round(spots[i][2][3], 2)}\n"

        self.spot_list["state"] = tk.NORMAL
        self.spot_list.delete(1.0, tk.END)
        self.spot_list.insert(tk.END, text)
        self.spot_list["state"] = tk.DISABLED

        coords = []
        for spot in spots:
            coords.append([spot[0], spot[1]])

        ax = plt.subplot()
        ax.imshow(frame)
        for i in range(len(coords)):
            c = plt.Circle((coords[i][0], coords[i][1]), 5, fill=False, color="white")
            plt.text(coords[i][0] + 4, coords[i][1] - 4, str(i + 1), size=10, color="white")
            ax.add_artist(c)
        self.ready_for_evolve = True
        self.toggle_button_state(self.analysis_mode.get())
        plt.title("Initial State")
        plt.show()

    def discard_spots(self):
        self.initial_state = None
        self.spot_list["state"] = tk.NORMAL
        self.spot_list.delete(1.0, tk.END)
        self.spot_list.update()
        self.spot_list["state"] = tk.DISABLED

    def analyse(self):
        if self.analysis_mode.get() == "Evolve Initial State":
            self.analyse_evolve()
        elif self.analysis_mode.get() == "Continuous":
            self.analyse_continuous()

    def analyse_evolve(self):
        global progress_win
        self.cancel_flag = False
        if self.field_selected.get() == self.field_list[0]:
            data = np.copy(self.proc_data_tracking)
            self.tracking_data_trk = []
        elif self.field_selected.get() == self.field_list[1]:
            data = np.copy(self.proc_data_coloc1)
            self.tracking_data_c1 = []
        elif self.field_selected.get() == self.field_list[2]:
            data = np.copy(self.proc_data_coloc2)
            self.tracking_data_c2 = []

        tracking_data = []
        previous_coordinates = [[[spot[0], spot[1]] for spot in self.initial_state]]

        shape = np.shape(data)
        length = shape[2]
        start = int(self.init_frame.get())

        for i in range(len(self.initial_state)):
            tracking_data.append(TrackingData())
            tracking_data[-1].start_frame = start

        progress_win = ProgressWin("Analysing Movie", "Tracking spots...")
        progress_win.progress["maximum"] = length - start

        for frame in range(start, length):
            detection_field = data[:, :, frame]

            new_coordinates = []
            for index, region in enumerate(previous_coordinates[-1]):
                coords = region
                new_region = tirf.track(detection_field, coords[0], coords[1])
                if new_region != "Failed":
                    tracking_data[index].fitting_params.append(new_region[2])
                    if new_region[0] < 0 or new_region[0] > shape[1]:
                        new_region[0] = previous_coordinates[-1][index][0]
                    if new_region[1] < 0 or new_region[1] > shape[0]:
                        new_region[1] = previous_coordinates[-1][index][1]
                    new_coordinates.append([new_region[0], new_region[1]])
                    tracking_data[index].coordinates.append([new_region[0], new_region[1]])
                else:
                    tracking_data[index].coordinates.append(None)
                    tracking_data[index].fitting_params.append(None)
                    new_coordinates.append(previous_coordinates[-1][index])
            previous_coordinates.append(new_coordinates)

            progress_win.progress.step()
            progress_win.progress.update()

        progress_win.handle_close()

        if self.field_selected.get() == self.field_list[0]:
            self.tracking_data_trk = trajectories
        elif self.field_selected.get() == self.field_list[1]:
            self.tracking_data_c1 = trajectories
        elif self.field_selected.get() == self.field_list[2]:
            self.tracking_data_c2 = trajectories

        try:
            self.display_tracking()
        except:
            """ Failed """
        try:
            self.display_coloc1()
        except:
            """ Failed """
        try:
            self.display_coloc2()
        except:
            """ Failed """

    def analyse_continuous(self):
        global progress_win
        self.cancel_flag = False
        self.saved_flag = False
        fitting_params = {
            "detection threshold": [float(self.detection.get())],
            "averaging distance": [float(self.averaging.get())],
            "minimum kernel residual": [float(self.residual.get())],
            "minimum sigma": [float(self.minisig.get())],
            "maximum sigma": [float(self.maxisig.get())],
            "minimum intensity": [int(self.intensity.get())],
            "minimum gauss amplitude": [int(self.amplitude.get())],
            "eccentricity threshold": [float(self.eccentric.get())],
            "minimum gauss residual": [float(self.gaussresid.get())],
        }

        maximum_displacement_threshold = float(self.max_displacement.get())
        if self.units.get() == "Nanometres (nm)":
            maximum_displacement_threshold = maximum_displacement_threshold / float(self.pixel_size.get())

        region = None
        region_type = None

        if self.field_selected.get() == self.field_list[0]:
            data = np.copy(self.proc_data_tracking)
            norm = self.global_norm_trk.get()
            if self.region_initial[0] is not None and self.region_final[0] is not None:
                region = (self.region_initial[0], self.region_final[0])
                region_type = self.invert_selection_flag_trk.get()
        elif self.field_selected.get() == self.field_list[1]:
            data = np.copy(self.proc_data_coloc1)
            norm = self.global_norm_c1.get()
            if self.region_initial[1] is not None and self.region_final[1] is not None:
                region = (self.region_initial[1], self.region_final[1])
                region_type = self.invert_selection_flag_c1.get()
        elif self.field_selected.get() == self.field_list[2]:
            data = np.copy(self.proc_data_coloc2)
            norm = self.global_norm_c2.get()
            if self.region_initial[2] is not None and self.region_final[2] is not None:
                region = (self.region_initial[2], self.region_final[2])
                region_type = self.invert_selection_flag_c2.get()


        max_frame_index_norm = np.unravel_index(np.argmax(data), data.shape)[2]
        max_frame_norm = data[:, :, max_frame_index_norm]
        maximum = np.max(max_frame_norm)
        max_frame_norm = (max_frame_norm / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
        max_frame_norm = np.clip(max_frame_norm, 0, 255)
        max_frame_norm[:, :] = self.adjust_contrast(0, 0, max_frame_norm[:, :],
                                                    self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                                    single_channel=True)

        print(f"Maximum frame index: {max_frame_index_norm}")

        if region is not None:
            if region[0][0] > region[1][0]:
                region[0][0], region[1][0] = region[1][0], region[0][0]
            if region[0][1] > region[1][1]:
                region[0][1], region[1][1] = region[1][1], region[0][1]

        tracking_data = []

        start = int(self.start_frame.get())
        stop = int(self.stop_frame.get())

        progress_win = ProgressWin("Analysing Movie", "Tracking spots...")
        progress_win.progress["maximum"] = stop - start + 1
        ax = plt.subplot()
        for frame in range(start, stop):
            if self.cancel_flag:
                progress_win.progress.stop()
                progress_win.handle_close()
                return
            fov = data[:, :, frame]

            if norm == 0:
                maximum = np.max(fov)
            elif norm == 1:
                maximum = np.max(data)
            fov = (fov / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
            fov = np.clip(fov, 0, 255)
            fov[:, :] = self.adjust_contrast(0, 0, fov[:, :],
                                             self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                             single_channel=True)
            if norm == 0:
                maximum = np.max(fov)
            elif norm == 1:
                maximum = np.max(max_frame_norm)

            fov = (fov / maximum) * 255
            # plt.imshow(fov)
            # plt.show()

            tracking_data.append(tirf.continuous_track(0, fov, fitting_params, 0, region=region, inverted=region_type))
            try:
                progress_win.progress.step(1)
                progress_win.progress.update()
            except:
                """ process was terminated """

        progress_win.progress.stop()
        progress_win.progress["mode"] = "indeterminate"
        progress_win.label["text"] = "Correlating frames - Building Trajectories..."
        progress_win.progress.start()

        trajectories = []

        # for some reason nothing gets tracked if the first frames contain no spots so we'll insert one off screen as
        # to not affect the analysis if this is the case.
        for valid in range(len(tracking_data)):
            if len(tracking_data[valid]) == 0:
                tracking_data[valid] = [[-10, -10, [1.4, 1.4, 255, 10]]]

        previous_state = tracking_data[0]

        for i in range(len(previous_state)):
            trajectories.append(TrackingData())
            trajectories[-1].coordinates.append([previous_state[i][0], previous_state[i][1]])
            trajectories[-1].fitting_params.append(previous_state[i][2])
            trajectories[-1].start_frame = start
            trajectories[-1].frame_list.append(start)

        for frame in range(1, stop - start):
            if self.cancel_flag:
                progress_win.progress.stop()
                progress_win.handle_close()
                return
            progress_win.progress.update()
            used_spot_indices = []
            for spot_index, spot in enumerate(tracking_data[frame]):
                coords = [spot[0], spot[1]]
                displacement_list = []
                indices = []

                for prev_spot_index, prev_spot in enumerate(trajectories):
                    if frame + start - prev_spot.frame_list[-1] <= int(self.max_dark_time.get()) + 1:
                        prev_coord_data = prev_spot.coordinates[-1]
                        prev_coords = [prev_coord_data[0], prev_coord_data[1]]
                        displacement = np.sqrt((coords[0] - prev_coords[0]) ** 2 + (coords[1] - prev_coords[1]) ** 2)
                        displacement_list.append(displacement)
                        indices.append(prev_spot_index)
                    else:
                        trajectories[prev_spot_index].termination_reason = "Maximum dark time exceeded."

                if len(displacement_list) > 0:
                    minimum_disp = min(displacement_list)
                    min_disp_index = displacement_list.index(minimum_disp)

                    if minimum_disp < maximum_displacement_threshold and indices[
                        min_disp_index] not in used_spot_indices:
                        trajectories[indices[min_disp_index]].coordinates.append(coords)
                        trajectories[indices[min_disp_index]].fitting_params.append(spot[2])
                        trajectories[indices[min_disp_index]].frame_list.append(frame + start)
                        used_spot_indices.append(indices[min_disp_index])

                    else:
                        trajectories.append(TrackingData())
                        trajectories[-1].coordinates.append(coords)
                        trajectories[-1].fitting_params.append(spot[2])
                        trajectories[-1].start_frame = frame + start
                        trajectories[-1].frame_list.append(frame + start)

                else:
                    """ No spots left to compare to """

        progress_win.progress.stop()
        progress_win.handle_close()

        if self.field_selected.get() == self.field_list[0]:
            self.tracking_data_trk = []
        elif self.field_selected.get() == self.field_list[1]:
            self.tracking_data_c1 = []
        elif self.field_selected.get() == self.field_list[2]:
            self.tracking_data_c2 = []

        backup = list(np.copy(trajectories))
        trajectories = []
        for i in range(len(backup)):
            if len(backup[i].coordinates) >= int(self.min_duration.get()):
                trajectories.append(backup[i])

        if self.field_selected.get() == self.field_list[0]:
            self.tracking_data_trk = trajectories
        elif self.field_selected.get() == self.field_list[1]:
            self.tracking_data_c1 = trajectories
        elif self.field_selected.get() == self.field_list[2]:
            self.tracking_data_c2 = trajectories

        try:
            self.display_tracking()
        except:
            """ Failed """
        try:
            self.display_coloc1()
        except:
            """ Failed """
        try:
            self.display_coloc2()
        except:
            """ Failed """

        self.update_trajectories()
        self.display_trajectory()
        self.include_track_button["state"] = tk.NORMAL

    def analyse_continuous_devmode(self):
        global progress_win
        self.cancel_flag = False
        self.saved_flag = False
        fitting_params = {
            "detection threshold": [float(self.detection.get())],
            "averaging distance": [float(self.averaging.get())],
            "minimum kernel residual": [float(self.residual.get())],
            "minimum sigma": [float(self.minisig.get())],
            "maximum sigma": [float(self.maxisig.get())],
            "minimum intensity": [int(self.intensity.get())],
            "minimum gauss amplitude": [int(self.amplitude.get())],
            "eccentricity threshold": [float(self.eccentric.get())],
            "minimum gauss residual": [float(self.gaussresid.get())],
        }

        maximum_displacement_threshold = float(self.max_displacement.get())
        if self.units.get() == "Nanometres (nm)":
            maximum_displacement_threshold = maximum_displacement_threshold / float(self.pixel_size.get())

        region = None
        region_type = None

        if self.field_selected.get() == self.field_list[0]:
            data = np.copy(self.proc_data_tracking)
            norm = self.global_norm_trk.get()
            if self.region_initial[0] is not None and self.region_final[0] is not None:
                region = (self.region_initial[0], self.region_final[0])
                region_type = self.invert_selection_flag_trk.get()
        elif self.field_selected.get() == self.field_list[1]:
            data = np.copy(self.proc_data_coloc1)
            norm = self.global_norm_c1.get()
            if self.region_initial[1] is not None and self.region_final[1] is not None:
                region = (self.region_initial[1], self.region_final[1])
                region_type = self.invert_selection_flag_c1.get()
        elif self.field_selected.get() == self.field_list[2]:
            data = np.copy(self.proc_data_coloc2)
            norm = self.global_norm_c2.get()
            if self.region_initial[2] is not None and self.region_final[2] is not None:
                region = (self.region_initial[2], self.region_final[2])
                region_type = self.invert_selection_flag_c2.get()


        max_frame_index_norm = np.unravel_index(np.argmax(data), data.shape)[2]
        max_frame_norm = data[:, :, max_frame_index_norm]
        maximum = np.max(max_frame_norm)
        max_frame_norm = (max_frame_norm / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
        max_frame_norm = np.clip(max_frame_norm, 0, 255)
        max_frame_norm[:, :] = self.adjust_contrast(0, 0, max_frame_norm[:, :],
                                                    self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                                    single_channel=True)

        print(f"Maximum frame index: {max_frame_index_norm}")

        if region is not None:
            if region[0][0] > region[1][0]:
                region[0][0], region[1][0] = region[1][0], region[0][0]
            if region[0][1] > region[1][1]:
                region[0][1], region[1][1] = region[1][1], region[0][1]

        tracking_data = []

        start = int(self.start_frame.get())
        stop = int(self.stop_frame.get())

        progress_win = ProgressWin("Analysing Movie", "Tracking spots...")
        progress_win.progress["maximum"] = stop - start + 1
        ax = plt.subplot()
        for frame in range(start, stop):
            if self.cancel_flag:
                progress_win.progress.stop()
                progress_win.handle_close()
                return
            fov = data[:, :, frame]

            if norm == 0:
                maximum = np.max(fov)
            elif norm == 1:
                maximum = np.max(data)
            fov = (fov / maximum) * 255 * self.display_brightness[self.field_dict[self.field_selected.get()]].get() / 20
            fov = np.clip(fov, 0, 255)
            fov[:, :] = self.adjust_contrast(0, 0, fov[:, :],
                                             self.display_contrast[self.field_dict[self.field_selected.get()]].get(),
                                             single_channel=True)
            if norm == 0:
                maximum = np.max(fov)
            elif norm == 1:
                maximum = np.max(max_frame_norm)

            fov = (fov / maximum) * 255
            # plt.imshow(fov)
            # plt.show()

            tracking_data.append(tirf.continuous_track(0, fov, fitting_params, 0, region=region, inverted=region_type))
            try:
                progress_win.progress.step(1)
                progress_win.progress.update()
            except:
                """ process was terminated """

        progress_win.progress.stop()
        progress_win.progress["mode"] = "indeterminate"
        progress_win.label["text"] = "Correlating frames - Building Trajectories..."
        progress_win.progress.start()

        trajectories = []

        # for some reason nothing gets tracked if the first frames contain no spots so we'll insert one off screen as
        # to not affect the analysis if this is the case.
        for valid in range(len(tracking_data)):
            if len(tracking_data[valid]) == 0:
                tracking_data[valid] = [[-10, -10, [1.4, 1.4, 255, 10]]]

        previous_state = tracking_data[0]

        for i in range(len(previous_state)):
            trajectories.append(TrackingData())
            trajectories[-1].coordinates.append([previous_state[i][0], previous_state[i][1]])
            trajectories[-1].fitting_params.append(previous_state[i][2])
            trajectories[-1].start_frame = start
            trajectories[-1].frame_list.append(start)

        for frame in range(1, stop - start):
            if self.cancel_flag:
                progress_win.progress.stop()
                progress_win.handle_close()
                return
            progress_win.progress.update()
            used_spot_indices = []
            deletion_indices = []
            for spot_index, spot in enumerate(tracking_data[frame]):
                coords = [spot[0], spot[1]]
                displacement_list = []
                indices = []

                for prev_spot_index, prev_spot in enumerate(trajectories):
                    if frame + start - prev_spot.frame_list[-1] <= int(self.max_dark_time.get()) + 1:
                        prev_coord_data = prev_spot.coordinates[-1]
                        prev_coords = [prev_coord_data[0], prev_coord_data[1]]
                        displacement = np.sqrt((coords[0] - prev_coords[0]) ** 2 + (coords[1] - prev_coords[1]) ** 2)
                        displacement_list.append(displacement)
                        indices.append(prev_spot_index)
                    else:
                        trajectories[prev_spot_index].termination_reason = "Maximum dark time exceeded."


                # this logic is why blank frames stop new tracks being added
                if len(displacement_list) > 0:
                    minimum_disp = min(displacement_list)
                    min_disp_index = displacement_list.index(minimum_disp)

                    if minimum_disp < maximum_displacement_threshold and indices[
                        min_disp_index] not in used_spot_indices:
                        trajectories[indices[min_disp_index]].coordinates.append(coords)
                        trajectories[indices[min_disp_index]].fitting_params.append(spot[2])
                        trajectories[indices[min_disp_index]].frame_list.append(frame + start)
                        used_spot_indices.append(indices[min_disp_index])

                    # because to add a new track, the displacement list can't be empty
                    else:
                        trajectories.append(TrackingData())
                        trajectories[-1].coordinates.append(coords)
                        trajectories[-1].fitting_params.append(spot[2])
                        trajectories[-1].start_frame = frame + start
                        trajectories[-1].frame_list.append(frame + start)

                else:
                    trajectories.append(TrackingData())
                    trajectories[-1].coordinates.append(coords)
                    trajectories[-1].fitting_params.append(spot[2])
                    trajectories[-1].start_frame = frame + start
                    trajectories[-1].frame_list.append(frame + start)

            for prev_spot_index, prev_spot in enumerate(trajectories):
                if (frame + start) - prev_spot.frame_list[-1] >= int(self.max_dark_time.get()) + 1:
                    deletion_indices.append(prev_spot_index)

            print(len(trajectories))
            print(deletion_indices)
            print("\n")
            deletion_indices = list(set(deletion_indices))
            print(deletion_indices)
            print(len(deletion_indices))
            input("\nNext?\n")


            for idx in deletion_indices:
                del trajectories[idx]

        progress_win.progress.stop()
        progress_win.handle_close()

        if self.field_selected.get() == self.field_list[0]:
            self.tracking_data_trk = []
        elif self.field_selected.get() == self.field_list[1]:
            self.tracking_data_c1 = []
        elif self.field_selected.get() == self.field_list[2]:
            self.tracking_data_c2 = []

        backup = list(np.copy(trajectories))
        trajectories = []
        for i in range(len(backup)):
            if len(backup[i].coordinates) >= int(self.min_duration.get()):
                trajectories.append(backup[i])

        if self.field_selected.get() == self.field_list[0]:
            self.tracking_data_trk = trajectories
        elif self.field_selected.get() == self.field_list[1]:
            self.tracking_data_c1 = trajectories
        elif self.field_selected.get() == self.field_list[2]:
            self.tracking_data_c2 = trajectories

        try:
            self.display_tracking()
        except:
            """ Failed """
        try:
            self.display_coloc1()
        except:
            """ Failed """
        try:
            self.display_coloc2()
        except:
            """ Failed """

        self.update_trajectories()
        self.display_trajectory()
        self.include_track_button["state"] = tk.NORMAL

    def cancel_tracking(self):
        self.cancel_flag = True

    def quick_scroll(self):
        self.current_trajectory.set(self.quick_current.get())

    def display_trajectory(self):
        try:
            name = self.current_trajectory.get()
            index = self.trajectory_list.index(name)
            if self.field_selected.get() == self.field_list[0]:
                data = self.tracking_data_trk
            elif self.field_selected.get() == self.field_list[1]:
                data = self.tracking_data_c1
            elif self.field_selected.get() == self.field_list[2]:
                data = self.tracking_data_c2

            traj = data[index]
            x = [c[0] for c in traj.coordinates]
            y = [c[1] for c in traj.coordinates]

            start_frame = traj.start_frame
            if self.field_selected.get() == self.field_list[0]:
                self.slider_trk.set(start_frame)
                self.display_tracking()
            elif self.field_selected.get() == self.field_list[1]:
                self.slider_c1.set(start_frame)
                self.display_coloc1()
            elif self.field_selected.get() == self.field_list[2]:
                self.slider_c2.set(start_frame)
                self.display_coloc2()

            nm_scale = float(self.pixel_size.get())

            for i in range(len(x)):
                x[i] = x[i] * nm_scale
                y[i] = y[i] * nm_scale

            min_x, min_y = np.min(x), np.min(y)
            for i in range(len(x)):
                x[i] = x[i] - min_x
                y[i] = y[i] - min_y

            self.plot_trajectory_figure.clf()
            self.plot_trajectory_canvas.draw()

            if self.dark_mode.get() == 1:
                bg_col = "#222222"
                box_col = "blue"
                text_col = "white"
                grid_col = "#333333"
                line_col = "lightgray"
            else:
                bg_col = "white"
                box_col = "black"
                text_col = "black"
                grid_col = "#dddddd"
                line_col = "black"

            self.plot_trajectory_figure.set_facecolor(bg_col)
            self.plot_trajectory_subfig = self.plot_trajectory_figure.add_subplot(111)
            self.plot_trajectory_subfig.set_facecolor(bg_col)
            self.plot_trajectory_subfig.spines['bottom'].set_color(box_col)
            self.plot_trajectory_subfig.spines['top'].set_color(box_col)
            self.plot_trajectory_subfig.spines['left'].set_color(box_col)
            self.plot_trajectory_subfig.spines['right'].set_color(box_col)
            self.plot_trajectory_subfig.xaxis.label.set_color(text_col)
            self.plot_trajectory_subfig.yaxis.label.set_color(text_col)
            self.plot_trajectory_subfig.tick_params(axis='x', colors=text_col, labelsize=self.g_fonts["axis"])
            self.plot_trajectory_subfig.tick_params(axis='y', colors=text_col, labelsize=self.g_fonts["axis"])
            if self.plot_grids.get() == 1:
                self.plot_trajectory_subfig.grid(color=grid_col)

            self.plot_trajectory_subfig.plot(x, y, linewidth=1, color=line_col)
            self.plot_trajectory_subfig.text(x[0], y[0], "Start", size=9, color="green")
            self.plot_trajectory_subfig.text(x[-1], y[-1], "End", size=9, color="red")
            self.plot_trajectory_subfig.axis("square")
            self.plot_trajectory_subfig.set_ylim(max(self.plot_trajectory_subfig.get_ylim()),
                                                 min(self.plot_trajectory_subfig.get_ylim()))

            self.plot_trajectory_subfig.set_xlabel("x displacement (nm)", color=text_col, size=self.g_fonts["labels"])
            self.plot_trajectory_subfig.set_ylabel("y displacement (nm)", color=text_col, size=self.g_fonts["labels"])
            self.plot_trajectory_subfig.set_title(name, size=self.g_fonts["title"], color=text_col)
            self.plot_trajectory_canvas.draw()
            try:
                self.calculate_trajectory_info(traj)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ Error calculating info """
        except ValueError:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            self.plot_trajectory_figure.clf()
            self.plot_trajectory_canvas.draw()
            self.trajectory_info["text"] = ""
            self.trajectory_info.update()

    def calculate_trajectory_info(self, trajectory):

        if trajectory.include_export:
            self.include_track_var.set(1)
            self.include_track_button.update()
        else:
            self.include_track_var.set(0)
            self.include_track_button.update()

        self.msd_proportion.set(trajectory.MSD_prop)

        name = self.current_trajectory.get()
        term_reason = trajectory.termination_reason

        parameters = trajectory.fitting_params
        coordinates = trajectory.coordinates
        frames = trajectory.frame_list

        sigma_x = [p[0] for p in parameters]
        sigma_y = [p[1] for p in parameters]
        amp = [p[2] for p in parameters]
        residual = [p[3] for p in parameters]

        ecc_x = []
        ecc_y = []
        for i in range(len(sigma_x)):
            ecc_x.append(sigma_x[i] / sigma_y[i])
            ecc_y.append(sigma_y[i] / sigma_x[i])

        start = frames[0]
        end = frames[-1]

        displacement = []
        for i in range(len(coordinates) - 1):
            disp = np.sqrt(
                (coordinates[i + 1][0] - coordinates[i][0]) ** 2 + (coordinates[i + 1][1] - coordinates[i][1]) ** 2)
            displacement.append(disp * float(self.pixel_size.get()))
        mean_disp = sum(displacement) / len(displacement)
        min_disp = min(displacement)
        max_disp = max(displacement)
        min_sigma_x = min(sigma_x)
        min_sigma_y = min(sigma_y)
        max_sigma_x = max(sigma_x)
        max_sigma_y = max(sigma_y)
        min_amp, max_residual = min(amp), max(residual)
        min_ecc_x, min_ecc_y, max_ecc_x, max_ecc_y = min(ecc_x), min(ecc_y), max(ecc_x), max(ecc_y)
        total_dist = np.sum(displacement)
        total_displacement = np.sqrt((coordinates[0][0] - coordinates[-1][0]) ** 2 +
                                     (coordinates[0][1] - coordinates[-1][1]) ** 2) * float(self.pixel_size.get())


        frame_times = np.array(frames[1:]) * float(self.frame_interval.get())
        first = frame_times[0]
        frame_times = frame_times - first

        new_MSD = self.calculate_msd(coordinates)
        length = int((float(self.msd_proportion.get()) / 100) * len(new_MSD))
        if length > 2:
            gradient, intercept = linear_regression(frame_times[1:length], new_MSD[:length])
        else:
            gradient, intercept = 0, 0

        DFC_series = []
        try:
            for get_diff_std in range(5, int(float(self.msd_proportion.get())) + 1):
                f_length = int((get_diff_std / 100) * len(new_MSD))
                if f_length > 2:
                    gradient, intercept = linear_regression(frame_times[1:f_length], new_MSD[:f_length])
                else:
                    gradient, intercept = 0, 0
                if f_length > 1:
                    DFC_series.append(gradient / 4)
            np_DFC = np.array(DFC_series)
            np_DFC_nonzero = np_DFC[np_DFC != 0]
            DFC_stddev = np.std(np_DFC_nonzero, ddof=1)
            DFC_stddev = round(DFC_stddev, 6)
            if len(np_DFC_nonzero) == 0:
                DFC_stddev = "N/A"
        except:
            print(traceback.format_exc())
            DFC_stddev = "N/A"

        dark_times = []
        for i in range(len(frames) - 1):
            if frames[i + 1] - frames[i] > 1:
                dark_times.append(frames[i + 1] - frames[i])

        diffusion_coefficient = gradient / 4

        if len(dark_times) > 0:
            max_dark_time = max(dark_times) - 1
        else:
            max_dark_time = 0
        blink_count = len(dark_times)

        try:
            diff_coeff_percent_error = round((DFC_stddev*int(self.msd_uncertainty_degree.get())) / (diffusion_coefficient) * 100, 3)
        except:
            print(traceback.format_exc())
            diff_coeff_percent_error = "N/A"

        label_text = f"Name: {name}\nFirst Frame: {start}\nLast Frame: {end}\nTermination Reason: {term_reason}\n" \
                     f"Mean Displacement: {round(mean_disp, 2)} nm / frame\nMin. Displacement: {round(min_disp, 2)} nm / frame\nMax. Displacement: " \
                     f"{round(max_disp, 2)} nm / frame\nMax. Dark Duration: {max_dark_time} frames\nBlink Count: {blink_count}\n" \
                     f"Min. Gaussian Sigma (x,y): {round(min_sigma_x, 2)}, {round(min_sigma_y, 2)}\n" \
                     f"Max. Gaussian Sigma (x,y): {round(max_sigma_x, 2)}, {round(max_sigma_x, 2)}\n" \
                     f"Min. Gaussian Amplitude: {round(min_amp, 1)}\nMax. Gaussian Residual: {round(max_residual, 2)}\n" \
                     f"Min. Eccentricity (x/y, y/x): {round(min_ecc_x, 2)}, {round(min_ecc_y, 2)}\n" \
                     f"Max. Eccentricity (x/y, y/x): {round(max_ecc_x, 2)}, {round(max_ecc_x, 2)}\nEnd-to-End Displacement: " \
                     f"{round(total_displacement, 2)} nm\nTotal Distance Covered: {round(total_dist, 2)} nm\n" \
                     f"Diffusion Coeffecient: {round(diffusion_coefficient, 5)} um^2 / s ± {diff_coeff_percent_error}% ({int(self.msd_uncertainty_degree.get())} std.dev.)\n" \
                     f"MSD Fit Solver Diff. Coeff. Std.Dev: {DFC_stddev} um^2 / s"
        self.trajectory_info["text"] = label_text
        self.trajectory_info.update()

        self.current_trajectory_series = TrajectorySeries()
        self.current_trajectory_series.name = name
        self.current_trajectory_series.frame_list = frames
        self.current_trajectory_series.coordinates = coordinates
        self.current_trajectory_series.sigma_x = sigma_x
        self.current_trajectory_series.sigma_y = sigma_y
        self.current_trajectory_series.amplitude = amp
        self.current_trajectory_series.residual = residual
        self.current_trajectory_series.eccentricity_x = ecc_x
        self.current_trajectory_series.eccentricity_y = ecc_y
        self.current_trajectory_series.displacement = displacement
        self.current_trajectory_series.msd = new_MSD
        self.current_trajectory_series.offset_time = frame_times

        if self.raw_data_tracking is not None:
            self.current_trajectory_series.trace = self.calculate_trace(self.raw_data_tracking, trajectory)
        if self.raw_data_coloc1 is not None:
            self.current_trajectory_series.trace_coloc1 = self.calculate_trace(self.raw_data_coloc1, trajectory)
        if self.raw_data_coloc2 is not None:
            self.current_trajectory_series.trace_coloc2 = self.calculate_trace(self.raw_data_coloc2, trajectory)

    def open_r2_win(self):
        global r2_win
        r2_win = RSquaredWin()

    def msd_test_plot(self):
        name = self.current_trajectory.get()
        index = self.trajectory_list.index(name)
        if self.field_selected.get() == self.field_list[0]:
            data = self.tracking_data_trk
        elif self.field_selected.get() == self.field_list[1]:
            data = self.tracking_data_c1
        elif self.field_selected.get() == self.field_list[2]:
            data = self.tracking_data_c2

        traj = data[index]

        coords = traj.coordinates
        frames = traj.frame_list
        frame_times = np.array(frames[1:]) * float(self.frame_interval.get())
        msd = self.calculate_msd(coords)

        R_squared_values = []
        DFC_values = []
        props = []
        for proportion in range(5, 100):
            props.append(proportion)

            length = int((proportion / 100) * len(msd))
            if length > 2:
                gradient, intercept = linear_regression(frame_times[1:length], msd[:length])
            else:
                gradient, intercept = 0, 0

            msd_fit = []
            for t_fit in range(len(frame_times)):
                msd_fit.append(gradient * frame_times[t_fit] + intercept)

            DFC_values.append(gradient / 4)

            if length > 2:
                r2 = r2_score(msd_fit[:length], msd[:length])
            else:
                r2 = 0

            R_squared_values.append(r2)

        r2_derivative = np.gradient(R_squared_values, props)

        return props, R_squared_values, r2_derivative, DFC_values, traj.MSD_prop

    def msd_autofit(self):
        global progress_win
        if self.field_selected.get() == self.field_list[0]:
            data = self.tracking_data_trk
        elif self.field_selected.get() == self.field_list[1]:
            data = self.tracking_data_c1
        elif self.field_selected.get() == self.field_list[2]:
            data = self.tracking_data_c2

        progress_win = ProgressWin("MSD AutoFit", "Finding optimal MSD fit...")
        progress_win.cancel_button.place_forget()
        progress_win.progress["maximum"] = len(data) + 1

        for index, trajectory in enumerate(data):
            progress_win.label["text"] = f"Finding optimal MSD fit ({index + 1} / {len(data)})..."
            progress_win.progress.step(1)
            progress_win.window.update()
            trajectory.include_export = True

            try:
                coords = trajectory.coordinates
                frames = trajectory.frame_list
                frame_times = np.array(frames[1:]) * float(self.frame_interval.get())
                msd = self.calculate_msd(coords)

                R_squared_values = []
                DFC_values = []
                props = []
                for proportion in range(int(self.minimum_msd_prop.get()), 100):
                    props.append(proportion)

                    length = int((proportion / 100) * len(msd))
                    if length > 2:
                        gradient, intercept = linear_regression(frame_times[1:length], msd[:length])
                    else:
                        gradient, intercept = 0, 0

                    msd_fit = []
                    for t_fit in range(len(frame_times)):
                        msd_fit.append(gradient * frame_times[t_fit] + intercept)

                    DFC_values.append(gradient / 4)

                    if length > 2:
                        r2 = r2_score(msd_fit[:length], msd[:length])
                    else:
                        r2 = 0

                    R_squared_values.append(r2)

                minimum_point = np.min(R_squared_values)
                minimum_idx = R_squared_values.index(minimum_point)
                minimum_prop = props[minimum_idx]



                length = int((minimum_prop / 100) * len(msd))
                if length > 2:
                    gradient, intercept = linear_regression(frame_times[1:length], msd[:length])
                else:
                    gradient, intercept = 0, 0

                r2_derivative = np.gradient(R_squared_values, props)

                final_idx = len(r2_derivative) - 1
                for prop_idx in range(len(r2_derivative)):
                    length = int((props[prop_idx] / 100) * len(msd))
                    if r2_derivative[prop_idx] != 0:
                        if r2_derivative[prop_idx] < float(self.msd_gradient_threshold_r2.get()) and length > int(self.minimum_msd_timepoints.get()):
                            final_idx = prop_idx - 1
                            break

                final_fit_point = props[final_idx]
                final_diff_coeff = DFC_values[final_fit_point]

                mean_diff_coeff_before_minimum = np.mean(DFC_values[:minimum_idx])

                try:
                    int(self.msd_stationary_r2_threshold.get())
                except:
                    easygui.msgbox(title="Error!", msg="Min. Diff. Contains an invalid quantity!")
                    return

                if (minimum_point < int(self.msd_stationary_r2_threshold.get()) and mean_diff_coeff_before_minimum < float(self.msd_min_diff.get())) or final_diff_coeff < float(self.msd_min_diff.get()):
                    final_fit_point = minimum_prop
                    final_idx = minimum_idx
                    print(f"Trajectory {index + 1} is stationary")
                    if self.exclude_stationary.get() == 1:
                        trajectory.include_export = False
                    if trajectory.termination_reason is not None:
                        if "stationary" not in trajectory.termination_reason:
                            trajectory.termination_reason += " Spot may be stationary."
                    else:
                        trajectory.termination_reason = "Spot may be stationary."

                trajectory.MSD_prop = final_fit_point

                print(f"Trajectory: {index + 1}, minimum = {minimum_point}")

                if self.exclude_based_on_error.get() == 1:
                    diffusion_coefficient = DFC_values[final_idx]
                    DFC_series = DFC_values[:final_idx+1]
                    np_DFC = np.array(DFC_series)
                    np_DFC_nonzero = np_DFC[np_DFC != 0]
                    DFC_stddev = np.std(np_DFC_nonzero, ddof=1)
                    diff_coeff_percent_error = round((DFC_stddev * int(self.msd_uncertainty_degree.get())) / (diffusion_coefficient) * 100, 5)
                    print(round(diffusion_coefficient, 5), diff_coeff_percent_error)
                    if diffusion_coefficient == 0:
                        trajectory.include_export = False
                    if abs(diff_coeff_percent_error) > int(self.msd_maximum_error.get()):
                        trajectory.include_export = False

                # Negative diffusion coefficients can't exist:
                diffusion_coefficient = DFC_values[final_idx]
                if diffusion_coefficient < 0:
                    trajectory.include_export = False


            except:
                print("Optimization failed on this track")
                trajectory.include_export = False

        progress_win.progress.stop()
        progress_win.handle_close()
        progress_win = None
        self.display_trajectory()
        self.saved_flag = False


    def calculate_msd(self, position):
        cumulative_MSD = []
        for time_step in range(1, len(position)-1):
            sumSD = 0
            count = 0
            for index in range(len(position)):
                try:
                    pos1 = position[index]
                    pos2 = position[index + time_step]
                except IndexError:
                    break
                pix_displacement = np.sqrt((pos2[0] - pos1[0])**2 + (pos2[1] - pos1[1])**2)
                nm_displacement = (pix_displacement * float(self.pixel_size.get()))
                sumSD += (nm_displacement / 1000)**2
                count += 1
            try:
                MSD = sumSD / (len(position) - time_step)
                # print(count, len(position) - time_step)
                cumulative_MSD.append(MSD)
            except ZeroDivisionError:
                break
        return cumulative_MSD

    def export_to_xl(self, auto=False, xlsx_obj=None, xlsx_path=None, trcs_path=None):
        global progress_win

        if not auto:
            export_template_path = easygui.filesavebox(msg="Save excel file for export",
                                                       default=default_dir + "*.xlsx")
            if not export_template_path:
                return
            if export_template_path[-5:] != ".xlsx":
                export_template_path += ".xlsx"
        if self.field_selected.get() == self.field_list[0]:
            data = np.copy(self.tracking_data_trk)
        elif self.field_selected.get() == self.field_list[1]:
            data = np.copy(self.tracking_data_c1)
        elif self.field_selected.get() == self.field_list[2]:
            data = np.copy(self.tracking_data_c2)

        if len(data) == 0:
            easygui.msgbox(title="Error!", msg="No data to export!")
            return

        progress_win = ProgressWin("Export in progress", "Exporting data to excel...")
        progress_win.cancel_button.place_forget()
        progress_win.progress["maximum"] = len(data)

        if not auto:
            export_template = self.create_blank_template()
        else:
            export_template = xlsx_obj
        raw_sheet = export_template["Raw Data"]
        msd_sheet = export_template["MSD"]

        column = 1
        column2 = 2
        row2 = 5

        tr_info = []
        all_traces = []
        all_fits = []

        for export in range(len(data)):

            try:
                progress_win.progress.step(1)
                progress_win.progress.update()
            except:
                """ process was terminated """

            trajectory = data[export]

            if not trajectory.include_export:
                print("Trajectory " + str(export + 1) + " is not selected for export. Skipping...")
                continue

            term_reason = trajectory.termination_reason

            parameters = trajectory.fitting_params
            coordinates = trajectory.coordinates
            frames = trajectory.frame_list

            sigma_x = [p[0] for p in parameters]
            sigma_y = [p[1] for p in parameters]
            amp = [p[2] for p in parameters]
            residual = [p[3] for p in parameters]

            ecc_x = []
            ecc_y = []
            for i in range(len(sigma_x)):
                ecc_x.append(sigma_x[i] / sigma_y[i])
                ecc_y.append(sigma_y[i] / sigma_x[i])

            start = frames[0]
            end = frames[-1]
            length = end - start
            length = length * float(self.frame_interval.get())
            bleaching_time = length

            displacement = []
            for i in range(len(coordinates) - 1):
                disp = np.sqrt(
                    (coordinates[i + 1][0] - coordinates[i][0]) ** 2 + (coordinates[i + 1][1] - coordinates[i][1]) ** 2)
                displacement.append(disp * float(self.pixel_size.get()))
            mean_disp = sum(displacement) / len(displacement)
            min_disp = min(displacement)
            max_disp = max(displacement)
            min_sigma_x = min(sigma_x)
            min_sigma_y = min(sigma_y)
            max_sigma_x = max(sigma_x)
            max_sigma_y = max(sigma_y)
            min_amp, max_residual = min(amp), max(residual)
            min_ecc_x, min_ecc_y, max_ecc_x, max_ecc_y = min(ecc_x), min(ecc_y), max(ecc_x), max(ecc_y)
            total_dist = np.sum(displacement)
            total_displacement = np.sqrt((coordinates[0][0] - coordinates[-1][0]) ** 2 +
                                         (coordinates[0][1] - coordinates[-1][1]) ** 2) * float(self.pixel_size.get())


            frame_times = np.array(frames[1:]) * float(self.frame_interval.get())
            first = frame_times[0]
            frame_times = frame_times - first

            new_MSD = self.calculate_msd(coordinates)
            length = int((float(trajectory.MSD_prop) / 100) * len(new_MSD))
            if length > 2:
                gradient, intercept = linear_regression(frame_times[1:length], new_MSD[:length])
            else:
                gradient, intercept = 0, 0
            dark_times = []
            for i in range(len(frames) - 1):
                if frames[i + 1] - frames[i] > 1:
                    dark_times.append(frames[i + 1] - frames[i])

            diffusion_coefficient = gradient / 4

            if len(dark_times) > 0:
                max_dark_time = max(dark_times) - 1
            else:
                max_dark_time = 0
            blink_count = len(dark_times)

            DFC_series = []
            try:
                for get_diff_std in range(5, int(float(self.msd_proportion.get())) + 1):
                    f_length = int((get_diff_std / 100) * len(new_MSD))
                    if f_length > 2:
                        gradient, intercept = linear_regression(frame_times[1:f_length], new_MSD[:f_length])
                    else:
                        gradient, intercept = 0, 0
                    if f_length > 1:
                        DFC_series.append(gradient / 4)
                np_DFC = np.array(DFC_series)
                np_DFC_nonzero = np_DFC[np_DFC != 0]
                print(np_DFC_nonzero)
                DFC_stddev = round(np.std(np_DFC_nonzero, ddof=1), 8)
                if len(np_DFC_nonzero) == 0:
                    DFC_stddev = "N/A"
            except:
                print(traceback.format_exc())
                DFC_stddev = "N/A"


            raw_sheet.cell(row=1, column=column).value = "Name:"
            raw_sheet.cell(row=1, column=column + 1).value = "Path " + str(export + 1)
            raw_sheet.cell(row=2, column=column).value = "First frame:"
            raw_sheet.cell(row=2, column=column + 1).value = start
            raw_sheet.cell(row=3, column=column).value = "Last frame:"
            raw_sheet.cell(row=3, column=column + 1).value = end
            raw_sheet.cell(row=4, column=column).value = "Mean displacement (um):"
            raw_sheet.cell(row=4, column=column + 1).value = round(mean_disp / 1000, 6)
            raw_sheet.cell(row=5, column=column).value = "Min. displacement (um):"
            raw_sheet.cell(row=5, column=column + 1).value = round(min_disp / 1000, 6)
            raw_sheet.cell(row=6, column=column).value = "Max. displacement (um):"
            raw_sheet.cell(row=6, column=column + 1).value = round(max_disp / 1000, 6)
            raw_sheet.cell(row=7, column=column).value = "Max. dark time (frames):"
            raw_sheet.cell(row=7, column=column + 1).value = max_dark_time
            raw_sheet.cell(row=8, column=column).value = "Blink count:"
            raw_sheet.cell(row=8, column=column + 1).value = blink_count
            raw_sheet.cell(row=9, column=column).value = "Min. Gauss. sigma [x, y] (pixels):"
            raw_sheet.cell(row=9, column=column + 1).value = round(min_sigma_x, 3)
            raw_sheet.cell(row=9, column=column + 2).value = round(min_sigma_y, 3)
            raw_sheet.cell(row=10, column=column).value = "Max. Gauss. sigma [x, y] (pixels):"
            raw_sheet.cell(row=10, column=column + 1).value = round(max_sigma_x, 3)
            raw_sheet.cell(row=10, column=column + 2).value = round(max_sigma_y, 3)
            raw_sheet.cell(row=11, column=column).value = "Min. Gauss. amplitude:"
            raw_sheet.cell(row=11, column=column + 1).value = round(min_amp, 6)
            raw_sheet.cell(row=12, column=column).value = "Max. Gauss. residual:"
            raw_sheet.cell(row=12, column=column + 1).value = round(max_residual, 6)
            raw_sheet.cell(row=13, column=column).value = "Min. Eccentricity (x/y, y/x):"
            raw_sheet.cell(row=13, column=column + 1).value = round(min_ecc_x, 3)
            raw_sheet.cell(row=13, column=column + 2).value = round(min_ecc_y, 3)
            raw_sheet.cell(row=14, column=column).value = "Max. Eccentricity (x/y, y/x):"
            raw_sheet.cell(row=14, column=column + 1).value = round(max_ecc_x, 3)
            raw_sheet.cell(row=14, column=column + 2).value = round(max_ecc_y, 3)
            raw_sheet.cell(row=15, column=column).value = "End-to-end displacement (um):"
            raw_sheet.cell(row=15, column=column + 1).value = round(total_displacement / 1000, 6)
            raw_sheet.cell(row=16, column=column).value = "Total 'taught' path length (um):"
            raw_sheet.cell(row=16, column=column + 1).value = round(total_dist / 1000, 6)
            raw_sheet.cell(row=17, column=column).value = "MSD fitting percentage:"
            raw_sheet.cell(row=17, column=column + 1).value = trajectory.MSD_prop
            raw_sheet.cell(row=18, column=column).value = "Diffusion Coefficient (um^2/s):"
            raw_sheet.cell(row=18, column=column + 1).value = round(diffusion_coefficient, 6)
            raw_sheet.cell(row=19, column=column).value = "Std. Dev. (um^2/s):"
            raw_sheet.cell(row=19, column=column + 1).value = DFC_stddev

            msd_sheet.cell(row=1, column=column2).value = "Path " + str(export + 1)
            msd_sheet.cell(row=2, column=column2).value = round(diffusion_coefficient, 6)
            msd_sheet.cell(row=3, column=column2).value = bleaching_time

            try:
                if self.raw_data_tracking is not None:
                    trace = self.calculate_trace(self.raw_data_tracking, trajectory)
                if self.raw_data_coloc1 is not None:
                    trace_coloc1 = self.calculate_trace(self.raw_data_coloc1, trajectory)
                if self.raw_data_coloc2 is not None:
                    trace_coloc2 = self.calculate_trace(self.raw_data_coloc2, trajectory)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                print("error occurred while calculating trace: " + str(export + 1))
                trace = [0, 0]

            row = 22
            raw_sheet.cell(row=row - 1, column=column).value = "Frame"
            raw_sheet.cell(row=row - 1, column=column + 1).value = "Time (s)"
            raw_sheet.cell(row=row - 1, column=column + 2).value = "x coord. (px)"
            raw_sheet.cell(row=row - 1, column=column + 3).value = "y coord. (px)"
            raw_sheet.cell(row=row - 1, column=column + 4).value = "displacement (um)"
            raw_sheet.cell(row=row - 1, column=column + 5).value = "msd (um)"
            raw_sheet.cell(row=row - 1, column=column + 6).value = "trace (A.U.)"

            for time_point in range(len(frames)):
                try:
                    raw_sheet.cell(row=row + time_point, column=column).value = frames[time_point]
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point, column=column + 1).value = frame_times[time_point]
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point, column=column + 2).value = coordinates[time_point][0]
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point, column=column + 3).value = coordinates[time_point][1]
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point + 1, column=column + 4).value = round(
                        displacement[time_point] / 1000, 5)
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point + 1, column=column + 5).value = new_MSD[time_point]
                except:
                    """ Failed! """

                try:
                    raw_sheet.cell(row=row + time_point, column=column + 6).value = trace[time_point]
                except:
                    """ Failed! """

                try:
                    msd_sheet.cell(row=row2 + time_point, column=column2).value = new_MSD[time_point]
                except:
                    """ Failed! """

            tr_info.append(["Cyanine 5", 0, len(trace), export + 1, [], None, None, None, []])
            all_traces.append(trace)
            fit = []
            for i in range(len(trace)):
                fit.append(0)
            all_fits.append(fit)

            column += 10
            column2 += 1
        if not auto:
            try:
                export_template.save(export_template_path)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                easygui.msgbox(title="Error!", msg="Warning! an error occurred while attempting to save experiment:\n\n" + str(traceback.format_exc()))
                try:
                    progress_win.progress.step(1)
                    progress_win.progress.stop()
                    progress_win.handle_close()
                except:
                    """ No progress bar window open """
        else:
            try:
                export_template.save(xlsx_path)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                easygui.msgbox(title="Error!", msg="Warning! an error occurred while attempting to save experiment:\n\n" + str(traceback.format_exc()))
                try:
                    progress_win.progress.step(1)
                    progress_win.progress.stop()
                    progress_win.handle_close()
                except:
                    """ No progress bar window open """
        try:
            progress_win.progress.stop()
            progress_win.handle_close()
        except:
            """ Already closed """

        if not auto:
            save_path = easygui.filesavebox(title="Save traces to data file", default="*.trcs")
            if not save_path:
                return
            if save_path[-5:] != ".trcs":
                save_path = save_path + ".trcs"
            with open(save_path, "wb") as save_trcs:
                pickle.dump(tr_info, save_trcs)
                pickle.dump(all_traces, save_trcs)
                pickle.dump(all_fits, save_trcs)
            self.save_document()
        if auto:
            try:
                if auto_track_win.save_traces.get() == 1:
                    with open(trcs_path, "wb") as save_trcs:
                        pickle.dump(tr_info, save_trcs)
                        pickle.dump(all_traces, save_trcs)
                        pickle.dump(all_fits, save_trcs)
            except:
                """ Failed to save traces """
    @staticmethod
    def create_blank_template():
        wb = opxl.Workbook()
        ws = wb.active
        ws.title = "Raw Data"
        ws2 = wb.create_sheet("MSD")
        ws2.title = "MSD"
        ws2.cell(row=1, column=1).value = "name"
        ws2.cell(row=2, column=1).value = "diffusion coeff"
        ws2.cell(row=3, column=1).value = "bleaching time (s)"
        ws2.cell(row=4, column=2).value = "msd cumulative"
        return wb

    def toggle_button_state(self, mode):
        if self.current_field_length != 1:
            if not self.ready_for_evolve and mode == "Evolve Initial State":
                self.start_tracking_button["state"] = tk.DISABLED
            elif not self.ready_for_evolve and mode == "Continuous":
                self.start_tracking_button["state"] = tk.NORMAL

            if self.ready_for_evolve:
                self.start_tracking_button["state"] = tk.NORMAL
        else:
            self.start_tracking_button["state"] = tk.DISABLED

    def update_units(self, new_unit):
        unit_key = {
            "Pixels (px)": "px",
            "Nanometres (nm)": "nm",
        }

        self.unitl1["text"] = unit_key[new_unit]
        self.unitl1.update()

        if unit_key[new_unit] == "nm":
            quant = float(self.max_displacement.get())
            quant = quant * float(self.pixel_size.get())
            self.max_displacement.delete(0, tk.END)
            self.max_displacement.insert(tk.END, round(quant, 1))
        else:
            quant = float(self.max_displacement.get())
            quant = quant / float(self.pixel_size.get())
            self.max_displacement.delete(0, tk.END)
            self.max_displacement.insert(tk.END, round(quant, 1))

    def create_mask(self, sigma_x, sigma_y, sx=11, sy=11):
        """ Mask spot in terms of its x and y standard deviation, by setting the radii of an ellipse to 3 x sigma
            We can capture at least 80% of the spot. """

        mask = np.ones((sy, sx))
        centre_x = int(sx / 2)
        centre_y = int(sy / 2)

        ellipse_radius_x = sigma_x * 1.5
        ellipse_radius_y = sigma_y * 1.5

        for x in range(sx):
            for y in range(sy):
                ax = x
                ay = y
                boundary_condition = (ax - centre_x) ** 2 / ellipse_radius_x ** 2 + (
                        ay - centre_y) ** 2 / ellipse_radius_y ** 2
                if boundary_condition <= 1:
                    mask[y][x] = 0

        return mask

    def calculate_trace(self, data, traject):
        trajectory = traject
        coordinates = trajectory.coordinates
        mask = self.create_mask(1.6, 1.6)

        trace = []

        for ind, fr in enumerate(trajectory.frame_list):
            try:
                detection_field = np.copy(data[:, :, fr])
                h, w = np.shape(detection_field)[0], np.shape(detection_field)[1]
            except:
                error = True
                print("error occured here")

            coords = coordinates[ind]

            x_low = 5
            x_hi = 6
            y_low = 5
            y_hi = 6

            if coords[0] < 5:
                x_low = coords[0]
            if coords[0] > w - 6:
                x_hi = w - coords[0]
            if coords[1] < 5:
                y_low = coords[1]
            if coords[1] > h - 6:
                y_hi = h - coords[1]

            spot_grid = detection_field[int(coords[1] - y_low):int(coords[1] + y_hi),
                        int(coords[0] - x_low):int(coords[0] + x_hi)]

            if coords[0] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[:, shape_diff[1]:]
            if coords[1] < 10:
                shape_diff = list(np.array(np.shape(mask)) - np.array(np.shape(spot_grid)))
                mask = mask[shape_diff[0]:, :]
            if coords[0] > h - 10:
                mask = mask[:, :np.shape(spot_grid)[1]]
            if coords[1] > w - 10:
                mask = mask[:np.shape(spot_grid)[0], :]

            residual_matrix = spot_grid * mask
            mask_size = np.shape(mask)[0] * np.shape(mask)[1]
            mask_ratio = mask_size / np.sum(mask)

            trace_intensity = np.sum(spot_grid) / mask_size
            background = (np.sum(residual_matrix) * mask_ratio) / mask_size
            trace_corrected = trace_intensity - background

            trace.append(trace_corrected * 8)

        return trace

    @staticmethod
    def adjust_contrast(R, G, B, C, single_channel=False):
        H = 259
        L = 255
        factor = (H * (L + C)) / (L * (H - C))
        if not single_channel:
            r_prime = np.clip(factor * (R - 128) + 128, 0, 255)
            g_prime = np.clip(factor * (G - 128) + 128, 0, 255)
            b_prime = np.clip(factor * (B - 128) + 128, 0, 255)
            return r_prime, g_prime, b_prime
        else:
            b_prime = np.clip(factor * (B - 128) + 128, 0, 255)
            return b_prime

    @staticmethod
    def view_raw_data():
        global raw_stack_window
        try:
            raw_stack_window.window.destroy()
        except:
            """ Window was closed """
        raw_stack_window = RawStackTrack(mode="raw")

    @staticmethod
    def view_proc_data():
        global proc_stack_window
        try:
            proc_stack_window.window.destroy()
        except:
            """ Window was closed """
        proc_stack_window = RawStackTrack(mode="proc")

    def open_all_msds(self):
        global msd_win
        if self.field_selected.get() == self.field_list[0]:
            data = self.tracking_data_trk
        elif self.field_selected.get() == self.field_list[1]:
            data = self.tracking_data_c1
        elif self.field_selected.get() == self.field_list[2]:
            data = self.tracking_data_c2
        try:
            msd_win.handle_close()
        except:
            """ Window was closed """
        msd_win = MsdWin(data)

    def open_hist_win(self):
        global histwin
        if self.field_selected.get() == self.field_list[0]:
            data = self.tracking_data_trk
        elif self.field_selected.get() == self.field_list[1]:
            data = self.tracking_data_c1
        elif self.field_selected.get() == self.field_list[2]:
            data = self.tracking_data_c2
        try:
            histwin.handle_close()
        except:
            """ Failed to close window """
            try:
                histwin.window.destroy()
            except:
                """ Window was closed """
        histwin = HistWin(data)

    def automate_tracking(self):
        global auto_track_win
        self.automate_button["state"] = tk.DISABLED
        auto_track_win = AutoTrackWin()

    def open_font_win(self):
        global font_win
        try:
            font_win.handle_close()
        except:
            """ Failed """
        font_win = FontWin()

class RawStackTrack:
    def __init__(self, mode):
        self.window = tk.Tk()
        if mode == "raw":
            title_string = "Raw TIF Stack"
        else:
            title_string = "Enhanced TIF Stack"
        self.window.title(title_string)
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.canvas = tk.Canvas(master=self.window, bg="#222222", width=512, height=512)
        self.canvas.grid(row=0, column=0, columnspan=3)
        self.current_frame = tk.IntVar(master=self.window)

        self.mode = mode

        if tracking_win.field_selected.get() == tracking_win.field_list[0]:
            if mode == "raw":
                self.data = tracking_win.raw_data_tracking
            else:
                self.data = tracking_win.proc_data_tracking
        elif tracking_win.field_selected.get() == tracking_win.field_list[1]:
            if mode == "raw":
                self.data = tracking_win.raw_data_coloc1
            else:
                self.data = tracking_win.proc_data_coloc1
        elif tracking_win.field_selected.get() == tracking_win.field_list[2]:
            if mode == "raw":
                self.data = tracking_win.raw_data_coloc2
            else:
                self.data = tracking_win.proc_data_coloc2
        if self.data is not None:
            self.length = np.shape(self.data)[2]
            self.slider = tk.Scale(master=self.window, from_=0, to=self.length - 1, bg="#333333", fg="#cccccc",
                                   length=436,
                                   width=24, orient=tk.HORIZONTAL, command=self.display_frame,
                                   variable=self.current_frame)
            # change_col_hover_enterbox(self.slider, button_hv, "#333333")
            self.slider.grid(row=2, column=1, pady=4)
            self.back = tk.Button(master=self.window, text="<", bg="#222222", fg="white", padx=6, pady=12,
                                  command=self.dec)
            # change_col_hover(self.back, button_hv, "#222222")
            self.next = tk.Button(master=self.window, text=">", bg="#222222", fg="white", padx=6, pady=12,
                                  command=self.inc)
            # change_col_hover(self.next, button_hv, "#222222")
            self.normalize = tk.Button(master=self.window, text="Use Global Normalization", padx=10, pady=4,
                                       command=self.toggle_norm, bg="#222266", fg="#cccccc")
            # change_col_hover(self.normalize, button_hv, "#222266")
            self.normalize.grid(row=1, column=1, pady=4)
            self.back.grid(row=2, column=0, padx=8)
            self.next.grid(row=2, column=2, padx=8)

            self.colour_maps = ["Viridis", "Cividis", "Magma", "Inferno", "Plasma", "Greys"]
            self.map_dict = {
                "Viridis": mplib.cm.viridis,
                "Cividis": mplib.cm.cividis,
                "Magma": mplib.cm.magma,
                "Inferno": mplib.cm.inferno,
                "Plasma": mplib.cm.plasma,
                "Greys": mplib.cm.gray,
            }
            self.cm_selected = tk.StringVar(master=self.window)
            self.cm_selected.set(self.colour_maps[3])
            self.select_map = tk.OptionMenu(self.window, self.cm_selected, *self.colour_maps,
                                            command=lambda x: self.display_frame(self.current_frame.get()))
            self.select_map["bg"] = "#444f55"
            self.select_map["fg"] = "#cccccc"
            self.select_map.config(highlightbackground="#444444")
            self.select_map.place(x=370, y=520, width=140)

            self.global_normalize = False

            try:
                self.maximum = np.max(self.data)
            except:
                """ Error """

            self.display_frame(0)

        self.window.resizable(False, False)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """

    def inc(self):
        frame = int(float(self.current_frame.get()))
        if frame < self.length - 1:
            self.slider.set(frame + 1)

    def dec(self):
        frame = int(float(self.current_frame.get()))
        if frame > 0:
            self.slider.set(frame - 1)

    def display_frame(self, frame):
        global raw_img
        global proc_img
        frame = int(float(frame))
        if self.length is not None:
            frame_array = self.data[:, :, frame]
            h, w = np.shape(frame_array)[0], np.shape(frame_array)[1]
            if not self.global_normalize:
                maximum = np.max(frame_array)
            elif self.global_normalize:
                try:
                    maximum = self.maximum
                except:
                    """ Error """
            # frame_array = (frame_array / maximum) * 765
            # frame_rgb = np.zeros((h, w, 3))
            # blue = np.clip(np.copy(frame_array), 0, 255)
            # green = np.clip(np.copy(frame_array - 255), 0, 255)
            # red = np.clip(np.copy(frame_array - 510), 0, 255)
            # frame_rgb[:, :, 2] = blue - (green + red) / 2
            # frame_rgb[:, :, 1] = green - red / 1.01
            # frame_rgb[:, :, 0] = red

            norm = mplib.colors.Normalize(vmin=0, vmax=maximum)
            frame_rgb = self.map_dict[self.cm_selected.get()](norm(frame_array))
            frame_rgb = frame_rgb[:, :, :3] * 255


            image = tirf.create_image(frame_rgb, 512, 512)
            if self.mode == "raw":
                raw_img = tirf.ImageTk.PhotoImage(master=self.canvas, image=image)
                self.canvas.create_image(2, 2, anchor="nw", image=raw_img)
            else:
                proc_img = tirf.ImageTk.PhotoImage(master=self.canvas, image=image)
                self.canvas.create_image(2, 2, anchor="nw", image=proc_img)

    def toggle_norm(self):
        frame = int(float(self.current_frame.get()))
        if self.global_normalize:
            self.normalize["bg"] = "#222266"
            self.normalize["fg"] = "#cccccc"
            self.normalize["relief"] = tk.RAISED
            self.global_normalize = False
            self.display_frame(frame)
            return
        elif not self.global_normalize:
            self.normalize["bg"] = "#33ff66"
            self.normalize["fg"] = "black"
            self.normalize["relief"] = tk.SUNKEN
            self.global_normalize = True
            self.display_frame(frame)


class ProgressWin:
    def __init__(self, title, msg, cancel_func=None):
        self.title = title
        self.msg = msg
        self.window = tk.Tk()
        self.window.title(self.title)
        self.window["bg"] = "#666666"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.geometry("300x120+800+400")
        self.window.resizable(False, False)
        self.window.overrideredirect(1)

        self.title_bar = tk.Label(master=self.window, bg="#222222", fg="#cccccc", text=title, justify="left", anchor=tk.W)
        self.title_bar.place(x=0, y=0, width=300, height=24)
        self.progress = ttk.Progressbar(master=self.window, orient="horizontal", mode="determinate", length=280)
        self.label = tk.Label(master=self.window, text=msg, bg="#666666", fg="white", anchor=tk.NW, justify="left")
        self.cancel_button = ttk.Button(master=self.window, text="Cancel", command=tracking_win.cancel_tracking)
        if cancel_func:
            self.cancel_button["command"] = cancel_func
        self.progress.place(x=10, y=30)
        self.label.place(x=30, y=58)
        self.cancel_button.place(x=116, y=82)
        self.minimize_button = tk.Button(master=self.window, text="_", relief=tk.FLAT, bg="#444444", fg="white", padx=2,
                                         command=self.minimize)
        self.minimize_button.place(x=278, y=2, height=20)
        self.title_bar.bind("<B1-Motion>", self.move_window)
        self.title_bar.bind("<Button-1>", self.get_init_click)
        self.title_click = [None, None]

    def get_init_click(self, event):
        self.title_click = [event.x, event.y]

    def move_window(self, event):
        x = event.x + self.window.winfo_x()
        y = event.y + self.window.winfo_y()
        self.window.geometry(f"300x120+{x - self.title_click[0]}+{y - self.title_click[1]}")

    def minimize(self):
        try:
            tracking_win.restore_minimized_task_button.place(x=610, y=850, width=120)
        except:
            """ Failed """
        self.window.state("withdrawn")

    def restore(self):
        self.window.state("normal")
        try:
            tracking_win.restore_minimized_task_button.place_forget()
        except:
            """ Failed """

    def handle_close(self):
        try:
            self.progress.stop()
        except:
            """ No progress bar """
        try:
            self.window.destroy()
        except:
            """ Window was closed """
        try:
            tracking_win.restore_minimized_task_button.place_forget()
        except:
            """ Failed """


class AutoSelectWin:
    def __init__(self, file_list):
        self.window = tk.Toplevel()
        self.window.title(f"Select Files for Automation")
        x_coord = auto_track_win.window.winfo_x()
        self.window.geometry(f"400x600+{x_coord+420}+150")
        self.window["bg"] = "#222233"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.file_list = file_list

        self.frame_base = tk.Frame(master=self.window, width=374, height=524, bg="#333344", bd=2, relief=tk.SUNKEN)
        self.frame_base.place(x=2, y=2)
        self.scrollframe = Sframe(master=self.frame_base, width=374, height=524, bg="#333344")
        self.scrollframe.pack(side="top", expand=1, fill="both")
        self.scrollframe.bind_arrow_keys(self.frame_base)
        self.scrollframe.bind_scroll_wheel(self.frame_base)
        self.link_widget = self.scrollframe.display_widget(tk.Frame)

        if len(self.file_list) < 17:
            s_height = 17*30+14
        else:
            s_height = len(self.file_list)*30+10

        self.widget_frame = tk.Frame(master=self.link_widget, bg="#333344", height=s_height, width=374)
        self.widget_frame.pack()

        self.label_widgets = []
        self.tickbox_variables = []
        self.tickbox_widgets = []

        for idx, filename in enumerate(self.file_list):
            self.label_widgets.append(tk.Label(master=self.widget_frame, text=filename[0], bg="#333344",fg="#cccccc"))
            self.tickbox_variables.append(tk.IntVar(master=self.window))
            if self.file_list[idx][1]:
                self.tickbox_variables[-1].set(1)
            elif not self.file_list[idx][1]:
                self.tickbox_variables[-1].set(0)
            self.tickbox_widgets.append(tk.Checkbutton(master=self.widget_frame, activebackground="#333344", bg="#333344",
                                                       onvalue=1, offvalue=0, variable=self.tickbox_variables[-1]))
            self.tickbox_widgets[-1].place(x=10, y=idx*30 + 10)
            self.label_widgets[-1].place(x=34, y=idx*30 + 11)

        self.select_all_button = tk.Button(master=self.window, text="Select All", bg="#222266", fg="#cccccc",
                                           padx=12, pady=1, command=self.select_all)
        change_col_hover(self.select_all_button, button_hv, "#222266")
        self.select_all_button.place(x=20, y=565)

        self.deselect_all_button = tk.Button(master=self.window, text="Deselect All", bg="#222222", fg="#cccccc",
                                           padx=5, pady=1, command=self.deselect_all)
        change_col_hover(self.deselect_all_button, button_hv, "#222222")
        self.deselect_all_button.place(x=116, y=565)

        self.cancel_button = tk.Button(master=self.window, text="Cancel", bg="#662222", fg="#cccccc",
                                             padx=5, pady=1, command=self.handle_close)
        change_col_hover(self.cancel_button, "#994444", "#662222")
        self.cancel_button.place(x=260, y=565)

        self.done_button = tk.Button(master=self.window, text="Done", bg="#338833", fg="#cccccc",
                                       padx=8, pady=1, command=self.done)
        change_col_hover(self.done_button, "#44aa44", "#338833")
        self.done_button.place(x=326, y=565)

    def select_all(self):
        for idx, var in enumerate(self.tickbox_variables):
            if var.get() == 0:
                self.tickbox_widgets[idx].invoke()

    def deselect_all(self):
        for idx, var in enumerate(self.tickbox_variables):
            if var.get() == 1:
                self.tickbox_widgets[idx].invoke()

    def done(self):
        is_at_least_one_file_true = False
        for idx, var in enumerate(self.tickbox_variables):
            if var.get() == 0:
                self.file_list[idx][1] = False
            elif var.get() == 1:
                self.file_list[idx][1] = True
                is_at_least_one_file_true = True
        auto_track_win.custom_file_selection = self.file_list
        if is_at_least_one_file_true:
            auto_track_win.custom_selection = True
        self.handle_close()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to destroy window """


class AutoTrackWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title(f"Automate Tracking Analysis")
        self.window.geometry("400x400+760+250")
        self.window["bg"] = "#222233"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        try:
            default_path = tracking_win.auto_file
        except:
            default_path = "C://"

        self.cancel_flag = False
        self.stats = {
            "current file": None,
            "current tracks": None,
            "total tracks": 0,
            "total tracks kept": 0,
        }


        tk.Frame(master=self.window, bg="#444455").place(x=6, y=14, width=388, height=1)
        tk.Label(master=self.window, text="File", bg="#222233", fg="#cccccc").place(x=16, y=6, height=20)
        self.dir_path_label = tk.Entry(master=self.window, bg="#444455", fg="#cccccc", relief=tk.SUNKEN)
        self.dir_path_label.insert(0, default_path)
        self.dir_path_label.place(x=8, y=28, width=324, height=24)

        self.browse_button = tk.Button(master=self.window, text="Browse", bg="#222222", fg="#cccccc", padx=2,
                                       command=self.browse_folder)
        change_col_hover(self.browse_button, button_hv, "#222222")
        self.browse_button.place(x=340, y=28, height=24)

        self.select_button = tk.Button(master=self.window, text="Select", bg="#222266", fg="#cccccc", padx=5,
                                       command=self.select_files)
        change_col_hover(self.select_button, button_hv, "#222266")
        self.select_button.place(x=340, y=58, height=24)

        self.save_exp = tk.IntVar(master=self.window)
        self.save_exp.set(1)
        self.save_exp_button = tk.Checkbutton(master=self.window, activebackground="#222233", bg="#222233",
                                            onvalue=1, offvalue=0, variable=self.save_exp)
        self.save_exp_button.place(x=8, y=60)
        tk.Label(master=self.window, text="Save experiment (.txp) files", bg="#222233", fg="#cccccc").place(x=32, y=63, height=20)

        self.save_excel = tk.IntVar(master=self.window)
        self.save_excel.set(1)
        self.save_excel_button = tk.Checkbutton(master=self.window, activebackground="#222233", bg="#222233",
                                              onvalue=1, offvalue=0, variable=self.save_excel, command=self.fix_checkbox)
        self.save_excel_button.place(x=8, y=90)
        tk.Label(master=self.window, text="Export experiment data to excel (.xlsx) files", bg="#222233", fg="#cccccc").place(x=32, y=93, height=20)

        self.save_traces = tk.IntVar(master=self.window)
        self.save_traces.set(1)
        self.save_traces_button = tk.Checkbutton(master=self.window, activebackground="#222233", bg="#222233",
                                                onvalue=1, offvalue=0, variable=self.save_traces, command=self.fix_checkbox)
        self.save_traces_button.place(x=8, y=120)
        tk.Label(master=self.window, text="Save intensity traces to traces (.trcs) files", bg="#222233", fg="#cccccc").place(x=32, y=123, height=20)

        self.autofit = tk.IntVar(master=self.window)
        self.autofit.set(1)
        self.autofit_button = tk.Checkbutton(master=self.window, activebackground="#333344", bg="#222233",
                                             onvalue=1, offvalue=0, variable=self.autofit)
        self.autofit_button.place(x=290, y=90)
        tk.Label(master=self.window, text="Auto-Fit MSDs", bg="#222233", fg="#cccccc").place(x=310, y=93, height=20)

        self.start_button = tk.Button(master=self.window, text="Begin Automated Analysis", bg="#338833", fg="white",
                                      padx=12, pady=2, command=self.start)
        change_col_hover(self.start_button, "#44aa44", "#338833")
        self.start_button.place(x=8, y=152)

        self.stop_button = ttk.Button(master=self.window, text="Stop Analysis", command=self.stop)
        self.stop_button["state"] = tk.DISABLED
        self.stop_button.place(x=190, y=152, width=120, height=27)

        tk.Frame(master=self.window, bg="#444455").place(x=6, y=196, width=388, height=1)
        tk.Label(master=self.window, text="Statistics", bg="#222233", fg="#cccccc").place(x=16, y=188, height=20)

        info = f"Current file: '{self.stats['current file']}'\n\nNumber of tracks found: {self.stats['current tracks']}" \
               f"\nTotal tracks found: {self.stats['total tracks']}\nTotal tracks kept for export: {self.stats['total tracks kept']}"
        self.stats_label = tk.Label(master=self.window, text=info, anchor=tk.NW, padx=6, pady=2, bg="#444455",
                                    fg="#dddddd", bd=2, relief=tk.SUNKEN, wraplength=380, justify="left")
        self.stats_label.place(x=8, y=210, width=384, height=118)

        self.progress = ttk.Progressbar(master=self.window, orient="horizontal", mode="determinate", length=384)
        self.prog_label = tk.Label(master=self.window, text="", bg="#222233", fg="white", anchor=tk.NW, justify="left")

        self.custom_selection = False
        self.custom_file_selection = []
        self.browse_folder(auto=True)

    def select_files(self):
        global auto_select_win
        if len(self.custom_file_selection) == 0:
            return
        try:
            auto_select_win.handle_close()
        except:
            """ Window wasn't open """
        auto_select_win = AutoSelectWin(self.custom_file_selection)

    def start(self):
        if not self.custom_selection:
            folder_path = self.dir_path_label.get()
            tif_files = []
            for file in os.listdir(folder_path):
                if file.endswith(".tif"):
                    tif_files.append(os.path.join(folder_path, file))
            if len(tif_files) == 0:
                self.stop_button["state"] = tk.DISABLED
                easygui.msgbox(title="Error!", msg="There are no .tif files in the currently selected folder.")
                return
        else:
            folder_path = self.dir_path_label.get()
            tif_files = [os.path.join(folder_path, f[0]) for f in self.custom_file_selection if f[1]]

        for f in tif_files:
            print(f)

        self.stop_button["state"] = tk.NORMAL
        self.start_button["state"] = tk.DISABLED
        self.browse_button["state"] = tk.DISABLED
        self.select_button["state"] = tk.DISABLED
        self.progress.place(x=8, y=340)
        self.prog_label.place(x=10, y=368)

        self.progress["maximum"] = len(tif_files) + 1

        for idx, file_path in enumerate(tif_files):
            self.progress.step(1)
            self.progress.update()

            if self.cancel_flag:
                self.cancel_flag = False
                break

            try:
                if not self.cancel_flag:
                    self.prog_label["text"] = f"Loading files ({idx + 1}/{len(tif_files)})..."
                    self.prog_label.update()
                tracking_win.load_file(file=file_path, auto=True)
                if not self.cancel_flag:
                    self.prog_label["text"] = f"Analysing files ({idx + 1}/{len(tif_files)})..."
                    self.prog_label.update()
            except:
                print("\nAn error occurred while loading TIF stack! Full traceback shown below:\n\n")
                print(traceback.format_exc())
            try:
                tracking_win.analyse_continuous()
            except:
                print("\nAn error occurred during automated tracking! Full traceback shown below:\n\n")
                print(traceback.format_exc())
            try:
                if self.autofit.get() == 1:
                    tracking_win.msd_autofit()
            except:
                print("\nAn error occurred during automated MSD fitting! Full traceback shown below:\n\n")
                print(traceback.format_exc())

            data = tracking_win.tracking_data_trk
            print(f"Found {len(data)} tracks!")

            if len(data) == 0:
                print("\nContinued to next file instead of saving because there are no tracks!\n")
                continue

            try:
                self.stats["current file"] = file_path
                self.stats["current tracks"] = len(data)
                kept_count = 0
                for traj in data:
                    if traj.include_export:
                        kept_count += 1
                self.stats["total tracks"] += len(data)
                self.stats["total tracks kept"] += kept_count

                self.stats_label["text"] = f"Current file: '{self.stats['current file']}'\n\nNumber of tracks found: {self.stats['current tracks']}" \
                                           f"\nTotal tracks found: {self.stats['total tracks']}\nTotal tracks kept for export: {self.stats['total tracks kept']}"
                self.stats_label.update()
            except:
                print("An error occurred while updating statistics! Full traceback shown below:\n\n")
                print(traceback.format_exc())

            if self.save_exp.get() == 1:
                try:
                    path = file_path[:-4] + ".txp"
                    tracking_win.save_document(auto=True, autopath=path)
                except:
                    print("An error occurred while saving .txp! Full traceback shown below:\n\n")
                    print(traceback.format_exc())

            if self.save_excel.get() == 1:
                try:
                    path2 = file_path[:-4] + ".xlsx"
                    path3 = file_path[:-4] + ".trcs"
                    export_template = self.create_blank_template()
                    tracking_win.export_to_xl(auto=True, xlsx_obj=export_template, xlsx_path=path2, trcs_path=path3)
                except:
                    print("An error occurred while saving .xlsx and .trcs! Full traceback shown below:\n\n")
                    print(traceback.format_exc())


        self.stop_button["state"] = tk.DISABLED
        self.start_button["state"] = tk.NORMAL
        self.browse_button["state"] = tk.NORMAL
        self.select_button["state"] = tk.NORMAL
        self.progress.place_forget()
        self.prog_label.place_forget()
        self.progress.stop()

    def fix_checkbox(self):
        if self.save_excel.get() == 0 and self.save_traces.get() == 1:
            self.save_traces_button.invoke()

    def create_blank_template(self):
        wb = opxl.Workbook()
        ws = wb.active
        ws.title = "Raw Data"
        ws2 = wb.create_sheet("MSD")
        ws2.title = "MSD"
        ws2.cell(row=1, column=1).value = "name"
        ws2.cell(row=2, column=1).value = "diffusion coeff"
        ws2.cell(row=3, column=1).value = "bleaching time (s)"
        ws2.cell(row=4, column=2).value = "msd cumulative"

        return wb

    def stop(self):
        self.cancel_flag = True
        self.prog_label["text"] = "Analysis will stop after the current file has finished..."
        self.progress.stop()

    def browse_folder(self, auto=False):
        if not auto:
            path = easygui.diropenbox(msg="Open a folder containing .tif files to be analysed.", default=default_dir)
        else:
            path = self.dir_path_label.get()
        if path:
            tracking_win.auto_file = path
            self.dir_path_label.delete(0, tk.END)
            self.dir_path_label.insert(0, path)

            folder_path = self.dir_path_label.get()
            file_list = []
            for file in os.listdir(folder_path):
                if file.endswith(".tif"):
                    file_list.append(os.path.join(folder_path, file))
            if not auto:
                if len(file_list) == 0:
                    easygui.msgbox(title="Error!", msg="There are no .tif files in the currently selected folder.")
                    return

            self.custom_file_selection = [[os.path.split(file)[1], True] for file in file_list]

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to destroy window """
        tracking_win.automate_button["state"] = tk.NORMAL


class HistWin:
    def __init__(self, tracks):
        self.window = tk.Toplevel()
        self.window.title(f"Diffusion Histogram")
        self.window.geometry("1100x662+410+180")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.colours = ["#0088ff", "#ff9900", "#00ff44", "#ff0000", "#cc00ff", "#555500", "#ff8888", "#555555"]
        self.output_stats = []

        self.tracks = tracks
        self.diff_dist = []
        self.backup_dist = []
        self.x_axis = np.linspace(0, 1, 1000)
        self.cancel_flag = False

        self.frame = tk.Frame(master=self.window)
        self.frame.place(x=6, y=76)
        self.figure = plt.Figure(figsize=(10.88, 5.38), dpi=100)
        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.95, bottom=0.1, left=0.06, right=0.98)

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)

        self.toolbar = NavigationToolbar2Tk(self.canvas, self.window)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack()

        self.field_list = ["1 Component", "2 Components", "3 Components", "4 Components", "5 Components", "6 Components",
                           "7 Components", "8 Components"]

        self.field_selected = tk.StringVar(master=self.window)
        self.field_selected.set(self.field_list[0])
        self.select_field = tk.OptionMenu(self.window, self.field_selected, *self.field_list, command=self.replot)
        self.select_field["bg"] = "#444f55"
        self.select_field["fg"] = "#cccccc"
        self.select_field.config(highlightbackground="#444444")
        self.select_field.place(x=10, y=6, width=150, height=26)

        tk.Label(master=self.window, text="Bin count", bg="#444444", fg="#cccccc").place(x=180, y=10, height=20)
        self.bin_count_entry = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.bin_count_entry["validatecommand"] = (self.bin_count_entry.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.bin_count_entry.insert(tk.END, 40)
        change_col_hover_enterbox(self.bin_count_entry, button_hv, "#333333")
        self.bin_count_entry.place(x=248, y=8, width=40, height=21)

        self.replot_button = tk.Button(master=self.window, text="Re-plot", padx=13, pady=2, bg="#222266", fg="#cccccc",
                                       command=lambda: self.replot(None))
        change_col_hover(self.replot_button, button_hv, "#222266")
        self.replot_button.place(x=307, y=8, height=21)

        self.plot_component_ratios = tk.Button(master=self.window, text="Ratios", padx=5, pady=2,
                                               bg="#333377", fg="#cccccc", command=self.plot_components)
        change_col_hover(self.plot_component_ratios, button_hv, "#333377")
        self.plot_component_ratios.place(x=390, y=8, height=21)

        self.plot_validation_button = tk.Button(master=self.window, text="Plot σ/D", padx=6, pady=2,
                                               bg="#333377", fg="#cccccc", command=self.plot_validation)
        change_col_hover(self.plot_validation_button, button_hv, "#333377")
        self.plot_validation_button.place(x=456, y=8, height=21)


        tk.Label(master=self.window, text="Covariance Type", bg="#444444", fg="#cccccc").place(x=10, y=37, height=20)
        self.field_list3 = ["full", "diag", "spherical"]
        self.field_selected3 = tk.StringVar(master=self.window)
        self.field_selected3.set(self.field_list3[0])
        self.select_field3 = tk.OptionMenu(self.window, self.field_selected3, *self.field_list3, command=self.replot)
        self.select_field3["bg"] = "#444f55"
        self.select_field3["fg"] = "#cccccc"
        self.select_field3.config(highlightbackground="#444444")
        self.select_field3.place(x=110, y=35, width=120, height=26)

        tk.Label(master=self.window, text="Initializer", bg="#444444", fg="#cccccc").place(x=248, y=37, height=20)
        self.field_list2 = ["Comp 1, ", "Comp 2, ", "Comp 3, ", "Comp 4, ", "Comp 5, ", "Comp 6, ", "Comp 7, ", "Comp 8, "]
        self.field_selected2 = tk.StringVar(master=self.window)
        self.field_selected2.set("Select...")
        self.select_field2 = tk.OptionMenu(self.window, self.field_selected2, *self.field_list2, command=self.bind_mouse)
        self.select_field2["bg"] = "#444f55"
        self.select_field2["fg"] = "#cccccc"
        self.select_field2.config(highlightbackground="#444444")
        self.select_field2.place(x=304, y=35, width=140, height=26)

        self.manual = tk.IntVar(master=self.window)
        self.manual.set(0)
        self.manual_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                            onvalue=1, offvalue=0, variable=self.manual)
        self.manual_button.place(x=450, y=35)
        tk.Label(master=self.window, text="Manual", bg="#444444", fg="#cccccc").place(x=476, y=38, height=20)

        tk.Frame(master=self.window, bg="#888888").place(x=530, y=6, width=1, height=62)

        self.chain_experiments_button = tk.Button(master=self.window, text="Chain Experiments", bg="#446688",
                                                  fg="#cccccc", padx=4, command=self.chain_experiments)
        change_col_hover(self.chain_experiments_button, "#6688aa", "#446688")
        self.chain_experiments_button.place(x=538, y=8, height=21)

        self.recalc_msds = tk.IntVar(master=self.window)
        self.recalc_msds.set(0)
        self.recalc_msds_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                                 onvalue=1, offvalue=0, variable=self.recalc_msds)
        self.recalc_msds_button.place(x=658, y=5)
        tk.Label(master=self.window, text="Recalculate MSD fits", bg="#444444", fg="#cccccc").place(x=684, y=8, height=20)

        self.exclude_if_stationary = tk.IntVar(master=self.window)
        self.exclude_if_stationary.set(0)
        self.exclude_if_stationary_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                                 onvalue=1, offvalue=0, variable=self.exclude_if_stationary)
        self.exclude_if_stationary_button.place(x=538, y=52, height=20)
        tk.Label(master=self.window, text="Exclude if stationary", bg="#444444", fg="#cccccc").place(x=564, y=53, height=20)

        self.exclude_if_error = tk.IntVar(master=self.window)
        self.exclude_if_error.set(0)
        self.exclude_if_error_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                                           onvalue=1, offvalue=0, variable=self.exclude_if_error)
        self.exclude_if_error_button.place(x=538, y=30, height=20)
        tk.Label(master=self.window, text="Exclude if error  >                 %", bg="#444444", fg="#cccccc").place(x=564, y=31, height=20)

        self.error_entry = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.error_entry["validatecommand"] = (self.error_entry.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.error_entry.insert(tk.END, int(tracking_win.msd_maximum_error.get()))
        change_col_hover_enterbox(self.error_entry, button_hv, "#333333")
        self.error_entry.place(x=668, y=30, width=40, height=21)

        tk.Frame(master=self.window, bg="#888888").place(x=810, y=6, width=1, height=62)

        self.dark_mode = tk.IntVar(master=self.window)
        self.dark_mode.set(int(tracking_win.dark_mode.get()))
        self.dark_mode_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                               onvalue=1, offvalue=0, variable=self.dark_mode, command=lambda: self.replot(None))
        self.dark_mode_button.place(x=820, y=7, height=20)
        tk.Label(master=self.window, text="Dark Theme", bg="#444444", fg="#cccccc").place(x=844, y=8, height=20)

        self.export_button = tk.Button(master=self.window, text="Export to Excel", bg="#446688",
                                                  fg="#cccccc", padx=10, command=self.export)
        change_col_hover(self.export_button, "#6688aa", "#446688")
        self.export_button.place(x=820, y=36, height=32)

        tk.Frame(master=self.window, bg="#888888").place(x=935, y=6, width=1, height=62)

        self.remove_outliers_button = tk.Button(master=self.window, text="Remove Outliers", bg="#222222",
                                       fg="#cccccc", padx=10, command=self.remove_outliers)
        change_col_hover(self.remove_outliers_button, button_hv, "#222222")
        self.remove_outliers_button.place(x=945, y=8, height=28)

        self.info_icon = tk.PhotoImage(master=self.window, file=cwd + "/icons/info.png")
        info_1 = tk.Button(master=self.window, image=self.info_icon, bg="#444444", relief=tk.FLAT,
                           activebackground='#777788',
                           command=lambda: tracking_win.info_message(title="Remove Outliers",
                                                             message="MSD fitting seems to produce a skewed distribution biased slightly in favour "
                                                                     "of diffusion coefficients greater than the true value. The Gaussian mixture "
                                                                     "model is sensistive to even a small number of outliers if they deviate significantly "
                                                                     "from the main distribution. This tool will reject diffusion coefficients from "
                                                                     "the current distribution if they are greater than the total mean by more "
                                                                     "standard deviations than the value in the threshold box below. This can be undone."
                                                             ))
        change_col_hover(info_1, "#666677", "#444444")
        info_1.place(x=1070, y=12, width=19, height=19)

        tk.Label(master=self.window, text="Threshold", bg="#444444", fg="#cccccc").place(x=940, y=46, height=20)
        self.outlier_threshold = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.outlier_threshold["validatecommand"] = (self.outlier_threshold.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.outlier_threshold.insert(tk.END, 3)
        change_col_hover_enterbox(self.outlier_threshold, button_hv, "#333333")
        self.outlier_threshold.place(x=1004, y=44, width=30, height=21)

        self.undo_remove_button = tk.Button(master=self.window, text="Undo", bg="#222222",
                                                fg="#cccccc", padx=4, command=self.undo_remove_outliers)
        change_col_hover(self.undo_remove_button, "#226622", "#222222")
        self.undo_remove_button.place(x=1045, y=44, height=21)

        tk.Frame(master=self.window, bg="#888888").place(x=728, y=30, width=1, height=38)

        tk.Label(master=self.window, text="         Lock", bg="#444444", fg="#cccccc").place(x=738, y=29)
        tk.Label(master=self.window, text="Constraints", bg="#444444", fg="#cccccc").place(x=738, y=47)
        self.lock_constraints = tk.IntVar(master=self.window)
        self.lock_constraints.set(0)
        self.lock_constraints_button = tk.Checkbutton(master=self.window, activebackground="#444444", bg="#444444",
                                                      onvalue=1, offvalue=0, variable=self.lock_constraints)
        self.lock_constraints_button.place(x=738, y=30, height=20)

        self.capture_mouse = None
        self.RMS_method = False
        self.window.update()

        self.get_diff_distribution()

        self.np_dist = np.array(self.diff_dist)
        self.gmm_data = self.np_dist.reshape(-1, 1)
        self.np_dist.sort()

        self.fits, self.error_state = self.fit_gmm()
        self.plot_histogram_and_fits()

    def plot_validation(self):
        fits, total_fit, stats = self.get_gaussians(self.fits)
        x = []
        y = []
        for i in range(len(stats)):
            x.append(stats[i][0])
            y.append(stats[i][1])
        plt.plot(x, y, marker="o", markersize=5, linewidth=0)
        a, c = np.polyfit(x, y, 1)
        r2 = r2_score((a*np.array(x) + c), y)
        plt.plot(x, a*np.array(x) + c)
        plt.title(f"Component Validation (Grad = {a}), r^2 = {round(r2, 5)}")
        plt.xlabel("Component Diffusivity (µm² / s)")
        plt.ylabel("Component σ")
        plt.show()

    def remove_outliers(self):
        mean = np.mean(self.diff_dist)
        stdev = np.std(self.diff_dist)

        new_dist = []
        for idx in range(len(self.diff_dist)):
            if not self.diff_dist[idx] > mean + stdev * int(self.outlier_threshold.get()):
                new_dist.append(self.diff_dist[idx])

        self.diff_dist = list(new_dist)
        self.np_dist = np.array(self.diff_dist)
        self.gmm_data = self.np_dist.reshape(-1, 1)
        self.np_dist.sort()

        self.replot(None)

    def undo_remove_outliers(self):
        self.diff_dist = list(self.backup_dist)
        self.np_dist = np.array(self.diff_dist)
        self.gmm_data = self.np_dist.reshape(-1, 1)
        self.np_dist.sort()

        self.replot(None)

    def cancel_chain(self):
        self.cancel_flag = True

    def chain_experiments(self):
        self.cancel_flag = False
        folder = easygui.diropenbox(title="Open folder containing TIF stacks from 1 single molecule experiment",
                                    default=default_dir)
        if not folder:
            return

        self.RMS_method = easygui.indexbox(msg="Use Alternative 'RMS method' for calculating the diffusion coefficient? "
                                               "'RMS method' calculates the root mean squre displacement of each track "
                                               "with a locked time step of 5 and calcualtes the diffusion coefficient "
                                               "by rearranging the equation 'x_RMS = sqrt(4Dt)' to 'D = x_RMS² / 4t'. "
                                               "Filtering using the MSD fit method will still be applied. ",
                                           title="Diffusion Coefficient Calculation",
                                           choices=["    Default    ", "RMS Method"],
                                           default_choice=0)
        if self.RMS_method is None:
            return

        exp_files = []
        for file in os.listdir(folder):
            if file.endswith(".txp"):
                exp_files.append(os.path.join(folder, file))

        full_distribution = []
        for idx, filename in enumerate(exp_files):
            if self.cancel_flag:
                break
            try:
                tracking_win.open_document_auto(filename)
                if self.recalc_msds.get() == 1:
                    if not self.lock_constraints.get() == 1:
                        if tracking_win.exclude_stationary.get() == 1 and self.exclude_if_stationary.get() == 0:
                            tracking_win.exclude_stationary_button.invoke()
                        elif tracking_win.exclude_stationary.get() == 0 and self.exclude_if_stationary.get() == 1:
                            tracking_win.exclude_stationary_button.invoke()
                        if tracking_win.exclude_based_on_error.get() == 1 and self.exclude_if_error.get() == 0:
                            tracking_win.exclude_based_on_error_button.invoke()
                        if tracking_win.exclude_based_on_error.get() == 0 and self.exclude_if_error.get() == 1:
                            tracking_win.exclude_based_on_error_button.invoke()
                        tracking_win.msd_maximum_error.delete(0, tk.END)
                        tracking_win.msd_maximum_error.insert(0, int(self.error_entry.get()))
                    tracking_win.msd_autofit()
                file_dist = self.get_diff_distribution_auto()
                full_distribution += file_dist
            except:
                print("\n")
                print(traceback.format_exc())
                print("\n")

        if self.cancel_flag and len(full_distribution) == 0:
            return

        self.diff_dist = full_distribution
        self.backup_dist = list(self.diff_dist)
        self.np_dist = np.array(self.diff_dist)
        self.gmm_data = self.np_dist.reshape(-1, 1)
        self.np_dist.sort()

        self.fits, self.error_state = self.fit_gmm()
        self.figure.clf()
        self.canvas.draw()
        self.plot_histogram_and_fits()


    def get_init_means(self):
        name = self.field_selected.get()
        index = self.field_list.index(name)
        try:
            means = []
            for idx in range(index + 1):
                mean = float(self.field_list2[idx].split(",")[1][1:])
                means.append([mean])
        except:
            easygui.msgbox(title="Warning!", msg="Warning! Not all means for all components have been manually initialized. Switching to automatic.")
            self.manual_button.invoke()
            return None

    def bind_mouse(self, null):
        if self.manual.get() == 0:
            self.manual_button.invoke()
        self.capture_mouse = self.figure.canvas.mpl_connect('button_press_event', self.get_means_from_click)

    def get_means_from_click(self, event):
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        mouse_xpos = event.xdata
        mouse_xpos = round(mouse_xpos, 6)

        name = self.field_selected2.get()
        index = self.field_list2.index(name)
        self.field_list2[index] = f"Comp {index + 1}, {mouse_xpos}"
        self.select_field2.destroy()
        self.field_selected2.set(self.field_list2[index])
        self.select_field2 = tk.OptionMenu(self.window, self.field_selected2, *self.field_list2, command=self.bind_mouse)
        self.select_field2["bg"] = "#444f55"
        self.select_field2["fg"] = "#cccccc"
        self.select_field2.config(highlightbackground="#444444")
        self.select_field2.place(x=304, y=35, width=140, height=26)

    def plot_components(self):
        dist = []
        name = self.field_selected.get()
        index = self.field_list.index(name)
        components = index + 1

        fits, total_fit, stats = self.get_gaussians(self.fits)

        weights = []
        labels = []
        for idx in range(components):
            mean = float(stats[idx][0])
            labels.append(str(round(mean, 4)))
            weights.append(np.sum(fits[idx]))

        print(weights)
        print(labels)
        plt.bar(x=labels, height=weights)
        plt.title("Component Proportions")
        plt.xlabel("Component (µm² / s)")
        plt.ylabel("Component Integral (NOT NORMALIZED)")
        plt.show()

    def replot(self, null):
        self.figure.clf()
        self.canvas.draw()
        self.fits, self.error_state = self.fit_gmm()
        self.plot_histogram_and_fits()

    def plot_histogram_and_fits(self):
        fits, total_fit, stats = self.get_gaussians(self.fits)
        if fits is None:
            return

        if self.dark_mode.get() == 1:
            bg_col = "#222222"
            fig_col = "#333333"
            highlight_col = "blue"
            text_col = "white"
            total_fit_col = "white"
            grid_col = "#333333"
        else:
            bg_col = "white"
            fig_col = "white"
            highlight_col = "black"
            text_col = "black"
            total_fit_col = "black"
            grid_col = "#cccccc"

        self.figure.set_facecolor(fig_col)
        plotter = self.figure.add_subplot(111)
        plotter.set_facecolor(bg_col)
        plotter.spines['bottom'].set_color(highlight_col)
        plotter.spines['top'].set_color(highlight_col)
        plotter.spines['left'].set_color(highlight_col)
        plotter.spines['right'].set_color(highlight_col)
        plotter.xaxis.label.set_color(text_col)
        plotter.yaxis.label.set_color(text_col)
        plotter.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        plotter.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        plotter.set_xlabel("Diffusion Coefficient (µm² / s)", color=text_col, size=tracking_win.g_fonts["labels"])
        plotter.set_ylabel("Frequency Density", color=text_col, size=tracking_win.g_fonts["labels"])
        plotter.set_title(f"Diffusion Coefficient Distribution (N = {len(self.diff_dist)})", color=text_col, size=tracking_win.g_fonts["title"])
        plotter.grid(color=grid_col)
        plotter.hist(self.diff_dist, bins=int(self.bin_count_entry.get()), density=True, alpha=0.5, ec="#9900ff", color="#5500aa")

        leg = ["Overall Fit"]
        for i in range(len(stats)):
            text = f"Mean coeff: {round(stats[i][0], 4)}, Std.Dev. = {round(stats[i][1], 4)}"
            leg.append(text)

        plotter.plot(self.x_axis, total_fit, color=total_fit_col, linewidth=2)
        for idx in range(len(fits)):
            plotter.plot(self.x_axis, fits[idx], linewidth=1, color=self.colours[idx])
        plotter.legend(leg, loc="upper right")
        self.canvas.draw()

        if self.error_state:
            easygui.msgbox(title="Warning!", msg="Warning! Inconsistency detected in clustering! You may be trying to "
                                                 "fit to many components to the data or the number of samples is not "
                                                 "statistically significant.")
        self.window.attributes('-topmost', False)

    def gauss(self, x, a, x0, sigma):
        return a * np.exp(-(x - x0) ** 2 / (2 * sigma ** 2))

    def get_gaussians(self, gmm_fit):
        fits = []
        name = self.field_selected.get()
        index = self.field_list.index(name)
        components = index + 1

        if gmm_fit is None:
            return None, None, None

        weights = gmm_fit.weights_
        means = gmm_fit.means_
        covars = gmm_fit.covariances_
        stats = []
        features = []

        start = np.min(self.np_dist)
        end = np.max(self.np_dist)
        self.x_axis = np.linspace(start, end, 1000)

        for fit_index in range(components):
            features.append([float(means[fit_index].ravel()), float(np.sqrt(covars[fit_index]).ravel()), float(weights[fit_index].ravel())])
            stats.append([round(float(means[fit_index].ravel()), 8), round(float(np.sqrt(covars[fit_index]).ravel()), 8)])

        stats.sort()
        features.sort()

        for fit_index in range(components):
            init_fit = features[fit_index][2] * scipy_stats.norm.pdf(self.np_dist, features[fit_index][0], features[fit_index][1]).ravel()
            scale = np.max(init_fit)
            features[fit_index][2] = scale
            fit = self.gauss(self.x_axis, scale, features[fit_index][0], features[fit_index][1]).ravel()
            fits.append(fit)

        total_fit = np.copy(fits[0])
        for idx in range(1, len(fits)):
            total_fit += fits[idx]

        self.output_stats = features

        return fits, total_fit, stats

    def fit_gmm(self):
        if len(self.gmm_data) == 0:
            easygui.msgbox(title="Error!", msg="No data to plot!")
            self.window.attributes('-topmost', False)
            return None, False
        try:
            means_sets = []
            name = self.field_selected.get()
            index = self.field_list.index(name)
            components = index + 1
            if self.manual.get() == 1:
                initial_means = self.get_init_means()
            else:
                initial_means = None
            for repeats in range(20):
                gmm = mixture.GaussianMixture(n_components=components,
                                              covariance_type=self.field_selected3.get(),
                                              means_init=initial_means)

                gmm.fit(self.gmm_data)
                means = gmm.means_
                f_means = []
                for idx in range(len(means)):
                    f_means.append(means[idx][0])
                f_means.sort()
                means_sets.append(f_means)

            means_sets = np.array(means_sets)

            stddevs = []
            percentage_errors = []
            for comp in range(components):
                std = np.std(means_sets[:, comp])
                m = np.mean(means_sets[:, comp])
                stddevs.append(std)
                percentage_errors.append((std / m)*100)

            print(percentage_errors)
            np_error = np.array(percentage_errors)
            error_state = False
            if np.any(np_error > 5):
                error_state = True
            return gmm, error_state
        except:
            easygui.msgbox(title="Error!", msg=f"An error occured:\n\n{traceback.format_exc()}")
            return None

    def get_diff_distribution(self):
        global progress_win
        self.diff_dist = []

        progress_win = ProgressWin("Preparing Data", "Calculating all coefficients...")
        progress_win.cancel_button.place_forget()
        progress_win.progress["maximum"] = len(self.tracks) + 1

        for index, trajectory in enumerate(self.tracks):

            progress_win.label["text"] = f"Calculating all coefficients ({index + 1} / {len(self.tracks)})..."
            progress_win.progress.step(1)
            progress_win.window.update()

            coordinates = trajectory.coordinates
            frames = trajectory.frame_list

            frame_times = np.array(frames[1:]) * float(tracking_win.frame_interval.get())
            first = frame_times[0]
            frame_times = frame_times - first

            new_MSD = tracking_win.calculate_msd(coordinates)
            length = int((float(trajectory.MSD_prop) / 100) * len(new_MSD))
            if length > 2:
                gradient, intercept = linear_regression(frame_times[1:length], new_MSD[:length])
            else:
                gradient, intercept = 0, 0
            dark_times = []
            for i in range(len(frames) - 1):
                if frames[i + 1] - frames[i] > 1:
                    dark_times.append(frames[i + 1] - frames[i])

            diffusion_coefficient = gradient / 4
            if trajectory.include_export:
                self.diff_dist.append(diffusion_coefficient)

        self.backup_dist = list(self.diff_dist)
        progress_win.progress.stop()
        progress_win.handle_close()
        progress_win = None

    def get_diff_distribution_auto(self):
        global progress_win
        if tracking_win.field_selected.get() == tracking_win.field_list[0]:
            data = tracking_win.tracking_data_trk
        elif tracking_win.field_selected.get() == tracking_win.field_list[1]:
            data = tracking_win.tracking_data_c1
        elif tracking_win.field_selected.get() == tracking_win.field_list[2]:
            data = tracking_win.tracking_data_c2

        auto_diff_dist = []

        progress_win = ProgressWin("Preparing Data", "Calculating all coefficients...", cancel_func=self.cancel_chain)
        progress_win.progress["maximum"] = len(data) + 1

        for index, trajectory in enumerate(data):

            if self.cancel_flag:
                break

            progress_win.label["text"] = f"Calculating all coefficients ({index + 1} / {len(data)})..."
            progress_win.progress.step(1)
            progress_win.window.update()

            coordinates = trajectory.coordinates
            frames = trajectory.frame_list

            frame_times = np.array(frames[1:]) * float(tracking_win.frame_interval.get())
            first = frame_times[0]
            frame_times = frame_times - first

            new_MSD = tracking_win.calculate_msd(coordinates)
            length = int((float(trajectory.MSD_prop) / 100) * len(new_MSD))
            if length > 2:
                gradient, intercept = linear_regression(frame_times[1:length], new_MSD[:length])
            else:
                gradient, intercept = 0, 0
            dark_times = []
            for i in range(len(frames) - 1):
                if frames[i + 1] - frames[i] > 1:
                    dark_times.append(frames[i + 1] - frames[i])

            diffusion_coefficient = gradient / 4

            # RMS method
            if self.RMS_method == 1:
                displacement = []
                for idx in range(len(coordinates) - 5):
                    displacement.append(np.sqrt((coordinates[idx + 5][0] - coordinates[idx][0]) ** 2 + (coordinates[idx + 5][1] - coordinates[idx][1]) ** 2))
                displacement = np.array(displacement)
                displacement = displacement * 160
                displacement = displacement / 1000
                x_RMS = np.sqrt(np.mean(displacement ** 2))
                diffusion_coefficient = (x_RMS ** 2 - intercept) / (4 * 5 * float(tracking_win.frame_interval.get()))

            if trajectory.include_export:
                auto_diff_dist.append(diffusion_coefficient)

        progress_win.progress.stop()
        progress_win.handle_close()
        progress_win = None
        return auto_diff_dist

    def export(self):
        path = easygui.filesavebox(msg="Save excel file for export", default=default_dir + "*.xlsx")
        if not path:
            return
        if path[-5:] != ".xlsx":
            path += ".xlsx"
        try:
            workbook = self.create_workbook()
            sheet = workbook["Diffusion Distribution"]

            fits, total_fit, stats = self.get_gaussians(self.fits)

            hist, bins = np.histogram(self.diff_dist, density=True, bins=int(self.bin_count_entry.get()))

            for idx, coeff in enumerate(self.diff_dist):
                sheet.cell(row=5+idx, column=2).value = coeff

            for idx in range(len(hist)):
                sheet.cell(row=5+idx, column=4).value = hist[idx]
                sheet.cell(row=5+idx, column=4).font = Font(name="Calibri", size=11, color="ffffff")
                sheet.cell(row=5+idx, column=4).fill = PatternFill("solid", fgColor="5500aa")
                sheet.cell(row=5 + idx, column=3).value = bins[idx]

            for idx in range(len(self.output_stats)):
                sheet.cell(row=3, column=9+idx).value = idx + 1
                sheet.cell(row=6, column=9+idx).value = self.output_stats[idx][0]
                sheet.cell(row=7, column=9+idx).value = self.output_stats[idx][1]
                sheet.cell(row=8, column=9+idx).value = self.output_stats[idx][2]

                sheet.cell(row=12, column=9+idx).value = f"Fit {idx + 1}"
                if idx == len(self.output_stats) - 1:
                    sheet.cell(row=12, column=10 + idx).value = "Total Fit"
                for x_idx in range(len(self.x_axis)):
                    sheet.cell(row=14+x_idx, column=7).value = self.x_axis[x_idx]
                for fit_idx in range(len(self.x_axis)):
                    sheet.cell(row=14+fit_idx, column=9+idx).value = round(fits[idx][fit_idx], 8)
                    sheet.cell(row=14+fit_idx, column=9+idx).fill = PatternFill("solid", fgColor=self.colours[idx][1:])
                    if idx == len(self.output_stats) - 1:
                        sheet.cell(row=14 + fit_idx, column=10 + idx).value = round(total_fit[fit_idx], 8)
                        sheet.cell(row=14 + fit_idx, column=10 + idx).fill = PatternFill("solid", fgColor="000000")
                        sheet.cell(row=14 + fit_idx, column=10 + idx).font = Font(name="Calibri", size=11, color="ffffff")

            histogram_xl = BarChart()
            cats = Reference(sheet, min_col=3, min_row=5, max_row=5+len(hist))
            data = Reference(sheet, min_col=4, min_row=5, max_row=5+len(hist))
            histogram_xl.add_data(data)
            histogram_xl.set_categories(cats)
            histogram_xl.title = "Diffusion Coefficient Distribution"
            histogram_xl.x_axis.title = "Diffusion Coeffficient (um^2/s)"
            histogram_xl.y_axis.title = "Frequency Density"
            histogram_xl.height = 10
            histogram_xl.width = 30

            sheet.add_chart(histogram_xl, "R3")

            scatter_xl = ScatterChart()
            x_data = Reference(sheet, min_col=7, min_row=14, max_row=14+len(self.x_axis))
            y_data = []
            series = []
            titles = ["Component 1", "Component 2", "Component 3", "Component 4", "Component 5", "Component 6", "Component 7", "Component 8", "Total Fit"]
            for idx in range(len(self.output_stats)+1):
                y_data.append(Reference(sheet, min_col=9+idx, min_row=14, max_row=14+len(self.x_axis)))
                series.append(Series(values=y_data[-1], xvalues=x_data, title=titles[idx]))
                scatter_xl.append(series[-1])
            scatter_xl.title = "Gaussian Fits"
            scatter_xl.x_axis.title = "Diffusion Coeffficient (um^2/s)"
            scatter_xl.y_axis.title = "Gaussian"
            scatter_xl.height = 10
            scatter_xl.width = 30

            sheet.add_chart(scatter_xl, "R24")

            sheet.sheet_view.zoomScale = 80
            workbook.save(path)
            easygui.msgbox(title="Export Complete!", msg="Export completed successfully.")
        except:
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="Warning! An error occurred while exporting! Check console for information.")

    @staticmethod
    def create_workbook():
        wb = opxl.Workbook()
        ws = wb.active
        ws.title = "Diffusion Distribution"

        ws.cell(row=3, column=2).value = "Distribution"
        ws.cell(row=3, column=2).font = Font(name="Calibri", size=11, bold=True)
        wb["Diffusion Distribution"].column_dimensions["B"].width = len(ws.cell(row=3, column=2).value) + 1
        ws.cell(row=3, column=4).value = "Histogram"
        ws.cell(row=3, column=4).font = Font(name="Calibri", size=11, bold=True)
        wb["Diffusion Distribution"].column_dimensions["C"].width = len(ws.cell(row=3, column=4).value) + 2
        ws.cell(row=3, column=3).value = "Bins"
        ws.cell(row=3, column=3).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=3, column=7).value = "Component"
        ws.cell(row=3, column=7).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=5, column=7).value = "Gaussian Parameters"
        ws.cell(row=5, column=7).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=6, column=7).value = "mean (diff. coeff.)"
        ws.cell(row=7, column=7).value = "standard deviation"
        ws.cell(row=8, column=7).value = "amplitude"
        ws.cell(row=12, column=7).value = "x coordinates below"
        ws.cell(row=12, column=7).font = Font(name="Calibri", size=11, bold=True)
        wb["Diffusion Distribution"].column_dimensions["G"].width = len(ws.cell(row=5, column=7).value) + 2
        return wb

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to close window """


class RSquaredWin:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title(f"Coefficient of Determination of {tracking_win.current_trajectory_series.name} MSD fit")
        self.window.geometry("800x940+560+40")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.figure = plt.Figure(figsize=(8, 9), dpi=100)
        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.95, bottom=0.08, left=0.12, right=0.97, hspace=0.32, wspace=0.2)

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.window)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.window)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack(fill="both")

        props, r2, r2_deriv, dfc_func, final_prop = tracking_win.msd_test_plot()

        data = [r2, r2_deriv, dfc_func]
        titles = ["R^2 as a Function of Fitting Percentage", "Derivative of R^2", "Diffusion Coefficient as a Function of Fitting Percentage"]
        ylabels = ["Coeff. Determination (R^2)", "d(R^2) / dx", "Diff. Coeff. (um^2 / s)"]

        if tracking_win.dark_mode.get() == 1:
            bg_col = "#222222"
            box_col = "blue"
            text_col = "white"
            grid_col = "#333333"
            line_col = "lightgray"
            fig_col = "#333333"
        else:
            bg_col = "white"
            box_col = "black"
            text_col = "black"
            grid_col = "#dddddd"
            line_col = "black"
            fig_col = "white"

        self.figure.set_facecolor(fig_col)
        for subplot in range(3):
            ax = self.figure.add_subplot(3, 1, subplot+1)
            ax.set_facecolor(bg_col)
            ax.spines['bottom'].set_color(box_col)
            ax.spines['top'].set_color(box_col)
            ax.spines['left'].set_color(box_col)
            ax.spines['right'].set_color(box_col)
            ax.xaxis.label.set_color(text_col)
            ax.yaxis.label.set_color(text_col)
            ax.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            ax.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            ax.set_xlabel("Fitting Percentage", color=text_col, size=tracking_win.g_fonts["labels"])
            ax.set_ylabel(ylabels[subplot], color=text_col, size=tracking_win.g_fonts["labels"])
            ax.set_title(titles[subplot], color=text_col, size=tracking_win.g_fonts["title"])
            if tracking_win.plot_grids.get() == 1:
                ax.grid(color=grid_col)
            ax.plot(props, data[subplot], color=line_col)
            ax.axvline(final_prop, color="red", linestyle="--")
        self.canvas.draw()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to destroy window """


class MsdWin:
    def __init__(self, tracks):
        self.window = tk.Tk()
        self.window.title("MSD Fits")
        self.window.geometry("1300x940+310+40")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, True)

        self.tracks = tracks
        self.page_index = 0
        self.tracks_per_page = 20
        self.max_pages = int(len(tracks) / self.tracks_per_page)
        if len(tracks) % self.tracks_per_page != 0:
            self.max_pages += 1
        if len(tracks) < self.tracks_per_page:
            self.tracks_per_page = len(tracks)
        self.tracks_to_show = tracks[:self.tracks_per_page]

        self.page_label = tk.Label(master=self.window, text=f"Page  {self.page_index + 1} / {self.max_pages}",
                                   bg="#444444", fg="#cccccc", font="TkDefaultFont 12")
        self.page_label.place(x=600, y=12)

        self.back = tk.Button(master=self.window, text="<", bg="#222266", fg="white", padx=10, pady=2, command=self.page_decrement)
        self.next = tk.Button(master=self.window, text=">", bg="#222266", fg="white", padx=10, pady=2, command=self.page_increment)
        change_col_hover(self.back, button_hv, "#222266")
        change_col_hover(self.next, button_hv, "#222266")
        self.back.place(x=530, y=12)
        self.next.place(x=720, y=12)

        self.figure = plt.Figure(figsize=(12.68, 4*self.tracks_per_page), dpi=100)
        self.grid_spec = gridspec.GridSpec(self.tracks_per_page, 2, width_ratios=[1, 2])
        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.99, bottom=0.01, left=0.02, right=0.98, hspace=0.32, wspace=0.15)

        self.frame_base = tk.Frame(master=self.window, width=1288, height=828, bg="#333333", bd=2,
                                   relief=tk.SUNKEN)
        self.frame_base.place(x=4, y=54)
        self.scrollframe = Sframe(master=self.frame_base, width=1268, height=818, bg="#333333")
        self.scrollframe.pack(side="top", expand=1, fill="both")
        self.scrollframe.bind_arrow_keys(self.frame_base)
        self.scrollframe.bind_scroll_wheel(self.frame_base)
        self.link_widget = self.scrollframe.display_widget(tk.Frame)

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.link_widget)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.window)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack(fill="both")
        self.window.update()
        self.plot_msds()

    def page_increment(self):
        if self.page_index < self.max_pages - 1:
            self.page_index += 1
            self.tracks_to_show = self.tracks[self.page_index*self.tracks_per_page:self.page_index*self.tracks_per_page+self.tracks_per_page]
            self.plot_msds()
            self.scrollframe.scroll_to_top()
            self.page_label["text"] = f"Page  {self.page_index + 1} / {self.max_pages}"
            self.page_label.update()

    def page_decrement(self):
        if self.page_index > 0:
            self.page_index -= 1
            self.tracks_to_show = self.tracks[self.page_index*self.tracks_per_page:self.page_index*self.tracks_per_page+self.tracks_per_page]
            self.plot_msds()
            self.scrollframe.scroll_to_top()
            self.page_label["text"] = f"Page  {self.page_index + 1} / {self.max_pages}"
            self.page_label.update()

    def plot_msds(self):
        self.figure.clf()
        self.canvas.draw()

        for subplot in range(len(self.tracks_to_show)):

            trajectory = self.tracks_to_show[subplot]

            delta_t, msd, fit, grad, intercept, diff_coeff, error, length = self.calculate_msd_data(trajectory)
            prop_fit = fit[:length]

            if tracking_win.dark_mode.get() == 1:
                bg_col = "#222222"
                box_col = "blue"
                text_col = "white"
                grid_col = "#333333"
                line_col = "white"
                fig_col = "#333333"
            else:
                bg_col = "white"
                box_col = "black"
                text_col = "black"
                grid_col = "#dddddd"
                line_col = "black"
                fig_col = "white"

            self.figure.set_facecolor(fig_col)
            plot_area = self.figure.add_subplot(self.grid_spec[subplot*2 + 1])
            plot_area.set_facecolor(bg_col)
            plot_area.spines['bottom'].set_color(box_col)
            plot_area.spines['top'].set_color(box_col)
            plot_area.spines['left'].set_color(box_col)
            plot_area.spines['right'].set_color(box_col)
            plot_area.xaxis.label.set_color(text_col)
            plot_area.yaxis.label.set_color(text_col)
            plot_area.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            plot_area.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            plot_area.set_xlabel("Time (s)", color=text_col, size=tracking_win.g_fonts["labels"])
            plot_area.set_ylabel("MSD (um^2)", color=text_col, size=tracking_win.g_fonts["labels"])
            plot_area.set_title(f"MSD Fit {(self.page_index*self.tracks_per_page + subplot + 1)}, m = {round(grad, 5)}, "
                                f"c = {round(intercept, 5)}, D = {round(diff_coeff, 5)} ± {error}%", color=text_col, size=tracking_win.g_fonts["title"])
            if tracking_win.plot_grids.get() == 1:
                plot_area.grid(color=grid_col)
            plot_area.plot(delta_t[1:], msd, marker="o", markersize=2, linewidth=2)
            plot_area.plot(delta_t[1:], fit[1:], linewidth=1, linestyle="--", color="red")
            plot_area.plot(delta_t[1:length], prop_fit[1:], linewidth=1, linestyle="--", color=line_col)
            if tracking_win.plot_grids.get() == 1:
                try:
                    if abs(error) < int(tracking_win.msd_maximum_error.get()):
                        plot_area.text(0.02, 0.98, "✓", ha="left", va="top", transform=plot_area.transAxes, color="green", size=18)
                    else:
                        plot_area.text(0.02, 0.98, "x", ha="left", va="top", transform=plot_area.transAxes, color="red", size=18)
                except:
                    print(traceback.format_exc())
                    plot_area.text(0.02, 0.98, "x", ha="left", va="top", transform=plot_area.transAxes, color="red", size=18)

            self.plot_track(trajectory, subplot)

        self.figure.tight_layout()
        self.canvas.draw()

    def plot_track(self, traj, subplot):

        x = [c[0] for c in traj.coordinates]
        y = [c[1] for c in traj.coordinates]

        nm_scale = float(tracking_win.pixel_size.get())
        for i in range(len(x)):
            x[i] = x[i] * nm_scale
            y[i] = y[i] * nm_scale

        min_x, min_y = np.min(x), np.min(y)
        for i in range(len(x)):
            x[i] = x[i] - min_x
            y[i] = y[i] - min_y

        if tracking_win.dark_mode.get() == 1:
            bg_col = "#222222"
            box_col = "blue"
            text_col = "white"
            grid_col = "#333333"
            line_col = "lightgray"
        else:
            bg_col = "white"
            box_col = "black"
            text_col = "black"
            grid_col = "#dddddd"
            line_col = "black"

        plot_trajectory_subfig = self.figure.add_subplot(self.grid_spec[subplot*2])
        plot_trajectory_subfig.set_facecolor(bg_col)
        plot_trajectory_subfig.spines['bottom'].set_color(box_col)
        plot_trajectory_subfig.spines['top'].set_color(box_col)
        plot_trajectory_subfig.spines['left'].set_color(box_col)
        plot_trajectory_subfig.spines['right'].set_color(box_col)
        plot_trajectory_subfig.xaxis.label.set_color(text_col)
        plot_trajectory_subfig.yaxis.label.set_color(text_col)
        plot_trajectory_subfig.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        plot_trajectory_subfig.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        if tracking_win.plot_grids.get() == 1:
            plot_trajectory_subfig.grid(color=grid_col)

        plot_trajectory_subfig.plot(x, y, linewidth=1, color=line_col)
        plot_trajectory_subfig.text(x[0], y[0], "Start", size=9, color="green")
        plot_trajectory_subfig.text(x[-1], y[-1], "End", size=9, color="red")
        plot_trajectory_subfig.axis("square")
        plot_trajectory_subfig.set_ylim(max(plot_trajectory_subfig.get_ylim()), min(plot_trajectory_subfig.get_ylim()))

        plot_trajectory_subfig.set_xlabel("x displacement (nm)", color=text_col, size=tracking_win.g_fonts["labels"])
        plot_trajectory_subfig.set_ylabel("y displacement (nm)", color=text_col, size=tracking_win.g_fonts["labels"])
        plot_trajectory_subfig.set_title(f"Track {(self.page_index*self.tracks_per_page + subplot + 1)}", size=tracking_win.g_fonts["title"], color=text_col)


    def calculate_msd_data(self, traj):
        frames = traj.frame_list
        coordinates = traj.coordinates
        proportion = traj.MSD_prop

        frame_times = np.array(frames[1:]) * float(tracking_win.frame_interval.get())
        first = frame_times[0]
        frame_times = frame_times - first

        new_MSD = tracking_win.calculate_msd(coordinates)
        length = int((proportion / 100) * len(new_MSD))
        if length > 2:
            gradient, intercept = linear_regression(frame_times[1:length], new_MSD[:length])
        else:
            gradient, intercept = 0, 0

        grad, inter = gradient, intercept
        diffusion_coefficient = gradient / 4

        msd_fit = []
        for t_fit in range(len(frame_times)):
            msd_fit.append(gradient * frame_times[t_fit] + intercept)

        DFC_series = []
        try:
            for get_diff_std in range(5, int(proportion) + 1):
                f_length = int((get_diff_std / 100) * len(new_MSD))
                if f_length > 2:
                    gradient, intercept = linear_regression(frame_times[1:f_length], new_MSD[:f_length])
                else:
                    gradient, intercept = 0, 0
                if f_length > 1:
                    DFC_series.append(gradient / 4)
            np_DFC = np.array(DFC_series)
            np_DFC_nonzero = np_DFC[np_DFC != 0]
            DFC_stddev = np.std(np_DFC_nonzero, ddof=1)
            DFC_stddev = round(DFC_stddev, 6)
            if len(np_DFC_nonzero) == 0:
                DFC_stddev = "N/A"
        except:
            print(traceback.format_exc())
            DFC_stddev = "N/A"

        try:
            diff_coeff_percent_error = round((DFC_stddev*int(tracking_win.msd_uncertainty_degree.get())) / (diffusion_coefficient) * 100, 1)
        except:
            print(traceback.format_exc())
            diff_coeff_percent_error = "N/A"

        return frame_times, new_MSD, msd_fit, grad, inter, diffusion_coefficient, diff_coeff_percent_error, length

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed to destroy window """


class GraphWin:
    def __init__(self):
        self.win2 = None
        self.window = tk.Tk()
        self.window.title(tracking_win.current_trajectory_series.name + " - Time Series")
        self.window.geometry("1600x900+150+50")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        self.trace = tracking_win.current_trajectory_series.trace
        self.tc1 = tracking_win.current_trajectory_series.trace_coloc1
        self.tc2 = tracking_win.current_trajectory_series.trace_coloc2

        name = tracking_win.current_trajectory_series.name
        frames = range(tracking_win.current_trajectory_series.frame_list[0],
                       tracking_win.current_trajectory_series.frame_list[-1] + 1)
        sigma_x = self.insert_NaN(tracking_win.current_trajectory_series.frame_list,
                                  tracking_win.current_trajectory_series.sigma_x)
        sigma_y = self.insert_NaN(tracking_win.current_trajectory_series.frame_list,
                                  tracking_win.current_trajectory_series.sigma_y)
        amp = self.insert_NaN(tracking_win.current_trajectory_series.frame_list,
                              tracking_win.current_trajectory_series.amplitude)
        residual = self.insert_NaN(tracking_win.current_trajectory_series.frame_list,
                                   tracking_win.current_trajectory_series.residual)
        ecc = []
        for i in range(len(tracking_win.current_trajectory_series.eccentricity_x)):
            if tracking_win.current_trajectory_series.eccentricity_x[i] < 1:
                ecc.append(tracking_win.current_trajectory_series.eccentricity_x[i])
            else:
                ecc.append(tracking_win.current_trajectory_series.eccentricity_y[i])
        ecc = self.insert_NaN(tracking_win.current_trajectory_series.frame_list, ecc)

        displacement = self.insert_NaN(tracking_win.current_trajectory_series.frame_list,
                                       tracking_win.current_trajectory_series.displacement)

        length = int((float(tracking_win.msd_proportion.get()) / 100) * len(tracking_win.current_trajectory_series.msd))
        if length > 2:
            gradient, intercept = linear_regression(tracking_win.current_trajectory_series.offset_time[1:length],
                                                               tracking_win.current_trajectory_series.msd[:length])
        else:
            gradient, intercept = 0, 0
        offset_time, msd = tracking_win.current_trajectory_series.offset_time[1:], tracking_win.current_trajectory_series.msd

        msd_fit = []
        for t_fit in range(len(offset_time)):
            msd_fit.append(gradient * offset_time[t_fit] + intercept)

        msd_fit_prop = []
        for t_fit in range(len(offset_time[:length])):
            msd_fit_prop.append(gradient * offset_time[t_fit] + intercept)

        if tracking_win.dark_mode.get() == 1:
            fig_col = "#333333"
            bg_col = "#444444"
        else:
            bg_col = "white"
            fig_col = "white"

        self.window["bg"] = bg_col

        self.frame1 = tk.Frame(master=self.window, width=480, height=280)
        self.frame1.place(x=10, y=10)
        self.fig1 = plt.Figure(figsize=(4.8, 2.80), dpi=100)
        self.canvas1 = FigureCanvasTkAgg(self.fig1, master=self.frame1)
        self.canvas1.draw()
        self.canvas1.get_tk_widget().pack(side=tk.TOP)
        self.fig1.set_facecolor(fig_col)
        self.fig1.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)
        self.frame7 = tk.Frame(master=self.window, width=300, height=280)
        self.frame7.place(x=490, y=10)
        self.fig7 = plt.Figure(figsize=(3, 2.80), dpi=100)
        self.canvas7 = FigureCanvasTkAgg(self.fig7, master=self.frame7)
        self.canvas7.draw()
        self.canvas7.get_tk_widget().pack(side=tk.TOP)
        self.fig7.set_facecolor(fig_col)
        self.fig7.subplots_adjust(bottom=0.16, left=0.16, right=0.9, top=0.9)

        self.frame2 = tk.Frame(master=self.window, width=780, height=280)
        self.frame2.place(x=10, y=300)
        self.fig2 = plt.Figure(figsize=(7.8, 2.80), dpi=100)
        self.canvas2 = FigureCanvasTkAgg(self.fig2, master=self.frame2)
        self.canvas2.draw()
        self.canvas2.get_tk_widget().pack(side=tk.TOP)
        self.fig2.set_facecolor(fig_col)
        self.fig2.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)

        self.frame3 = tk.Frame(master=self.window, width=780, height=280)
        self.frame3.place(x=10, y=600)
        self.fig3 = plt.Figure(figsize=(7.8, 2.80), dpi=100)
        self.canvas3 = FigureCanvasTkAgg(self.fig3, master=self.frame3)
        self.canvas3.draw()
        self.canvas3.get_tk_widget().pack(side=tk.TOP)
        self.fig3.set_facecolor(fig_col)
        self.fig3.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)

        self.frame4 = tk.Frame(master=self.window, width=780, height=280)
        self.frame4.place(x=800, y=10)
        self.fig4 = plt.Figure(figsize=(7.8, 2.80), dpi=100)
        self.canvas4 = FigureCanvasTkAgg(self.fig4, master=self.frame4)
        self.canvas4.draw()
        self.canvas4.get_tk_widget().pack(side=tk.TOP)
        self.fig4.set_facecolor(fig_col)
        self.fig4.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)

        self.frame5 = tk.Frame(master=self.window, width=780, height=280)
        self.frame5.place(x=800, y=300)
        self.fig5 = plt.Figure(figsize=(7.8, 2.80), dpi=100)
        self.canvas5 = FigureCanvasTkAgg(self.fig5, master=self.frame5)
        self.canvas5.draw()
        self.canvas5.get_tk_widget().pack(side=tk.TOP)
        self.fig5.set_facecolor(fig_col)
        self.fig5.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)

        self.frame6 = tk.Frame(master=self.window, width=780, height=280)
        self.frame6.place(x=800, y=600)
        self.fig6 = plt.Figure(figsize=(7.8, 2.80), dpi=100)
        self.canvas6 = FigureCanvasTkAgg(self.fig6, master=self.frame6)
        self.canvas6.draw()
        self.canvas6.get_tk_widget().pack(side=tk.TOP)
        self.fig6.set_facecolor(fig_col)
        self.fig6.subplots_adjust(bottom=0.16, left=0.12, right=0.94, top=0.9)

        self.plot(frames, displacement, "Time (frames)", "Displacement (nm)", "Displacement Magnitude",
                  self.fig1, self.canvas1)
        self.plot(frames, residual, "Time (frames)", "Normalized Intensity", "Gasussian Residual Variation with Time",
                  self.fig2, self.canvas2)
        self.plot(frames, amp, "Time (frames)", "Normalized Intensity", "Gaussian Amplitude Variation with Time",
                  self.fig3, self.canvas3)
        self.plot(offset_time, None, "Time (seconds)", "MSD (um^2)",
                  f"Mean Squared Distance (y = {round(gradient, 4)} x + {round(intercept, 4)}), D = {round(gradient / 4, 6)} um^2/s",
                  self.fig4, self.canvas4, dual=True, sets=[msd, msd_fit, offset_time[:length], msd_fit_prop], fit=True)
        self.plot(frames, None, "Time (frames)", "Sigma (pixels) / Ecc ratio",
                  "Gaussian Spread in x {blue} and y {orange}, "
                  "Eccentricity (min/maj) {green}",
                  self.fig5, self.canvas5, dual=True, sets=[sigma_x, sigma_y, ecc])
        self.plot(list(range(len(self.trace))), self.trace, "Time (frames)", "Intensity (A.U.)",
                  "Fluorescence Time Trace",
                  self.fig6, self.canvas6)
        self.plot(frames, displacement, "Velocity (nm / frame)", "Frequency", "Velocity distribution", self.fig7,
                  self.canvas7, type="hist")

        self.all_traces_button = tk.Button(master=self.window, text="Traces", padx=10, pady=1, bg="#336677", fg="white",
                                           command=self.traces_window)
        change_col_hover(self.all_traces_button, "#448899", "#336677")
        self.all_traces_button.place(x=1518, y=601)
        self.window.update()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Window was closed """
        try:
            self.win2.destroy()
        except:
            """ Window was closed """

    def traces_window(self):
        self.win2 = tk.Tk()
        self.win2.title(tracking_win.current_trajectory_series.name + " - Colocalization and Traces")
        self.win2.geometry("800x900+150+50")
        self.win2.protocol("WM_DELETE_WINDOW", self.handle_close_2)
        self.win2.attributes('-topmost', True)
        self.win2.resizable(False, False)
        self.win2["bg"] = "#444444"

        frame = tk.Frame(master=self.win2, width=780, height=880)
        frame.place(x=10, y=10)
        fig = plt.Figure(figsize=(7.8, 8.8), dpi=100)
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP)
        fig.subplots_adjust(bottom=0.05, left=0.12, right=0.94, top=0.97, hspace=0.32)

        fig.clf()
        canvas.draw()

        trcs = [self.trace, self.tc1, self.tc2]
        titles = ["Primary Tracking Field", "Colocalization Field 1", "Colocalization Field 2"]

        if tracking_win.dark_mode.get() == 1:
            bg_col = "#222222"
            box_col = "blue"
            text_col = "white"
            grid_col = "#333333"
        else:
            bg_col = "white"
            box_col = "black"
            text_col = "black"
            grid_col = "#dddddd"

        fig.set_facecolor(bg_col)
        for subplot in range(3):
            ax = fig.add_subplot(3, 1, subplot + 1)
            ax.set_facecolor(bg_col)
            ax.spines['bottom'].set_color(box_col)
            ax.spines['top'].set_color(box_col)
            ax.spines['left'].set_color(box_col)
            ax.spines['right'].set_color(box_col)
            ax.xaxis.label.set_color(box_col)
            ax.yaxis.label.set_color(box_col)
            ax.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            ax.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
            if tracking_win.plot_grids.get() == 1:
                ax.grid(color=grid_col)
            try:
                ax.plot(trcs[subplot], linewidth=2)
            except:
                print("Warning error occured: Full traceback shown below:")
                print(traceback.format_exc())
                """ No Coloc. Trace """
            ax.set_xlabel("Time (frames)", color=text_col, size=tracking_win.g_fonts["labels"])
            ax.set_ylabel("Intensity (A.U.)", color=text_col, size=tracking_win.g_fonts["labels"])
            ax.set_title(titles[subplot], size=tracking_win.g_fonts["title"], color=text_col)
        canvas.draw()

    def handle_close_2(self):
        try:
            self.win2.destroy()
        except:
            """ Failed """

    @staticmethod
    def insert_NaN(frames, series, mode="nan"):
        new_series = []
        for time in range(len(frames) - 1):
            diff = frames[time + 1] - frames[time]
            if diff == 1:
                new_series.append(series[time])
            else:
                new_series.append(series[time])
                for i in range(diff - 1):
                    if mode == "nan":
                        new_series.append(np.nan)
                    elif mode == "same":
                        if new_series[-1] != None:
                            new_series.append(new_series[-1])
                        else:
                            new_series.append(0)
        new_series.append(series[-1])
        return new_series

    @staticmethod
    def plot(series_x, series_y, xlabel, ylabel, title, fig, canv, dual=False, sets=[], type="plot", fit=False):

        if tracking_win.dark_mode.get() == 1:
            bg_col = "#222222"
            box_col = "blue"
            text_col = "white"
            grid_col = "#333333"
            line_col = "white"
        else:
            bg_col = "white"
            box_col = "black"
            text_col = "black"
            grid_col = "#dddddd"
            line_col = "black"

        ax = fig.add_subplot(111)
        ax.set_facecolor(bg_col)
        ax.spines['bottom'].set_color(box_col)
        ax.spines['top'].set_color(box_col)
        ax.spines['left'].set_color(box_col)
        ax.spines['right'].set_color(box_col)
        ax.xaxis.label.set_color(text_col)
        ax.yaxis.label.set_color(text_col)
        ax.tick_params(axis='x', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        ax.tick_params(axis='y', colors=text_col, labelsize=tracking_win.g_fonts["axis"])
        if type == "plot" and tracking_win.plot_grids.get() == 1:
            ax.grid(color=grid_col)
        ax.set_xlabel(xlabel, color=text_col, size=tracking_win.g_fonts["labels"])
        ax.set_ylabel(ylabel, color=text_col, size=tracking_win.g_fonts["labels"])
        ax.set_title(title, size=tracking_win.g_fonts["title"], color=text_col)
        if type == "plot":
            if not dual:
                ax.plot(series_x, series_y, marker="o", markersize=2)
            else:
                if not fit:
                    for set in range(len(sets)):
                        ax.plot(series_x, sets[set], marker="o", markersize=2)
                else:
                    ax.plot(series_x, sets[0], marker="o", markersize=2)
                    ax.plot(series_x, sets[1], linewidth=1, linestyle="--", color="red")
                    ax.plot(sets[2], sets[3], linewidth=1, linestyle="--", color=line_col)
        else:
            ax.hist(series_y, bins=10, rwidth=0.9)
        canv.draw()


class TrackingData:
    """ Data format for a single spot during tracking coordinates are in the form of a list where each index is
     one frame in the moview as the spot is tracked. Fitting parameters is a list where each index contains a list of
     parameters of the gaussian fit for the spot during the indexed frame which should correlate to the same index as
     the coordinates. Termination reason is a string describing why the spot tracking was terminated for this spot.
     start and end are the frames when tracking was started and ended for the spot respectively. """

    def __init__(self):
        self.frame_list = []
        self.coordinates = []
        self.fitting_params = []
        self.termination_reason = None
        self.start_frame = None
        self.end_frame = None
        self.include_export = True
        self.MSD_prop = 100


class TrajectorySeries:
    """ For storing the data series before plotting """

    def __init__(self):
        self.name = None
        self.frame_list = []
        self.coordinates = []
        self.sigma_x = []
        self.sigma_y = []
        self.amplitude = []
        self.residual = []
        self.eccentricity_x = []
        self.eccentricity_y = []
        self.displacement = []
        self.msd = []
        self.offset_time = []
        self.trace = []
        self.trace_coloc1 = []
        self.trace_coloc2 = []


class EnlargedView:
    def __init__(self, title, shape, function, callback):
        self.window = tk.Tk()
        self.window.title(title)
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)

        self.callback = callback
        self.canvas = tk.Canvas(master=self.window, bg="#222222", width=int(shape[1] * 1.75), height=int(shape[0] * 1.75))
        self.canvas.bind("<ButtonRelease-1>", self.callback)
        self.canvas.pack()
        self.function = function
        self.window.update()

    def handle_close(self):
        try:
            tracking_win.popout_state = [False, False, False]
        except:
            """ Failed """
        try:
            self.window.destroy()
        except:
            """ Failed """
        self.function()


class FontWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title("Font Sizes")
        self.window.geometry("240x150+500+700")
        self.window["bg"] = "#444444"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)

        tk.Label(master=self.window, text="Title Size", bg="#444444", fg="#cccccc").place(x=20, y=10, height=20)
        self.title = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.title["validatecommand"] = (self.title.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.title.insert(tk.END, tracking_win.g_fonts["title"])
        change_col_hover_enterbox(self.title, button_hv, "#333333")
        self.title.place(x=150, y=8, width=40, height=21)

        tk.Label(master=self.window, text="Axis Labels Size", bg="#444444", fg="#cccccc").place(x=20, y=40, height=20)
        self.labels = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.labels["validatecommand"] = (self.labels.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.labels.insert(tk.END, tracking_win.g_fonts["labels"])
        change_col_hover_enterbox(self.labels, button_hv, "#333333")
        self.labels.place(x=150, y=38, width=40, height=21)

        tk.Label(master=self.window, text="Axis Tick Marks Size", bg="#444444", fg="#cccccc").place(x=20, y=70, height=20)
        self.ticks = tk.Entry(master=self.window, bg="#333333", fg="white", validate="key")
        self.ticks["validatecommand"] = (self.ticks.register(tracking_win.validate_int_track), "%P", "%d", "%s")
        self.ticks.insert(tk.END, tracking_win.g_fonts["axis"])
        change_col_hover_enterbox(self.ticks, button_hv, "#333333")
        self.ticks.place(x=150, y=68, width=40, height=21)

        self.cancel = tk.Button(master=self.window, text="Cancel", padx=7, pady=2, bg="#662222", fg="white", command=self.handle_close)
        change_col_hover(self.cancel, "#993333", "#662222")
        self.cancel.place(x=10, y=110)

        self.done = tk.Button(master=self.window, text="Apply", padx=10, pady=2, bg="#333377", fg="white",
                              command=self.apply)
        change_col_hover(self.done, button_hv, "#333377")
        self.done.place(x=170, y=110)

        self.defaults = tk.Button(master=self.window, text="Default", padx=2, pady=2, bg="#222222", fg="white",
                              command=self.defaults)
        change_col_hover(self.defaults, button_hv, "#222222")
        self.defaults.place(x=80, y=110)

    def defaults(self):
        self.title.delete(0, tk.END)
        self.labels.delete(0, tk.END)
        self.ticks.delete(0, tk.END)
        self.title.insert(0, 11)
        self.labels.insert(0, 10)
        self.ticks.insert(0, 8)

    def apply(self):
        tracking_win.g_fonts = {
            "title": int(self.title.get()),
            "labels": int(self.labels.get()),
            "axis": int(self.ticks.get())
        }
        tracking_win.display_trajectory()
        self.handle_close()

    def handle_close(self):
        try:
            self.window.destroy()
        except:
            """ Failed """

def open_tracking_window():
    global tracking_win
    set_GUI_state(tk.DISABLED)
    load_button["state"] = tk.DISABLED
    load_pickle_button["state"] = tk.DISABLED
    preferences_button["state"] = tk.DISABLED
    import_raw_button["state"] = tk.DISABLED
    tracking_win = TrackingUI()


# end of trackXpress extension library --------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def open_time_series():
    global graph_window
    if tracking_win.current_trajectory_series.name is None:
        return
    try:
        graph_window.window.destroy()
    except:
        """ Window was closed """
    graph_window = GraphWin()


def restore_defaults():
    global preferences
    global default_dir
    preferences = {
        "Trace skip number": 10,
        "Default directory": "",
        "GUI colours": "#555555",
        "Subsampling": False,
        "kernel": "gauss",
        "pre-gauss": False,
        "Subtraction amount": 10,
        "488nm model dir": "100ms_300fr_Protein_high_SNR_model2",
        "561nm model dir": "100ms_300fr_Protein_low_SNR_model2",
        "640nm model dir": "100ms_300fr_Organic_high_SNR_model4",
        "Fitting tool window size": 6,
        "Fitting tool threshold": 3,
        "Fitting tool window stride": 2,
        "Fit convolve": True,
        "Fluorophore config": ["Cyanine 5", "mCherry", "mEGFP"],
        "Sum view frame fraction": [0.2, 0.2, 0.2],
        "Intensity target": [4, 3, 6],
        "Custom NeuralNet integration": False,
        "Custom NeuralNet settings": [],
        "Calibration optimizer settings": {
            "xmin": -256,
            "ymin": -256,
            "xmax": 640,
            "ymax": 640,
            "xstep": 96,
            "ystep": 96,
            "minscale": 80,
            "maxscale": 130,
            "scalestep": 20,
            "maxiter": 100,
        },
        "Restore Defaults": "",
    }
    default_dir = "N:/"
    save_preferences()


def show_info():
    easygui.msgbox(title="FluoroTensor Information", msg="Version: FluoroTensor " + version + " (" + update_date + ")" +
                   "\n\n" + "Notes:\n" + update_notes)


def toggle_custom_integration():
    choice = easygui.choicebox(msg="Enable or disable custom neural network integration.", title="Preferences",
                               choices=["Enable", "Disable"])
    if choice == "Enable":
        preferences["Custom NeuralNet integration"] = True
    elif choice == "Disable":
        preferences["Custom NeuralNet integration"] = False


def toggle_show_fit():
    global show_fit
    if show_fit:
        show_fit = False
        plot_trace(active_trace_list[current_trace])
        return
    elif not show_fit:
        show_fit = True
        plot_trace(active_trace_list[current_trace])


def custom_integration_settings():
    global custom_network_win
    custom_network_win = CustomNetWin()
    custom_network_win.window.mainloop()


def toggle_fret_mode():
    global fret_mode, current_trace, active_trace_list
    if not fret_mode:
        fret_mode = True
        easygui.msgbox(title="FRET mode activated", msg="FRET mode has been activated. Traces displayed in the main "
                       "graph (lower) are filtered to be the marker spot and the corresponding secondary FRET trace "
                       "will be shown above.")
        if Cy5_trace_count > 0:
            active_trace_list = []
            active_trace_list += Cy5_sublist
        current_trace = 0
        plot_trace(active_trace_list[current_trace])
        update_infobox()
        return
    else:
        fret_mode = False
        active_trace_list = []
        active_trace_list += all_sublist
        current_trace = 0
        plot_trace(active_trace_list[current_trace])
        update_infobox()


def open_fret_window():
    global fret_window
    if len(unique_ref_list) == 0:
        easygui.msgbox(title="Warning!", msg="No data to analyse!")
        return
    set_GUI_state(tk.DISABLED)
    fret_win = FRETWin()


def set_GUI_state(state):
    # Due to limitations in easygui instancing, disable all buttons to prevent crashes while an easygui
    # window is in use. function takes argument (state) which should be called as tk.DISABLED or tk.NORMAL
    global load_from_excel
    open_tracking_button["state"] = state
    if state == tk.DISABLED:
        export_button["state"] = state
        fitting_button["state"] = state
        ai_fit_button["state"] = state
        show_ai_fit_button["state"] = state
        current_fit_button["state"] = state
        edit_fit_button["state"] = state
        advanced_fit_button["state"] = state
    else:
        if ready_to_export:
            export_button["state"] = state
            fitting_button["state"] = state
            ai_fit_button["state"] = state
            show_ai_fit_button["state"] = state
            current_fit_button["state"] = state
            edit_fit_button["state"] = state
            advanced_fit_button["state"] = state
        if used_neural_network:
            fitting_button["state"] = state
            ai_fit_button["state"] = state

    if not is_data_loaded:
        state = tk.DISABLED
    load_button["state"] = state
    import_raw_button["state"] = state
    if load_from_excel:
        retrieve_button["state"] = state
    save_button["state"] = state
    filter_button["state"] = state
    correct_button["state"] = state
    preferences_button["state"] = state
    next_button["state"] = state
    back_button["state"] = state
    skip_forward_button["state"] = state
    skip_back_button["state"] = state
    jump_trace_button["state"] = state
    # subsample_button["state"] = state
    load_pickle_button["state"] = state
    display_stats_button["state"] = state
    display_stats_filtered_button["state"] = state
    # correlation_button["state"] = state
    integration_button["state"] = state
    # integ_scan_button["state"] = state
    kernel_button["state"] = state
    neural_network_button["state"] = state
    trim_button["state"] = state
    auto_trim_button["state"] = state
    fret_analysis_button["state"] = state
    if not is_data_loaded:
        preferences_button["state"] = tk.NORMAL
        load_button["state"] = tk.NORMAL
        load_pickle_button["state"] = tk.NORMAL
        import_raw_button["state"] = tk.NORMAL


def shutdown():
    print("FluoroTensor is shutting down...")
    try:
        root.destroy()
    except:
        """ Could not destroy root window """
    try:
        raw_gui.handle_close(forced=True)
    except:
        """ Failed to destroy window """
    try:
        tracking_win.handle_close(forced=True)
    except:
        """ Failed to destroy window """

    print("Terminating python interpreter, console will close shortly...")
    sys.exit()


preferences_function_map = {
    "Trace skip number": edit_trace_skip,
    "Default directory": set_default_dir,
    "GUI colours": set_gui_col,
    "Subsampling": set_subsample,
    "kernel": change_kernel,
    "pre-gauss": pregauss,
    "Subtraction amount": change_subtraction,
    "488nm model dir": pick_green_dir,
    "561nm model dir": pick_yellow_dir,
    "640nm model dir": pick_red_dir,
    "Fitting tool window size": change_fitting_window,
    "Fitting tool threshold": change_fitting_threshold,
    "Fitting tool window stride": change_fitting_stride,
    "Fit convolve": enable_fit_conv,
    "Fluorophore config": set_fluoro_names,
    "Sum view frame fraction": set_frame_fraction,
    "Intensity target": set_intensity_target,
    "Custom NeuralNet integration": toggle_custom_integration,
    "Custom NeuralNet settings": custom_integration_settings,
    "Calibration optimizer settings": calibration_optimizer_settings,
    "Restore Defaults": restore_defaults,
}

button_bg = "#222222"
button_fg = "#eeeeee"
button_hv = "#444499"


def change_col_hover(button, active_col, inactive_col):
    button.bind("<Enter>", func=lambda f: configure_hover_on(button, active_col, inactive_col))
    button.bind("<Leave>", func=lambda f: configure_hover_off(button, active_col, inactive_col))


def change_col_hover_enterbox(button, active_col, inactive_col):
    button.bind("<Enter>", func=lambda f: configure_hover_on_enter(button, active_col, inactive_col))
    button.bind("<Leave>", func=lambda f: configure_hover_off_enter(button, active_col, inactive_col))


def configure_hover_on(button, on_col, off_col):
    button_active = check_button_activity(button)
    if not button["relief"] == tk.SUNKEN:
        if button_active:
            button.config(background=on_col)
        else:
            button.config(background=off_col)


def configure_hover_off(button, on_col, off_col):
    button_active = check_button_activity(button)
    if not button["relief"] == tk.SUNKEN:
        if button_active:
            button.config(background=off_col)
        else:
            button.config(background=off_col)


def configure_hover_on_enter(button, on_col, off_col):
            button.config(background=on_col)


def configure_hover_off_enter(button, on_col, off_col):
            button.config(background=off_col)


def check_button_activity(button):
    if button["state"] == tk.NORMAL:
        return True
    else:
        return False


plot_space = tk.Frame(master=root, width=800, height=900)
plot_space_2 = tk.Frame(master=root, width=800, height=400)
gui_space = tk.Frame(master=root, width=600, height=900, bg=colours[3])
plot_space.grid(row=1, column=0, padx=5, pady=4)
gui_space.grid(row=0, column=1, rowspan=2, padx=5, pady=4)
plot_space_2.grid(row=0, column=0, padx=5, pady=4)

trace_fig2 = plt.Figure(figsize=(9, 2.5), dpi=100)
trace_canv2 = FigureCanvasTkAgg(trace_fig2, master=plot_space_2)
trace_canv2.draw()
trace_canv2.get_tk_widget().pack(side=tk.TOP)
trace_fig2.set_facecolor("#333333")

trace_figure = plt.Figure(figsize=(9, 6), dpi=100)
trace_canvas = FigureCanvasTkAgg(trace_figure, master=plot_space)
trace_canvas.draw()
trace_canvas.get_tk_widget().pack(side=tk.TOP)
trace_figure.set_facecolor("#333333")

toolbar = NavigationToolbar2Tk(trace_canvas, plot_space)
toolbar.config(background="#eeeeee")
for w in toolbar.winfo_children():
    w.config(background="#eeeeee")
toolbar._message_label.config(background="#eeeeee")
toolbar.update()
trace_canvas.get_tk_widget().pack()

trace_figure.subplots_adjust(bottom=0.1, left=0.1, right=0.98, top=0.95)
trace_fig2.subplots_adjust(bottom=0.12, left=0.1, right=0.98, top=0.88)

startup_message = "Program successfully started.\n\n\nThe following neural networks are assigned:\n\n" + \
                  "488nm predictor: " + str(preferences["488nm model dir"]) + \
                  "\n561nm predictor: " + str(preferences["561nm model dir"]) + \
                  "\n640nm predictor: " + str(preferences["640nm model dir"])


status = tk.Message(master=gui_space, text=startup_message, anchor=tk.NW, bg="#444444", padx=10,
                    pady=10, relief=tk.SUNKEN, width=550, bd=2, fg='#dddddd')
status.place(x=20, y=600, width=560, height=280)

current_info = tk.Message(master=gui_space, text="Current trace: " + str(current_trace) + "\nFluorophore: " +
                          str(current_fluorophore) + "\nStep count: " + str(current_step_count) +
                          "\nTrace length: " + str(current_trace_length), anchor=tk.NW, bg="#444444", padx=10, pady=10,
                          relief=tk.RIDGE, width=340, bd=2, fg="#dddddd")
current_info.place(x=230, y=10, width=360, height=200)

step_box = tk.Message(master=gui_space, text="", bg="#222222", fg="#dddddd", padx=5, pady=5, relief=tk.RIDGE, width=80,
                      bd=2)
step_box.config(font=('lucida console', 36, 'bold'))
step_box.place(x=488, y=12, width=100, height=60)

file_tab = tk.Message(master=gui_space, relief=tk.RIDGE, width=180, bg="#666666", fg="#cccccc", padx=10, pady=2, bd=2,
                      text="Import\n\n\n\n\n\n\nSerialisation (Save/Load)", anchor=tk.NW)
file_tab.place(x=10, y=10, width=200, height=200)

tracking_tab = tk.Message(master=gui_space, relief=tk.RIDGE, width=180, bg="#666666", fg="#cccccc", padx=10, pady=2, bd=2,
                          text="SM Tracking", anchor=tk.NW)
tracking_tab.place(x=10, y=220, width=200, height=80)

analysis_tab = tk.Message(master=gui_space, relief=tk.RIDGE, width=180, bg="#666666", fg="#cccccc", padx=10, pady=2, bd=2,
                          text="Export\n\n\nTrace Processing", anchor=tk.NW)
analysis_tab.place(x=10, y=310, width=200, height=210)

misc_tab = tk.Message(master=gui_space, text="Settings / Misc.", anchor=tk.NW, bg="#666666", padx=10, pady=2, bd=2,
                      relief=tk.RIDGE, width=340, fg="#cccccc")
misc_tab.place(x=230, y=220, width=360, height=80)

machine_tab = tk.Message(master=gui_space, text="AI Step Detection",
                         anchor=tk.NW, bg="#666666", padx=10, pady=2, bd=2,
                         relief=tk.RIDGE, width=160, fg="#cccccc")
machine_tab.place(x=230, y=310, width=180, height=210)

algorithm_tab = tk.Message(master=gui_space, text="Algorithmic Fitting",
                           anchor=tk.NW, bg="#666666", padx=10, pady=2, bd=2,
                           relief=tk.RIDGE, width=340, fg="#cccccc")
algorithm_tab.place(x=410, y=310, width=180, height=210)


open_tracking_button = tk.Button(master=gui_space, text="TrackXpress Add-in", padx=54, pady=12, bd=2,
                                 command=open_tracking_window, bg=button_bg, fg=button_fg)
change_col_hover(open_tracking_button, button_hv, button_bg)

import_raw_button = tk.Button(master=gui_space, text="Import Raw", padx=54, pady=1, bd=2,
                              command=create_raw_gui, bg=button_bg, fg=button_fg)
change_col_hover(import_raw_button, button_hv, button_bg)

load_button = tk.Button(master=gui_space, text="Import Spreadsheet", padx=33, pady=1, borderwidth=2,
                        command=create_workbook_object, bg=button_bg, fg=button_fg)
change_col_hover(load_button, button_hv, button_bg)

retrieve_button = tk.Button(master=gui_space, text="Retrieve Data", padx=50, pady=1, borderwidth=2,
                            command=create_metadata_list, bg=button_bg, fg=button_fg)
change_col_hover(retrieve_button, button_hv, button_bg)

save_button = tk.Button(master=gui_space, text="Save Data to File", padx=42, pady=2, borderwidth=2,
                        command=pickle_data, bg=button_bg, fg=button_fg)
change_col_hover(save_button, button_hv, button_bg)

export_button = tk.Button(master=gui_space, text="Export to Excel", padx=46, pady=2, borderwidth=2,
                          command=export_to_excel, bg=button_bg, fg=button_fg)
change_col_hover(export_button, button_hv, button_bg)

filter_button = tk.Button(master=gui_space, text="Filter Traces", padx=35, pady=12, borderwidth=2,
                          command=change_filter, bg=button_bg, fg=button_fg)
change_col_hover(filter_button, button_hv, button_bg)

correct_button = tk.Button(master=gui_space, text="Amend Step Count", padx=24, pady=4, borderwidth=2,
                           command=make_correction, bg=button_bg, fg=button_fg)
change_col_hover(correct_button, button_hv, button_bg)

ai_fit_button = tk.Button(master=gui_space, text="AI Trace Fit", padx=4, pady=4, borderwidth=2,
                          command=ai_fit_trace, bg=button_bg, fg=button_fg)
change_col_hover(ai_fit_button, button_hv, button_bg)

show_ai_fit_button = tk.Button(master=gui_space, text="Activations", padx=4, pady=4, borderwidth=2,
                               command=show_ai_activations, bg=button_bg, fg=button_fg)
change_col_hover(show_ai_fit_button, button_hv, button_bg)

preferences_button = tk.Button(master=gui_space, text="Preferences", padx=45, pady=12, borderwidth=2,
                               command=preference_update, bg="#333377", fg=button_fg)
change_col_hover(preferences_button, button_hv, "#333377")

next_button = tk.Button(master=gui_space, text="Next Trace -->", padx=16, pady=6, borderwidth=2,
                        command=next_trace, bg="#333377", fg=button_fg)
change_col_hover(next_button, button_hv, "#333377")

back_button = tk.Button(master=gui_space, text="<-- Previous Trace", padx=6, pady=6, borderwidth=2,
                        command=previous_trace, bg="#333377", fg=button_fg)
change_col_hover(back_button, button_hv, "#333377")

skip_forward_button = tk.Button(master=gui_space, text=str(preferences["Trace skip number"])+" -->>", padx=6, pady=6,
                                borderwidth=2, command=skip_next, bg=button_bg, fg=button_fg)
change_col_hover(skip_forward_button, button_hv, button_bg)

skip_back_button = tk.Button(master=gui_space, text="<<-- "+str(preferences["Trace skip number"]), padx=6, pady=6,
                             borderwidth=2, command=skip_back, bg=button_bg, fg=button_fg)
change_col_hover(skip_back_button, button_hv, button_bg)

jump_trace_button = tk.Button(master=gui_space, text="Jump to Trace", padx=6, pady=6, borderwidth=2,
                              command=jump_to_trace, bg="#bb9922", fg="black")
change_col_hover(jump_trace_button, "#ffcc33", "#bb9922")

# subsample_button = tk.Button(master=gui_space, text=" Enable Subsampling", padx=6, pady=6, borderwidth=2,
#                              command=toggle_subsampling, bg=button_bg, fg="black")
# change_col_hover(subsample_button, button_hv, button_bg)

load_pickle_button = tk.Button(master=gui_space, text="Load Data File", padx=48, pady=2, borderwidth=2,
                               command=unpickle_data, bg=button_bg, fg=button_fg)
change_col_hover(load_pickle_button, button_hv, button_bg)

display_stats_button = tk.Button(master=gui_space, text="Distributions", padx=10, pady=5, borderwidth=2,
                                 command=display_stats, bg="#ff7722", fg="black")
change_col_hover(display_stats_button, "#ffaa33", "#ff7722")

display_stats_filtered_button = tk.Button(master=gui_space, text="Filtered", padx=15, pady=5, borderwidth=2,
                                 command=display_stats_filtered, bg="#ff7722", fg="black")
change_col_hover(display_stats_filtered_button, "#ffaa33", "#ff7722")

# correlation_button = tk.Button(master=gui_space, text="Autocorrelation", padx=6, pady=6, borderwidth=2,
#                                command=auto_correlation)

integration_button = tk.Button(master=gui_space, text="Integration", padx=56, pady=2, borderwidth=2,
                               command=integration, bg=button_bg, fg=button_fg)
change_col_hover(integration_button, button_hv, button_bg)

integ_inc_button = tk.Button(master=gui_space, text="+", padx=11, pady=1, borderwidth=2,
                             command=integration_inc, bg=button_bg, fg=button_fg)
change_col_hover(integ_inc_button, "#007700", button_bg)

integ_dec_button = tk.Button(master=gui_space, text="-", padx=12, pady=1, borderwidth=2,
                             command=integration_dec, bg=button_bg, fg=button_fg)
change_col_hover(integ_dec_button, "#770000", button_bg)

integ_sub_reset_button = tk.Button(master=gui_space, text="Reset", padx=20, pady=1, borderwidth=2,
                                   command=integ_subtract_reset, bg=button_bg, fg=button_fg)
change_col_hover(integ_sub_reset_button, button_hv, button_bg)

# integ_scan_button = tk.Button(master=gui_space, text="Scan", padx=10, pady=6, borderwidth=2,
#                               command=integral_scan)

kernel_button = tk.Button(master=gui_space, text="Convolution", padx=52, pady=2, borderwidth=2,
                          command=convolution, bg=button_bg, fg=button_fg)
change_col_hover(kernel_button, button_hv, button_bg)

neural_network_button = tk.Button(master=gui_space, text="Detect Steps", padx=42, pady=4, borderwidth=2,
                                  command=neural_detection, bg="#773333", fg=button_fg)
change_col_hover(neural_network_button, "#aa4444", "#773333")

trim_button = tk.Button(master=gui_space, text="Trim Trace", padx=15, pady=1, borderwidth=2, command=trim_trace,
                        bg=button_bg, fg=button_fg)
change_col_hover(trim_button, "#770000", button_bg)

undo_trim_button = tk.Button(master=gui_space, text="Undo", padx=8, pady=1, borderwidth=2, command=undo_trim,
                             bg=button_bg, fg=button_fg)
change_col_hover(undo_trim_button, "#007700", button_bg)

auto_trim_button = tk.Button(master=gui_space, text="Smart Trim", padx=14, pady=1, borderwidth=2, command=auto_trim,
                             bg=button_bg, fg=button_fg)
change_col_hover(auto_trim_button, "#770000", button_bg)

undo_auto_trim_button = tk.Button(master=gui_space, text="Undo", padx=8, pady=1, borderwidth=2, command=undo_auto_trim,
                                  bg=button_bg, fg=button_fg)
change_col_hover(undo_auto_trim_button, "#007700", button_bg)

fitting_button = tk.Button(master=gui_space, text="Calculate Fits & Plateaus", padx=9, pady=2, borderwidth=2,
                           bg=button_bg, fg=button_fg, command=calculate_all_fits)
change_col_hover(fitting_button, button_hv, button_bg)

current_fit_button = tk.Button(master=gui_space, text="Recalculate Current Trace", padx=7, pady=2, borderwidth=2,
                               bg=button_bg, fg=button_fg, command=calculate_current_fit)
change_col_hover(current_fit_button, button_hv, button_bg)

edit_fit_button = tk.Button(master=gui_space, text="Fit Editing Mode", padx=30, pady=2, borderwidth=2,
                            bg=button_bg, fg=button_fg, command=toggle_editing_mode)
change_col_hover(edit_fit_button, button_hv, button_bg)

advanced_fit_button = tk.Button(master=gui_space, text="Advanced Fit", padx=40, pady=2, borderwidth=2,
                                bg=button_bg, fg=button_fg, command=select_advanced)
change_col_hover(advanced_fit_button, button_hv, button_bg)

snr_button = tk.Button(master=gui_space, text="Plot aSNR", padx=8, pady=2, borderwidth=2,
                       bg="#ff7722", fg="black", command=plot_snr)
change_col_hover(snr_button, "#ffaa33", "#ff7722")

use_ai_var = tk.IntVar(master=root)
use_ai_var.set(1)
use_ai_checkbutton = tk.Checkbutton(master=gui_space, bg="#666666", variable=use_ai_var, onvalue=1,
                                    offvalue=0, activebackground="#666666", text="AI Mode")


info_button = tk.Button(master=gui_space, text="Info", padx=10, pady=2, borderwidth=2,
                        bg=button_bg, fg=button_fg, command=show_info)
change_col_hover(info_button, button_hv, button_bg)


toggle_fit_button = ttk.Checkbutton(master=plot_space, text="Show Fit", command=toggle_show_fit)
fret_mode_button = ttk.Checkbutton(master=plot_space_2, text="FRET display", command=toggle_fret_mode)

toggle_fit_button.state(["!alternate"])
toggle_fit_button.state(["selected"])
toggle_fit_button.place(x=820, y=5)

fret_mode_button.state(["!alternate"])
fret_mode_button.state(["!selected"])
fret_mode_button.place(x=800, y=5)

fret_analysis_button = tk.Button(master=plot_space_2, text="Analyse FRET Traces", bg="#333377", padx=8, pady=0,
                                 fg="#cccccc", command=open_fret_window)
change_col_hover(fret_analysis_button, button_hv, "#333377")
fret_analysis_button.place(x=660, y=5, height=22)

about_button = tk.Button(master=plot_space_2, text="About", bg="#222222", padx=12, pady=0,
                         fg="#cccccc", command=view_licence)
change_col_hover(about_button, button_hv, "#222222")
about_button.place(x=586, y=5, height=22)


def add_to_unique():
    if is_data_loaded:
        entry = []
        entry.append(list(trace_info[active_trace_list[current_trace]]))
        entry.append(list(all_traces[active_trace_list[current_trace]]))
        entry.append(list(all_fits[active_trace_list[current_trace]]))
        unique_ref_list.append(entry)
        entry = []
        if fret_mode:
            try:
                entry.append(list(trace_info[active_trace_list[current_trace]+1]))
                entry.append(list(all_traces[active_trace_list[current_trace]+1]))
                entry.append(list(all_fits[active_trace_list[current_trace]+1]))
                unique_ref_list.append(entry)
            except:
                print(traceback.format_exc())
        print("added", active_trace_list[current_trace] + 1, "to persistent list")
        status["text"] = "Added current trace (" + str(active_trace_list[current_trace] + 1) + ") to persistent list. " + \
                         "List now contains " + str(len(unique_ref_list)) + " traces, fits, and information sets."
        status.update()


def add_all_in_filter_to_unique():
    if is_data_loaded:
        for index in active_trace_list:
            entry = []
            entry.append(list(trace_info[index]))
            entry.append(list(all_traces[index]))
            entry.append(list(all_fits[index]))
            unique_ref_list.append(entry)
        print("added all traces in current selection to list")
        status["text"] = "Added all traces in filtered selection to persistent list. Filter = " + str(fluoro_filter)
        status.update()


def clear_unique():
    global unique_ref_list
    print("cleared persistent list")
    unique_ref_list = []
    status["text"] = "Cleared persistent list of all traces, fits and information sets."
    status.update()


def remove_from_unique():
    if len(unique_ref_list) > 0:
        print("removed most recently added trace from persistent list")
        unique_ref_list.pop(-1)
        status["text"] = "removed most recently added trace from persistent list"
        status.update()


def save_unique():
    if len(unique_ref_list) > 0:
        inf_save = []
        tr_save = []
        fit_save = []
        for index in range(len(unique_ref_list)):
            inf_save.append(unique_ref_list[index][0])
            tr_save.append(unique_ref_list[index][1])
            fit_save.append(unique_ref_list[index][2])
        path = easygui.filesavebox(msg="Save unique list of 1 step traces", default=default_dir)
        if path:
            ext = path[-5:]
            if ext == ".trcs":
                path = path[:-5]
            with open(path + ".trcs", "wb") as save:
                pickle.dump(inf_save, save)
                pickle.dump(tr_save, save)
                pickle.dump(fit_save, save)
            print("saved unique list")
            status["text"] = "Successfully saved persistent list to '" + str(path) + "'"
            status.update()


def plot_unique():
    if len(unique_ref_list) < 1:
        return

    number = len(unique_ref_list)
    rows_and_columns = int(np.sqrt(number)) + 1
    if np.sqrt(number) == int(np.sqrt(number)):
        rows_and_columns = int(np.sqrt(number))

    trace_figure.clf()
    trace_canvas.draw()
    for subplot in range(len(unique_ref_list)):
        plot_area = trace_figure.add_subplot(rows_and_columns, rows_and_columns, subplot + 1)
        plot_area.set_facecolor("#222222")
        plot_area.spines['bottom'].set_color('blue')
        plot_area.spines['top'].set_color('blue')
        plot_area.spines['left'].set_color('blue')
        plot_area.spines['right'].set_color('blue')
        plot_area.xaxis.label.set_color('white')
        plot_area.yaxis.label.set_color('white')
        plot_area.tick_params(axis='x', colors='white', labelsize=int(10/np.sqrt(rows_and_columns)))
        plot_area.tick_params(axis='y', colors='white', labelsize=int(10/np.sqrt(rows_and_columns)))
        plot_area.grid(color="#333333")
        plot_area.plot(unique_ref_list[subplot][1], linewidth=1)
        plot_area.plot(unique_ref_list[subplot][2], linewidth=1)
    trace_canvas.draw()
    status["text"] = "Plotted all traces in persistent list on the graph pane. Press next trace to go back to standard trace view"
    status.update()


add_to_unique_save_button = ttk.Button(master=plot_space, text="Add Trace", command=add_to_unique)
clear_unique_button = ttk.Button(master=plot_space, text="Clear List", command=clear_unique)
remove_from_save_unique_button = ttk.Button(master=plot_space, text="Remove Last", command=remove_from_unique)
add_all_to_unique_button = ttk.Button(master=plot_space, text="Add All in Filter", command=add_all_in_filter_to_unique)
save_unique_button = ttk.Button(master=plot_space, text="Save", command=save_unique)
plot_unique_button = ttk.Button(master=plot_space, text="Plot", command=plot_unique)


import_raw_button.place(x=20, y=30)
open_tracking_button.place(x=20, y=240, width=180)
load_button.place(x=20, y=60)
retrieve_button["state"] = tk.DISABLED
retrieve_button.place(x=20, y=90)
save_button["state"] = tk.DISABLED
save_button.place(x=20, y=136)
load_pickle_button.place(x=20, y=168)
export_button["state"] = tk.DISABLED
export_button.place(x=20, y=330)
filter_button["state"] = tk.DISABLED
filter_button.place(x=410, y=240)
correct_button["state"] = tk.DISABLED
correct_button.place(x=240, y=370)
ai_fit_button["state"] = tk.DISABLED
ai_fit_button.place(x=240, y=410)
show_ai_fit_button["state"] = tk.DISABLED
show_ai_fit_button.place(x=321, y=410)
preferences_button["state"] = tk.NORMAL
preferences_button.place(x=240, y=240)
back_button["state"] = tk.DISABLED
back_button.place(x=102, y=540)
next_button["state"] = tk.DISABLED
next_button.place(x=230, y=540)
skip_forward_button["state"] = tk.DISABLED
skip_forward_button.place(x=365, y=540)
skip_back_button["state"] = tk.DISABLED
skip_back_button.place(x=20, y=540)
jump_trace_button["state"] = tk.DISABLED
jump_trace_button.place(x=486, y=540)
# subsample_button["state"] = tk.DISABLED
# subsample_button.place(x=388, y=400)
display_stats_button["state"] = tk.DISABLED
display_stats_button.place(x=20, y=480)
display_stats_filtered_button["state"] = tk.DISABLED
display_stats_filtered_button.place(x=122, y=480)
# correlation_button["state"] = tk.DISABLED
# correlation_button.place(x=260, y=400)
kernel_button["state"] = tk.DISABLED
kernel_button.place(x=20, y=375, width=180)
integration_button["state"] = tk.DISABLED
integration_button.place(x=20, y=410, width=180)
integ_inc_button["state"] = tk.DISABLED
integ_dec_button["state"] = tk.DISABLED
integ_sub_reset_button["state"] = tk.DISABLED
integ_inc_button.place(x=161, y=446)
integ_dec_button.place(x=116, y=446)
integ_sub_reset_button.place(x=20, y=446)
# integ_scan_button["state"] = tk.DISABLED
# integ_scan_button.place(x=20, y=450)
neural_network_button["state"] = tk.DISABLED
neural_network_button.place(x=240, y=330)
trim_button["state"] = tk.DISABLED
trim_button.place(x=240, y=480)
undo_trim_button["state"] = tk.DISABLED
undo_trim_button.place(x=344, y=480)
auto_trim_button["state"] = tk.DISABLED
auto_trim_button.place(x=240, y=448)
undo_auto_trim_button["state"] = tk.DISABLED
undo_auto_trim_button.place(x=344, y=448)
fitting_button["state"] = tk.DISABLED
current_fit_button["state"] = tk.DISABLED
edit_fit_button["state"] = tk.DISABLED
fitting_button.place(x=420, y=330)
current_fit_button.place(x=420, y=366)
edit_fit_button.place(x=420, y=402)
advanced_fit_button["state"] = tk.DISABLED
advanced_fit_button.place(x=420, y=438)
snr_button.place(x=500, y=474)
use_ai_checkbutton.place(x=418, y=475)
info_button.place(x=528, y=602)

add_to_unique_save_button.place(x=260, y=610)
remove_from_save_unique_button.place(x=340, y=610)
add_all_to_unique_button.place(x=421, y=610)
clear_unique_button.place(x=540, y=610)
save_unique_button.place(x=620, y=610)
plot_unique_button.place(x=700, y=610)

gui_space["bg"] = preferences["GUI colours"]
gui_space.update()

# if preferences["Subsampling"]:
#     subsample_button["relief"] = tk.SUNKEN
#     subsample_button["bg"] = "#22ff55"
#     subsample_button["text"] = "Disable Subsampling"
#
#
# if not preferences["Subsampling"]:
#     subsample_button["relief"] = tk.RAISED
#     subsample_button["bg"] = "#eeeeee"
#     subsample_button["text"] = " Enable Subsampling"

root.protocol("WM_DELETE_WINDOW", shutdown)
load_calibration()
root.mainloop()
root.mainloop()
root.mainloop()
